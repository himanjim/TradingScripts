from openpyxl import Workbook

def print_statistics(moving_average_strategy_responses, file, ma_strategy_excel_location):
    no_of_tradable_matches = 0
    no_of_sma_greater_than_lma = 0
    no_of_correct_rsi = 0
    no_of_correct_trend = 0
    no_of_mas_diverging = 0
    no_of_stock_price_appropriately_placed_between_yearly_highs_lows = 0
    # no_of_macd_greater_than_0 = 0
    # no_of_macd_greater_than_9_day_ema = 0
    no_of_stock_price_greater_than_mas = 0
    no_of_lma_high_slope = 0
    no_of_macd_high_slope = 0
    no_of_high_volumes = 0

    wb = Workbook ()

    # grab the active worksheet
    ws = wb.active

    ws['B2'] = 'STOCK'
    ws['C2'] = 'STRATEGY'
    ws['D2'] = 'TRADABLE'
    ws['E2'] = 'SMA>LMA'
    ws['F2'] = 'CORRECT RSI'
    ws['G2'] = 'CORRECT TREND'
    ws['H2'] = 'MAS DIVERGING'
    ws['I2'] = 'STOCK<>YEARLY HIGHS LOWS'
    ws['J2'] = 'STOCK>MAS'
    ws['K2'] = 'LMA SLOPE'
    ws['L2'] = 'HIGH VOLS'
    ws['M2'] = 'DAYS BACK'
    ws['N2'] = 'STOCK PRICE'
    ws['O2'] = 'RSI'

    row_count = 3
    for moving_average_strategy_response in moving_average_strategy_responses:

        ws['B' + str (row_count)] = moving_average_strategy_response.stock_id
        ws['C' + str (row_count)] = str (moving_average_strategy_response.ma_strategy_name.name)
        ws['M' + str (row_count)] = moving_average_strategy_response.days_back_when_stock_price_less_than_sma
        ws['N' + str (row_count)] = moving_average_strategy_response.current_day_current_price
        ws['O' + str (row_count)] = moving_average_strategy_response.rsi

        if moving_average_strategy_response.is_strategy_tradable():
            no_of_tradable_matches += 1
            ws['D' + str (row_count)] = 1
        else:
            ws['D' + str (row_count)] = 0

        if moving_average_strategy_response.sma_greater_than_lma :
            no_of_sma_greater_than_lma += 1
            ws['E' + str (row_count)] = 1
        else:
            ws['E' + str (row_count)] = 0

        if moving_average_strategy_response.correct_rsi is None or moving_average_strategy_response.correct_rsi :
            no_of_correct_rsi += 1
            ws['F' + str (row_count)] = 1
        else:
            ws['F' + str (row_count)] = 0

        if moving_average_strategy_response.correct_trend:
            no_of_correct_trend += 1
            ws['G' + str (row_count)] = 1
        else:
            ws['G' + str (row_count)] = 0

        if moving_average_strategy_response.mas_diverging is None or moving_average_strategy_response.mas_diverging :
            no_of_mas_diverging += 1
            ws['H' + str (row_count)] = 1
        else:
            ws['H' + str (row_count)] = 0


        if moving_average_strategy_response.stock_price_appropriately_placed_between_yearly_highs_lows is None or moving_average_strategy_response.stock_price_appropriately_placed_between_yearly_highs_lows :
            no_of_stock_price_appropriately_placed_between_yearly_highs_lows += 1
            ws['I' + str (row_count)] = 1
        else:
            ws['I' + str (row_count)] = 0

        # if  moving_average_strategy_response.macd[0]['MACD'][
        #             0] > 0:
        #     no_of_macd_greater_than_0 += 1
        #
        # if  moving_average_strategy_response.macd_greater_than_9_day_ema==True:
        #     no_of_macd_greater_than_9_day_ema += 1

        if moving_average_strategy_response.stock_price_greater_than_mas :
            no_of_stock_price_greater_than_mas += 1
            ws['J' + str (row_count)] = 1
        else:
            ws['J' + str (row_count)] = 0

        if moving_average_strategy_response.lma_high_slope  or moving_average_strategy_response.lma_high_slope is None:
            no_of_lma_high_slope += 1
            ws['K' + str (row_count)] = 1
        else:
            ws['K' + str (row_count)] = 0

        # if moving_average_strategy_response.macd_high_slope :
        #     no_of_macd_high_slope += 1

        if moving_average_strategy_response.high_volumes :
            no_of_high_volumes += 1
            ws['L' + str (row_count)] = 1
        else:
            ws['L' + str (row_count)] = 0

        row_count += 1

    wb.save (ma_strategy_excel_location)

    file.write('\n\nStats start @#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$')
    file.write('\nNo of strategies tested:'+str(len(moving_average_strategy_responses)))
    file.write('\nNo of tradable matches:'+str(no_of_tradable_matches))
    file.write('\nNo_of_sma_greater_than_lma:'+str(no_of_sma_greater_than_lma))
    file.write('\nNo_of_correct_rsi:'+str(no_of_correct_rsi))
    file.write('\nNo_of_correct_trend:'+str(no_of_correct_trend))
    file.write('\nNo_of_mas_diverging:'+str(no_of_mas_diverging))
    file.write('\nNo_of_stock_price_appropriately_placed_between_yearly_highs_lows:'+str(no_of_stock_price_appropriately_placed_between_yearly_highs_lows))
    # file.write('No_of_macd_greater_than_0:'+str(no_of_macd_greater_than_0))
    # file.write('No_of_macd_greater_than_9_day_ema:'+str(no_of_macd_greater_than_9_day_ema))
    file.write('\nNo_of_stock_price_greater_than_mas:'+str(no_of_stock_price_greater_than_mas))
    file.write('\nNo_of_lma_high_slope:'+str(no_of_lma_high_slope))
    # file.write('No_of_macd_high_slope:'+str(no_of_macd_high_slope))
    file.write('\nNo_of_high_volumes:'+str(no_of_high_volumes))
    file.write ('\nStats end @#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$\n\n')


