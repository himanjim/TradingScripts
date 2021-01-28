from openpyxl import Workbook

import ScrapUtils as sutil


def print_statistics(stocks_pattern_recognition_responses, market_previous_trend, file,
                     correct_cs_pattern_excel_location, nifty_50_stocks, lots, margins):
    no_of_correct_patterns = 0
    no_of_tradable_patterns = 0
    no_of_strong_correct_trend = 0
    no_of_weak_correct_trend = 0
    no_of_high_volumes = 0
    no_of_correct_rsi = 0
    no_of_correct_rsi_14_9_period_sma = 0
    no_of_correct_resistance = 0
    no_of_correct_support = 0
    no_of_correct_risk_reward_ratio = 0
    no_of_trend_same_as_market = 0
    no_of_correct_candle_length = 0

    wb = Workbook ()

    # grab the active worksheet
    ws = wb.active

    ws['B2'] = 'STOCK'
    ws['C2'] = 'PATTERN'
    ws['D2'] = 'TRADABLE'
    ws['E2'] = 'PATTERN MATCH'
    ws['F2'] = 'CORRECT TREND'
    ws['G2'] = 'HIGH VOLUME'
    ws['H2'] = 'CORRECT RSI'
    ws['I2'] = 'LOT'
    ws['J2'] = 'MARGIN'
    ws['K2'] = 'SL'
    ws['L2'] = 'CORRECT RISK REWARD RATIO'
    ws['M2'] = 'POINTS'
    ws['N2'] = 'RISK REWARD RATIO'
    ws['O2'] = 'SUPPORT'
    ws['P2'] = 'RESISTANCE'
    ws['Q2'] = 'VOL'
    ws['R2'] = 'AV.VOL'
    ws['S2'] = 'CANDLE LENGTH OK'
    ws['T2'] = 'VOLAT SL'
    ws['U2'] = 'NIFTY50'
    ws['V2'] = 'HIGH VOL'

    row_count=3
    for stocks_pattern_recognition_response in stocks_pattern_recognition_responses:

        if stocks_pattern_recognition_response.pattern_match:
            no_of_correct_patterns += 1
            ws['E' + str (row_count)] = 1

            ws['B'+str(row_count)] = stocks_pattern_recognition_response.stock_id
            ws['C'+str(row_count)] = str(stocks_pattern_recognition_response.pattern_name.name)
            ws['M'+str(row_count)] = stocks_pattern_recognition_response.points
            ws['N'+str(row_count)] = stocks_pattern_recognition_response.risk_reward_ratio
            ws['O'+str(row_count)] = stocks_pattern_recognition_response.support
            ws['P'+str(row_count)] = stocks_pattern_recognition_response.resistance
            ws['Q' + str (row_count)] = stocks_pattern_recognition_response.current_day_volumes
            ws['R' + str (row_count)] = stocks_pattern_recognition_response.last_10_day_average_volumes
            ws['T' + str (row_count)] = str (stocks_pattern_recognition_response.volatility_stop_loss)

            is_nifty_stock = False
            for stock in nifty_50_stocks:
                if stocks_pattern_recognition_response.stock_id == stock[sutil.STOCK_ID]:
                    is_nifty_stock = True
                    break

            ws['U' + str (row_count)] = [0, 1][is_nifty_stock]
            ws['V' + str (row_count)] = [0, 1][
                stocks_pattern_recognition_response.current_day_volumes > stocks_pattern_recognition_response.last_10_day_average_volumes]

            if stocks_pattern_recognition_response.is_pattern_tradable():
                no_of_tradable_patterns += 1
                ws['D' + str (row_count)] = 1
            else:
                ws['D' + str (row_count)] = 0

            if stocks_pattern_recognition_response.strong_correct_trend :
                no_of_strong_correct_trend += 1

            if stocks_pattern_recognition_response.weak_correct_trend :
                no_of_weak_correct_trend += 1

            if stocks_pattern_recognition_response.previous_trend is None or stocks_pattern_recognition_response.strong_correct_trend or stocks_pattern_recognition_response.weak_correct_trend:
                ws['F' + str (row_count)] = 1
            else:
                ws['F' + str (row_count)] = 0

            if stocks_pattern_recognition_response.high_volumes :
                no_of_high_volumes += 1
                ws['G' + str (row_count)] = 1
            else:
                ws['G' + str (row_count)] = 0

            if stocks_pattern_recognition_response.correct_rsi :
                no_of_correct_rsi += 1
                ws['H' + str (row_count)] = 1
            else:
                ws['H' + str (row_count)] = 0

            # if stocks_pattern_recognition_response.correct_rsi_14_9_period_SMA :
            #     no_of_correct_rsi_14_9_period_sma += 1
            #     ws['I' + str (row_count)] = 1
            # else:
            #     ws['I' + str (row_count)] = 0

            ws['I' + str (row_count)] = lots[stocks_pattern_recognition_response.stock_id]

            if stocks_pattern_recognition_response.correct_resistance:
                no_of_correct_resistance += 1

            if stocks_pattern_recognition_response.correct_support:
                no_of_correct_support += 1

            # if stocks_pattern_recognition_response.correct_resistance or stocks_pattern_recognition_response.correct_support:
            #     ws['J' + str (row_count)] = 1
            # else:
            #     ws['J' + str (row_count)] = 0

            ws['J' + str (row_count)] = margins[stocks_pattern_recognition_response.stock_id]

            # if stocks_pattern_recognition_response.pattern_trend_same_as_market_trend:
            #     no_of_trend_same_as_market += 1
            #     ws['K' + str (row_count)] = 1
            # else:
            #     ws['K' + str (row_count)] = 0

            last_day_price = stocks_pattern_recognition_response.fetched_dataset[-1]

            ws['K' + str (row_count)] = lots[stocks_pattern_recognition_response.stock_id] * \
                                        [last_day_price['high'] - last_day_price['close'],
                                         last_day_price['close'] - last_day_price['low']][
                                            stocks_pattern_recognition_response.action.value == 1]

            if stocks_pattern_recognition_response.correct_risk_reward_ratio :
                no_of_correct_risk_reward_ratio += 1
                ws['L' + str (row_count)] = 1
            else:
                ws['L' + str (row_count)] = 0

            if stocks_pattern_recognition_response.correct_candle_length :
                no_of_correct_candle_length += 1
                ws['S' + str (row_count)] = 1
            else:
                ws['S' + str (row_count)] = 0

            row_count += 1

    wb.save (correct_cs_pattern_excel_location)

    file.write('\nStats start @#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$')
    file.write('\nNo of patterns tested:'+str(len(stocks_pattern_recognition_responses)))
    file.write('\nNo of tradable patterns:'+str(no_of_tradable_patterns))
    file.write('\nNo of correct patterns:'+str(no_of_correct_patterns))
    file.write('\nNo of strong correct trend:'+str(no_of_strong_correct_trend))
    file.write('\nNo of weak correct trend:'+str(no_of_weak_correct_trend))
    file.write('\nNo of high volumes:'+str(no_of_high_volumes))
    file.write('\nNo of correct rsi:'+str(no_of_correct_rsi))
    file.write('\nNo of correct rsi 14_9_period_sma:'+str(no_of_correct_rsi_14_9_period_sma))
    file.write('\nNo of correct resistance:'+str(no_of_correct_resistance))
    file.write('\nNo of correct support:'+str(no_of_correct_support))
    file.write('\nNo of trend same as market:'+str(no_of_trend_same_as_market))
    file.write('\nNo of correct risk reward ratio:'+str(no_of_correct_risk_reward_ratio))
    file.write ('\nNo of correct candle length:' + str (no_of_correct_candle_length))
    file.write('\nStats end @#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$\n')


