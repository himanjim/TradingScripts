"""
ATM short-straddle backtester + robustness optimizer (NIFTY / SENSEX, DTE=0/1).
================================================================================

This file does two things, selected by the RUN_MODE setting in the RUN CONTROL
block at the top (no command line -- just edit and press Run in PyCharm):

  1. SINGLE BACKTEST   -> RUN_MODE = "backtest"
     Runs one backtest using the parameters/env-vars below and writes a detailed
     multi-sheet Excel report. Identical to the original workflow.

  2. ROBUSTNESS OPTIMIZER -> RUN_MODE = "optimize"
     Searches the six tunables with Optuna (TPE), maximizing ROBUSTNESS = the
     fraction of profitable MONTHS first, then profitable days. Loads the option
     data ONCE and re-simulates per trial, printing progress every few trials.

Optimized parameters:
     ENTRY_TIME_IST, LOSS_LIMIT_RUPEES_BY_ATTEMPT, PROFIT_PROTECT_TRIGGER_RUPEES,
     MAX_REATTEMPTS, PROFIT_TARGET_PCT, REENTRY_DELAY_BY_ATTEMPT

--------------------------------------------------------------------------------
QUICK SMALL-SAMPLE RUN (do this first to confirm everything works end-to-end):
--------------------------------------------------------------------------------
In the RUN CONTROL block at the top, set:
     RUN_MODE          = "optimize"
     OPT_TRIALS        = 5
     SAMPLE_MAX_PICKLES = 3      # read only 3 pickle files (also shrinks the
                                 # Kite underlying download)
     SAMPLE_MAX_DAYS    = 20     # simulate only the 20 most recent day-groups
     OPT_PROGRESS_EVERY = 1      # print after every trial
...then press Run. It should finish in a minute or two and print load progress,
all five trial lines, and a BEST CONFIG block.

For the real search, set:
     OPT_TRIALS = 300
     OPT_CV_FOLDS = 5
     SAMPLE_MAX_PICKLES = None
     SAMPLE_MAX_DAYS = None

Dependencies: pandas, openpyxl, optuna  (pip install optuna).
"""

import os
from pathlib import Path
import glob
import time
from dataclasses import dataclass
from datetime import datetime, date, time as dtime, timedelta
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd

import Trading_2024.OptionTradeUtils as oUtils

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore

try:
    import pytz  # type: ignore
except Exception:
    pytz = None  # type: ignore

try:
    from dateutil.relativedelta import relativedelta  # type: ignore
except Exception:
    relativedelta = None  # type: ignore


# =============================================================================
# RUN CONTROL  -- EDIT THESE, THEN JUST RUN THE FILE (no command line needed)
# =============================================================================
# What to do when you run this file:
#   "backtest" -> one backtest with the params below + Excel report (original flow)
#   "optimize" -> Optuna robustness search over the six tunables
RUN_MODE = "optimize"

# --- Optimizer settings (used only when RUN_MODE == "optimize") ---
OPT_TRIALS = 50            # number of optimization trials
OPT_CV_FOLDS = 5            # 1 = score on full sample; >1 = walk-forward block robustness
OPT_PROGRESS_EVERY = 5      # (retained for compatibility) per-trial stats now print EVERY trial
OPT_SEED = 42              # RNG seed for reproducible searches

# --- Small-sample smoke test (set BOTH to None for a real run) ---
# For a quick end-to-end check, e.g. SAMPLE_MAX_PICKLES = 3 and SAMPLE_MAX_DAYS = 20.
# SAMPLE_MAX_PICKLES also shrinks the Kite underlying download, so the smoke test
# stays fast without touching anything else.
SAMPLE_MAX_PICKLES = None   # e.g. 3  -> read only the first N pickle files
SAMPLE_MAX_DAYS = None      # e.g. 20 -> simulate only the most recent N day-groups

# (Objective weights / guards live with the optimizer further down: OPT_W_MONTH,
#  OPT_W_DAY, OPT_W_PNL, OPT_CV_PENALTY, OPT_MIN_DAYS, OPT_MIN_MONTHS.)

# =============================================================================
# USER CONFIG
# =============================================================================
# PICKLES_DIR = r"G:\My Drive\Trading\Dhan_Historical_Options_Data_New"
PICKLES_DIR = r"G:\My Drive\Trading\Historical_Options_Data"
ENTRY_TIME_IST = os.getenv("ENTRY_TIME_IST", "10:00")  # "HH:MM"

def _safe_fname_part(s: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in s)

def _get_downloads_folder() -> str:
    """
    Returns the current user's default Downloads folder.
    Falls back to home directory if Downloads is not found.
    """
    downloads = Path.home() / "Downloads"
    return str(downloads if downloads.exists() else Path.home())

# --- Generic integer-list parser used for DTE and re-entry delay settings ---
def _parse_int_list(env_val, default_list):
    """
    Parse a comma-separated integer list from an environment variable.

    Examples:
        ALLOWED_DTE="0,1"              -> [0, 1]
        REENTRY_DELAY_BY_ATTEMPT="1,5" -> [1, 5]

    If parsing fails or the env var is blank, the supplied default is used.
    """
    if env_val:
        try:
            vals = [int(round(float(x))) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass
    return list(default_list)


def _parse_pct_value(x) -> float:
    """
    Parse a percentage value into decimal form.

    Accepted user formats:
        10      -> 0.10  (10%)
        "10%"   -> 0.10  (10%)
        0.10    -> 0.10  (10%)
        "0.10"  -> 0.10  (10%)

    This lets you configure percentages naturally from environment variables
    while keeping calculations internally consistent.
    """
    s = str(x).strip().replace("%", "")
    if s == "":
        raise ValueError("blank percentage value")

    v = float(s)

    # If user entered 10, treat it as 10%; if user entered 0.10, keep it as 10%.
    if abs(v) > 1.0:
        v = v / 100.0

    if v < 0:
        raise ValueError("percentage cannot be negative")

    return float(v)


def _parse_pct_list(env_val, default_list):
    """
    Parse comma-separated percentages into decimal form.

    Examples:
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15" -> [0.10, 0.12, 0.15]
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10"     -> [0.10]
    """
    if env_val:
        try:
            vals = [_parse_pct_value(x) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass

    return [_parse_pct_value(x) for x in default_list]


def _fmt_int_list(lst) -> str:
    """Format integer-list settings for the output filename."""
    return "-".join(str(int(v)) for v in lst) if lst else "off"


def _fmt_pct_value(v: float) -> str:
    """Format a decimal percentage such as 0.10 as '10pct' for filenames/logs."""
    return f"{v * 100:.2f}".rstrip("0").rstrip(".") + "pct"


def _fmt_pct_list(lst) -> str:
    """Format a list of decimal percentages for the output filename."""
    return "-".join(_fmt_pct_value(float(v)) for v in lst) if lst else "off"


def _parse_float_env(env_name: str, default_value: float) -> float:
    """
    Parse a positive/zero floating-point rupee setting from an environment variable.

    If parsing fails, the supplied default is used. A value <= 0 is treated by
    the relevant logic as disabled where applicable.
    """
    raw = os.getenv(env_name)
    if raw is None or str(raw).strip() == "":
        return float(default_value)
    try:
        return float(str(raw).replace(",", "").strip())
    except Exception:
        return float(default_value)


def _fmt_rupee_value(v: float) -> str:
    """Format rupee values compactly for output filenames/logs."""
    return str(int(round(float(v))))


# =============================================================================
# RISK CONFIGURATION AS % OF ENTRY PREMIUM
# =============================================================================
# IMPORTANT:
# The original version used absolute rupee values in these two variables.
# This version intentionally keeps the same variable names for compatibility
# with your existing env-var workflow, but the meaning is now PERCENTAGE.
#
# Base for percentage calculation:
#     entry_premium_sum_rupees = (entry_ce + entry_pe) * qty
#
# Stop-loss threshold:
#     loss_limit_rupees = LOSS_LIMIT_% * entry_premium_sum_rupees
#
# Profit-protect threshold/giveback:
#     G = PROFIT_PROTECT_% * entry_premium_sum_rupees
#
# Example:
#     If CE+PE premium collected = 120 and qty = 325,
#     entry_premium_sum_rupees = 39,000.
#     10% stop-loss = 3,900.
#     30% profit-protect trigger/giveback = 11,700.
# =============================================================================

# --- Per-attempt STOP-LOSS as % of premium collected on that attempt ---
# Index 0 = first entry, 1 = first re-entry, etc.
# Attempts beyond the list reuse the LAST value.
#
# Default: 10% for every attempt.
#
# Env examples:
#     LOSS_LIMIT_RUPEES_BY_ATTEMPT="10"
#     LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15"
#     LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10,0.12,0.15"
LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_pct_list(
    os.getenv("LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
    [10, 10, 10, 10, 10, 10, 10, 10, 10, 10],
)


def loss_limit_pct_for_attempt(attempt_idx: int) -> float:
    """Return the stop-loss percentage, in decimal form, for the given attempt."""
    s = LOSS_LIMIT_RUPEES_BY_ATTEMPT
    if not s:
        return 0.0
    return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])


# --- Allowed days-to-expiry to trade: [0,1]=expiry day + day before; [0]=expiry only ---
ALLOWED_DTE = _parse_int_list(os.getenv("ALLOWED_DTE"), [0])

# --- Profit-protect threshold/giveback as % of premium collected on that attempt ---
# Default: 30%.
#
# Env examples:
#     PROFIT_PROTECT_TRIGGER_RUPEES="30"
#     PROFIT_PROTECT_TRIGGER_RUPEES="0.30"
#
# Current logic uses the same rupee amount for:
#     1. arming profit-protect once peak P&L reaches G
#     2. exiting when current P&L falls to peak - G
PROFIT_PROTECT_TRIGGER_RUPEES = _parse_pct_value(os.getenv("PROFIT_PROTECT_TRIGGER_RUPEES", "30"))

# --- Absolute daily circuit breaker -------------------------------------------------
# Once cumulative realized NET P&L for the current underlying/day reaches this
# loss, no further re-entry is allowed for that day.
#
# Default: Rs. 30,000 loss. Set MAX_DAILY_LOSS_RUPEES=0 to disable.
MAX_DAILY_LOSS_RUPEES = _parse_float_env("MAX_DAILY_LOSS_RUPEES", 30000.0)

# --- Absolute cap on the percentage-based per-attempt stop-loss ----------------------
# The stop-loss is still calculated as:
#     LOSS_LIMIT_% * entry_premium_sum
# But it is capped at this absolute rupee value.
#
# Effective stop-loss per attempt:
#     min(LOSS_LIMIT_% * entry_premium_sum, MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
#
# Default: Rs. 3,000 loss per attempt. Set to 0 to disable the cap.
MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_float_env("MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT", 3000.0)

MAX_REATTEMPTS = int(os.getenv("MAX_REATTEMPTS", "10"))  # 1 = only one re-entry

# --- Per-DAY profit target as a fraction of premium collected on the CURRENT attempt ---
# When an attempt's profit reaches PROFIT_TARGET_PCT * (CE+PE)*qty, it exits at the
# target and NO further trades are taken that day. 0 disables. e.g. 0.70 = 70%.
PROFIT_TARGET_PCT = float(os.getenv("PROFIT_TARGET_PCT", "0.70"))
# --- Per-attempt RE-ENTRY GAP in minutes (index 0 = gap before 1st re-entry, 1 = before 2nd, ...) ---
# Attempts beyond the list reuse the LAST value. Override via env comma list, e.g.
# REENTRY_DELAY_BY_ATTEMPT="10,15,20".
REENTRY_DELAY_BY_ATTEMPT = _parse_int_list(
    os.getenv("REENTRY_DELAY_BY_ATTEMPT"),
    [1, 1, 1, 1, 5,5, 10, 10, 10, 10],
)

def reentry_delay_for_attempt(attempt_idx: int) -> int:
    s = REENTRY_DELAY_BY_ATTEMPT
    if not s:
        return 0
    return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])

_DEFAULT_OUT = os.path.join(
    _get_downloads_folder(),
    f"short_straddle_backtest_reattempt{_safe_fname_part(ENTRY_TIME_IST)}"
    f"_SLpct_{_safe_fname_part(_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    f"_DTE_{_safe_fname_part('-'.join(str(d) for d in ALLOWED_DTE))}"
    f"_PPTpct_{_safe_fname_part(_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES))}"
    f"_DailyMaxLoss_{_safe_fname_part(_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES))}"
    f"_StopCap_{_safe_fname_part(_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    f"_MR_{_safe_fname_part(str(MAX_REATTEMPTS))}"
    f"_PT_{_safe_fname_part(str(int(round(PROFIT_TARGET_PCT * 100))))}pct"
    f"_RDM_{_safe_fname_part(_fmt_int_list(REENTRY_DELAY_BY_ATTEMPT))}.xlsx"
)

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", _DEFAULT_OUT)

FAIL_ON_PICKLE_ERROR = os.getenv("FAIL_ON_PICKLE_ERROR", "0").strip() == "1"

SESSION_START_IST = dtime(9, 15)
SESSION_END_IST = dtime(15, 30)

# LOOKBACK_MONTHS is AUTO by default: the script backtests the full date range
# present in the option pickles. If you explicitly set LOOKBACK_MONTHS to a
# number, that number acts as an optional manual cap.
LOOKBACK_MONTHS_RAW = os.getenv("LOOKBACK_MONTHS", "AUTO").strip()
LOOKBACK_MONTHS: Optional[int]
if LOOKBACK_MONTHS_RAW.upper() in ("", "AUTO", "ALL", "MAX", "FULL"):
    LOOKBACK_MONTHS = None
else:
    LOOKBACK_MONTHS = int(float(LOOKBACK_MONTHS_RAW))

QTY_UNITS = {"NIFTY": 325, "SENSEX": 100}
TRADEABLE = set(QTY_UNITS.keys())

STRIKE_STEP = {"NIFTY": 50, "SENSEX": 100}

# =============================================================================
# TRANSACTION CHARGES (Zerodha F&O Options — NSE)
# =============================================================================
# Each short-straddle attempt = 4 executed orders (sell CE, sell PE, buy CE, buy PE)
BROKERAGE_PER_ORDER       = 20.0       # ₹20 flat per executed order
ORDERS_PER_TRADE          = 4          # sell CE + sell PE + buy CE + buy PE
STT_SELL_PCT              = 0.001      # 0.1% on sell-side premium
EXCHANGE_TXN_PCT          = 0.0003553  # 0.03553% on premium (NSE options)
SEBI_PER_CRORE            = 10.0       # ₹10 per crore of turnover
STAMP_BUY_PCT             = 0.00003    # 0.003% on buy-side premium
IPFT_PER_CRORE            = 0.010      # ₹0.01 per crore (on premium)
GST_PCT                   = 0.18       # 18% on (brokerage + txn charges + SEBI)
INCLUDE_TRANSACTION_COSTS = True       # set False to disable

UNDERLYING_KITE = {
    "NIFTY": {"exchange": "NSE", "tradingsymbol": "NIFTY 50"},
    "SENSEX": {"exchange": "BSE", "tradingsymbol": "SENSEX"},
}

MAX_DAYS_PER_CHUNK = 25
MAX_ATTEMPTS = 5
SLEEP_BETWEEN_CALLS_SEC = 0.20


# =============================================================================
# HELPERS
# =============================================================================
def parse_hhmm(s: str) -> dtime:
    hh, mm = s.strip().split(":")
    return dtime(int(hh), int(mm))

ENTRY_TIME = parse_hhmm(ENTRY_TIME_IST)

def ist_tz():
    if ZoneInfo is not None:
        return ZoneInfo("Asia/Kolkata")
    if pytz is not None:
        return pytz.timezone("Asia/Kolkata")
    return "Asia/Kolkata"

def ensure_ist(series_or_scalar) -> Any:
    tz = ist_tz()
    dt = pd.to_datetime(series_or_scalar, errors="coerce")
    if isinstance(dt, pd.Series):
        if dt.dt.tz is None:
            return dt.dt.tz_localize(tz)
        return dt.dt.tz_convert(tz)
    if getattr(dt, "tzinfo", None) is None:
        return dt.tz_localize(tz)
    return dt.tz_convert(tz)

def normalize_underlying(name: str) -> Optional[str]:
    if not isinstance(name, str):
        return None
    u = name.upper().strip()
    if "SENSEX" in u:
        return "SENSEX"
    if "BANKNIFTY" in u or "NIFTY BANK" in u:
        return "BANKNIFTY"
    if "NIFTY" in u:
        return "NIFTY"
    return None

def round_to_step(x: float, step: int) -> int:
    return int(round(x / step) * step)

def build_minute_index(day_d: date, start_t: dtime, end_t: dtime) -> pd.DatetimeIndex:
    tz = ist_tz()
    start = pd.Timestamp(datetime.combine(day_d, start_t), tz=tz)
    end = pd.Timestamp(datetime.combine(day_d, end_t), tz=tz)
    return pd.date_range(start=start, end=end, freq="1min")

def asof_close(df: pd.DataFrame, ts: pd.Timestamp) -> float:
    if df.empty:
        return float("nan")
    d = df[["date", "close"]].dropna().copy()
    d["date"] = ensure_ist(d["date"])
    d = d.sort_values("date").set_index("date")
    loc = d.index.get_indexer([ts], method="pad")
    if loc[0] == -1:
        return float("nan")
    return float(d.iloc[loc[0]]["close"])

def compute_window_start(end_day: date, months: int) -> date:
    if relativedelta is not None:
        return (pd.Timestamp(end_day) - relativedelta(months=months)).date()
    return (pd.Timestamp(end_day) - pd.Timedelta(days=30 * months)).date()


def determine_backtest_window_start(min_day_seen: date, end_day: date) -> date:
    """
    Determine the backtest start date.

    Default behaviour: use the earliest usable option-data date found in the
    pickles, i.e. the maximum available backtest period.

    Optional override: if LOOKBACK_MONTHS is set to a numeric value through the
    environment, use the later of:
        1. earliest option-data date; and
        2. end_day - LOOKBACK_MONTHS
    so the script never requests data before the pickles actually start.
    """
    if LOOKBACK_MONTHS is None:
        return min_day_seen

    capped_start = compute_window_start(end_day, LOOKBACK_MONTHS)
    return max(min_day_seen, capped_start)

# =============================================================================
# TRANSACTION COST CALCULATOR
# =============================================================================
def compute_trade_charges(
    entry_ce: float, entry_pe: float,
    exit_ce: float, exit_pe: float,
    qty: int,
) -> float:
    """
    Compute total Zerodha transaction charges for one short-straddle attempt.

    Entry = SELL CE + SELL PE  (2 orders, sell side)
    Exit  = BUY  CE + BUY  PE (2 orders, buy side)

    Returns total charges in rupees (always positive).
    """
    if not INCLUDE_TRANSACTION_COSTS:
        return 0.0

    # Turnover values (in rupees)
    entry_turnover = (entry_ce + entry_pe) * qty   # sell side
    exit_turnover  = (exit_ce + exit_pe) * qty     # buy side
    total_turnover = entry_turnover + exit_turnover

    # 1. Brokerage: ₹20 × 4 orders
    brokerage = BROKERAGE_PER_ORDER * ORDERS_PER_TRADE

    # 2. STT: 0.1% on sell-side premium only (entry for short straddle)
    stt = entry_turnover * STT_SELL_PCT

    # 3. Exchange transaction charges: 0.03553% on both sides
    txn_charges = total_turnover * EXCHANGE_TXN_PCT

    # 4. SEBI charges: ₹10 per crore on total turnover
    sebi = total_turnover * SEBI_PER_CRORE / 1_00_00_000

    # 5. Stamp duty: 0.003% on buy side only (exit for short straddle)
    stamp = exit_turnover * STAMP_BUY_PCT

    # 6. IPFT: ₹0.01 per crore on premium (both sides)
    ipft = total_turnover * IPFT_PER_CRORE / 1_00_00_000

    # 7. GST: 18% on (brokerage + transaction charges + SEBI charges)
    gst = (brokerage + txn_charges + sebi) * GST_PCT

    total_charges = brokerage + stt + txn_charges + sebi + stamp + ipft + gst
    return round(total_charges, 2)

# =============================================================================
# Kite historical helpers
# =============================================================================
def _iter_chunks_by_date(from_dt: datetime, to_dt: datetime, days_per_chunk: int) -> List[Tuple[datetime, datetime]]:
    if from_dt > to_dt:
        raise ValueError("from_dt must be <= to_dt")
    chunks: List[Tuple[datetime, datetime]] = []
    cur = from_dt.date()
    end_d = to_dt.date()
    while cur <= end_d:
        chunk_end = min(cur + timedelta(days=days_per_chunk - 1), end_d)
        c_from = from_dt if cur == from_dt.date() else datetime.combine(cur, SESSION_START_IST)
        c_to = to_dt if chunk_end == end_d else datetime.combine(chunk_end, SESSION_END_IST)
        chunks.append((c_from, c_to))
        cur = chunk_end + timedelta(days=1)
    return chunks

def _kite_instruments_cached(kite, exchange: str, cache: Dict[str, List[Dict]]) -> List[Dict]:
    ex = exchange.upper().strip()
    if ex not in cache:
        print(f"[STEP] Loading instruments dump for {ex} ...")
        cache[ex] = kite.instruments(ex)
        print(f"[INFO] {ex} instruments: {len(cache[ex])}")
    return cache[ex]

def get_instrument_token(kite, exchange: str, tradingsymbol: str, cache: Dict[str, List[Dict]]) -> int:
    ex = exchange.upper().strip()
    wanted = tradingsymbol.strip().upper()
    for r in _kite_instruments_cached(kite, ex, cache):
        if str(r.get("tradingsymbol", "")).upper() == wanted:
            return int(r["instrument_token"])
    raise ValueError(f"Instrument not found on {ex}: '{tradingsymbol}'")

def fetch_history_minute(kite, instrument_token: int, from_dt: datetime, to_dt: datetime, label: str) -> List[Dict]:
    interval = "minute"
    chunks = _iter_chunks_by_date(from_dt, to_dt, MAX_DAYS_PER_CHUNK)
    rows_all: List[Dict] = []
    print(f"[INFO] Fetch {label} token={instrument_token} chunks={len(chunks)} {from_dt} -> {to_dt}")
    for i, (c_from, c_to) in enumerate(chunks, start=1):
        last_err = None
        for attempt in range(1, MAX_ATTEMPTS + 1):
            try:
                rows = kite.historical_data(
                    instrument_token=instrument_token,
                    from_date=c_from,
                    to_date=c_to,
                    interval=interval,
                    continuous=False,
                    oi=False,
                )
                rows_all.extend(rows)
                last_err = None
                break
            except Exception as e:
                last_err = e
                time.sleep(min(8.0, 1.5 * attempt))
        if last_err is not None:
            print(f"[ERROR] {label} chunk {i}/{len(chunks)} failed: {c_from}->{c_to}: {last_err}")
        time.sleep(SLEEP_BETWEEN_CALLS_SEC)
    return rows_all

def rows_to_df(rows: List[Dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["date", "open", "high", "low", "close", "volume"])
    df = pd.DataFrame(rows)
    df["date"] = ensure_ist(df["date"])
    df = df.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    return df


# =============================================================================
# DATA STRUCTURES
# =============================================================================
@dataclass
class TradeRow:
    day: date
    underlying: str
    trade_seq: int
    expiry: date
    days_to_expiry: int
    atm_strike: int
    qty_units: int
    entry_time: str
    exit_time: str
    exit_reason: str
    entry_underlying: float
    ce_symbol: str
    pe_symbol: str
    entry_ce: float
    entry_pe: float
    exit_ce: float
    exit_pe: float
    exit_pnl_gross: float   # P&L before charges
    txn_charges: float      # total transaction charges for this attempt
    exit_pnl: float         # net P&L after deducting charges
    eod_pnl: float
    max_profit: float
    max_loss: float
    max_profit_before_exit: float   # peak profit reached before this trade exited

    # Premium/risk diagnostics for percentage-based risk rules
    entry_premium_sum: float                 # (entry_ce + entry_pe) * qty
    stop_pct: float                          # stop-loss % of entry premium, decimal form; 0.10 = 10%
    uncapped_stop_rupees: float              # percentage-based stop before absolute cap
    stop_cap_rupees: float                   # configured absolute cap; <=0 means cap disabled
    stop_rupees: float                       # effective rupee stop after cap
    profit_protect_trigger_pct: float         # profit-protect % of entry premium, decimal form; 0.30 = 30%
    profit_protect_trigger_rupees: float      # computed rupee profit-protect trigger/giveback
    daily_realized_pnl_after_trade: float     # cumulative net P&L after this attempt
    daily_loss_limit_rupees: float            # configured daily loss circuit breaker
    daily_loss_limit_hit: bool                # True means no further trades for that day


# =============================================================================
# PASS-1: nearest expiry per (underlying, day)
# =============================================================================
def scan_pickles_pass1(pickle_paths: List[str]) -> Tuple[date, Dict[Tuple[str, date], date], date]:
    max_day_seen: Optional[date] = None
    min_day_seen: Optional[date] = None
    min_expiry_map: Dict[Tuple[str, date], date] = {}

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            for c in ("date", "name", "expiry", "type"):
                if c not in df.columns:
                    raise ValueError(f"Missing column '{c}' in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")]
            if d2.empty:
                continue

            d2 = d2[["date", "name", "expiry"]].copy()
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2 = d2.dropna(subset=["underlying", "day", "expiry_date"])

            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            file_min_day = d2["day"].min()
            file_max_day = d2["day"].max()
            max_day_seen = file_max_day if (max_day_seen is None or file_max_day > max_day_seen) else max_day_seen
            min_day_seen = file_min_day if (min_day_seen is None or file_min_day < min_day_seen) else min_day_seen

            grp = d2.groupby(["underlying", "day"], sort=False)["expiry_date"].min()
            for (und, dy), ex in grp.items():
                key = (und, dy)
                if key not in min_expiry_map or ex < min_expiry_map[key]:
                    min_expiry_map[key] = ex

            print(f"[PASS1 OK] {os.path.basename(p)} option_days={d2['day'].nunique()}")

        except Exception as e:
            msg = f"[PASS1 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    if max_day_seen is None or min_day_seen is None:
        raise RuntimeError("No usable option data found in pickles (PASS1) for tradeable underlyings.")

    return max_day_seen, min_expiry_map, min_day_seen


# =============================================================================
# Underlying download
# =============================================================================
def download_underlyings(kite, day_start: date, day_end: date) -> Dict[str, pd.DataFrame]:
    cache: Dict[str, List[Dict]] = {}
    from_dt = datetime.combine(day_start, SESSION_START_IST)
    to_dt = datetime.combine(day_end, SESSION_END_IST)

    out: Dict[str, pd.DataFrame] = {}
    for und, meta in UNDERLYING_KITE.items():
        token = get_instrument_token(kite, meta["exchange"], meta["tradingsymbol"], cache)
        rows = fetch_history_minute(kite, token, from_dt, to_dt, label=f"{meta['exchange']}:{meta['tradingsymbol']}")
        df = rows_to_df(rows)
        df["day"] = df["date"].dt.tz_convert(ist_tz()).dt.date
        out[und] = df
        print(f"[UNDERLYING OK] {und}: candles={len(df)} days={df['day'].nunique()}")
    return out


# =============================================================================
# Simulation helpers
# =============================================================================
def _pick_symbol(day_opt: pd.DataFrame, strike: int, opt_type: str) -> Optional[str]:
    sub = day_opt[(day_opt["strike_int"] == strike) & (day_opt["option_type"] == opt_type)]
    if sub.empty:
        return None
    syms = sorted(sub["instrument"].astype(str).unique().tolist())
    return syms[0] if syms else None

def _build_leg_series(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex,
                      strike: int, opt_type: str, symbol: str,
                      price_col: str = "close", do_ffill: bool = True) -> pd.Series:
    sub = day_opt[
        (day_opt["strike_int"] == strike) &
        (day_opt["option_type"] == opt_type) &
        (day_opt["instrument"].astype(str) == symbol)
    ][["date", price_col]].dropna()

    if sub.empty:
        return pd.Series(index=idx_all, dtype="float64")

    sub = sub.copy()
    sub["date"] = ensure_ist(sub["date"])
    sub = sub.sort_values("date").drop_duplicates(subset=["date"], keep="last").set_index("date")
    s = sub[price_col].astype(float).reindex(idx_all)
    return s.ffill() if do_ffill else s


def build_price_book(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex):
    """
    Precompute, ONCE per trading day, the minute-indexed (close/high/low) series
    for every (strike, option_type) present -- exactly as _build_leg_series would
    produce them with do_ffill=False. The optimizer rebuilds the same parameter-
    independent series on every trial; caching them here turns ~300x of pandas
    reindex+tz work into a single dict lookup, which is the whole speed-up.

    Returns:
        book    : {(strike_int, opt_type, price_col): pd.Series on idx_all (raw)}
        symbols : {(strike_int, opt_type): chosen instrument symbol}
    Numerically identical to _build_leg_series (verified by regression test).
    """
    book: Dict[Tuple[int, str, str], pd.Series] = {}
    symbols: Dict[Tuple[int, str], str] = {}
    if day_opt is None or day_opt.empty:
        return book, symbols

    for (strike, opt), sub in day_opt.groupby(["strike_int", "option_type"], sort=False):
        # Same instrument-selection rule as _pick_symbol (first sorted symbol).
        syms = sorted(sub["instrument"].astype(str).unique().tolist())
        if not syms:
            continue
        sym = syms[0]
        symbols[(int(strike), str(opt))] = sym

        sub_sym = sub[sub["instrument"].astype(str) == sym][["date", "close", "high", "low"]].copy()
        sub_sym["date"] = ensure_ist(sub_sym["date"])
        sub_sym = (sub_sym.sort_values("date")
                          .drop_duplicates(subset=["date"], keep="last")
                          .set_index("date"))
        for col in ("close", "high", "low"):
            book[(int(strike), str(opt), col)] = sub_sym[col].astype(float).reindex(idx_all)
    return book, symbols


def _leg_from_book(book, idx_all, strike: int, opt_type: str, price_col: str) -> pd.Series:
    """Cached-series accessor mirroring _build_leg_series(..., do_ffill=False)."""
    s = book.get((int(strike), str(opt_type), price_col))
    if s is None:
        return pd.Series(index=idx_all, dtype="float64")
    return s

# =============================================================================
# PARAMS (everything the optimizer can vary, threaded explicitly into the sim
# instead of being read from module globals -- so we can run many backtests in
# one process without re-importing). The module globals above are still used to
# build the DEFAULT params for the normal single-run main() workflow.
# =============================================================================
@dataclass
class Params:
    entry_time: dtime                       # ENTRY_TIME_IST, as a datetime.time
    loss_limit_pct_by_attempt: List[float]  # LOSS_LIMIT_RUPEES_BY_ATTEMPT (decimals)
    profit_protect_pct: float               # PROFIT_PROTECT_TRIGGER_RUPEES (decimal)
    max_reattempts: int                     # MAX_REATTEMPTS
    profit_target_pct: float                # PROFIT_TARGET_PCT (decimal)
    reentry_delay_by_attempt: List[int]     # REENTRY_DELAY_BY_ATTEMPT (minutes)
    # not optimized here, but consumed by the sim -- kept on Params so the sim
    # never reaches back into globals:
    max_daily_loss_rupees: float
    max_loss_limit_cap_rupees: float

    def loss_limit_pct_for_attempt(self, attempt_idx: int) -> float:
        s = self.loss_limit_pct_by_attempt
        if not s:
            return 0.0
        return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])

    def reentry_delay_for_attempt(self, attempt_idx: int) -> int:
        s = self.reentry_delay_by_attempt
        if not s:
            return 0
        return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])


def default_params() -> "Params":
    """Build Params from the module-level globals (preserves env-var behaviour)."""
    return Params(
        entry_time=ENTRY_TIME,
        loss_limit_pct_by_attempt=list(LOSS_LIMIT_RUPEES_BY_ATTEMPT),
        profit_protect_pct=float(PROFIT_PROTECT_TRIGGER_RUPEES),
        max_reattempts=int(MAX_REATTEMPTS),
        profit_target_pct=float(PROFIT_TARGET_PCT),
        reentry_delay_by_attempt=list(REENTRY_DELAY_BY_ATTEMPT),
        max_daily_loss_rupees=float(MAX_DAILY_LOSS_RUPEES),
        max_loss_limit_cap_rupees=float(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT),
    )


def simulate_day_multi_trades(
    *,
    und: str,
    dy: date,
    expiry: date,
    day_opt: pd.DataFrame,
    underlying_day: pd.DataFrame,
    params: "Params",
    price_book: Optional[dict] = None,   # precomputed per-day series (perf cache)
    symbols: Optional[dict] = None,      # precomputed (strike,opt)->symbol
    idx_all: Optional[pd.DatetimeIndex] = None,
) -> Tuple[List[TradeRow], List[Dict[str, Any]]]:

    results: List[TradeRow] = []
    skipped: List[Dict[str, Any]] = []

    # Use the precomputed per-day price book if the caller supplied one (the
    # optimizer does, so the heavy series construction happens once). If not
    # (e.g. main()'s single run, or a direct call), build it on the fly here so
    # behaviour is identical either way.
    if idx_all is None:
        idx_all = build_minute_index(dy, SESSION_START_IST, SESSION_END_IST)
    if price_book is None or symbols is None:
        price_book, symbols = build_price_book(day_opt, idx_all)
    session_end_ts = idx_all[-1]

    qty = int(QTY_UNITS[und])
    step = int(STRIKE_STEP[und])

    # Profit-protect is now percentage-based, so the actual rupee value is not
    # known until CE/PE entry prices are available for the current attempt.
    profit_protect_pct = float(params.profit_protect_pct)
    profit_protect_enabled = profit_protect_pct > 0.0

    cur_entry_ts = pd.Timestamp(datetime.combine(dy, params.entry_time), tz=ist_tz())
    trade_seq = 1

    # Cumulative realized NET P&L for this underlying/day. Used for the daily
    # loss circuit breaker. Charges are included through exit_pnl.
    daily_realized_pnl = 0.0
    daily_loss_limit_enabled = params.max_daily_loss_rupees > 0

    while cur_entry_ts <= session_end_ts:
        if daily_loss_limit_enabled and daily_realized_pnl <= -float(params.max_daily_loss_rupees):
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq,
                "reason": (
                    f"Daily loss limit hit before next entry: "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={params.max_daily_loss_rupees:.2f}"
                ),
            })
            break

        u_px = asof_close(underlying_day, cur_entry_ts)
        if pd.isna(u_px):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": f"No underlying price at entry {cur_entry_ts.strftime('%H:%M')}"})
            break

        atm = round_to_step(float(u_px), step)

        ce_sym = symbols.get((atm, "CE"))
        pe_sym = symbols.get((atm, "PE"))
        if not ce_sym or not pe_sym:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "ATM CE/PE not available in pickle band"})
            break

        # Close series (used for entry pricing, profit-protect tracking, and reporting)
        # Raw close series for exact entry validation -- from the precomputed book.
        ce_close_raw = _leg_from_book(price_book, idx_all, atm, "CE", "close")
        pe_close_raw = _leg_from_book(price_book, idx_all, atm, "PE", "close")

        # Forward-filled close series for post-entry tracking/reporting
        ce_close = ce_close_raw.ffill()
        pe_close = pe_close_raw.ffill()

        # High/Low series (used only to detect STOPLOSS intraminute extremes)
        ce_high = _leg_from_book(price_book, idx_all, atm, "CE", "high")
        ce_low = _leg_from_book(price_book, idx_all, atm, "CE", "low")
        pe_high = _leg_from_book(price_book, idx_all, atm, "PE", "high")
        pe_low = _leg_from_book(price_book, idx_all, atm, "PE", "low")

        if cur_entry_ts not in idx_all:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": "Entry timestamp not in session index"})
            break

        ce_entry = ce_close_raw.loc[cur_entry_ts]
        pe_entry = pe_close_raw.loc[cur_entry_ts]
        monitor_start_ts = pd.Timestamp(cur_entry_ts) + pd.Timedelta(minutes=1)
        if monitor_start_ts > session_end_ts:
            break

        if pd.isna(ce_entry) or pd.isna(pe_entry):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "No CE/PE price at entry (after ffill)"})
            break

        # ---------------------------------------------------------------------
        # Percentage-based risk basis for THIS attempt
        # ---------------------------------------------------------------------
        # For every entry/re-entry, compute the premium collected in rupees.
        # Stop-loss and profit-protect thresholds are derived from this value.
        #
        # Example:
        #   entry_ce=70, entry_pe=50, qty=325
        #   entry_premium_sum = (70 + 50) * 325 = 39,000
        #   10% stop-loss = 3,900
        #   30% profit-protect threshold/giveback = 11,700
        # ---------------------------------------------------------------------
        entry_premium_sum = (float(ce_entry) + float(pe_entry)) * qty

        loss_limit_pct = params.loss_limit_pct_for_attempt(trade_seq - 1)
        uncapped_loss_limit_rupees = float(loss_limit_pct * entry_premium_sum)

        # Absolute cap on the percentage-based stop-loss.
        # Example: 10% of premium may be Rs. 4,500, but with a Rs. 3,000 cap
        # the effective stop used by the simulator is Rs. 3,000.
        stop_cap_rupees = float(params.max_loss_limit_cap_rupees)
        if stop_cap_rupees > 0:
            loss_limit_rupees = float(min(uncapped_loss_limit_rupees, stop_cap_rupees))
        else:
            loss_limit_rupees = float(uncapped_loss_limit_rupees)

        # G is the same variable used by the existing profit-protect logic:
        #   - profit-protect arms when peak P&L >= G
        #   - profit-protect exits when current P&L <= peak - G
        G = float(profit_protect_pct * entry_premium_sum)

        # Close-based PnL (same as before)
        pnl_close_all = (float(ce_entry) - ce_close) * qty + (float(pe_entry) - pe_close) * qty
        pnl = pnl_close_all.loc[monitor_start_ts:].dropna()  # keep 'pnl' as close-based for profit-protect

        # STOPLOSS worst-case PnL candidates within each minute:
        #  A) CE high, PE low
        pnl_ceHigh_peLow_all = (float(ce_entry) - ce_high) * qty + (float(pe_entry) - pe_low) * qty
        #  B) CE low, PE high
        pnl_ceLow_peHigh_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_high) * qty

        # Worst-case PnL per minute among (close, A, B)
        pnl_sl_all = pd.concat([pnl_close_all, pnl_ceHigh_peLow_all, pnl_ceLow_peHigh_all], axis=1).min(axis=1)
        pnl_sl = pnl_sl_all.loc[monitor_start_ts:].dropna()

        if pnl.empty:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "PnL series empty after entry"})
            break

        eod_ts = pnl.index[-1]
        eod_pnl = float(pnl.iloc[-1])

        max_profit = float(max(0.0, pnl.max()))
        max_loss = float(min(0.0, pnl.min()))

        # STOPLOSS uses the attempt-specific rupee value after applying the
        # absolute per-attempt cap.
        stop_hit = pnl_sl <= -loss_limit_rupees
        stop_ts = pnl_sl.index[stop_hit.to_numpy().argmax()] if stop_hit.any() else None

        protect_ts = None
        if profit_protect_enabled:
            peak = pnl.cummax()
            armed = peak >= G
            trail = peak - G
            protect_hit = armed & (pnl <= trail)
            protect_ts = pnl.index[protect_hit.to_numpy().argmax()] if protect_hit.any() else None

        # --- Per-day PROFIT TARGET: % of premium collected on this attempt ---
        # When reached, this trade exits at the target AND no further trades are
        # taken for the day (PROFIT_TARGET is excluded from the re-entry rule below).
        target_ts = None
        target_rupees = None
        if params.profit_target_pct > 0.0:
            target_rupees = params.profit_target_pct * entry_premium_sum
            # best-case (favourable) intrabar profit: both legs bought back at their lows
            pnl_best_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_low) * qty
            pnl_tp = pd.concat([pnl_close_all, pnl_best_all], axis=1).max(axis=1)
            pnl_tp = pnl_tp.loc[monitor_start_ts:].dropna()
            tp_hit = pnl_tp >= float(target_rupees)
            target_ts = pnl_tp.index[tp_hit.to_numpy().argmax()] if tp_hit.any() else None

        # Earliest triggered exit wins; on identical timestamps prefer the more
        # conservative outcome: STOPLOSS, then PROFIT_TARGET, then PROFIT_PROTECT.
        exit_ts = eod_ts
        exit_reason = "EOD"
        _candidates = []
        if stop_ts is not None:
            _candidates.append((stop_ts, 0, "STOPLOSS"))
        if target_ts is not None:
            _candidates.append((target_ts, 1, "PROFIT_TARGET"))
        if protect_ts is not None:
            _candidates.append((protect_ts, 2, "PROFIT_PROTECT"))
        if _candidates:
            _candidates.sort(key=lambda c: (c[0], c[1]))
            exit_ts, _, exit_reason = _candidates[0]

        if exit_reason == "STOPLOSS":
            exit_pnl_gross = -float(loss_limit_rupees)
        elif exit_reason == "PROFIT_TARGET":
            exit_pnl_gross = float(target_rupees)
        else:
            exit_pnl_gross = float(pnl.loc[exit_ts])

        # Peak (close-based) profit reached during this trade's life, up to its exit
        pnl_pre_exit = pnl.loc[:exit_ts]
        max_profit_before_exit = float(max(0.0, pnl_pre_exit.max())) if len(pnl_pre_exit) else 0.0

        exit_ce = float(ce_close.loc[exit_ts]) if pd.notna(ce_close.loc[exit_ts]) else float("nan")
        exit_pe = float(pe_close.loc[exit_ts]) if pd.notna(pe_close.loc[exit_ts]) else float("nan")

        txn_charges = compute_trade_charges(
            entry_ce=float(ce_entry), entry_pe=float(pe_entry),
            exit_ce=exit_ce if not pd.isna(exit_ce) else 0.0,
            exit_pe=exit_pe if not pd.isna(exit_pe) else 0.0,
            qty=qty,
        )
        exit_pnl = exit_pnl_gross - txn_charges

        # Update cumulative realized NET P&L for the day. This is checked
        # before allowing any further re-entry.
        daily_realized_pnl += float(exit_pnl)
        daily_loss_limit_hit = bool(
            daily_loss_limit_enabled and daily_realized_pnl <= -float(params.max_daily_loss_rupees)
        )

        dte = int((expiry - dy).days)

        results.append(
            TradeRow(
                day=dy,
                underlying=und,
                trade_seq=trade_seq,
                expiry=expiry,
                days_to_expiry=dte,
                atm_strike=int(atm),
                qty_units=qty,
                entry_time=pd.Timestamp(cur_entry_ts).strftime("%H:%M"),
                exit_time=pd.Timestamp(exit_ts).strftime("%H:%M"),
                exit_reason=exit_reason,
                entry_underlying=float(u_px),
                ce_symbol=ce_sym,
                pe_symbol=pe_sym,
                entry_ce=float(ce_entry),
                entry_pe=float(pe_entry),
                exit_ce=exit_ce,
                exit_pe=exit_pe,
                exit_pnl_gross=exit_pnl_gross,
                txn_charges=txn_charges,
                exit_pnl=exit_pnl,
                eod_pnl=eod_pnl,
                max_profit=max_profit,
                max_loss=max_loss,
                max_profit_before_exit=max_profit_before_exit,
                entry_premium_sum=float(entry_premium_sum),
                stop_pct=float(loss_limit_pct),
                uncapped_stop_rupees=float(uncapped_loss_limit_rupees),
                stop_cap_rupees=float(stop_cap_rupees),
                stop_rupees=float(loss_limit_rupees),
                profit_protect_trigger_pct=float(profit_protect_pct),
                profit_protect_trigger_rupees=float(G),
                daily_realized_pnl_after_trade=float(daily_realized_pnl),
                daily_loss_limit_rupees=float(params.max_daily_loss_rupees),
                daily_loss_limit_hit=bool(daily_loss_limit_hit),
            )
        )

        if daily_loss_limit_hit:
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq + 1,
                "reason": (
                    f"No re-entry: daily loss limit hit after trade_seq={trade_seq}; "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={params.max_daily_loss_rupees:.2f}"
                ),
            })
            break

        if exit_reason in ("STOPLOSS", "PROFIT_PROTECT") and (trade_seq - 1) < params.max_reattempts:
            delay_min = params.reentry_delay_for_attempt(trade_seq - 1)  # gap before this re-entry
            trade_seq += 1
            cur_entry_ts = pd.Timestamp(exit_ts) + pd.Timedelta(minutes=delay_min)
            if cur_entry_ts > session_end_ts:
                break
            continue

        break

    return results, skipped


# =============================================================================
# PASS-2: process each pickle and simulate trades for days where this expiry is nearest
# =============================================================================
@dataclass
class DayGroup:
    """One simulatable (underlying, day, expiry) unit with its option + underlying
    slices already extracted. Built ONCE; reused across every optimizer trial.
    The price_book/symbols/idx_all are filled lazily on first simulation and then
    reused, so the expensive per-day series construction happens only once."""
    und: str
    dy: date
    expiry: date
    day_opt: pd.DataFrame
    underlying_day: pd.DataFrame
    price_book: Optional[dict] = None
    symbols: Optional[dict] = None
    idx_all: Optional[pd.DatetimeIndex] = None


def build_day_groups(
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
    max_pickles: Optional[int] = None,   # small-sample: only read this many files
    max_days: Optional[int] = None,      # small-sample: keep only the most recent N day-groups
) -> Tuple[List[DayGroup], List[Dict[str, Any]]]:
    """
    Parse pickles and slice them into per-day groups. This is the EXPENSIVE part
    (disk I/O + parsing) and is parameter-independent, so the optimizer runs it
    ONCE and re-simulates the returned groups for every trial.

    Progress: prints a counter after each pickle so a long load is visible.
    Small-sample: `max_pickles` limits how many files are read; `max_days` keeps
    only the most recent N day-groups after loading. Both are for quick smoke
    tests and should be left as None for a real run.
    """
    groups: List[DayGroup] = []
    skipped_rows: List[Dict[str, Any]] = []
    processed_day_keys: set = set()  # prevent double-count of the same (und,day,expiry) across files

    # Small-sample shortcut: only look at the first `max_pickles` files.
    if max_pickles is not None and max_pickles > 0:
        pickle_paths = pickle_paths[:max_pickles]

    total_files = len(pickle_paths)
    for fi, p in enumerate(pickle_paths, start=1):
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                print(f"[LOAD {fi}/{total_files}] {os.path.basename(p)}: empty (skipped)")
                continue

            needed_cols = ["date", "name", "type", "option_type", "strike", "expiry", "instrument", "high", "low", "close"]
            missing = [c for c in needed_cols if c not in df.columns]
            if missing:
                raise ValueError(f"Missing columns {missing} in {p}")

            # Keep only OPTION rows and the columns we actually use.
            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")][needed_cols].copy()
            if d2.empty:
                continue

            # Normalize types/keys used downstream.
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            if d2.empty:
                continue

            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2["strike_num"] = pd.to_numeric(d2["strike"], errors="coerce")
            d2["strike_int"] = d2["strike_num"].round().astype("Int64")
            d2["option_type"] = d2["option_type"].astype(str).str.upper()

            d2 = d2.dropna(subset=["day", "underlying", "expiry_date", "strike_int", "close"])
            d2["strike_int"] = d2["strike_int"].astype(int)
            d2 = d2[d2["expiry_date"] >= d2["day"]]          # drop stale (already-expired) rows
            if d2.empty:
                continue
            d2 = d2[(d2["day"] >= window_start) & (d2["day"] <= window_end)]  # date window
            if d2.empty:
                continue

            # One simulatable unit per (underlying, day, expiry), but only when this
            # expiry is the NEAREST expiry for that (underlying, day) -- i.e. the
            # contract we'd actually trade that day.
            for (und, dy, ex), g in d2.groupby(["underlying", "day", "expiry_date"], sort=False):
                if min_expiry_map.get((und, dy)) != ex:
                    continue
                day_key = (und, dy, ex)
                if day_key in processed_day_keys:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex,
                                         "reason": "Duplicate (underlying,day,expiry) across pickles; skipped to avoid double-count"})
                    continue
                processed_day_keys.add(day_key)

                # Attach the underlying minute-bars for this day so the simulator
                # doesn't have to re-filter the full series on every trial.
                uday = underlying_data.get(und)
                if uday is None:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "No underlying series downloaded"})
                    continue
                uday = uday[uday["day"] == dy]
                if uday.empty:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "Underlying missing for day"})
                    continue

                groups.append(DayGroup(und=und, dy=dy, expiry=ex,
                                       day_opt=g.copy(), underlying_day=uday.copy()))

            print(f"[LOAD {fi}/{total_files}] {os.path.basename(p)} grouped "
                  f"(day-groups so far: {len(groups)})", flush=True)

        except Exception as e:
            msg = f"[LOAD {fi}/{total_files} WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    # Deterministic order so re-simulation is reproducible across trials.
    groups.sort(key=lambda gr: (gr.dy, gr.und))

    # Small-sample shortcut: keep only the most recent N day-groups.
    if max_days is not None and max_days > 0 and len(groups) > max_days:
        groups = groups[-max_days:]
        print(f"[LOAD] small-sample: keeping most recent {len(groups)} day-groups")

    # Pre-build each day's price book NOW (one-time, parameter-independent), so
    # every optimizer trial is uniformly fast and the progress ETA is honest from
    # the very first trial instead of trial 1 being anomalously slow.
    if groups:
        print(f"[LOAD] precomputing per-day price books for {len(groups)} day-groups ...", flush=True)
        for gi, gr in enumerate(groups, start=1):
            gr.idx_all = build_minute_index(gr.dy, SESSION_START_IST, SESSION_END_IST)
            gr.price_book, gr.symbols = build_price_book(gr.day_opt, gr.idx_all)
            if gi % 50 == 0 or gi == len(groups):
                print(f"[LOAD] price book {gi}/{len(groups)}", flush=True)

    print(f"[LOAD] done: {len(groups)} day-groups ready", flush=True)
    return groups, skipped_rows


def simulate_groups(params: "Params", groups: List[DayGroup]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Run the per-day simulation over pre-built groups for ONE parameter set.
    The per-day price book is built on the first trial and cached on the group,
    so every subsequent trial only does the (cheap) parameter-dependent work."""
    all_trades: List[Dict[str, Any]] = []
    sim_skips: List[Dict[str, Any]] = []
    for gr in groups:
        # Build the parameter-INDEPENDENT price book once, then reuse it forever.
        if gr.price_book is None:
            gr.idx_all = build_minute_index(gr.dy, SESSION_START_IST, SESSION_END_IST)
            gr.price_book, gr.symbols = build_price_book(gr.day_opt, gr.idx_all)
        trades, skips = simulate_day_multi_trades(
            und=gr.und, dy=gr.dy, expiry=gr.expiry,
            day_opt=gr.day_opt, underlying_day=gr.underlying_day, params=params,
            price_book=gr.price_book, symbols=gr.symbols, idx_all=gr.idx_all,
        )
        all_trades.extend([t.__dict__ for t in trades])
        sim_skips.extend(skips)

    all_df = pd.DataFrame(all_trades)
    if not all_df.empty:
        all_df = all_df.sort_values(["day", "underlying", "trade_seq"]).reset_index(drop=True)
    return all_df, pd.DataFrame(sim_skips)


def process_pickles_generate_trades(
    params: "Params",
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Single-run path used by main(): build groups, then simulate once."""
    groups, parse_skips = build_day_groups(pickle_paths, min_expiry_map, underlying_data, window_start, window_end)
    all_df, sim_skips = simulate_groups(params, groups)

    skip_df = pd.concat([pd.DataFrame(parse_skips), sim_skips], ignore_index=True) \
        if (parse_skips or not sim_skips.empty) else pd.DataFrame()
    if not skip_df.empty:
        if "day" not in skip_df.columns:
            skip_df["day"] = pd.NaT
        if "underlying" not in skip_df.columns:
            skip_df["underlying"] = pd.NA
        skip_df = skip_df.sort_values(["day", "underlying"], na_position="last").reset_index(drop=True)

    return all_df, skip_df


# =============================================================================
# Actual trades: one underlying per day (nearest expiry), include all re-entries for that underlying/day
# =============================================================================
def pick_actual_underlying_by_day(min_expiry_map: Dict[Tuple[str, date], date]) -> Dict[date, str]:
    by_day: Dict[date, List[Tuple[date, str]]] = {}
    for (und, dy), ex in min_expiry_map.items():
        if und not in TRADEABLE:
            continue

        dte = int((ex - dy).days)
        if dte not in ALLOWED_DTE:
            continue

        by_day.setdefault(dy, []).append((ex, und))

    out: Dict[date, str] = {}
    for dy, lst in by_day.items():
        # nearest expiry first; if tied, prefer NIFTY
        lst_sorted = sorted(lst, key=lambda t: (t[0], 0 if t[1] == "NIFTY" else 1))
        out[dy] = lst_sorted[0][1]
    return out

def build_actual_trades_df(all_trades_df: pd.DataFrame, min_expiry_map: Dict[Tuple[str, date], date]) -> pd.DataFrame:
    if all_trades_df.empty:
        return pd.DataFrame()

    actual_underlying = pick_actual_underlying_by_day(min_expiry_map)

    m = all_trades_df.copy()
    m["actual_underlying_for_day"] = m["day"].map(actual_underlying)

    # keep only days for which a 0/1-DTE actual underlying exists
    m = m[m["actual_underlying_for_day"].notna()]

    # keep only the selected underlying for that day
    m = m[m["underlying"] == m["actual_underlying_for_day"]]

    # keep only 0- and 1-DTE rows
    # keep only 0- and 1-DTE rows
    m = m[m["days_to_expiry"].isin(ALLOWED_DTE)]

    # keep all reattempts for the one selected underlying on that day
    m = m.drop(columns=["actual_underlying_for_day"])
    m = m.sort_values(["day", "trade_seq"]).reset_index(drop=True)

    # 1 if net exit PnL is positive, else 0
    m["is_exit_pnl_positive"] = (m["exit_pnl"] > 0).astype(int)

    return m


# =============================================================================
# Excel output
# =============================================================================
def _autosize_columns_safe(ws) -> None:
    # Safe autosize even when the sheet is "empty-ish"
    try:
        max_col = ws.max_column or 0
        if max_col <= 0:
            return
        for col_idx in range(1, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = 0
            for row_idx in range(1, min(ws.max_row or 1, 2000) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
    except Exception:
        # Never fail the whole run just because autosize misbehaved
        return

def _color_actual_trades_by_date(wb, actual_trades_df) -> None:
    """Shade rows so all attempts on the same calendar date share one colour,
    alternating between two soft fills as the date changes (visual grouping)."""
    if actual_trades_df is None or actual_trades_df.empty:
        return
    if "actual_trades" not in wb.sheetnames:
        return
    cols = list(actual_trades_df.columns)
    if "day" not in cols:
        return
    from openpyxl.styles import PatternFill
    ws = wb["actual_trades"]
    ncols = len(cols)
    fills = [
        PatternFill(fill_type="solid", fgColor="E8F0FE"),  # light blue
        PatternFill(fill_type="solid", fgColor="FFF3E0"),  # light amber
    ]
    days = actual_trades_df["day"].tolist()
    color_idx = 0
    prev_day = None
    first = True
    for i, d in enumerate(days):
        if first:
            first = False
        elif d != prev_day:
            color_idx ^= 1
        prev_day = d
        fill = fills[color_idx]
        excel_row = i + 2  # header occupies row 1
        for c in range(1, ncols + 1):
            ws.cell(row=excel_row, column=c).fill = fill


def write_excel(all_trades_df: pd.DataFrame, actual_trades_df: pd.DataFrame, skipped_df: pd.DataFrame) -> None:
    out_dir = os.path.dirname(os.path.abspath(OUTPUT_XLSX))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    piv_exit = pd.DataFrame()
    piv_eod_first = pd.DataFrame()
    monthwise_summary = pd.DataFrame()
    if not all_trades_df.empty:
        piv_exit = all_trades_df.pivot_table(index="day", columns="underlying", values="exit_pnl", aggfunc="sum").reset_index()

        first = all_trades_df[all_trades_df["trade_seq"] == 1]
        piv_eod_first = first.pivot_table(index="day", columns="underlying", values="eod_pnl", aggfunc="sum").reset_index()

        inst = all_trades_df.copy()
        inst["is_win_exit"] = inst["exit_pnl"] > 0
        inst["is_stoploss"] = inst["exit_reason"].astype(str).str.upper().eq("STOPLOSS")
        inst["is_profit_protect"] = inst["exit_reason"].astype(str).str.upper().eq("PROFIT_PROTECT")
        instrument_summary = (
            inst.groupby("underlying", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                win_rate_exit_pct=("is_win_exit", lambda s: 100.0 * s.mean()),
                stoploss_rate_pct=("is_stoploss", lambda s: 100.0 * s.mean()),
                profit_protect_rate_pct=("is_profit_protect", lambda s: 100.0 * s.mean()),
                avg_max_profit=("max_profit", "mean"),
                avg_max_loss=("max_loss", "mean"),
                worst_max_loss=("max_loss", "min"),
            )
            .sort_values("total_exit_pnl", ascending=False)
            .reset_index(drop=True)
        )
    else:
        instrument_summary = pd.DataFrame()

    if not actual_trades_df.empty:
        tmp = actual_trades_df.copy()
        tmp["month"] = pd.to_datetime(tmp["day"]).dt.to_period("M").astype(str)

        # Existing trade-level monthly summary
        monthwise_summary = (
            tmp.groupby("month", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                winning_trades=("is_exit_pnl_positive", "sum"),
            )
        )
        monthwise_summary["losing_trades"] = monthwise_summary["trades"] - monthwise_summary["winning_trades"]
        monthwise_summary["win_rate_pct"] = (
                100.0 * monthwise_summary["winning_trades"] / monthwise_summary["trades"]
        ).round(2)

        # New: daily PnL inside each month
        daily_tmp = (
            tmp.groupby(["month", "day"], as_index=False)
            .agg(daily_pnl=("exit_pnl", "sum"))
        )

        loss_day_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                avg_loss_on_loss_days=(
                    "daily_pnl",
                    lambda s: float(s[s < 0].mean()) if (s < 0).any() else 0.0
                ),
                max_loss_in_a_day=(
                    "daily_pnl",
                    lambda s: float(s.min()) if len(s) else 0.0
                ),
            )
        )

        # Date on which the worst (maximum-loss) day occurred, per month
        worst_rows = daily_tmp.loc[daily_tmp.groupby("month")["daily_pnl"].idxmin()]
        worst_day = worst_rows[["month", "day"]].rename(columns={"day": "max_loss_day_date"})

        monthwise_summary = monthwise_summary.merge(loss_day_stats, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(worst_day, on="month", how="left")

        # Place the date column right after the max-loss value column
        _cols = list(monthwise_summary.columns)
        if "max_loss_day_date" in _cols and "max_loss_in_a_day" in _cols:
            _cols.remove("max_loss_day_date")
            _cols.insert(_cols.index("max_loss_in_a_day") + 1, "max_loss_day_date")
            monthwise_summary = monthwise_summary[_cols]
    else:
        monthwise_summary = pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as xw:
        all_trades_df.to_excel(xw, sheet_name="all_trades_backtested", index=False)
        actual_trades_df.to_excel(xw, sheet_name="actual_trades", index=False)
        monthwise_summary.to_excel(xw, sheet_name="monthwise_summary", index=False)
        piv_exit.to_excel(xw, sheet_name="exit_pnl_pivot", index=False)
        piv_eod_first.to_excel(xw, sheet_name="eod_pnl_first_trade_pivot", index=False)
        instrument_summary.to_excel(xw, sheet_name="instrument_summary", index=False)
        skipped_df.to_excel(xw, sheet_name="skipped", index=False)

        wb = xw.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _autosize_columns_safe(ws)

        _color_actual_trades_by_date(wb, actual_trades_df)

    print(f"[DONE] Excel written: {OUTPUT_XLSX}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")

    print(f"[INFO] Pickles found: {len(paths)}")

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)

    lookback_label = "AUTO/full pickle range" if LOOKBACK_MONTHS is None else f"{LOOKBACK_MONTHS} months cap"

    print(f"[INFO] Data day-range seen: {min_day_seen} -> {end_day}")
    print(f"[INFO] Backtest window: {window_start} -> {end_day} ({lookback_label})")
    print(f"[INFO] Stoploss %/attempt: {_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Per-attempt stop cap: Rs {_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Daily max loss: Rs {_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES)} | "
          f"ProfitProtect trigger/giveback %: {_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES)} | "
          f"Re-entry delay min/attempt: {REENTRY_DELAY_BY_ATTEMPT} | Allowed DTE: {ALLOWED_DTE}")
    print(f"[INFO] Day profit target: {PROFIT_TARGET_PCT:.0%} of premium (0 = disabled)")
    print(f"[INFO] Tradeables: {sorted(TRADEABLE)}")
    print(f"[INFO] Output: {OUTPUT_XLSX}")

    print("[STEP] Initializing Kite ...")
    kite = oUtils.intialize_kite_api()
    print("[OK] Kite ready.")

    underlying_data = download_underlyings(kite, window_start, end_day)

    all_trades_df, skipped_df = process_pickles_generate_trades(
        default_params(), paths, min_expiry_map, underlying_data, window_start, end_day
    )

    actual_trades_df = build_actual_trades_df(all_trades_df, min_expiry_map)

    write_excel(all_trades_df, actual_trades_df, skipped_df)

    if not all_trades_df.empty:
        print(all_trades_df.groupby("underlying")[["exit_pnl"]].describe())
    else:
        print("[WARN] No completed trades. Check 'skipped' sheet for reasons.")


# =============================================================================
# ROBUSTNESS OPTIMIZER  (Optuna TPE + optional walk-forward CV)
# =============================================================================
# Goal: maximize ROBUSTNESS, defined primarily as the fraction of PROFITABLE
# MONTHS (then profitable days), not raw total PnL. A config that earns a little
# every month is preferred over one that makes a fortune in two months and
# bleeds the rest.
#
# Tunables exposed to the optimizer (the 6 requested), with the variable-length
# per-attempt lists parameterized compactly as (base, step) so the search space
# stays low-dimensional and the per-attempt schedule stays monotone/sensible:
#   ENTRY_TIME_IST                -> minutes-from-open (discrete grid)
#   LOSS_LIMIT_RUPEES_BY_ATTEMPT  -> sl_base_pct + n*sl_step_pct  (rising stops)
#   PROFIT_PROTECT_TRIGGER_RUPEES -> single pct (0 disables)
#   MAX_REATTEMPTS                -> int
#   PROFIT_TARGET_PCT             -> single pct (0 disables)
#   REENTRY_DELAY_BY_ATTEMPT      -> delay_base + n*delay_step (minutes)
# =============================================================================

# ---- objective weights / guards (override via env if desired) ----
OPT_W_MONTH = _parse_float_env("OPT_W_MONTH", 1.0)    # weight on profitable-month ratio
OPT_W_DAY = _parse_float_env("OPT_W_DAY", 0.30)       # weight on profitable-day ratio
OPT_W_PNL = _parse_float_env("OPT_W_PNL", 0.10)       # tiny tie-break toward positive PnL
OPT_CV_PENALTY = _parse_float_env("OPT_CV_PENALTY", 0.50)  # penalty * std across CV folds
OPT_MIN_DAYS = int(_parse_float_env("OPT_MIN_DAYS", 30))   # guard: need enough data
OPT_MIN_MONTHS = int(_parse_float_env("OPT_MIN_MONTHS", 3))


def _inr(x: float) -> str:
    """Format a rupee amount with Indian digit grouping, ASCII-safe for Windows
    consoles (e.g. 1234567 -> 'Rs.12,34,567'). Avoids the unicode rupee sign so
    it never throws UnicodeEncodeError in a cp1252 terminal."""
    try:
        n = int(round(float(x)))
    except (ValueError, TypeError):
        return "Rs.0"
    sign = "-" if n < 0 else ""
    s = str(abs(n))
    if len(s) <= 3:
        body = s
    else:
        last3, rest = s[-3:], s[:-3]
        parts = []
        while len(rest) > 2:               # group remaining digits in pairs (Indian style)
            parts.insert(0, rest[-2:]); rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        body = ",".join(parts) + "," + last3
    return f"Rs.{sign}{body}"


def robustness_metrics(actual_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Per-day and per-month profitability stats from an actual_trades frame.

    Returns the headline robustness ratios PLUS rupee P/L summaries used in the
    per-trial print: net P/L, mean monthly P/L, median monthly P/L, worst month,
    and the full month->P/L series (`monthly`) for the best-trial breakdown.
    """
    if actual_df is None or actual_df.empty:
        return {"n_days": 0, "n_months": 0, "prof_day_ratio": 0.0,
                "prof_month_ratio": 0.0, "total_pnl": 0.0, "mean_month": 0.0,
                "median_month": 0.0, "worst_day": 0.0, "worst_month": 0.0,
                "monthly": pd.Series(dtype="float64")}
    t = actual_df.copy()
    d = pd.to_datetime(t["day"])
    daily = t.groupby(d.dt.date)["exit_pnl"].sum()
    monthly = t.groupby(d.dt.to_period("M"))["exit_pnl"].sum().sort_index()
    return {
        "n_days": int(len(daily)),
        "n_months": int(len(monthly)),
        "prof_day_ratio": float((daily > 0).mean()),
        "prof_month_ratio": float((monthly > 0).mean()),
        "total_pnl": float(daily.sum()),          # NET P/L across the whole sample
        "mean_month": float(monthly.mean()),      # MEAN monthly P/L
        "median_month": float(monthly.median()),  # MEDIAN monthly P/L
        "worst_day": float(daily.min()),
        "worst_month": float(monthly.min()),
        "monthly": monthly,                       # full month -> P/L (for breakdown)
    }


def _score_from_metrics(m: Dict[str, float]) -> float:
    """Single-sample robustness score. Guards against thin data."""
    if m["n_days"] < OPT_MIN_DAYS or m["n_months"] < OPT_MIN_MONTHS:
        return -1.0
    pnl_norm = m["total_pnl"] / (abs(m["total_pnl"]) + 1_000_000.0)  # bounded (-1, 1)
    return (OPT_W_MONTH * m["prof_month_ratio"]
            + OPT_W_DAY * m["prof_day_ratio"]
            + OPT_W_PNL * pnl_norm)


def _cv_score(actual_df: pd.DataFrame, folds: int) -> float:
    """
    Walk-forward-style robustness: split the months into `folds` CONTIGUOUS time
    blocks and reward configs that stay profitable in EVERY block, not just on
    average. Score = mean(block profitable-month ratio) - penalty*std. This
    resists overfitting to one lucky regime.
    """
    if actual_df is None or actual_df.empty:
        return -1.0
    t = actual_df.copy()
    d = pd.to_datetime(t["day"])
    monthly = t.groupby(d.dt.to_period("M"))["exit_pnl"].sum().sort_index()
    if len(monthly) < max(OPT_MIN_MONTHS, folds):
        return -1.0
    # global activity guard
    daily = t.groupby(d.dt.date)["exit_pnl"].sum()
    if len(daily) < OPT_MIN_DAYS:
        return -1.0

    months = list(monthly.items())
    block_ratios: List[float] = []
    n = len(months)
    for k in range(folds):
        lo = (k * n) // folds
        hi = ((k + 1) * n) // folds
        block = months[lo:hi]
        if not block:
            continue
        wins = sum(1 for _, v in block if v > 0)
        block_ratios.append(wins / len(block))
    if not block_ratios:
        return -1.0
    s = pd.Series(block_ratios)
    return float(s.mean() - OPT_CV_PENALTY * s.std(ddof=0))


def _params_from_trial(trial, base: "Params") -> "Params":
    """Map an Optuna trial to a Params object. Lists are built from (base, step)."""
    # ENTRY_TIME: discrete grid 09:20..13:30 in 5-min steps
    entry_min = trial.suggest_int("entry_minute_from_0920", 0, 250, step=5)
    eh, em = divmod(9 * 60 + 20 + entry_min, 60)
    entry_time = dtime(eh, em)

    # rising per-attempt stop-loss schedule
    sl_base = trial.suggest_float("sl_base_pct", 0.05, 0.40)
    sl_step = trial.suggest_float("sl_step_pct", 0.0, 0.10)

    max_reattempts = trial.suggest_int("max_reattempts", 0, 10)
    profit_protect = trial.suggest_float("profit_protect_pct", 0.0, 0.80)
    profit_target = trial.suggest_float("profit_target_pct", 0.20, 0.95)

    # rising per-attempt re-entry delay schedule (minutes)
    delay_base = trial.suggest_int("reentry_delay_base_min", 1, 20)
    delay_step = trial.suggest_int("reentry_delay_step_min", 0, 10)

    n_slots = max_reattempts + 1
    sl_list = [round(min(0.95, sl_base + i * sl_step), 4) for i in range(n_slots)]
    delay_list = [int(delay_base + i * delay_step) for i in range(n_slots)]

    return Params(
        entry_time=entry_time,
        loss_limit_pct_by_attempt=sl_list,
        profit_protect_pct=float(profit_protect),
        max_reattempts=int(max_reattempts),
        profit_target_pct=float(profit_target),
        reentry_delay_by_attempt=delay_list,
        max_daily_loss_rupees=base.max_daily_loss_rupees,
        max_loss_limit_cap_rupees=base.max_loss_limit_cap_rupees,
    )


def optimize(
    groups: List[DayGroup],
    min_expiry_map: Dict[Tuple[str, date], date],
    n_trials: int = 100,
    cv_folds: int = 1,
    seed: int = 42,
    progress_every: int = 5,   # print a progress line every this many trials
):
    """
    Run the Optuna study over pre-built groups.

    cv_folds <= 1 : score on the full sample (profitable-month-first).
    cv_folds  > 1 : contiguous-block walk-forward robustness score.

    Progress: a callback prints after every `progress_every` trials, showing this
    trial's score, the best-so-far score, the best config's profitable-month
    ratio, and elapsed/ETA so a long search is never silent.
    """
    import optuna
    import time as _time
    optuna.logging.set_verbosity(optuna.logging.WARNING)  # we do our own printing

    base = default_params()  # supplies the non-optimized fields (daily cap, stop cap)

    def objective(trial):
        # 1) turn the trial's suggestions into a concrete Params object
        params = _params_from_trial(trial, base)
        # 2) re-simulate all cached day-groups for this parameter set (the cheap part)
        all_df, _ = simulate_groups(params, groups)
        # 3) reduce to the actually-traded book and measure robustness
        actual_df = build_actual_trades_df(all_df, min_expiry_map)
        m = robustness_metrics(actual_df)
        # 4) stash SCALAR diagnostics so the progress callback can print them
        #    (Optuna user_attrs must be JSON-serializable -> no pandas objects).
        for k in ("n_days", "n_months", "prof_day_ratio", "prof_month_ratio",
                  "total_pnl", "mean_month", "median_month", "worst_day", "worst_month"):
            trial.set_user_attr(k, m[k])
        # month -> P/L as a plain dict so we can print a month-wise breakdown
        trial.set_user_attr("monthly_pnl", {str(p): float(v) for p, v in m["monthly"].items()})
        # 5) the score Optuna maximizes
        if cv_folds and cv_folds > 1:
            return _cv_score(actual_df, cv_folds)
        return _score_from_metrics(m)

    study = optuna.create_study(direction="maximize",
                                sampler=optuna.samplers.TPESampler(seed=seed))

    # ---- live progress callback ----
    # Prints a rich stats line on EVERY trial: the score, net P/L, mean & median
    # monthly P/L, profitable-month and profitable-day ratios, and the worst
    # month -- plus elapsed/ETA. Whenever a new best appears, it also prints the
    # full month-by-month P/L of that best config so you can eyeball consistency.
    start = _time.time()

    def _progress(study_, trial_):
        n_done = trial_.number + 1
        ua = trial_.user_attrs
        val = trial_.value if trial_.value is not None else float("nan")
        elapsed = _time.time() - start
        eta = (elapsed / n_done) * (n_trials - n_done)
        try:
            best_val = study_.best_value
            best_num = study_.best_trial.number
        except Exception:
            best_val, best_num = float("nan"), -1

        n_mo = int(ua.get("n_months", 0))
        pmr = float(ua.get("prof_month_ratio", 0.0))
        prof_mo = int(round(pmr * n_mo))
        print(
            f"[TRIAL {n_done:>4}/{n_trials}] score={val:+.4f} "
            f"net={_inr(ua.get('total_pnl', 0))} "
            f"mean/mo={_inr(ua.get('mean_month', 0))} "
            f"med/mo={_inr(ua.get('median_month', 0))} "
            f"prof_mo={pmr*100:3.0f}%({prof_mo}/{n_mo}) "
            f"prof_day={float(ua.get('prof_day_ratio', 0.0))*100:3.0f}% "
            f"worst_mo={_inr(ua.get('worst_month', 0))} "
            f"| best={best_val:+.4f} | {elapsed:5.0f}s eta={eta:6.0f}s",
            flush=True,
        )

        # New best -> print the month-wise P/L breakdown of the best config.
        if trial_.number == best_num:
            mp = ua.get("monthly_pnl", {})
            if mp:
                cells = [f"{k}:{_inr(v)}" for k, v in sorted(mp.items())]
                # wrap a few months per line so it stays readable
                print("   >>> NEW BEST -- month-wise net P/L:", flush=True)
                for i in range(0, len(cells), 4):
                    print("       " + "   ".join(cells[i:i + 4]), flush=True)

    print(f"[OPT] starting {n_trials} trials over {len(groups)} day-groups "
          f"(cv_folds={cv_folds}) ...", flush=True)
    study.optimize(objective, n_trials=n_trials, callbacks=[_progress], show_progress_bar=False)

    best = study.best_trial
    print("\n================ BEST CONFIG ================")
    print(f"score={best.value:.4f}  (cv_folds={cv_folds})")
    bp = _params_from_trial(_FrozenTrialView(best), base)
    print(f"ENTRY_TIME_IST              = {bp.entry_time.strftime('%H:%M')}")
    print(f"LOSS_LIMIT_RUPEES_BY_ATTEMPT= {[round(x,4) for x in bp.loss_limit_pct_by_attempt]}")
    print(f"PROFIT_PROTECT_TRIGGER      = {bp.profit_protect_pct:.4f}")
    print(f"MAX_REATTEMPTS              = {bp.max_reattempts}")
    print(f"PROFIT_TARGET_PCT           = {bp.profit_target_pct:.4f}")
    print(f"REENTRY_DELAY_BY_ATTEMPT    = {bp.reentry_delay_by_attempt}")
    print("---- robustness of best (full sample) ----")
    ba = best.user_attrs
    print(f"  net P/L            = {_inr(ba.get('total_pnl', 0))}")
    print(f"  mean monthly P/L   = {_inr(ba.get('mean_month', 0))}")
    print(f"  median monthly P/L = {_inr(ba.get('median_month', 0))}")
    print(f"  worst month        = {_inr(ba.get('worst_month', 0))}")
    n_mo = int(ba.get('n_months', 0))
    pmr = float(ba.get('prof_month_ratio', 0.0))
    print(f"  profitable months  = {int(round(pmr*n_mo))}/{n_mo} ({pmr*100:.1f}%)")
    print(f"  profitable days    = {float(ba.get('prof_day_ratio', 0.0))*100:.1f}% of {int(ba.get('n_days', 0))} days")
    mp = ba.get("monthly_pnl", {})
    if mp:
        print("  month-wise net P/L:")
        for k, v in sorted(mp.items()):
            flag = "" if v > 0 else "   <-- loss"
            print(f"     {k}: {_inr(v)}{flag}")

    # Ready-to-paste config block so you can drop the winner straight into
    # the RUN CONTROL / param section for a confirmation backtest.
    print("\n---- paste into your single-run params to verify ----")
    print(f"ENTRY_TIME_IST = \"{bp.entry_time.strftime('%H:%M')}\"")
    print(f"# LOSS_LIMIT schedule (per attempt): {[round(x,4) for x in bp.loss_limit_pct_by_attempt]}")
    print(f"# PROFIT_PROTECT_TRIGGER_RUPEES (pct): {bp.profit_protect_pct:.4f}")
    print(f"# MAX_REATTEMPTS: {bp.max_reattempts}")
    print(f"# PROFIT_TARGET_PCT: {bp.profit_target_pct:.4f}")
    print(f"# REENTRY_DELAY_BY_ATTEMPT (min): {bp.reentry_delay_by_attempt}")
    return study, bp


class _FrozenTrialView:
    """Lets us rebuild Params from a finished trial's params via suggest_* calls."""
    def __init__(self, trial):
        self._p = dict(trial.params)

    def suggest_int(self, name, *a, **k):
        return int(self._p[name])

    def suggest_float(self, name, *a, **k):
        return float(self._p[name])


def run_optimizer(n_trials: int, cv_folds: int,
                  max_pickles: Optional[int] = None,
                  max_days: Optional[int] = None,
                  progress_every: int = 5,
                  seed: int = 42):
    """
    End-to-end optimizer entrypoint: load data ONCE (pickles + Kite underlyings),
    cache the day-groups, then run the Optuna search over them.

    The phases are printed as banners so you always know where it is:
        [PHASE 1] scan pickles      [PHASE 2] download underlyings
        [PHASE 3] build day-groups  [PHASE 4] optimize
    """
    print("[PHASE 1] Scanning pickles for date range and nearest expiries ...", flush=True)
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")
    if max_pickles:
        # scan only the same subset we will load, so the window matches the sample
        paths = paths[:max_pickles]
    print(f"[PHASE 1] {len(paths)} pickle file(s) in scope", flush=True)

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)
    print(f"[PHASE 1] window: {window_start} -> {end_day}", flush=True)

    print("[PHASE 2] Initializing Kite and downloading underlyings ...", flush=True)
    kite = oUtils.intialize_kite_api()
    underlying_data = download_underlyings(kite, window_start, end_day)

    print("[PHASE 3] Building (and caching) day-groups ...", flush=True)
    groups, _ = build_day_groups(paths, min_expiry_map, underlying_data,
                                 window_start, end_day,
                                 max_pickles=max_pickles, max_days=max_days)
    if not groups:
        raise RuntimeError("No day-groups built; nothing to optimize. Check window / pickles.")

    print(f"[PHASE 4] Optimizing: {n_trials} trials, cv_folds={cv_folds} ...", flush=True)
    return optimize(groups, min_expiry_map, n_trials=n_trials, cv_folds=cv_folds,
                    progress_every=progress_every, seed=seed)


# =============================================================================
# ENTRYPOINT
# =============================================================================
# =============================================================================
# ENTRYPOINT  -- behaviour is driven entirely by the RUN CONTROL block at the
# top of this file. Just press Run in PyCharm; no command-line arguments.
# =============================================================================
if __name__ == "__main__":
    if RUN_MODE == "optimize":
        run_optimizer(
            n_trials=OPT_TRIALS,
            cv_folds=OPT_CV_FOLDS,
            max_pickles=SAMPLE_MAX_PICKLES,
            max_days=SAMPLE_MAX_DAYS,
            progress_every=OPT_PROGRESS_EVERY,
            seed=OPT_SEED,
        )
    elif RUN_MODE == "backtest":
        main()
    else:
        raise SystemExit(f"Unknown RUN_MODE={RUN_MODE!r}; set it to 'backtest' or 'optimize'.")