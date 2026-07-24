import os
import hashlib
from pathlib import Path
import glob
import time
from dataclasses import dataclass
from datetime import datetime, date, time as dtime, timedelta
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore

try:
    import pytz  # type: ignore
except Exception:
    pytz = None  # type: ignore

try:
    from dateutil.relativedelta import relativedelta  # type: ignore
except Exception:
    relativedelta = None  # type: ignore


# =============================================================================
# USER CONFIG
# =============================================================================
# ---------------------------------------------------------------------------
# CONFIGURATION SOURCE: external property file
# ---------------------------------------------------------------------------
# Every tunable setting lives in a simple KEY=VALUE property file so it can be
# changed WITHOUT editing this script. Path defaults to
# "straddle_config_dte0_underlying_optimized.properties" next to this file;
# override with the STRADDLE_CONFIG environment variable. Values are pushed
# into the process environment so all os.getenv(...) reads below pick them up. A real
# environment variable that is already set takes precedence over the file.
def _load_property_file() -> str:
    cfg_path = os.getenv(
        "STRADDLE_CONFIG",
        os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "straddle_config_dte0_underlying_optimized.properties",
        ),
    )
    if not os.path.exists(cfg_path):
        print(f"[CONFIG] Property file not found at {cfg_path}; using built-in defaults.")
        return cfg_path
    loaded = 0
    with open(cfg_path, "r", encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith(";"):
                continue
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            key, val = key.strip(), val.strip()
            # strip optional surrounding quotes
            if len(val) >= 2 and val[0] == val[-1] and val[0] in ("'", '"'):
                val = val[1:-1]
            if key and key not in os.environ:   # real env vars win over the file
                os.environ[key] = val
                loaded += 1
    print(f"[CONFIG] Loaded {loaded} setting(s) from {cfg_path}")
    return cfg_path

PROPERTY_FILE_PATH = _load_property_file()
BUILD_ID = "dte0-dhan-compatible-v3.1-20260723"

# PICKLES_DIR = r"G:\My Drive\Trading\Historical_Options_Data"
PICKLES_DIR = os.getenv("PICKLES_DIR", r"G:\My Drive\Trading\Dhan_Historical_Options_Data_New")
ENTRY_TIME_IST = os.getenv("ENTRY_TIME_IST", "11:55")  # "HH:MM"
# EXIT_TIME_IST is the strategy time filter / square-off cutoff.
# No fresh entry or re-entry is initiated at or after this time. If an attempt is
# still open at this time, it is closed at the cutoff with exit_reason="TIME_EXIT".
EXIT_TIME_IST = os.getenv("EXIT_TIME_IST", "15:30")  # "HH:MM"

def _safe_fname_part(s: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in s)

def _get_downloads_folder() -> str:
    """
    Returns the current user's default Downloads folder.
    Falls back to home directory if Downloads is not found.
    """
    downloads = Path.home() / "Downloads"
    return str(downloads if downloads.exists() else Path.home())

# --- Generic integer-list parser used for DTE and re-entry delay settings ---
def _parse_int_list(env_val, default_list):
    """
    Parse a comma-separated integer list from an environment variable.

    Examples:
        ALLOWED_DTE="0"                -> [0]
        REENTRY_DELAY_BY_ATTEMPT="1,5" -> [1, 5]

    If parsing fails or the env var is blank, the supplied default is used.
    """
    if env_val:
        try:
            vals = [int(round(float(x))) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass
    return list(default_list)


def _parse_pct_value(x) -> float:
    """
    Parse a percentage value into decimal form.

    Accepted user formats:
        10      -> 0.10  (10%)
        "10%"   -> 0.10  (10%)
        0.10    -> 0.10  (10%)
        "0.10"  -> 0.10  (10%)

    This lets you configure percentages naturally from environment variables
    while keeping calculations internally consistent.
    """
    s = str(x).strip().replace("%", "")
    if s == "":
        raise ValueError("blank percentage value")

    v = float(s)

    # If user entered 10, treat it as 10%; if user entered 0.10, keep it as 10%.
    if abs(v) > 1.0:
        v = v / 100.0

    if v < 0:
        raise ValueError("percentage cannot be negative")

    return float(v)


def _parse_pct_list(env_val, default_list):
    """
    Parse comma-separated percentages into decimal form.

    Examples:
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15" -> [0.10, 0.12, 0.15]
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10"     -> [0.10]
    """
    if env_val:
        try:
            vals = [_parse_pct_value(x) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass

    return [_parse_pct_value(x) for x in default_list]


def _fmt_int_list(lst) -> str:
    """Format integer-list settings for the output filename."""
    return "-".join(str(int(v)) for v in lst) if lst else "off"


def _fmt_pct_value(v: float) -> str:
    """Format a decimal percentage such as 0.10 as '10pct' for filenames/logs."""
    return f"{v * 100:.2f}".rstrip("0").rstrip(".") + "pct"


def _fmt_pct_list(lst) -> str:
    """Format a list of decimal percentages for the output filename."""
    return "-".join(_fmt_pct_value(float(v)) for v in lst) if lst else "off"


def _parse_float_env(env_name: str, default_value: float) -> float:
    """
    Parse a positive/zero floating-point rupee setting from an environment variable.

    If parsing fails, the supplied default is used. A value <= 0 is treated by
    the relevant logic as disabled where applicable.
    """
    raw = os.getenv(env_name)
    if raw is None or str(raw).strip() == "":
        return float(default_value)
    try:
        return float(str(raw).replace(",", "").strip())
    except Exception:
        return float(default_value)


def _fmt_rupee_value(v: float) -> str:
    """Format rupee values compactly for output filenames/logs."""
    return str(int(round(float(v))))


LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_pct_list(
    os.getenv("LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
    [0.2487, 0.2824, 0.3162, 0.3499, 0.3837, 0.4174, 0.4512],
)


def loss_limit_pct_for_attempt(attempt_idx: int) -> float:
    """Return the stop-loss percentage, in decimal form, for the given attempt."""
    s = LOSS_LIMIT_RUPEES_BY_ATTEMPT
    if not s:
        return 0.0
    return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])


# --- Allowed days-to-expiry to trade. This optimized build is DTE-0 only. ---
ALLOWED_DTE = _parse_int_list(os.getenv("ALLOWED_DTE"), [0])
STRICT_DTE0_ONLY = os.getenv("STRICT_DTE0_ONLY", "1").strip() != "0"
if STRICT_DTE0_ONLY and sorted(set(ALLOWED_DTE)) != [0]:
    raise ValueError(
        "DTE-0-only scope violation: ALLOWED_DTE must contain exactly 0 "
        "when STRICT_DTE0_ONLY=1"
    )

def _parse_optional_expected_count(raw_value: Any) -> Optional[int]:
    text = str(raw_value if raw_value is not None else "").strip().upper()
    if text in ("", "0", "AUTO", "NONE", "OFF"):
        return None
    value = int(float(text))
    if value <= 0:
        return None
    return value


def _parse_optional_expected_signature(raw_value: Any) -> Optional[str]:
    text = str(raw_value if raw_value is not None else "").strip().lower()
    if text in ("", "0", "auto", "none", "off"):
        return None
    return text


EXPECTED_OPPORTUNITY_COUNT = _parse_optional_expected_count(
    os.getenv("EXPECTED_OPPORTUNITY_COUNT", "AUTO")
)
EXPECTED_OPPORTUNITY_SIGNATURE = _parse_optional_expected_signature(
    os.getenv("EXPECTED_OPPORTUNITY_SIGNATURE", "AUTO")
)
REQUIRE_FULL_OPPORTUNITY_COVERAGE = (
    os.getenv("REQUIRE_FULL_OPPORTUNITY_COVERAGE", "0").strip() == "1"
)

# --- Profit-protect threshold/giveback as % of premium collected on that attempt ---
# Default: 30%.
#
# Env examples:
#     PROFIT_PROTECT_TRIGGER_RUPEES="30"
#     PROFIT_PROTECT_TRIGGER_RUPEES="0.30"
#
# Current logic uses the same rupee amount for:
#     1. arming profit-protect once peak P&L reaches G
#     2. exiting when current P&L falls to peak - G
PROFIT_PROTECT_TRIGGER_RUPEES = _parse_pct_value(os.getenv("PROFIT_PROTECT_TRIGGER_RUPEES", 0.254741))

# --- Absolute daily circuit breaker -------------------------------------------------
# Once cumulative realized NET P&L for the current underlying/day reaches this
# loss, no further re-entry is allowed for that day.
#
# Default: Rs. 30,000 loss. Set MAX_DAILY_LOSS_RUPEES=0 to disable.
MAX_DAILY_LOSS_RUPEES = _parse_float_env("MAX_DAILY_LOSS_RUPEES", 30000.0)

# --- Absolute cap on the percentage-based per-attempt stop-loss ----------------------
# The stop-loss is still calculated as:
#     LOSS_LIMIT_% * entry_premium_sum
# But it is capped at this absolute rupee value.
#
# Effective stop-loss per attempt:
#     min(LOSS_LIMIT_% * entry_premium_sum, MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
#
# Default: Rs. 3,000 loss per attempt. Set to 0 to disable the cap.
MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_float_env("MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT", 3000.0)

MAX_REATTEMPTS = int(os.getenv("MAX_REATTEMPTS", "10"))  # 1 = only one re-entry

# --- Per-DAY profit target as a fraction of premium collected on the CURRENT attempt ---
# When an attempt's profit reaches PROFIT_TARGET_PCT * (CE+PE)*qty, it exits at the
# target and NO further trades are taken that day. 0 disables. e.g. 0.70 = 70%.
PROFIT_TARGET_PCT = float(os.getenv("PROFIT_TARGET_PCT", 0.7))
# --- Per-attempt RE-ENTRY GAP in minutes (index 0 = gap before 1st re-entry, 1 = before 2nd, ...) ---
# Attempts beyond the list reuse the LAST value. Override via env comma list, e.g.
# REENTRY_DELAY_BY_ATTEMPT="10,15,20".
REENTRY_DELAY_BY_ATTEMPT = _parse_int_list(
    os.getenv("REENTRY_DELAY_BY_ATTEMPT"),
    [6, 8, 10, 12, 14, 16, 18, 15, 15, 15, 15],
)

def reentry_delay_for_attempt(attempt_idx: int) -> int:
    s = REENTRY_DELAY_BY_ATTEMPT
    if not s:
        return 0
    return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])

_DEFAULT_OUT = os.path.join(
    _get_downloads_folder(),
    "short_straddle_dte0_underlying_optimized.xlsx",
)

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", _DEFAULT_OUT)

FAIL_ON_PICKLE_ERROR = os.getenv("FAIL_ON_PICKLE_ERROR", "0").strip() == "1"

SESSION_START_IST = dtime(9, 15)
SESSION_END_IST = dtime(15, 30)

# LOOKBACK_MONTHS is AUTO by default: the script backtests the full date range
# present in the option pickles. If you explicitly set LOOKBACK_MONTHS to a
# number, that number acts as an optional manual cap.
LOOKBACK_MONTHS_RAW = os.getenv("LOOKBACK_MONTHS", "AUTO").strip()
LOOKBACK_MONTHS: Optional[int]
if LOOKBACK_MONTHS_RAW.upper() in ("", "AUTO", "ALL", "MAX", "FULL"):
    LOOKBACK_MONTHS = None
else:
    LOOKBACK_MONTHS = int(float(LOOKBACK_MONTHS_RAW))

QTY_UNITS = {"NIFTY": 325, "SENSEX": 100}
TRADEABLE = set(QTY_UNITS.keys())

STRIKE_STEP = {"NIFTY": 50, "SENSEX": 100}

# =============================================================================
# TRANSACTION CHARGES (Zerodha F&O Options — NSE)
# =============================================================================
# Each short-straddle attempt = 4 executed orders (sell CE, sell PE, buy CE, buy PE)
BROKERAGE_PER_ORDER       = 20.0       # ₹20 flat per executed order
ORDERS_PER_TRADE          = 4          # sell CE + sell PE + buy CE + buy PE
STT_SELL_PCT              = 0.001      # 0.1% on sell-side premium
EXCHANGE_TXN_PCT          = 0.0003553  # 0.03553% on premium (NSE options)
SEBI_PER_CRORE            = 10.0       # ₹10 per crore of turnover
STAMP_BUY_PCT             = 0.00003    # 0.003% on buy-side premium
IPFT_PER_CRORE            = 0.010      # ₹0.01 per crore (on premium)
GST_PCT                   = 0.18       # 18% on (brokerage + txn charges + SEBI)
INCLUDE_TRANSACTION_COSTS = True       # set False to disable

UNDERLYING_KITE = {
    "NIFTY": {"exchange": "NSE", "tradingsymbol": "NIFTY 50"},
    "SENSEX": {"exchange": "BSE", "tradingsymbol": "SENSEX"},
}

MAX_DAYS_PER_CHUNK = 25
MAX_ATTEMPTS = 5
SLEEP_BETWEEN_CALLS_SEC = 0.20


# =============================================================================
# HELPERS
# =============================================================================
def parse_hhmm(s: str) -> dtime:
    hh, mm = s.strip().split(":")
    return dtime(int(hh), int(mm))

ENTRY_TIME = parse_hhmm(ENTRY_TIME_IST)
# Strategy cutoff time. The simulator will not initiate a new attempt at or
# after this time, and the current attempt will not be monitored beyond it.
EXIT_TIME = parse_hhmm(EXIT_TIME_IST)


@dataclass(frozen=True)
class UnderlyingStrategySettings:
    """Expiry-day execution and risk settings selected by underlying."""

    entry_time_ist: str
    exit_time_ist: str
    reentry_cutoff_time_ist: str
    entry_time: dtime
    exit_time: dtime
    reentry_cutoff_time: dtime
    loss_limit_pct_by_attempt: Tuple[float, ...]
    stop_cap_rupees: float
    profit_protect_pct: float
    profit_target_pct: float
    max_daily_loss_rupees: float
    max_reattempts: int
    reentry_delay_by_attempt: Tuple[int, ...]

    def loss_pct_for_attempt(self, attempt_idx: int) -> float:
        values = self.loss_limit_pct_by_attempt
        return float(values[min(attempt_idx, len(values) - 1)]) if values else 0.0

    def reentry_delay_for_attempt(self, attempt_idx: int) -> int:
        values = self.reentry_delay_by_attempt
        return int(values[min(attempt_idx, len(values) - 1)]) if values else 0


def _underlying_settings(prefix: str) -> UnderlyingStrategySettings:
    entry_ist = os.getenv(f"{prefix}_ENTRY_TIME_IST", ENTRY_TIME_IST)
    exit_ist = os.getenv(f"{prefix}_EXIT_TIME_IST", EXIT_TIME_IST)
    reentry_cutoff_ist = os.getenv(
        f"{prefix}_REENTRY_CUTOFF_TIME_IST",
        os.getenv("REENTRY_CUTOFF_TIME_IST", exit_ist),
    )
    return UnderlyingStrategySettings(
        entry_time_ist=entry_ist,
        exit_time_ist=exit_ist,
        reentry_cutoff_time_ist=reentry_cutoff_ist,
        entry_time=parse_hhmm(entry_ist),
        exit_time=parse_hhmm(exit_ist),
        reentry_cutoff_time=parse_hhmm(reentry_cutoff_ist),
        loss_limit_pct_by_attempt=tuple(
            _parse_pct_list(
                os.getenv(f"{prefix}_LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
                LOSS_LIMIT_RUPEES_BY_ATTEMPT,
            )
        ),
        stop_cap_rupees=_parse_float_env(
            f"{prefix}_MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT",
            MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT,
        ),
        profit_protect_pct=_parse_pct_value(
            os.getenv(
                f"{prefix}_PROFIT_PROTECT_TRIGGER_RUPEES",
                PROFIT_PROTECT_TRIGGER_RUPEES,
            )
        ),
        profit_target_pct=_parse_pct_value(
            os.getenv(f"{prefix}_PROFIT_TARGET_PCT", PROFIT_TARGET_PCT)
        ),
        max_daily_loss_rupees=_parse_float_env(
            f"{prefix}_MAX_DAILY_LOSS_RUPEES",
            MAX_DAILY_LOSS_RUPEES,
        ),
        max_reattempts=int(
            os.getenv(f"{prefix}_MAX_REATTEMPTS", str(MAX_REATTEMPTS))
        ),
        reentry_delay_by_attempt=tuple(
            _parse_int_list(
                os.getenv(f"{prefix}_REENTRY_DELAY_BY_ATTEMPT"),
                REENTRY_DELAY_BY_ATTEMPT,
            )
        ),
    )


NIFTY_SETTINGS = _underlying_settings("NIFTY")
SENSEX_SETTINGS = _underlying_settings("SENSEX")
SETTINGS_BY_UNDERLYING = {
    "NIFTY": NIFTY_SETTINGS,
    "SENSEX": SENSEX_SETTINGS,
}


def strategy_settings_for_trade(
    underlying: str,
    days_to_expiry: int,
) -> UnderlyingStrategySettings:
    if int(days_to_expiry) != 0:
        raise ValueError(
            f"DTE-0-only scope violation: received DTE={days_to_expiry} "
            f"for {underlying}"
        )
    return SETTINGS_BY_UNDERLYING[underlying]

def ist_tz():
    if ZoneInfo is not None:
        return ZoneInfo("Asia/Kolkata")
    if pytz is not None:
        return pytz.timezone("Asia/Kolkata")
    return "Asia/Kolkata"

def ensure_ist(series_or_scalar) -> Any:
    tz = ist_tz()
    dt = pd.to_datetime(series_or_scalar, errors="coerce")
    if isinstance(dt, pd.Series):
        if dt.dt.tz is None:
            return dt.dt.tz_localize(tz)
        return dt.dt.tz_convert(tz)
    if getattr(dt, "tzinfo", None) is None:
        return dt.tz_localize(tz)
    return dt.tz_convert(tz)

def normalize_underlying(name: str) -> Optional[str]:
    if not isinstance(name, str):
        return None
    u = name.upper().strip()
    if "SENSEX" in u:
        return "SENSEX"
    if "BANKNIFTY" in u or "NIFTY BANK" in u:
        return "BANKNIFTY"
    if "NIFTY" in u:
        return "NIFTY"
    return None

def round_to_step(x: float, step: int) -> int:
    return int(round(x / step) * step)

def build_minute_index(day_d: date, start_t: dtime, end_t: dtime) -> pd.DatetimeIndex:
    tz = ist_tz()
    start = pd.Timestamp(datetime.combine(day_d, start_t), tz=tz)
    end = pd.Timestamp(datetime.combine(day_d, end_t), tz=tz)
    return pd.date_range(start=start, end=end, freq="1min")

def asof_close(df: pd.DataFrame, ts: pd.Timestamp) -> float:
    if df.empty:
        return float("nan")
    d = df[["date", "close"]].dropna().copy()
    d["date"] = ensure_ist(d["date"])
    d = d.sort_values("date").set_index("date")
    loc = d.index.get_indexer([ts], method="pad")
    if loc[0] == -1:
        return float("nan")
    return float(d.iloc[loc[0]]["close"])

def compute_window_start(end_day: date, months: int) -> date:
    if relativedelta is not None:
        return (pd.Timestamp(end_day) - relativedelta(months=months)).date()
    return (pd.Timestamp(end_day) - pd.Timedelta(days=30 * months)).date()


def determine_backtest_window_start(min_day_seen: date, end_day: date) -> date:
    """
    Determine the backtest start date.

    Default behaviour: use the earliest usable option-data date found in the
    pickles, i.e. the maximum available backtest period.

    Optional override: if LOOKBACK_MONTHS is set to a numeric value through the
    environment, use the later of:
        1. earliest option-data date; and
        2. end_day - LOOKBACK_MONTHS
    so the script never requests data before the pickles actually start.
    """
    if LOOKBACK_MONTHS is None:
        return min_day_seen

    capped_start = compute_window_start(end_day, LOOKBACK_MONTHS)
    return max(min_day_seen, capped_start)

# =============================================================================
# TRANSACTION COST CALCULATOR
# =============================================================================
def compute_trade_charges(
    entry_ce: float, entry_pe: float,
    exit_ce: float, exit_pe: float,
    qty: int,
) -> float:
    """
    Compute total Zerodha transaction charges for one short-straddle attempt.

    Entry = SELL CE + SELL PE  (2 orders, sell side)
    Exit  = BUY  CE + BUY  PE (2 orders, buy side)

    Returns total charges in rupees (always positive).
    """
    if not INCLUDE_TRANSACTION_COSTS:
        return 0.0

    # Turnover values (in rupees)
    entry_turnover = (entry_ce + entry_pe) * qty   # sell side
    exit_turnover  = (exit_ce + exit_pe) * qty     # buy side
    total_turnover = entry_turnover + exit_turnover

    # 1. Brokerage: ₹20 × 4 orders
    brokerage = BROKERAGE_PER_ORDER * ORDERS_PER_TRADE

    # 2. STT: 0.1% on sell-side premium only (entry for short straddle)
    stt = entry_turnover * STT_SELL_PCT

    # 3. Exchange transaction charges: 0.03553% on both sides
    txn_charges = total_turnover * EXCHANGE_TXN_PCT

    # 4. SEBI charges: ₹10 per crore on total turnover
    sebi = total_turnover * SEBI_PER_CRORE / 1_00_00_000

    # 5. Stamp duty: 0.003% on buy side only (exit for short straddle)
    stamp = exit_turnover * STAMP_BUY_PCT

    # 6. IPFT: ₹0.01 per crore on premium (both sides)
    ipft = total_turnover * IPFT_PER_CRORE / 1_00_00_000

    # 7. GST: 18% on (brokerage + transaction charges + SEBI charges)
    gst = (brokerage + txn_charges + sebi) * GST_PCT

    total_charges = brokerage + stt + txn_charges + sebi + stamp + ipft + gst
    return round(total_charges, 2)

# =============================================================================
# Kite historical helpers
# =============================================================================
def _iter_chunks_by_date(from_dt: datetime, to_dt: datetime, days_per_chunk: int) -> List[Tuple[datetime, datetime]]:
    if from_dt > to_dt:
        raise ValueError("from_dt must be <= to_dt")
    chunks: List[Tuple[datetime, datetime]] = []
    cur = from_dt.date()
    end_d = to_dt.date()
    while cur <= end_d:
        chunk_end = min(cur + timedelta(days=days_per_chunk - 1), end_d)
        c_from = from_dt if cur == from_dt.date() else datetime.combine(cur, SESSION_START_IST)
        c_to = to_dt if chunk_end == end_d else datetime.combine(chunk_end, SESSION_END_IST)
        chunks.append((c_from, c_to))
        cur = chunk_end + timedelta(days=1)
    return chunks

def _kite_instruments_cached(kite, exchange: str, cache: Dict[str, List[Dict]]) -> List[Dict]:
    ex = exchange.upper().strip()
    if ex not in cache:
        print(f"[STEP] Loading instruments dump for {ex} ...")
        cache[ex] = kite.instruments(ex)
        print(f"[INFO] {ex} instruments: {len(cache[ex])}")
    return cache[ex]

def get_instrument_token(kite, exchange: str, tradingsymbol: str, cache: Dict[str, List[Dict]]) -> int:
    ex = exchange.upper().strip()
    wanted = tradingsymbol.strip().upper()
    for r in _kite_instruments_cached(kite, ex, cache):
        if str(r.get("tradingsymbol", "")).upper() == wanted:
            return int(r["instrument_token"])
    raise ValueError(f"Instrument not found on {ex}: '{tradingsymbol}'")

def fetch_history_minute(kite, instrument_token: int, from_dt: datetime, to_dt: datetime, label: str) -> List[Dict]:
    interval = "minute"
    chunks = _iter_chunks_by_date(from_dt, to_dt, MAX_DAYS_PER_CHUNK)
    rows_all: List[Dict] = []
    print(f"[INFO] Fetch {label} token={instrument_token} chunks={len(chunks)} {from_dt} -> {to_dt}")
    for i, (c_from, c_to) in enumerate(chunks, start=1):
        last_err = None
        for attempt in range(1, MAX_ATTEMPTS + 1):
            try:
                rows = kite.historical_data(
                    instrument_token=instrument_token,
                    from_date=c_from,
                    to_date=c_to,
                    interval=interval,
                    continuous=False,
                    oi=False,
                )
                rows_all.extend(rows)
                last_err = None
                break
            except Exception as e:
                last_err = e
                time.sleep(min(8.0, 1.5 * attempt))
        if last_err is not None:
            print(f"[ERROR] {label} chunk {i}/{len(chunks)} failed: {c_from}->{c_to}: {last_err}")
        time.sleep(SLEEP_BETWEEN_CALLS_SEC)
    return rows_all

def rows_to_df(rows: List[Dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["date", "open", "high", "low", "close", "volume"])
    df = pd.DataFrame(rows)
    df["date"] = ensure_ist(df["date"])
    df = df.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    return df


# =============================================================================
# DATA STRUCTURES
# =============================================================================
@dataclass
class TradeRow:
    day: date
    underlying: str
    trade_seq: int
    expiry: date
    days_to_expiry: int
    atm_strike: int
    qty_units: int
    entry_time: str
    exit_time: str
    exit_reason: str
    entry_underlying: float
    ce_symbol: str
    pe_symbol: str
    entry_ce: float
    entry_pe: float
    exit_ce: float
    exit_pe: float
    exit_pnl_gross: float   # P&L before charges
    txn_charges: float      # total transaction charges for this attempt
    exit_pnl: float         # net P&L after deducting charges
    eod_pnl: float
    max_profit: float
    max_loss: float
    max_profit_before_exit: float   # peak profit reached before this trade exited

    # Premium/risk diagnostics for percentage-based risk rules
    entry_premium_sum: float                 # (entry_ce + entry_pe) * qty
    stop_pct: float                          # stop-loss % of entry premium, decimal form; 0.10 = 10%
    uncapped_stop_rupees: float              # percentage-based stop before absolute cap
    stop_cap_rupees: float                   # configured absolute cap; <=0 means cap disabled
    stop_rupees: float                       # effective rupee stop after cap
    profit_protect_trigger_pct: float         # profit-protect % of entry premium, decimal form; 0.30 = 30%
    profit_protect_trigger_rupees: float      # computed rupee profit-protect trigger/giveback
    daily_realized_pnl_after_trade: float     # cumulative net P&L after this attempt
    daily_loss_limit_rupees: float            # configured daily loss circuit breaker
    daily_loss_limit_hit: bool                # True means no further trades for that day


# =============================================================================
# PASS-1: nearest expiry per (underlying, day)
# =============================================================================
def scan_pickles_pass1(pickle_paths: List[str]) -> Tuple[date, Dict[Tuple[str, date], date], date]:
    max_day_seen: Optional[date] = None
    min_day_seen: Optional[date] = None
    min_expiry_map: Dict[Tuple[str, date], date] = {}

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            for c in ("date", "name", "expiry", "type"):
                if c not in df.columns:
                    raise ValueError(f"Missing column '{c}' in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")]
            if d2.empty:
                continue

            d2 = d2[["date", "name", "expiry"]].copy()
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2 = d2.dropna(subset=["underlying", "day", "expiry_date"])

            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            file_min_day = d2["day"].min()
            file_max_day = d2["day"].max()
            max_day_seen = file_max_day if (max_day_seen is None or file_max_day > max_day_seen) else max_day_seen
            min_day_seen = file_min_day if (min_day_seen is None or file_min_day < min_day_seen) else min_day_seen

            grp = d2.groupby(["underlying", "day"], sort=False)["expiry_date"].min()
            for (und, dy), ex in grp.items():
                key = (und, dy)
                if key not in min_expiry_map or ex < min_expiry_map[key]:
                    min_expiry_map[key] = ex

            print(f"[PASS1 OK] {os.path.basename(p)} option_days={d2['day'].nunique()}")

        except Exception as e:
            if isinstance(e, ModuleNotFoundError):
                missing_module = getattr(e, "name", None) or str(e)
                raise RuntimeError(
                    f"Cannot deserialize {p}: Python module "
                    f"'{missing_module}' is missing. The Dhan pickles require "
                    "the same PyCharm interpreter/dependencies used to create "
                    "them (including pyarrow)."
                ) from e
            msg = f"[PASS1 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    if max_day_seen is None or min_day_seen is None:
        raise RuntimeError("No usable option data found in pickles (PASS1) for tradeable underlyings.")

    return max_day_seen, min_expiry_map, min_day_seen


# =============================================================================
# Underlying download
# =============================================================================
def download_underlyings(kite, day_start: date, day_end: date) -> Dict[str, pd.DataFrame]:
    cache: Dict[str, List[Dict]] = {}
    from_dt = datetime.combine(day_start, SESSION_START_IST)
    to_dt = datetime.combine(day_end, SESSION_END_IST)

    out: Dict[str, pd.DataFrame] = {}
    for und, meta in UNDERLYING_KITE.items():
        token = get_instrument_token(kite, meta["exchange"], meta["tradingsymbol"], cache)
        rows = fetch_history_minute(kite, token, from_dt, to_dt, label=f"{meta['exchange']}:{meta['tradingsymbol']}")
        df = rows_to_df(rows)
        df["day"] = df["date"].dt.tz_convert(ist_tz()).dt.date
        out[und] = df
        print(f"[UNDERLYING OK] {und}: candles={len(df)} days={df['day'].nunique()}")
    return out


def load_underlyings_from_pickles(
    pickle_paths: List[str],
    day_start: date,
    day_end: date,
) -> Dict[str, pd.DataFrame]:
    """
    Load the minute underlying candles already embedded in the supplied pickle
    files. This removes the external Kite dependency and keeps spot and option
    candles on the same source/timestamp convention.
    """
    chunks: Dict[str, List[pd.DataFrame]] = {und: [] for und in TRADEABLE}
    source_counts: Dict[str, Dict[str, int]] = {
        und: {"UNDERLYING rows": 0, "OPTION spot fallback": 0}
        for und in TRADEABLE
    }
    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            common_needed = {"date", "name", "type"}
            if not common_needed.issubset(df.columns):
                continue

            # Preferred schema: dedicated rows with type == UNDERLYING.
            underlying_mask = df["type"].astype(str).str.upper().eq("UNDERLYING")
            underlying_price_cols = {"open", "high", "low", "close", "volume"}
            if underlying_mask.any() and underlying_price_cols.issubset(df.columns):
                u = df.loc[
                    underlying_mask,
                    ["date", "name", "open", "high", "low", "close", "volume"],
                ].copy()
                source_label = "UNDERLYING rows"

            # Dhan A-compat schema: the pickle contains OPTION rows only, with
            # the synchronized underlying/index price repeated in `spot`.
            # The simulator only consumes the underlying close for ATM
            # selection, so derive one minute row per timestamp from `spot`.
            elif "spot" in df.columns:
                u = df.loc[
                    pd.to_numeric(df["spot"], errors="coerce").notna(),
                    ["date", "name", "spot"],
                ].copy()
                if not u.empty:
                    u["close"] = pd.to_numeric(u["spot"], errors="coerce")
                    # Some early Dhan files contain a different `spot` value
                    # on different option rows carrying the same minute
                    # timestamp. Never let option-row ordering choose ATM.
                    # The per-minute median is deterministic and resistant to
                    # isolated stale values.
                    u = (
                        u.dropna(subset=["date", "name", "close"])
                        .groupby(["date", "name"], as_index=False, sort=False)[
                            "close"
                        ]
                        .median()
                    )
                    u["open"] = u["close"]
                    u["high"] = u["close"]
                    u["low"] = u["close"]
                    u["volume"] = 0
                    u = u[
                        ["date", "name", "open", "high", "low", "close", "volume"]
                    ]
                source_label = "OPTION spot fallback"
            else:
                continue

            if u.empty:
                continue
            u["date"] = ensure_ist(u["date"])
            u["day"] = u["date"].dt.date
            u["underlying"] = u["name"].astype(str).map(normalize_underlying)
            u = u[
                u["underlying"].isin(TRADEABLE)
                & u["day"].between(day_start, day_end)
            ]
            for und, group in u.groupby("underlying", sort=False):
                und_key = str(und)
                minute_rows = (
                    group.drop(columns=["name", "underlying"])
                    .dropna(subset=["date", "close"])
                    .drop_duplicates(subset=["date"], keep="last")
                )
                if minute_rows.empty:
                    continue
                chunks[und_key].append(minute_rows)
                source_counts[und_key][source_label] += len(minute_rows)
        except Exception as exc:
            msg = f"[UNDERLYING WARN] {os.path.basename(p)} failed: {exc}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from exc
            print(msg)

    out: Dict[str, pd.DataFrame] = {}
    for und in sorted(TRADEABLE):
        if not chunks[und]:
            out[und] = pd.DataFrame(
                columns=["date", "open", "high", "low", "close", "volume", "day"]
            )
            continue
        frame = pd.concat(chunks[und], ignore_index=True)
        frame = (
            frame.drop_duplicates(subset=["date"], keep="last")
            .sort_values("date")
            .reset_index(drop=True)
        )
        out[und] = frame
        used_sources = ", ".join(
            f"{label}={count}"
            for label, count in source_counts[und].items()
            if count > 0
        )
        print(
            f"[UNDERLYING PICKLE OK] {und}: candles={len(frame)} "
            f"days={frame['day'].nunique()} sources[{used_sources}]"
        )
    return out


# =============================================================================
# Simulation helpers
# =============================================================================
def _pick_symbol(day_opt: pd.DataFrame, strike: int, opt_type: str) -> Optional[str]:
    sub = day_opt[(day_opt["strike_int"] == strike) & (day_opt["option_type"] == opt_type)]
    if sub.empty:
        return None
    syms = sorted(sub["instrument"].astype(str).unique().tolist())
    return syms[0] if syms else None

def _build_leg_series(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex,
                      strike: int, opt_type: str, symbol: str,
                      price_col: str = "close", do_ffill: bool = True) -> pd.Series:
    sub = day_opt[
        (day_opt["strike_int"] == strike) &
        (day_opt["option_type"] == opt_type) &
        (day_opt["instrument"].astype(str) == symbol)
    ][["date", price_col]].dropna()

    if sub.empty:
        return pd.Series(index=idx_all, dtype="float64")

    sub = sub.copy()
    sub["date"] = ensure_ist(sub["date"])
    sub = sub.sort_values("date").drop_duplicates(subset=["date"], keep="last").set_index("date")
    s = sub[price_col].astype(float).reindex(idx_all)
    return s.ffill() if do_ffill else s

def simulate_day_multi_trades(
    *,
    und: str,
    dy: date,
    expiry: date,
    day_opt: pd.DataFrame,
    underlying_day: pd.DataFrame,
) -> Tuple[List[TradeRow], List[Dict[str, Any]]]:

    results: List[TradeRow] = []
    skipped: List[Dict[str, Any]] = []

    idx_all = build_minute_index(dy, SESSION_START_IST, SESSION_END_IST)
    session_end_ts = idx_all[-1]
    dte = int((expiry - dy).days)
    settings = strategy_settings_for_trade(und, dte)

    # EXIT_TIME_IST is the strategy time filter.
    #
    # Earlier versions used EXIT_TIME_IST only as a "last new entry" gate while
    # still allowing the currently-open trade to run till SESSION_END_IST. That
    # made the Excel output look as if the filter was not working because trade
    # rows could still show exit_time after EXIT_TIME_IST.
    #
    # In this corrected version:
    #   1. no fresh entry/re-entry is initiated at or after EXIT_TIME_IST; and
    #   2. the active attempt is monitored only up to EXIT_TIME_IST.
    #
    # Therefore, when EXIT_TIME_IST is earlier than 15:30, the trade exits at the
    # cutoff if STOPLOSS / PROFIT_TARGET / PROFIT_PROTECT has not already fired.
    configured_exit_cutoff_ts = pd.Timestamp(
        datetime.combine(dy, settings.exit_time),
        tz=ist_tz(),
    )
    trade_end_ts = min(session_end_ts, configured_exit_cutoff_ts)
    configured_reentry_cutoff_ts = pd.Timestamp(
        datetime.combine(dy, settings.reentry_cutoff_time),
        tz=ist_tz(),
    )
    reentry_end_ts = min(trade_end_ts, configured_reentry_cutoff_ts)

    qty = int(QTY_UNITS[und])
    step = int(STRIKE_STEP[und])

    # Profit-protect is now percentage-based, so the actual rupee value is not
    # known until CE/PE entry prices are available for the current attempt.
    profit_protect_pct = float(settings.profit_protect_pct)
    profit_protect_enabled = profit_protect_pct > 0.0

    cur_entry_ts = pd.Timestamp(datetime.combine(dy, settings.entry_time), tz=ist_tz())
    trade_seq = 1

    # If the configured first entry itself is at/after EXIT_TIME_IST, the day
    # is skipped cleanly. This is intentional: EXIT_TIME_IST is the hard strategy
    # time filter, so a trade needs at least one minute of monitoring before it.
    if cur_entry_ts >= trade_end_ts:
        skipped.append({
            "day": dy,
            "underlying": und,
            "expiry": expiry,
            "trade_seq": trade_seq,
            "reason": (
                f"No entry: entry {settings.entry_time_ist} is at/after "
                f"configured exit {settings.exit_time_ist} for DTE={dte}"
            ),
        })
        return results, skipped

    # Cumulative realized NET P&L for this underlying/day. Used for the daily
    # loss circuit breaker. Charges are included through exit_pnl.
    daily_realized_pnl = 0.0
    daily_loss_limit_enabled = settings.max_daily_loss_rupees > 0

    while cur_entry_ts < trade_end_ts:
        if (
            daily_loss_limit_enabled
            and daily_realized_pnl <= -float(settings.max_daily_loss_rupees)
        ):
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq,
                "reason": (
                    f"Daily loss limit hit before next entry: "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={settings.max_daily_loss_rupees:.2f}"
                ),
            })
            break

        u_px = asof_close(underlying_day, cur_entry_ts)
        if pd.isna(u_px):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": f"No underlying price at entry {cur_entry_ts.strftime('%H:%M')}"})
            break

        atm = round_to_step(float(u_px), step)

        ce_sym = _pick_symbol(day_opt, atm, "CE")
        pe_sym = _pick_symbol(day_opt, atm, "PE")
        if not ce_sym or not pe_sym:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "ATM CE/PE not available in pickle band"})
            break

        # Close series (used for entry pricing, profit-protect tracking, and reporting)
        # Raw close series for exact entry validation
        ce_close_raw = _build_leg_series(day_opt, idx_all, atm, "CE", ce_sym, "close", do_ffill=False)
        pe_close_raw = _build_leg_series(day_opt, idx_all, atm, "PE", pe_sym, "close", do_ffill=False)

        # Forward-filled close series for post-entry tracking/reporting
        ce_close = ce_close_raw.ffill()
        pe_close = pe_close_raw.ffill()

        # High/Low series (used only to detect STOPLOSS intraminute extremes)
        ce_high = _build_leg_series(day_opt, idx_all, atm, "CE", ce_sym, "high", do_ffill=False)
        ce_low = _build_leg_series(day_opt, idx_all, atm, "CE", ce_sym, "low", do_ffill=False)
        pe_high = _build_leg_series(day_opt, idx_all, atm, "PE", pe_sym, "high", do_ffill=False)
        pe_low = _build_leg_series(day_opt, idx_all, atm, "PE", pe_sym, "low", do_ffill=False)

        if cur_entry_ts not in idx_all:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": "Entry timestamp not in session index"})
            break

        ce_entry = ce_close_raw.loc[cur_entry_ts]
        pe_entry = pe_close_raw.loc[cur_entry_ts]
        monitor_start_ts = pd.Timestamp(cur_entry_ts) + pd.Timedelta(minutes=1)
        if monitor_start_ts > trade_end_ts:
            break

        if pd.isna(ce_entry) or pd.isna(pe_entry):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "No CE/PE price at entry (after ffill)"})
            break

        # ---------------------------------------------------------------------
        # Percentage-based risk basis for THIS attempt
        # ---------------------------------------------------------------------
        # For every entry/re-entry, compute the premium collected in rupees.
        # Stop-loss and profit-protect thresholds are derived from this value.
        #
        # Example:
        #   entry_ce=70, entry_pe=50, qty=325
        #   entry_premium_sum = (70 + 50) * 325 = 39,000
        #   10% stop-loss = 3,900
        #   30% profit-protect threshold/giveback = 11,700
        # ---------------------------------------------------------------------
        entry_premium_sum = (float(ce_entry) + float(pe_entry)) * qty

        loss_limit_pct = settings.loss_pct_for_attempt(trade_seq - 1)
        uncapped_loss_limit_rupees = float(loss_limit_pct * entry_premium_sum)

        # Absolute cap on the percentage-based stop-loss.
        # Example: 10% of premium may be Rs. 4,500, but with a Rs. 3,000 cap
        # the effective stop used by the simulator is Rs. 3,000.
        stop_cap_rupees = float(settings.stop_cap_rupees)
        if stop_cap_rupees > 0:
            loss_limit_rupees = float(min(uncapped_loss_limit_rupees, stop_cap_rupees))
        else:
            loss_limit_rupees = float(uncapped_loss_limit_rupees)

        # G is the same variable used by the existing profit-protect logic:
        #   - profit-protect arms when peak P&L >= G
        #   - profit-protect exits when current P&L <= peak - G
        G = float(profit_protect_pct * entry_premium_sum)

        # Close-based PnL (same as before)
        pnl_close_all = (float(ce_entry) - ce_close) * qty + (float(pe_entry) - pe_close) * qty
        pnl = pnl_close_all.loc[monitor_start_ts:trade_end_ts].dropna()  # keep 'pnl' as close-based for profit-protect

        # STOPLOSS worst-case PnL candidates within each minute:
        #  A) CE high, PE low
        pnl_ceHigh_peLow_all = (float(ce_entry) - ce_high) * qty + (float(pe_entry) - pe_low) * qty
        #  B) CE low, PE high
        pnl_ceLow_peHigh_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_high) * qty

        # Worst-case PnL per minute among (close, A, B)
        pnl_sl_all = pd.concat([pnl_close_all, pnl_ceHigh_peLow_all, pnl_ceLow_peHigh_all], axis=1).min(axis=1)
        pnl_sl = pnl_sl_all.loc[monitor_start_ts:trade_end_ts].dropna()

        if pnl.empty:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "PnL series empty after entry"})
            break

        eod_ts = pnl.index[-1]
        eod_pnl = float(pnl.iloc[-1])

        # If EXIT_TIME_IST is earlier than market close and no risk/profit event
        # triggers before that, the attempt is closed at the configured cutoff.
        # The old "EOD" label is retained only when the monitoring horizon is
        # the real session end.
        default_exit_reason = "TIME_EXIT" if trade_end_ts < session_end_ts else "EOD"

        max_profit = float(max(0.0, pnl.max()))
        max_loss = float(min(0.0, pnl.min()))

        # STOPLOSS uses the attempt-specific rupee value after applying the
        # absolute per-attempt cap.
        stop_hit = pnl_sl <= -loss_limit_rupees
        stop_ts = pnl_sl.index[stop_hit.to_numpy().argmax()] if stop_hit.any() else None

        protect_ts = None
        if profit_protect_enabled:
            peak = pnl.cummax()
            armed = peak >= G
            trail = peak - G
            protect_hit = armed & (pnl <= trail)
            protect_ts = pnl.index[protect_hit.to_numpy().argmax()] if protect_hit.any() else None

        # --- Per-day PROFIT TARGET: % of premium collected on this attempt ---
        # When reached, this trade exits at the target AND no further trades are
        # taken for the day (PROFIT_TARGET is excluded from the re-entry rule below).
        target_ts = None
        target_rupees = None
        if settings.profit_target_pct > 0.0:
            target_rupees = settings.profit_target_pct * entry_premium_sum
            # best-case (favourable) intrabar profit: both legs bought back at their lows
            pnl_best_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_low) * qty
            pnl_tp = pd.concat([pnl_close_all, pnl_best_all], axis=1).max(axis=1)
            pnl_tp = pnl_tp.loc[monitor_start_ts:trade_end_ts].dropna()
            tp_hit = pnl_tp >= float(target_rupees)
            target_ts = pnl_tp.index[tp_hit.to_numpy().argmax()] if tp_hit.any() else None

        # Earliest triggered exit wins; on identical timestamps prefer the more
        # conservative outcome: STOPLOSS, then PROFIT_TARGET, then PROFIT_PROTECT.
        exit_ts = eod_ts
        exit_reason = default_exit_reason
        _candidates = []
        if stop_ts is not None:
            _candidates.append((stop_ts, 0, "STOPLOSS"))
        if target_ts is not None:
            _candidates.append((target_ts, 1, "PROFIT_TARGET"))
        if protect_ts is not None:
            _candidates.append((protect_ts, 2, "PROFIT_PROTECT"))
        if _candidates:
            _candidates.sort(key=lambda c: (c[0], c[1]))
            exit_ts, _, exit_reason = _candidates[0]

        if exit_reason == "STOPLOSS":
            exit_pnl_gross = -float(loss_limit_rupees)
        elif exit_reason == "PROFIT_TARGET":
            exit_pnl_gross = float(target_rupees)
        else:
            exit_pnl_gross = float(pnl.loc[exit_ts])

        # Peak (close-based) profit reached during this trade's life, up to its exit
        pnl_pre_exit = pnl.loc[:exit_ts]
        max_profit_before_exit = float(max(0.0, pnl_pre_exit.max())) if len(pnl_pre_exit) else 0.0

        exit_ce = float(ce_close.loc[exit_ts]) if pd.notna(ce_close.loc[exit_ts]) else float("nan")
        exit_pe = float(pe_close.loc[exit_ts]) if pd.notna(pe_close.loc[exit_ts]) else float("nan")

        txn_charges = compute_trade_charges(
            entry_ce=float(ce_entry), entry_pe=float(pe_entry),
            exit_ce=exit_ce if not pd.isna(exit_ce) else 0.0,
            exit_pe=exit_pe if not pd.isna(exit_pe) else 0.0,
            qty=qty,
        )
        exit_pnl = exit_pnl_gross - txn_charges

        # Update cumulative realized NET P&L for the day. This is checked
        # before allowing any further re-entry.
        daily_realized_pnl += float(exit_pnl)
        daily_loss_limit_hit = bool(
            daily_loss_limit_enabled
            and daily_realized_pnl <= -float(settings.max_daily_loss_rupees)
        )

        results.append(
            TradeRow(
                day=dy,
                underlying=und,
                trade_seq=trade_seq,
                expiry=expiry,
                days_to_expiry=dte,
                atm_strike=int(atm),
                qty_units=qty,
                entry_time=pd.Timestamp(cur_entry_ts).strftime("%H:%M"),
                exit_time=pd.Timestamp(exit_ts).strftime("%H:%M"),
                exit_reason=exit_reason,
                entry_underlying=float(u_px),
                ce_symbol=ce_sym,
                pe_symbol=pe_sym,
                entry_ce=float(ce_entry),
                entry_pe=float(pe_entry),
                exit_ce=exit_ce,
                exit_pe=exit_pe,
                exit_pnl_gross=exit_pnl_gross,
                txn_charges=txn_charges,
                exit_pnl=exit_pnl,
                eod_pnl=eod_pnl,
                max_profit=max_profit,
                max_loss=max_loss,
                max_profit_before_exit=max_profit_before_exit,
                entry_premium_sum=float(entry_premium_sum),
                stop_pct=float(loss_limit_pct),
                uncapped_stop_rupees=float(uncapped_loss_limit_rupees),
                stop_cap_rupees=float(stop_cap_rupees),
                stop_rupees=float(loss_limit_rupees),
                profit_protect_trigger_pct=float(profit_protect_pct),
                profit_protect_trigger_rupees=float(G),
                daily_realized_pnl_after_trade=float(daily_realized_pnl),
                daily_loss_limit_rupees=float(settings.max_daily_loss_rupees),
                daily_loss_limit_hit=bool(daily_loss_limit_hit),
            )
        )

        if daily_loss_limit_hit:
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq + 1,
                "reason": (
                    f"No re-entry: daily loss limit hit after trade_seq={trade_seq}; "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={settings.max_daily_loss_rupees:.2f}"
                ),
            })
            break

        if (
            exit_reason in ("STOPLOSS", "PROFIT_PROTECT")
            and (trade_seq - 1) < settings.max_reattempts
        ):
            delay_min = settings.reentry_delay_for_attempt(trade_seq - 1)
            trade_seq += 1
            cur_entry_ts = pd.Timestamp(exit_ts) + pd.Timedelta(minutes=delay_min)

            # A profile may stop admitting fresh re-entries before its hard
            # square-off time. The currently-open attempt still runs to
            # trade_end_ts, but the next attempt must begin before both limits.
            if cur_entry_ts >= reentry_end_ts:
                skipped.append({
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "reason": (
                        f"No re-entry: next entry time {pd.Timestamp(cur_entry_ts).strftime('%H:%M')} "
                        f"is at/after {und} DTE-0 re-entry cutoff "
                        f"{settings.reentry_cutoff_time_ist}"
                    ),
                })
                break
            continue

        break

    return results, skipped


# =============================================================================
# PASS-2: process each pickle and simulate trades for days where this expiry is nearest
# =============================================================================
def process_pickles_generate_trades(
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
) -> Tuple[pd.DataFrame, pd.DataFrame]:

    all_trades: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []

    # IMPORTANT: prevent double-count if same (und,day,expiry) appears in multiple files
    processed_day_keys: set[Tuple[str, date, date]] = set()

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            needed_cols = ["date", "name", "type", "option_type", "strike", "expiry", "instrument", "high", "low", "close"]
            missing = [c for c in needed_cols if c not in df.columns]
            if missing:
                raise ValueError(f"Missing columns {missing} in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")][needed_cols].copy()
            if d2.empty:
                continue

            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            if d2.empty:
                continue

            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2["strike_num"] = pd.to_numeric(d2["strike"], errors="coerce")
            d2["strike_int"] = d2["strike_num"].round().astype("Int64")  # safer than truncation
            d2["option_type"] = d2["option_type"].astype(str).str.upper()

            d2 = d2.dropna(subset=["day", "underlying", "expiry_date", "strike_int", "close"])
            d2["strike_int"] = d2["strike_int"].astype(int)

            # SAFETY: ignore stale rows where expiry is already before the trading day
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            # window filter
            d2 = d2[(d2["day"] >= window_start) & (d2["day"] <= window_end)]
            if d2.empty:
                continue

            # group by (und, day, expiry)
            for (und, dy, ex), g in d2.groupby(["underlying", "day", "expiry_date"], sort=False):
                key_ud = (und, dy)
                if key_ud not in min_expiry_map:
                    continue
                if min_expiry_map[key_ud] != ex:
                    continue
                # Enforce the configured DTE scope before the day is marked as
                # processed.  A pickle can contain several nearest-expiry days;
                # rejecting one non-DTE0 group must not discard a later expiry
                # day from the same file.
                dte = int((ex - dy).days)
                if dte not in ALLOWED_DTE:
                    continue

                day_key = (und, dy, ex)
                if day_key in processed_day_keys:
                    skipped_rows.append({
                        "day": dy, "underlying": und, "expiry": ex,
                        "reason": "Duplicate (underlying,day,expiry) encountered in multiple pickles; skipped to avoid double-count"
                    })
                    continue
                processed_day_keys.add(day_key)

                uday = underlying_data.get(und)
                if uday is None:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "No underlying series downloaded"})
                    continue
                uday = uday[uday["day"] == dy]
                if uday.empty:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "Underlying missing for day"})
                    continue

                trades, skips = simulate_day_multi_trades(
                    und=und,
                    dy=dy,
                    expiry=ex,
                    day_opt=g,
                    underlying_day=uday,
                )
                all_trades.extend([t.__dict__ for t in trades])
                skipped_rows.extend(skips)

            print(f"[PASS2 OK] {os.path.basename(p)} processed")

        except Exception as e:
            msg = f"[PASS2 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    all_df = pd.DataFrame(all_trades)
    if not all_df.empty:
        all_df = all_df.sort_values(["day", "underlying", "trade_seq"]).reset_index(drop=True)

    skip_df = pd.DataFrame(skipped_rows)
    if not skip_df.empty:
        if "day" not in skip_df.columns:
            skip_df["day"] = pd.NaT
        if "underlying" not in skip_df.columns:
            skip_df["underlying"] = pd.NA
        skip_df = skip_df.sort_values(["day", "underlying"], na_position="last").reset_index(drop=True)

    return all_df, skip_df


# =============================================================================
# Actual trades: one underlying per day (nearest expiry), include all re-entries for that underlying/day
# =============================================================================
def pick_actual_underlying_by_day(min_expiry_map: Dict[Tuple[str, date], date]) -> Dict[date, str]:
    by_day: Dict[date, List[Tuple[date, str]]] = {}
    for (und, dy), ex in min_expiry_map.items():
        if und not in TRADEABLE:
            continue

        dte = int((ex - dy).days)
        if dte not in ALLOWED_DTE:
            continue

        by_day.setdefault(dy, []).append((ex, und))

    out: Dict[date, str] = {}
    for dy, lst in by_day.items():
        # nearest expiry first; if tied, prefer NIFTY
        lst_sorted = sorted(lst, key=lambda t: (t[0], 0 if t[1] == "NIFTY" else 1))
        out[dy] = lst_sorted[0][1]
    return out


def build_source_opportunities_df(
    min_expiry_map: Dict[Tuple[str, date], date],
    window_start: date,
    window_end: date,
) -> pd.DataFrame:
    """
    Build the DTE0 opportunity set directly from the selected pickle folder.

    This is deliberately independent of whether a trade could be completed.
    It lets the final audit distinguish a different dataset from a sparse-data
    skip, and avoids hard-coding the old folder's 64-date signature when a new
    pickle archive is intentionally selected.
    """
    chosen = pick_actual_underlying_by_day(min_expiry_map)
    rows: List[Dict[str, Any]] = []
    for dy, und in sorted(chosen.items()):
        if dy < window_start or dy > window_end:
            continue
        expiry = min_expiry_map.get((und, dy))
        if expiry is None:
            continue
        dte = int((expiry - dy).days)
        if dte not in ALLOWED_DTE:
            continue
        rows.append(
            {
                "day": dy,
                "underlying": und,
                "qty_units": int(QTY_UNITS[und]),
                "expiry": expiry,
                "days_to_expiry": dte,
            }
        )
    return pd.DataFrame(rows)


def opportunity_signature(opportunities: pd.DataFrame) -> str:
    if opportunities.empty:
        return hashlib.sha256(b"").hexdigest()
    normalized = (
        opportunities.loc[:, ["day", "underlying", "qty_units"]]
        .drop_duplicates()
        .assign(
            day=lambda frame: pd.to_datetime(frame["day"]).dt.strftime(
                "%Y-%m-%d"
            )
        )
        .sort_values(["day", "underlying", "qty_units"])
    )
    signature_text = "\n".join(
        f"{row.day}|{row.underlying}|{int(row.qty_units)}"
        for row in normalized.itertuples(index=False)
    )
    return hashlib.sha256(signature_text.encode("utf-8")).hexdigest()


def build_actual_trades_df(all_trades_df: pd.DataFrame, min_expiry_map: Dict[Tuple[str, date], date]) -> pd.DataFrame:
    if all_trades_df.empty:
        return pd.DataFrame()

    actual_underlying = pick_actual_underlying_by_day(min_expiry_map)

    m = all_trades_df.copy()
    m["actual_underlying_for_day"] = m["day"].map(actual_underlying)

    # keep only days for which a 0/1-DTE actual underlying exists
    m = m[m["actual_underlying_for_day"].notna()]

    # keep only the selected underlying for that day
    m = m[m["underlying"] == m["actual_underlying_for_day"]]

    # keep only 0- and 1-DTE rows
    # keep only 0- and 1-DTE rows
    m = m[m["days_to_expiry"].isin(ALLOWED_DTE)]

    # keep all reattempts for the one selected underlying on that day
    m = m.drop(columns=["actual_underlying_for_day"])
    m = m.sort_values(["day", "trade_seq"]).reset_index(drop=True)

    # 1 if net exit PnL is positive, else 0
    m["is_exit_pnl_positive"] = (m["exit_pnl"] > 0).astype(int)

    return m


# =============================================================================
# Excel output
# =============================================================================
def _autosize_columns_safe(ws) -> None:
    # Safe autosize even when the sheet is "empty-ish"
    try:
        max_col = ws.max_column or 0
        if max_col <= 0:
            return
        for col_idx in range(1, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = 0
            for row_idx in range(1, min(ws.max_row or 1, 2000) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
    except Exception:
        # Never fail the whole run just because autosize misbehaved
        return

def _color_actual_trades_by_date(wb, actual_trades_df) -> None:
    """Shade rows so all attempts on the same calendar date share one colour,
    alternating between two soft fills as the date changes (visual grouping)."""
    if actual_trades_df is None or actual_trades_df.empty:
        return
    if "actual_trades" not in wb.sheetnames:
        return
    cols = list(actual_trades_df.columns)
    if "day" not in cols:
        return
    from openpyxl.styles import PatternFill
    ws = wb["actual_trades"]
    ncols = len(cols)
    fills = [
        PatternFill(fill_type="solid", fgColor="E8F0FE"),  # light blue
        PatternFill(fill_type="solid", fgColor="FFF3E0"),  # light amber
    ]
    days = actual_trades_df["day"].tolist()
    color_idx = 0
    prev_day = None
    first = True
    for i, d in enumerate(days):
        if first:
            first = False
        elif d != prev_day:
            color_idx ^= 1
        prev_day = d
        fill = fills[color_idx]
        excel_row = i + 2  # header occupies row 1
        for c in range(1, ncols + 1):
            ws.cell(row=excel_row, column=c).fill = fill


def write_excel(all_trades_df: pd.DataFrame, actual_trades_df: pd.DataFrame, skipped_df: pd.DataFrame) -> None:
    out_dir = os.path.dirname(os.path.abspath(OUTPUT_XLSX))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    piv_exit = pd.DataFrame()
    piv_eod_first = pd.DataFrame()
    monthwise_summary = pd.DataFrame()
    if not all_trades_df.empty:
        piv_exit = all_trades_df.pivot_table(index="day", columns="underlying", values="exit_pnl", aggfunc="sum").reset_index()

        first = all_trades_df[all_trades_df["trade_seq"] == 1]
        piv_eod_first = first.pivot_table(index="day", columns="underlying", values="eod_pnl", aggfunc="sum").reset_index()

        inst = all_trades_df.copy()
        inst["is_win_exit"] = inst["exit_pnl"] > 0
        inst["is_stoploss"] = inst["exit_reason"].astype(str).str.upper().eq("STOPLOSS")
        inst["is_profit_protect"] = inst["exit_reason"].astype(str).str.upper().eq("PROFIT_PROTECT")
        instrument_summary = (
            inst.groupby("underlying", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                win_rate_exit_pct=("is_win_exit", lambda s: 100.0 * s.mean()),
                stoploss_rate_pct=("is_stoploss", lambda s: 100.0 * s.mean()),
                profit_protect_rate_pct=("is_profit_protect", lambda s: 100.0 * s.mean()),
                avg_max_profit=("max_profit", "mean"),
                avg_max_loss=("max_loss", "mean"),
                worst_max_loss=("max_loss", "min"),
            )
            .sort_values("total_exit_pnl", ascending=False)
            .reset_index(drop=True)
        )
    else:
        instrument_summary = pd.DataFrame()

    if not actual_trades_df.empty:
        tmp = actual_trades_df.copy()
        tmp["month"] = pd.to_datetime(tmp["day"]).dt.to_period("M").astype(str)

        # Existing trade-level monthly summary
        monthwise_summary = (
            tmp.groupby("month", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                winning_trades=("is_exit_pnl_positive", "sum"),
            )
        )
        monthwise_summary["losing_trades"] = monthwise_summary["trades"] - monthwise_summary["winning_trades"]
        monthwise_summary["win_rate_pct"] = (
                100.0 * monthwise_summary["winning_trades"] / monthwise_summary["trades"]
        ).round(2)

        # New: daily PnL inside each month
        daily_tmp = (
            tmp.groupby(["month", "day"], as_index=False)
            .agg(daily_pnl=("exit_pnl", "sum"))
        )

        loss_day_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                avg_loss_on_loss_days=(
                    "daily_pnl",
                    lambda s: float(s[s < 0].mean()) if (s < 0).any() else 0.0
                ),
                max_loss_in_a_day=(
                    "daily_pnl",
                    lambda s: float(s.min()) if len(s) else 0.0
                ),
            )
        )

        # Count of profitable vs loss days within each month
        day_count_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                profitable_days=("daily_pnl", lambda s: int((s > 0).sum())),
                loss_days=("daily_pnl", lambda s: int((s < 0).sum())),
            )
        )

        # Date on which the worst (maximum-loss) day occurred, per month
        worst_rows = daily_tmp.loc[daily_tmp.groupby("month")["daily_pnl"].idxmin()]
        worst_day = worst_rows[["month", "day"]].rename(columns={"day": "max_loss_day_date"})

        monthwise_summary = monthwise_summary.merge(loss_day_stats, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(worst_day, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(day_count_stats, on="month", how="left")

        # Place the date column right after the max-loss value column
        _cols = list(monthwise_summary.columns)
        if "max_loss_day_date" in _cols and "max_loss_in_a_day" in _cols:
            _cols.remove("max_loss_day_date")
            _cols.insert(_cols.index("max_loss_in_a_day") + 1, "max_loss_day_date")
            monthwise_summary = monthwise_summary[_cols]

        # ---- Grand-total row across all months (totals for every column) ----
        if not monthwise_summary.empty:
            tot_trades = int(monthwise_summary["trades"].sum())
            tot_win = int(monthwise_summary["winning_trades"].sum())
            tot_lose = int(monthwise_summary["losing_trades"].sum())
            tot_pnl = float(monthwise_summary["total_exit_pnl"].sum())
            tot_prof_days = int(monthwise_summary["profitable_days"].sum())
            tot_loss_days = int(monthwise_summary["loss_days"].sum())
            neg_days = daily_tmp[daily_tmp["daily_pnl"] < 0]
            overall_worst_idx = daily_tmp["daily_pnl"].idxmin()
            total_row = {
                "month": "TOTAL",
                "trades": tot_trades,
                "total_exit_pnl": round(tot_pnl, 2),
                "avg_exit_pnl": round(tot_pnl / tot_trades, 2) if tot_trades else 0.0,
                "winning_trades": tot_win,
                "losing_trades": tot_lose,
                "win_rate_pct": round(100.0 * tot_win / tot_trades, 2) if tot_trades else 0.0,
                "avg_loss_on_loss_days": round(float(neg_days["daily_pnl"].mean()), 2) if len(neg_days) else 0.0,
                "max_loss_in_a_day": round(float(daily_tmp["daily_pnl"].min()), 2),
                "max_loss_day_date": daily_tmp.loc[overall_worst_idx, "day"],
                "profitable_days": tot_prof_days,
                "loss_days": tot_loss_days,
            }
            total_df = pd.DataFrame([total_row]).reindex(columns=monthwise_summary.columns)
            monthwise_summary = pd.concat([monthwise_summary, total_df], ignore_index=True)
    else:
        monthwise_summary = pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as xw:
        all_trades_df.to_excel(xw, sheet_name="all_trades_backtested", index=False)
        actual_trades_df.to_excel(xw, sheet_name="actual_trades", index=False)
        monthwise_summary.to_excel(xw, sheet_name="monthwise_summary", index=False)
        piv_exit.to_excel(xw, sheet_name="exit_pnl_pivot", index=False)
        piv_eod_first.to_excel(xw, sheet_name="eod_pnl_first_trade_pivot", index=False)
        instrument_summary.to_excel(xw, sheet_name="instrument_summary", index=False)
        skipped_df.to_excel(xw, sheet_name="skipped", index=False)

        wb = xw.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _autosize_columns_safe(ws)

        _color_actual_trades_by_date(wb, actual_trades_df)

    print(f"[DONE] Excel written: {OUTPUT_XLSX}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(
        f"[BUILD] {BUILD_ID} | script={os.path.abspath(__file__)} | "
        f"config={os.path.abspath(PROPERTY_FILE_PATH)} | "
        f"pickles={os.path.abspath(PICKLES_DIR)}"
    )
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")

    print(f"[INFO] Pickles found: {len(paths)}")

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)
    source_opportunities = build_source_opportunities_df(
        min_expiry_map,
        window_start,
        end_day,
    )
    if source_opportunities.empty:
        raise RuntimeError(
            "No DTE0 source opportunities were discovered in the configured "
            f"pickle folder: {PICKLES_DIR}"
        )
    source_opportunity_signature = opportunity_signature(source_opportunities)

    # Optional dataset pin. AUTO/blank means "use the folder discovered at
    # runtime"; an explicit count/signature is useful when a run must reproduce
    # one exact archive.
    if (
        EXPECTED_OPPORTUNITY_COUNT is not None
        and len(source_opportunities) != EXPECTED_OPPORTUNITY_COUNT
    ):
        raise RuntimeError(
            "Configured pickle folder does not match the pinned opportunity "
            f"set: expected {EXPECTED_OPPORTUNITY_COUNT}, discovered "
            f"{len(source_opportunities)} in {PICKLES_DIR}. Set "
            "EXPECTED_OPPORTUNITY_COUNT=AUTO only if this different dataset "
            "is intentional."
        )
    if (
        EXPECTED_OPPORTUNITY_SIGNATURE is not None
        and source_opportunity_signature != EXPECTED_OPPORTUNITY_SIGNATURE
    ):
        raise RuntimeError(
            "Configured pickle folder does not match the pinned opportunity "
            f"signature: expected {EXPECTED_OPPORTUNITY_SIGNATURE}, discovered "
            f"{source_opportunity_signature}. Set "
            "EXPECTED_OPPORTUNITY_SIGNATURE=AUTO only if this different "
            "dataset is intentional."
        )

    lookback_label = "AUTO/full pickle range" if LOOKBACK_MONTHS is None else f"{LOOKBACK_MONTHS} months cap"

    print(f"[INFO] Data day-range seen: {min_day_seen} -> {end_day}")
    print(f"[INFO] Backtest window: {window_start} -> {end_day} ({lookback_label})")
    for _profile_name, _settings in (
        ("NIFTY DTE-0", NIFTY_SETTINGS),
        ("SENSEX DTE-0", SENSEX_SETTINGS),
    ):
        print(
            f"[INFO] {_profile_name} profile | entry {_settings.entry_time_ist} | "
            f"exit {_settings.exit_time_ist} | re-entry cutoff "
            f"{_settings.reentry_cutoff_time_ist} | stops "
            f"{_fmt_pct_list(_settings.loss_limit_pct_by_attempt)} | "
            f"stop cap Rs {_fmt_rupee_value(_settings.stop_cap_rupees)} | "
            f"daily max loss Rs {_fmt_rupee_value(_settings.max_daily_loss_rupees)} | "
            f"profit protect {_fmt_pct_value(_settings.profit_protect_pct)} | "
            f"profit target {_settings.profit_target_pct:.2%} | "
            f"max re-entries {_settings.max_reattempts} | "
            f"delays {list(_settings.reentry_delay_by_attempt)}"
        )
    print(f"[INFO] Allowed DTE: {ALLOWED_DTE}")
    print(f"[INFO] Tradeables: {sorted(TRADEABLE)}")
    print(f"[INFO] Output: {OUTPUT_XLSX}")
    print(
        f"[SCOPE SOURCE] DTE0 opportunities={len(source_opportunities)} | "
        f"signature={source_opportunity_signature}"
    )

    print("[STEP] Loading embedded underlying candles from the same pickles ...")
    underlying_data = load_underlyings_from_pickles(paths, window_start, end_day)

    all_trades_df, skipped_df = process_pickles_generate_trades(
        paths, min_expiry_map, underlying_data, window_start, end_day
    )

    actual_trades_df = build_actual_trades_df(all_trades_df, min_expiry_map)

    # ---------------------------------------------------------------------
    # Hard scope validation. Profit may change; the opportunity set may not.
    # ---------------------------------------------------------------------
    if actual_trades_df.empty:
        raise RuntimeError("DTE-0 scope validation failed: no actual trades")
    if not actual_trades_df["days_to_expiry"].eq(0).all():
        invalid = sorted(actual_trades_df["days_to_expiry"].unique().tolist())
        raise RuntimeError(
            f"DTE-0 scope validation failed: observed DTE values {invalid}"
        )
    day_values = pd.to_datetime(actual_trades_df["day"]).dt.date
    expiry_values = pd.to_datetime(actual_trades_df["expiry"]).dt.date
    if not day_values.eq(expiry_values).all():
        raise RuntimeError(
            "DTE-0 scope validation failed: at least one expiry differs from day"
        )
    expected_qty = actual_trades_df["underlying"].map(QTY_UNITS)
    if expected_qty.isna().any() or not actual_trades_df["qty_units"].eq(expected_qty).all():
        raise RuntimeError("Quantity validation failed for NIFTY/SENSEX")

    actual_opportunities = (
        actual_trades_df.loc[:, ["day", "underlying", "qty_units"]]
        .drop_duplicates()
        .sort_values(["day", "underlying", "qty_units"])
    )
    actual_opportunity_signature = opportunity_signature(actual_opportunities)

    def _opportunity_keys(frame: pd.DataFrame) -> set[Tuple[str, str, int]]:
        return {
            (
                pd.Timestamp(row.day).strftime("%Y-%m-%d"),
                str(row.underlying),
                int(row.qty_units),
            )
            for row in frame.itertuples(index=False)
        }

    source_keys = _opportunity_keys(source_opportunities)
    actual_keys = _opportunity_keys(actual_opportunities)
    unexpected_keys = sorted(actual_keys - source_keys)
    missing_keys = sorted(source_keys - actual_keys)
    if unexpected_keys:
        raise RuntimeError(
            "Opportunity-set validation failed: actual trades contain "
            f"{len(unexpected_keys)} rows outside the configured DTE0 source "
            f"scope; sample={unexpected_keys[:5]}"
        )
    if missing_keys and REQUIRE_FULL_OPPORTUNITY_COVERAGE:
        raise RuntimeError(
            "Opportunity coverage validation failed: "
            f"{len(missing_keys)} of {len(source_keys)} selected DTE0 source "
            f"opportunities produced no completed trade; "
            f"sample={missing_keys[:5]}. Inspect the skipped sheet or set "
            "REQUIRE_FULL_OPPORTUNITY_COVERAGE=0 to retain sparse-data dates "
            "as explicit skips."
        )
    if missing_keys:
        print(
            f"[SCOPE WARN] {len(missing_keys)} of {len(source_keys)} selected "
            f"DTE0 source opportunities produced no completed trade; "
            f"sample={missing_keys[:5]}"
        )
        coverage_rows = pd.DataFrame(
            [
                {
                    "day": pd.Timestamp(day_text).date(),
                    "underlying": underlying,
                    "expiry": pd.Timestamp(day_text).date(),
                    "reason": (
                        "No completed actual trade for selected DTE0 source "
                        "opportunity; see other skipped rows for quote details"
                    ),
                }
                for day_text, underlying, _qty in missing_keys
            ]
        )
        skipped_df = pd.concat(
            [skipped_df, coverage_rows],
            ignore_index=True,
            sort=False,
        )
        if "day" in skipped_df.columns:
            skipped_df = skipped_df.sort_values(
                ["day", "underlying"],
                na_position="last",
            ).reset_index(drop=True)

    print(
        f"[CHECK] DTE-0 only | source_opportunities={len(source_keys)} | "
        f"traded_opportunities={len(actual_keys)} | missing={len(missing_keys)} "
        f"| source_signature={source_opportunity_signature} | "
        f"actual_signature={actual_opportunity_signature}"
    )

    write_excel(all_trades_df, actual_trades_df, skipped_df)

    if not all_trades_df.empty:
        print(all_trades_df.groupby("underlying")[["exit_pnl"]].describe())
    else:
        print("[WARN] No completed trades. Check 'skipped' sheet for reasons.")


if __name__ == "__main__":
    main()
