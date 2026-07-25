import os
from pathlib import Path
import glob
import time
from dataclasses import dataclass
from datetime import datetime, date, time as dtime, timedelta
from typing import Dict, List, Tuple, Optional, Any

import numpy as np
import pandas as pd

import Trading_2024.OptionTradeUtils as oUtils

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore

try:
    import pytz  # type: ignore
except Exception:
    pytz = None  # type: ignore

try:
    from dateutil.relativedelta import relativedelta  # type: ignore
except Exception:
    relativedelta = None  # type: ignore


# =============================================================================
# USER CONFIG
# =============================================================================
# ---------------------------------------------------------------------------
# CONFIGURATION SOURCE: external property file
# ---------------------------------------------------------------------------
# Every tunable setting lives in a simple KEY=VALUE property file so it can be
# changed WITHOUT editing this script. Path defaults to
# "straddle_config.properties" next to this file; override with the
# STRADDLE_CONFIG environment variable. Values are pushed into the process
# environment so all the os.getenv(...) reads below pick them up. A real
# environment variable that is already set takes precedence over the file.
def _load_property_file() -> str:
    cfg_path = os.getenv(
        "STRADDLE_CONFIG",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "configs/straddle_config_DTE_1_v2.properties"),
    )
    if not os.path.exists(cfg_path):
        print(f"[CONFIG] Property file not found at {cfg_path}; using built-in defaults.")
        return cfg_path
    loaded = 0
    with open(cfg_path, "r", encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith(";"):
                continue
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            key, val = key.strip(), val.strip()
            # strip optional surrounding quotes
            if len(val) >= 2 and val[0] == val[-1] and val[0] in ("'", '"'):
                val = val[1:-1]
            if key and key not in os.environ:   # real env vars win over the file
                os.environ[key] = val
                loaded += 1
    print(f"[CONFIG] Loaded {loaded} setting(s) from {cfg_path}")
    return cfg_path

PROPERTY_FILE_PATH = _load_property_file()

# PICKLES_DIR = r"G:\My Drive\Trading\Historical_Options_Data"
PICKLES_DIR = os.getenv("PICKLES_DIR", r"G:\My Drive\Trading\Dhan_Historical_Options_Data_New")
ENTRY_TIME_IST = os.getenv("ENTRY_TIME_IST", "11:55")  # "HH:MM"
# EXIT_TIME_IST is the strategy time filter / square-off cutoff.
# No fresh entry or re-entry is initiated at or after this time. If an attempt is
# still open at this time, it is closed at the cutoff with exit_reason="TIME_EXIT".
EXIT_TIME_IST = os.getenv("EXIT_TIME_IST", "15:30")  # "HH:MM"

def _safe_fname_part(s: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in s)

def _get_downloads_folder() -> str:
    """
    Returns the current user's default Downloads folder.
    Falls back to home directory if Downloads is not found.
    """
    downloads = Path.home() / "Downloads"
    return str(downloads if downloads.exists() else Path.home())

# --- Generic integer-list parser used for DTE and re-entry delay settings ---
def _parse_int_list(env_val, default_list):
    """
    Parse a comma-separated integer list from an environment variable.

    Examples:
        ALLOWED_DTE="0,1"              -> [0, 1]
        REENTRY_DELAY_BY_ATTEMPT="1,5" -> [1, 5]

    If parsing fails or the env var is blank, the supplied default is used.
    """
    if env_val:
        try:
            vals = [int(round(float(x))) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass
    return list(default_list)


def _parse_pct_value(x) -> float:
    """
    Parse a percentage value into decimal form.

    Accepted user formats:
        10      -> 0.10  (10%)
        "10%"   -> 0.10  (10%)
        0.10    -> 0.10  (10%)
        "0.10"  -> 0.10  (10%)

    This lets you configure percentages naturally from environment variables
    while keeping calculations internally consistent.
    """
    s = str(x).strip().replace("%", "")
    if s == "":
        raise ValueError("blank percentage value")

    v = float(s)

    # If user entered 10, treat it as 10%; if user entered 0.10, keep it as 10%.
    if abs(v) > 1.0:
        v = v / 100.0

    if v < 0:
        raise ValueError("percentage cannot be negative")

    return float(v)


def _parse_pct_list(env_val, default_list):
    """
    Parse comma-separated percentages into decimal form.

    Examples:
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15" -> [0.10, 0.12, 0.15]
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10"     -> [0.10]
    """
    if env_val:
        try:
            vals = [_parse_pct_value(x) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass

    return [_parse_pct_value(x) for x in default_list]


def _fmt_int_list(lst) -> str:
    """Format integer-list settings for the output filename."""
    return "-".join(str(int(v)) for v in lst) if lst else "off"


def _fmt_pct_value(v: float) -> str:
    """Format a decimal percentage such as 0.10 as '10pct' for filenames/logs."""
    return f"{v * 100:.2f}".rstrip("0").rstrip(".") + "pct"


def _fmt_pct_list(lst) -> str:
    """Format a list of decimal percentages for the output filename."""
    return "-".join(_fmt_pct_value(float(v)) for v in lst) if lst else "off"


def _parse_float_env(env_name: str, default_value: float) -> float:
    """
    Parse a positive/zero floating-point rupee setting from an environment variable.

    If parsing fails, the supplied default is used. A value <= 0 is treated by
    the relevant logic as disabled where applicable.
    """
    raw = os.getenv(env_name)
    if raw is None or str(raw).strip() == "":
        return float(default_value)
    try:
        return float(str(raw).replace(",", "").strip())
    except Exception:
        return float(default_value)


def _fmt_rupee_value(v: float) -> str:
    """Format rupee values compactly for output filenames/logs."""
    return str(int(round(float(v))))


LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_pct_list(
    os.getenv("LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
    [0.2487, 0.2824, 0.3162, 0.3499, 0.3837, 0.4174, 0.4512],
)


def loss_limit_pct_for_attempt(attempt_idx: int) -> float:
    """Return the stop-loss percentage, in decimal form, for the given attempt."""
    s = LOSS_LIMIT_RUPEES_BY_ATTEMPT
    if not s:
        return 0.0
    return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])


# --- Allowed days-to-expiry to trade: [0,1]=expiry day + day before; [0]=expiry only ---
ALLOWED_DTE = _parse_int_list(os.getenv("ALLOWED_DTE"), [0])

# --- Profit-protect threshold/giveback as % of premium collected on that attempt ---
# Default: 30%.
#
# Env examples:
#     PROFIT_PROTECT_TRIGGER_RUPEES="30"
#     PROFIT_PROTECT_TRIGGER_RUPEES="0.30"
#
# Current logic uses the same rupee amount for:
#     1. arming profit-protect once peak P&L reaches G
#     2. exiting when current P&L falls to peak - G
PROFIT_PROTECT_TRIGGER_RUPEES = _parse_pct_value(os.getenv("PROFIT_PROTECT_TRIGGER_RUPEES", 0.254741))

# --- Absolute daily circuit breaker -------------------------------------------------
# Once cumulative realized NET P&L for the current underlying/day reaches this
# loss, no further re-entry is allowed for that day.
#
# Default: Rs. 30,000 loss. Set MAX_DAILY_LOSS_RUPEES=0 to disable.
MAX_DAILY_LOSS_RUPEES = _parse_float_env("MAX_DAILY_LOSS_RUPEES", 30000.0)

# --- Absolute cap on the percentage-based per-attempt stop-loss ----------------------
# The stop-loss is still calculated as:
#     LOSS_LIMIT_% * entry_premium_sum
# But it is capped at this absolute rupee value.
#
# Effective stop-loss per attempt:
#     min(LOSS_LIMIT_% * entry_premium_sum, MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
#
# Default: Rs. 3,000 loss per attempt. Set to 0 to disable the cap.
MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_float_env("MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT", 3000.0)

MAX_REATTEMPTS = int(os.getenv("MAX_REATTEMPTS", "10"))  # 1 = only one re-entry

# --- Per-DAY profit target as a fraction of premium collected on the CURRENT attempt ---
# When an attempt's profit reaches PROFIT_TARGET_PCT * (CE+PE)*qty, it exits at the
# target and NO further trades are taken that day. 0 disables. e.g. 0.70 = 70%.
PROFIT_TARGET_PCT = float(os.getenv("PROFIT_TARGET_PCT", 0.7))
# --- Per-attempt RE-ENTRY GAP in minutes (index 0 = gap before 1st re-entry, 1 = before 2nd, ...) ---
# Attempts beyond the list reuse the LAST value. Override via env comma list, e.g.
# REENTRY_DELAY_BY_ATTEMPT="10,15,20".
REENTRY_DELAY_BY_ATTEMPT = _parse_int_list(
    os.getenv("REENTRY_DELAY_BY_ATTEMPT"),
    [6, 8, 10, 12, 14, 16, 18, 15, 15, 15, 15],
)

def reentry_delay_for_attempt(attempt_idx: int) -> int:
    s = REENTRY_DELAY_BY_ATTEMPT
    if not s:
        return 0
    return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])

# =============================================================================
# v2 SETTINGS
# =============================================================================
# Profit-protect is now asymmetric: the level that ARMS the trail and the
# give-back from the peak are separate. The original script forced them equal,
# which is why so much of the peak was handed back.
PP_ARM_PCT = _parse_pct_value(os.getenv("PP_ARM_PCT", PROFIT_PROTECT_TRIGGER_RUPEES))
PP_GIVEBACK_PCT = _parse_pct_value(os.getenv("PP_GIVEBACK_PCT", PROFIT_PROTECT_TRIGGER_RUPEES))

# Floor under the %-based stop, applied after MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT.
#     stop = clamp(pct * premium, STOP_FLOOR_RUPEES, MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
STOP_FLOOR_RUPEES = _parse_float_env("STOP_FLOOR_RUPEES", 0.0)

# --- Per-leg roll ------------------------------------------------------------
# When the position stop is about to fire, if one leg alone has lost more than
# LEG_ROLL_PCT of the premium collected, that leg is bought back and re-sold at
# the CURRENT at-the-money strike instead of closing the whole straddle. The
# untested leg keeps decaying. 0 disables.
LEG_ROLL_PCT = _parse_pct_value(os.getenv("LEG_ROLL_PCT", 0.0))
MAX_LEG_ROLLS = int(os.getenv("MAX_LEG_ROLLS", "0"))

# --- Re-entry after a profit target -----------------------------------------
REENTER_AFTER_PROFIT_TARGET = os.getenv("REENTER_AFTER_PROFIT_TARGET", "0").strip() == "1"

# --- Day-level trailing profit lock -----------------------------------------
# Once realised P&L for the day has been at least DAILY_LOCK_ARM_RUPEES, the day
# ends if it falls DAILY_LOCK_GIVEBACK_RUPEES below that peak. 0 disables.
DAILY_LOCK_ARM_RUPEES = _parse_float_env("DAILY_LOCK_ARM_RUPEES", 0.0)
DAILY_LOCK_GIVEBACK_RUPEES = _parse_float_env("DAILY_LOCK_GIVEBACK_RUPEES", 0.0)

# --- Entry movement gate -----------------------------------------------------
# Do not sell into a fast tape: postpone the entry minute by minute while the
# index has moved more than ENTRY_MOVE_MAX_PCT over the last
# ENTRY_MOVE_WINDOW_MIN minutes, up to ENTRY_MOVE_MAX_WAIT_MIN minutes.
ENTRY_MOVE_WINDOW_MIN = int(os.getenv("ENTRY_MOVE_WINDOW_MIN", "0"))
ENTRY_MOVE_MAX_PCT = _parse_pct_value(os.getenv("ENTRY_MOVE_MAX_PCT", 0.0))
ENTRY_MOVE_MAX_WAIT_MIN = int(os.getenv("ENTRY_MOVE_MAX_WAIT_MIN", "20"))

# --- Profit-target fill model ------------------------------------------------
# 0 (default, realistic): the target must be reached by the 1-minute CLOSE of the
#     straddle.
# 1 (legacy): the target may be filled at CE_low + PE_low of the same minute.
#     Both legs do not print their low at the same instant, so this prices a fill
#     that is not obtainable and inflates results.
PT_FILL_MODE = int(os.getenv("PT_FILL_MODE", "0"))

# --- Slippage ----------------------------------------------------------------
# Points given up per leg per side. Charged on entry, exit and every roll.
SLIPPAGE_POINTS = {
    "NIFTY": _parse_float_env("SLIPPAGE_POINTS_NIFTY", 0.0),
    "SENSEX": _parse_float_env("SLIPPAGE_POINTS_SENSEX", 0.0),
}
SLIPPAGE_STOP_MULT = _parse_float_env("SLIPPAGE_STOP_MULT", 1.0)

_DEFAULT_OUT = os.path.join(
    _get_downloads_folder(),
    f"short_straddle_v2_{_safe_fname_part(ENTRY_TIME_IST)}"
    f"_exit{_safe_fname_part(EXIT_TIME_IST)}"
    # f"_SLpct_{_safe_fname_part(_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    # f"_DTE_{_safe_fname_part('-'.join(str(d) for d in ALLOWED_DTE))}"
    f"_PPTpct_{_safe_fname_part(_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES))}"
    f"_DailyMaxLoss_{_safe_fname_part(_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES))}"
    f"_StopCap_{_safe_fname_part(_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    f"_MR_{_safe_fname_part(str(MAX_REATTEMPTS))}"
    # f"_PT_{_safe_fname_part(str(int(round(PROFIT_TARGET_PCT * 100))))}pct"
    f"_RDM_{_safe_fname_part(_fmt_int_list(REENTRY_DELAY_BY_ATTEMPT))}.xlsx"
)

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", _DEFAULT_OUT)

FAIL_ON_PICKLE_ERROR = os.getenv("FAIL_ON_PICKLE_ERROR", "0").strip() == "1"

SESSION_START_IST = dtime(9, 15)
SESSION_END_IST = dtime(15, 30)

# LOOKBACK_MONTHS is AUTO by default: the script backtests the full date range
# present in the option pickles. If you explicitly set LOOKBACK_MONTHS to a
# number, that number acts as an optional manual cap.
LOOKBACK_MONTHS_RAW = os.getenv("LOOKBACK_MONTHS", "AUTO").strip()
LOOKBACK_MONTHS: Optional[int]
if LOOKBACK_MONTHS_RAW.upper() in ("", "AUTO", "ALL", "MAX", "FULL"):
    LOOKBACK_MONTHS = None
else:
    LOOKBACK_MONTHS = int(float(LOOKBACK_MONTHS_RAW))

QTY_UNITS = {"NIFTY": 325, "SENSEX": 100}
TRADEABLE = set(QTY_UNITS.keys())

STRIKE_STEP = {"NIFTY": 50, "SENSEX": 100}

# =============================================================================
# TRANSACTION CHARGES (Zerodha F&O Options — NSE)
# =============================================================================
# Each short-straddle attempt = 4 executed orders (sell CE, sell PE, buy CE, buy PE)
BROKERAGE_PER_ORDER       = 20.0       # ₹20 flat per executed order
ORDERS_PER_TRADE          = 4          # sell CE + sell PE + buy CE + buy PE
STT_SELL_PCT              = 0.001      # 0.1% on sell-side premium
EXCHANGE_TXN_PCT          = 0.0003553  # 0.03553% on premium (NSE options)
SEBI_PER_CRORE            = 10.0       # ₹10 per crore of turnover
STAMP_BUY_PCT             = 0.00003    # 0.003% on buy-side premium
IPFT_PER_CRORE            = 0.010      # ₹0.01 per crore (on premium)
GST_PCT                   = 0.18       # 18% on (brokerage + txn charges + SEBI)
INCLUDE_TRANSACTION_COSTS = True       # set False to disable

UNDERLYING_KITE = {
    "NIFTY": {"exchange": "NSE", "tradingsymbol": "NIFTY 50"},
    "SENSEX": {"exchange": "BSE", "tradingsymbol": "SENSEX"},
}

MAX_DAYS_PER_CHUNK = 25
MAX_ATTEMPTS = 5
SLEEP_BETWEEN_CALLS_SEC = 0.20


# =============================================================================
# HELPERS
# =============================================================================
def parse_hhmm(s: str) -> dtime:
    hh, mm = s.strip().split(":")
    return dtime(int(hh), int(mm))

ENTRY_TIME = parse_hhmm(ENTRY_TIME_IST)
# Strategy cutoff time. The simulator will not initiate a new attempt at or
# after this time, and the current attempt will not be monitored beyond it.
EXIT_TIME = parse_hhmm(EXIT_TIME_IST)

def ist_tz():
    if ZoneInfo is not None:
        return ZoneInfo("Asia/Kolkata")
    if pytz is not None:
        return pytz.timezone("Asia/Kolkata")
    return "Asia/Kolkata"

def ensure_ist(series_or_scalar) -> Any:
    tz = ist_tz()
    dt = pd.to_datetime(series_or_scalar, errors="coerce")
    if isinstance(dt, pd.Series):
        if dt.dt.tz is None:
            return dt.dt.tz_localize(tz)
        return dt.dt.tz_convert(tz)
    if getattr(dt, "tzinfo", None) is None:
        return dt.tz_localize(tz)
    return dt.tz_convert(tz)

def normalize_underlying(name: str) -> Optional[str]:
    if not isinstance(name, str):
        return None
    u = name.upper().strip()
    if "SENSEX" in u:
        return "SENSEX"
    if "BANKNIFTY" in u or "NIFTY BANK" in u:
        return "BANKNIFTY"
    if "NIFTY" in u:
        return "NIFTY"
    return None

def round_to_step(x: float, step: int) -> int:
    return int(round(x / step) * step)

def build_minute_index(day_d: date, start_t: dtime, end_t: dtime) -> pd.DatetimeIndex:
    tz = ist_tz()
    start = pd.Timestamp(datetime.combine(day_d, start_t), tz=tz)
    end = pd.Timestamp(datetime.combine(day_d, end_t), tz=tz)
    return pd.date_range(start=start, end=end, freq="1min")

def asof_close(df: pd.DataFrame, ts: pd.Timestamp) -> float:
    if df.empty:
        return float("nan")
    d = df[["date", "close"]].dropna().copy()
    d["date"] = ensure_ist(d["date"])
    d = d.sort_values("date").set_index("date")
    loc = d.index.get_indexer([ts], method="pad")
    if loc[0] == -1:
        return float("nan")
    return float(d.iloc[loc[0]]["close"])

def compute_window_start(end_day: date, months: int) -> date:
    if relativedelta is not None:
        return (pd.Timestamp(end_day) - relativedelta(months=months)).date()
    return (pd.Timestamp(end_day) - pd.Timedelta(days=30 * months)).date()


def determine_backtest_window_start(min_day_seen: date, end_day: date) -> date:
    """
    Determine the backtest start date.

    Default behaviour: use the earliest usable option-data date found in the
    pickles, i.e. the maximum available backtest period.

    Optional override: if LOOKBACK_MONTHS is set to a numeric value through the
    environment, use the later of:
        1. earliest option-data date; and
        2. end_day - LOOKBACK_MONTHS
    so the script never requests data before the pickles actually start.
    """
    if LOOKBACK_MONTHS is None:
        return min_day_seen

    capped_start = compute_window_start(end_day, LOOKBACK_MONTHS)
    return max(min_day_seen, capped_start)

# =============================================================================
# TRANSACTION COST CALCULATOR
# =============================================================================
def compute_trade_charges(
    entry_ce: float, entry_pe: float,
    exit_ce: float, exit_pe: float,
    qty: int,
) -> float:
    """
    Compute total Zerodha transaction charges for one short-straddle attempt.

    Entry = SELL CE + SELL PE  (2 orders, sell side)
    Exit  = BUY  CE + BUY  PE (2 orders, buy side)

    Returns total charges in rupees (always positive).
    """
    if not INCLUDE_TRANSACTION_COSTS:
        return 0.0

    # Turnover values (in rupees)
    entry_turnover = (entry_ce + entry_pe) * qty   # sell side
    exit_turnover  = (exit_ce + exit_pe) * qty     # buy side
    total_turnover = entry_turnover + exit_turnover

    # 1. Brokerage: ₹20 × 4 orders
    brokerage = BROKERAGE_PER_ORDER * ORDERS_PER_TRADE

    # 2. STT: 0.1% on sell-side premium only (entry for short straddle)
    stt = entry_turnover * STT_SELL_PCT

    # 3. Exchange transaction charges: 0.03553% on both sides
    txn_charges = total_turnover * EXCHANGE_TXN_PCT

    # 4. SEBI charges: ₹10 per crore on total turnover
    sebi = total_turnover * SEBI_PER_CRORE / 1_00_00_000

    # 5. Stamp duty: 0.003% on buy side only (exit for short straddle)
    stamp = exit_turnover * STAMP_BUY_PCT

    # 6. IPFT: ₹0.01 per crore on premium (both sides)
    ipft = total_turnover * IPFT_PER_CRORE / 1_00_00_000

    # 7. GST: 18% on (brokerage + transaction charges + SEBI charges)
    gst = (brokerage + txn_charges + sebi) * GST_PCT

    total_charges = brokerage + stt + txn_charges + sebi + stamp + ipft + gst
    return round(total_charges, 2)

# =============================================================================
# Kite historical helpers
# =============================================================================
def _iter_chunks_by_date(from_dt: datetime, to_dt: datetime, days_per_chunk: int) -> List[Tuple[datetime, datetime]]:
    if from_dt > to_dt:
        raise ValueError("from_dt must be <= to_dt")
    chunks: List[Tuple[datetime, datetime]] = []
    cur = from_dt.date()
    end_d = to_dt.date()
    while cur <= end_d:
        chunk_end = min(cur + timedelta(days=days_per_chunk - 1), end_d)
        c_from = from_dt if cur == from_dt.date() else datetime.combine(cur, SESSION_START_IST)
        c_to = to_dt if chunk_end == end_d else datetime.combine(chunk_end, SESSION_END_IST)
        chunks.append((c_from, c_to))
        cur = chunk_end + timedelta(days=1)
    return chunks

def _kite_instruments_cached(kite, exchange: str, cache: Dict[str, List[Dict]]) -> List[Dict]:
    ex = exchange.upper().strip()
    if ex not in cache:
        print(f"[STEP] Loading instruments dump for {ex} ...")
        cache[ex] = kite.instruments(ex)
        print(f"[INFO] {ex} instruments: {len(cache[ex])}")
    return cache[ex]

def get_instrument_token(kite, exchange: str, tradingsymbol: str, cache: Dict[str, List[Dict]]) -> int:
    ex = exchange.upper().strip()
    wanted = tradingsymbol.strip().upper()
    for r in _kite_instruments_cached(kite, ex, cache):
        if str(r.get("tradingsymbol", "")).upper() == wanted:
            return int(r["instrument_token"])
    raise ValueError(f"Instrument not found on {ex}: '{tradingsymbol}'")

def fetch_history_minute(kite, instrument_token: int, from_dt: datetime, to_dt: datetime, label: str) -> List[Dict]:
    interval = "minute"
    chunks = _iter_chunks_by_date(from_dt, to_dt, MAX_DAYS_PER_CHUNK)
    rows_all: List[Dict] = []
    print(f"[INFO] Fetch {label} token={instrument_token} chunks={len(chunks)} {from_dt} -> {to_dt}")
    for i, (c_from, c_to) in enumerate(chunks, start=1):
        last_err = None
        for attempt in range(1, MAX_ATTEMPTS + 1):
            try:
                rows = kite.historical_data(
                    instrument_token=instrument_token,
                    from_date=c_from,
                    to_date=c_to,
                    interval=interval,
                    continuous=False,
                    oi=False,
                )
                rows_all.extend(rows)
                last_err = None
                break
            except Exception as e:
                last_err = e
                time.sleep(min(8.0, 1.5 * attempt))
        if last_err is not None:
            print(f"[ERROR] {label} chunk {i}/{len(chunks)} failed: {c_from}->{c_to}: {last_err}")
        time.sleep(SLEEP_BETWEEN_CALLS_SEC)
    return rows_all

def rows_to_df(rows: List[Dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["date", "open", "high", "low", "close", "volume"])
    df = pd.DataFrame(rows)
    df["date"] = ensure_ist(df["date"])
    df = df.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    return df


# =============================================================================
# DATA STRUCTURES
# =============================================================================
@dataclass
class TradeRow:
    day: date
    underlying: str
    trade_seq: int
    expiry: date
    days_to_expiry: int
    atm_strike: int
    qty_units: int
    entry_time: str
    exit_time: str
    exit_reason: str
    entry_underlying: float
    ce_symbol: str
    pe_symbol: str
    entry_ce: float
    entry_pe: float
    exit_ce: float
    exit_pe: float
    exit_pnl_gross: float   # P&L before charges
    txn_charges: float      # total transaction charges for this attempt
    exit_pnl: float         # net P&L after deducting charges
    eod_pnl: float
    max_profit: float
    max_loss: float
    max_profit_before_exit: float   # peak profit reached before this trade exited

    # Premium/risk diagnostics for percentage-based risk rules
    entry_premium_sum: float                 # (entry_ce + entry_pe) * qty
    stop_pct: float                          # stop-loss % of entry premium, decimal form; 0.10 = 10%
    uncapped_stop_rupees: float              # percentage-based stop before absolute cap
    stop_cap_rupees: float                   # configured absolute cap; <=0 means cap disabled
    stop_rupees: float                       # effective rupee stop after cap
    profit_protect_trigger_pct: float         # profit-protect % of entry premium, decimal form; 0.30 = 30%
    profit_protect_trigger_rupees: float      # computed rupee profit-protect trigger/giveback
    daily_realized_pnl_after_trade: float     # cumulative net P&L after this attempt
    daily_loss_limit_rupees: float            # configured daily loss circuit breaker
    daily_loss_limit_hit: bool                # True means no further trades for that day
    leg_rolls: int = 0                        # tested-leg rolls used in this attempt


# =============================================================================
# PASS-1: nearest expiry per (underlying, day)
# =============================================================================
def scan_pickles_pass1(pickle_paths: List[str]) -> Tuple[date, Dict[Tuple[str, date], date], date]:
    max_day_seen: Optional[date] = None
    min_day_seen: Optional[date] = None
    min_expiry_map: Dict[Tuple[str, date], date] = {}

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            for c in ("date", "name", "expiry", "type"):
                if c not in df.columns:
                    raise ValueError(f"Missing column '{c}' in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")]
            if d2.empty:
                continue

            d2 = d2[["date", "name", "expiry"]].copy()
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2 = d2.dropna(subset=["underlying", "day", "expiry_date"])

            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            file_min_day = d2["day"].min()
            file_max_day = d2["day"].max()
            max_day_seen = file_max_day if (max_day_seen is None or file_max_day > max_day_seen) else max_day_seen
            min_day_seen = file_min_day if (min_day_seen is None or file_min_day < min_day_seen) else min_day_seen

            grp = d2.groupby(["underlying", "day"], sort=False)["expiry_date"].min()
            for (und, dy), ex in grp.items():
                key = (und, dy)
                if key not in min_expiry_map or ex < min_expiry_map[key]:
                    min_expiry_map[key] = ex

            print(f"[PASS1 OK] {os.path.basename(p)} option_days={d2['day'].nunique()}")

        except Exception as e:
            msg = f"[PASS1 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    if max_day_seen is None or min_day_seen is None:
        raise RuntimeError("No usable option data found in pickles (PASS1) for tradeable underlyings.")

    return max_day_seen, min_expiry_map, min_day_seen


# =============================================================================
# Underlying download
# =============================================================================
def download_underlyings(kite, day_start: date, day_end: date) -> Dict[str, pd.DataFrame]:
    cache: Dict[str, List[Dict]] = {}
    from_dt = datetime.combine(day_start, SESSION_START_IST)
    to_dt = datetime.combine(day_end, SESSION_END_IST)

    out: Dict[str, pd.DataFrame] = {}
    for und, meta in UNDERLYING_KITE.items():
        token = get_instrument_token(kite, meta["exchange"], meta["tradingsymbol"], cache)
        rows = fetch_history_minute(kite, token, from_dt, to_dt, label=f"{meta['exchange']}:{meta['tradingsymbol']}")
        df = rows_to_df(rows)
        df["day"] = df["date"].dt.tz_convert(ist_tz()).dt.date
        out[und] = df
        print(f"[UNDERLYING OK] {und}: candles={len(df)} days={df['day'].nunique()}")
    return out


# =============================================================================
# Simulation helpers
# =============================================================================
def _pick_symbol(day_opt: pd.DataFrame, strike: int, opt_type: str) -> Optional[str]:
    sub = day_opt[(day_opt["strike_int"] == strike) & (day_opt["option_type"] == opt_type)]
    if sub.empty:
        return None
    syms = sorted(sub["instrument"].astype(str).unique().tolist())
    return syms[0] if syms else None

def _build_leg_series(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex,
                      strike: int, opt_type: str, symbol: str,
                      price_col: str = "close", do_ffill: bool = True) -> pd.Series:
    sub = day_opt[
        (day_opt["strike_int"] == strike) &
        (day_opt["option_type"] == opt_type) &
        (day_opt["instrument"].astype(str) == symbol)
    ][["date", price_col]].dropna()

    if sub.empty:
        return pd.Series(index=idx_all, dtype="float64")

    sub = sub.copy()
    sub["date"] = ensure_ist(sub["date"])
    sub = sub.sort_values("date").drop_duplicates(subset=["date"], keep="last").set_index("date")
    s = sub[price_col].astype(float).reindex(idx_all)
    return s.ffill() if do_ffill else s

def simulate_day_multi_trades(
    *,
    und: str,
    dy: date,
    expiry: date,
    day_opt: pd.DataFrame,
    underlying_day: pd.DataFrame,
) -> Tuple[List[TradeRow], List[Dict[str, Any]]]:
    """Simulate one underlying/day.

    v2 changes versus the original:
      * asymmetric profit-protect (separate arm level and give-back)
      * per-leg roll: shift only the tested leg instead of closing the straddle
      * realistic profit-target fill (close-based by default)
      * entry movement gate
      * day-level trailing profit lock
      * optional re-entry after a profit target
      * optional slippage
    """
    results: List[TradeRow] = []
    skipped: List[Dict[str, Any]] = []

    idx_all = build_minute_index(dy, SESSION_START_IST, SESSION_END_IST)
    session_end_ts = idx_all[-1]
    configured_exit_cutoff_ts = pd.Timestamp(datetime.combine(dy, EXIT_TIME), tz=ist_tz())
    trade_end_ts = min(session_end_ts, configured_exit_cutoff_ts)

    qty = int(QTY_UNITS[und])
    step = int(STRIKE_STEP[und])
    slip = float(SLIPPAGE_POINTS.get(und, 0.0))

    pos_of = {ts: i for i, ts in enumerate(idx_all)}
    end_i = pos_of[trade_end_ts]

    # underlying series on the minute grid
    u = underlying_day[["date", "close"]].dropna().copy()
    u["date"] = ensure_ist(u["date"])
    u = u.sort_values("date").drop_duplicates(subset=["date"], keep="last").set_index("date")
    spot = u["close"].astype(float).reindex(idx_all).ffill().to_numpy()

    # per-strike price matrices
    d = day_opt.copy()
    d["date"] = ensure_ist(d["date"])
    d = d[d["date"].isin(idx_all)]
    strikes = np.array(sorted(d["strike_int"].unique()), dtype=np.int64)
    if strikes.size == 0:
        skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": 1,
                        "reason": "No strikes in pickle for this day"})
        return results, skipped
    sidx = {int(s): i for i, s in enumerate(strikes)}
    n_t, n_s = len(idx_all), len(strikes)

    mats = {}
    for leg in ("CE", "PE"):
        for col in ("close", "high", "low"):
            mats[(leg, col)] = np.full((n_t, n_s), np.nan)
    sub = d[["date", "option_type", "strike_int", "close", "high", "low"]].dropna(subset=["close"])
    ti = sub["date"].map(pos_of).to_numpy()
    si = sub["strike_int"].map(sidx).to_numpy()
    ot = sub["option_type"].to_numpy()
    for leg in ("CE", "PE"):
        m = ot == leg
        if not m.any():
            continue
        for col in ("close", "high", "low"):
            mats[(leg, col)][ti[m], si[m]] = sub[col].to_numpy()[m]

    def ffill2d(a):
        out = a.copy()
        idx = np.where(np.isfinite(out), np.arange(out.shape[0])[:, None], 0)
        np.maximum.accumulate(idx, axis=0, out=idx)
        return out[idx, np.arange(out.shape[1])[None, :]]

    ce_raw, pe_raw = mats[("CE", "close")], mats[("PE", "close")]
    ce_c, pe_c = ffill2d(ce_raw), ffill2d(pe_raw)
    ce_h, ce_l = mats[("CE", "high")], mats[("CE", "low")]
    ph_, pl_ = mats[("PE", "high")], mats[("PE", "low")]

    def atm_index(s):
        k = int(round(s / step) * step)
        return sidx.get(k, -1)

    def sym(strike_i, leg):
        return _pick_symbol(day_opt, int(strikes[strike_i]), leg) or ""

    entry_i = pos_of.get(pd.Timestamp(datetime.combine(dy, ENTRY_TIME), tz=ist_tz()))
    if entry_i is None or entry_i >= end_i:
        skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": 1,
                        "reason": f"No entry: ENTRY_TIME_IST {ENTRY_TIME_IST} at/after {EXIT_TIME_IST}"})
        return results, skipped

    daily_realized = 0.0
    daily_peak = 0.0
    trade_seq = 1
    cur_i = entry_i
    dte = int((expiry - dy).days)

    while cur_i < end_i:
        if MAX_DAILY_LOSS_RUPEES > 0 and daily_realized <= -float(MAX_DAILY_LOSS_RUPEES):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": f"Daily loss limit hit: {daily_realized:.2f}"})
            break

        # ---- entry movement gate ------------------------------------------
        if ENTRY_MOVE_WINDOW_MIN > 0 and ENTRY_MOVE_MAX_PCT > 0:
            waited = 0
            while cur_i < end_i and waited < ENTRY_MOVE_MAX_WAIT_MIN:
                j = cur_i - ENTRY_MOVE_WINDOW_MIN
                if j < 0:
                    break
                s0, s1 = spot[j], spot[cur_i]
                if not (np.isfinite(s0) and np.isfinite(s1)):
                    break
                if abs(s1 - s0) / s1 <= ENTRY_MOVE_MAX_PCT:
                    break
                cur_i += 1
                waited += 1
            if cur_i >= end_i:
                break

        u_px = spot[cur_i]
        if not np.isfinite(u_px):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": "No underlying price at entry"})
            break
        k = atm_index(u_px)
        if k < 0:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": int(round(u_px / step) * step),
                            "reason": "ATM CE/PE not available in pickle band"})
            break

        ce_e, pe_e = ce_raw[cur_i, k], pe_raw[cur_i, k]
        if not (np.isfinite(ce_e) and np.isfinite(pe_e)) or ce_e <= 0 or pe_e <= 0:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": int(strikes[k]), "reason": "No CE/PE price at entry"})
            break

        entry_premium_sum = (float(ce_e) + float(pe_e)) * qty
        loss_limit_pct = loss_limit_pct_for_attempt(trade_seq - 1)
        uncapped = float(loss_limit_pct * entry_premium_sum)
        stop_rs = uncapped
        if MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT > 0:
            stop_rs = min(stop_rs, float(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT))
        if STOP_FLOOR_RUPEES > 0:
            stop_rs = max(stop_rs, float(STOP_FLOOR_RUPEES))

        arm_rs = float(PP_ARM_PCT * entry_premium_sum)
        give_rs = float(PP_GIVEBACK_PCT * entry_premium_sum)
        pt_rs = float(PROFIT_TARGET_PCT * entry_premium_sum) if PROFIT_TARGET_PCT > 0 else None
        leg_roll_rs = float(LEG_ROLL_PCT * entry_premium_sum)

        ck, pk = k, k
        c_e, p_e = float(ce_e), float(pe_e)
        realized = 0.0
        roll_cost = 0.0
        rolls = 0
        peak = 0.0
        armed = False
        exit_i = -1
        reason = "TIME_EXIT" if trade_end_ts < session_end_ts else "EOD"
        gross = 0.0

        t = cur_i + 1
        while t <= end_i:
            cc, pc = ce_c[t, ck], pe_c[t, pk]
            if not (np.isfinite(cc) and np.isfinite(pc)):
                t += 1
                continue

            pnl = realized + (c_e - cc) * qty + (p_e - pc) * qty

            worst = pnl
            a, b = ce_h[t, ck], pl_[t, pk]
            if np.isfinite(a) and np.isfinite(b):
                worst = min(worst, realized + (c_e - a) * qty + (p_e - b) * qty)
            a, b = ce_l[t, ck], ph_[t, pk]
            if np.isfinite(a) and np.isfinite(b):
                worst = min(worst, realized + (c_e - a) * qty + (p_e - b) * qty)

            best = pnl
            if PT_FILL_MODE == 1:
                a, b = ce_l[t, ck], pl_[t, pk]
                if np.isfinite(a) and np.isfinite(b):
                    best = max(best, realized + (c_e - a) * qty + (p_e - b) * qty)

            peak = max(peak, pnl)

            if worst <= -stop_rs:
                # roll the tested leg rather than concede the whole position
                rolled = False
                if LEG_ROLL_PCT > 0 and rolls < MAX_LEG_ROLLS:
                    ce_leg = (c_e - cc) * qty
                    pe_leg = (p_e - pc) * qty
                    nk = atm_index(spot[t]) if np.isfinite(spot[t]) else -1
                    if nk >= 0:
                        if ce_leg <= -leg_roll_rs and ce_leg < pe_leg and nk != ck:
                            npx = ce_raw[t, nk]
                            if np.isfinite(npx) and npx > 0:
                                realized += ce_leg
                                # a roll is 2 executed orders, not 4
                                roll_cost += (compute_trade_charges(c_e, 0.0, float(cc), 0.0, qty)
                                              - BROKERAGE_PER_ORDER * 2 * (1.0 + GST_PCT))
                                roll_cost += 2.0 * slip * qty
                                ck, c_e = nk, float(npx)
                                rolls += 1
                                rolled = True
                        elif pe_leg <= -leg_roll_rs and pe_leg < ce_leg and nk != pk:
                            npx = pe_raw[t, nk]
                            if np.isfinite(npx) and npx > 0:
                                realized += pe_leg
                                roll_cost += (compute_trade_charges(0.0, p_e, 0.0, float(pc), qty)
                                              - BROKERAGE_PER_ORDER * 2 * (1.0 + GST_PCT))
                                roll_cost += 2.0 * slip * qty
                                pk, p_e = nk, float(npx)
                                rolls += 1
                                rolled = True
                if not rolled:
                    exit_i, reason, gross = t, "STOPLOSS", -stop_rs
                    break
                t += 1
                continue

            if pt_rs is not None and best >= pt_rs:
                exit_i, reason, gross = t, "PROFIT_TARGET", pt_rs
                break

            if arm_rs > 0:
                if not armed and peak >= arm_rs:
                    armed = True
                if armed and pnl <= peak - give_rs:
                    exit_i, reason, gross = t, "PROFIT_PROTECT", pnl
                    break
            t += 1

        if exit_i < 0:
            tt = end_i
            while tt > cur_i and not (np.isfinite(ce_c[tt, ck]) and np.isfinite(pe_c[tt, pk])):
                tt -= 1
            if tt <= cur_i:
                break
            exit_i = tt
            gross = realized + (c_e - ce_c[tt, ck]) * qty + (p_e - pe_c[tt, pk]) * qty

        peak_before_exit = 0.0
        for tt in range(cur_i + 1, exit_i + 1):
            cc, pc = ce_c[tt, ck], pe_c[tt, pk]
            if np.isfinite(cc) and np.isfinite(pc):
                peak_before_exit = max(peak_before_exit,
                                       realized + (c_e - cc) * qty + (p_e - pc) * qty)

        x_ce = float(ce_c[exit_i, ck]) if np.isfinite(ce_c[exit_i, ck]) else float("nan")
        x_pe = float(pe_c[exit_i, pk]) if np.isfinite(pe_c[exit_i, pk]) else float("nan")

        txn = compute_trade_charges(entry_ce=c_e, entry_pe=p_e,
                                    exit_ce=0.0 if pd.isna(x_ce) else x_ce,
                                    exit_pe=0.0 if pd.isna(x_pe) else x_pe, qty=qty)
        txn += roll_cost + 4.0 * slip * qty
        if reason == "STOPLOSS":
            txn += 2.0 * slip * SLIPPAGE_STOP_MULT * qty
        exit_pnl = gross - txn

        daily_realized += float(exit_pnl)
        daily_peak = max(daily_peak, daily_realized)
        daily_hit = bool(MAX_DAILY_LOSS_RUPEES > 0 and daily_realized <= -float(MAX_DAILY_LOSS_RUPEES))

        eod = float("nan")
        for tt in range(end_i, cur_i, -1):
            cc, pc = ce_c[tt, ck], pe_c[tt, pk]
            if np.isfinite(cc) and np.isfinite(pc):
                eod = realized + (c_e - cc) * qty + (p_e - pc) * qty
                break

        seg = [realized + (c_e - ce_c[tt, ck]) * qty + (p_e - pe_c[tt, pk]) * qty
               for tt in range(cur_i + 1, end_i + 1)
               if np.isfinite(ce_c[tt, ck]) and np.isfinite(pe_c[tt, pk])]

        results.append(TradeRow(
            day=dy, underlying=und, trade_seq=trade_seq, expiry=expiry, days_to_expiry=dte,
            atm_strike=int(strikes[k]), qty_units=qty,
            entry_time=pd.Timestamp(idx_all[cur_i]).strftime("%H:%M"),
            exit_time=pd.Timestamp(idx_all[exit_i]).strftime("%H:%M"),
            exit_reason=reason, entry_underlying=float(u_px),
            ce_symbol=sym(ck, "CE"), pe_symbol=sym(pk, "PE"),
            entry_ce=float(ce_e), entry_pe=float(pe_e), exit_ce=x_ce, exit_pe=x_pe,
            exit_pnl_gross=float(gross), txn_charges=float(txn), exit_pnl=float(exit_pnl),
            eod_pnl=float(eod), max_profit=float(max(0.0, max(seg) if seg else 0.0)),
            max_loss=float(min(0.0, min(seg) if seg else 0.0)),
            max_profit_before_exit=float(peak_before_exit),
            entry_premium_sum=float(entry_premium_sum), stop_pct=float(loss_limit_pct),
            uncapped_stop_rupees=float(uncapped),
            stop_cap_rupees=float(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT),
            stop_rupees=float(stop_rs), profit_protect_trigger_pct=float(PP_ARM_PCT),
            profit_protect_trigger_rupees=float(arm_rs),
            daily_realized_pnl_after_trade=float(daily_realized),
            daily_loss_limit_rupees=float(MAX_DAILY_LOSS_RUPEES),
            daily_loss_limit_hit=daily_hit,
            leg_rolls=int(rolls),
        ))

        if daily_hit:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry,
                            "trade_seq": trade_seq + 1,
                            "reason": f"No re-entry: daily loss limit hit ({daily_realized:.2f})"})
            break

        if (DAILY_LOCK_GIVEBACK_RUPEES > 0 and daily_peak >= DAILY_LOCK_ARM_RUPEES
                and daily_realized <= daily_peak - DAILY_LOCK_GIVEBACK_RUPEES):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry,
                            "trade_seq": trade_seq + 1,
                            "reason": (f"No re-entry: daily profit lock "
                                       f"(peak={daily_peak:.0f}, now={daily_realized:.0f})")})
            break

        allow = reason in ("STOPLOSS", "PROFIT_PROTECT")
        if reason == "PROFIT_TARGET" and REENTER_AFTER_PROFIT_TARGET:
            allow = True
        if not allow or (trade_seq - 1) >= MAX_REATTEMPTS:
            break

        delay_min = reentry_delay_for_attempt(trade_seq - 1)
        trade_seq += 1
        cur_i = exit_i + delay_min
        if cur_i >= end_i:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": f"No re-entry: next entry at/after {EXIT_TIME_IST}"})
            break

    return results, skipped


# =============================================================================
# PASS-2: process each pickle and simulate trades for days where this expiry is nearest
# =============================================================================
def process_pickles_generate_trades(
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
) -> Tuple[pd.DataFrame, pd.DataFrame]:

    all_trades: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []

    # IMPORTANT: prevent double-count if same (und,day,expiry) appears in multiple files
    processed_day_keys: set[Tuple[str, date, date]] = set()

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            needed_cols = ["date", "name", "type", "option_type", "strike", "expiry", "instrument", "high", "low", "close"]
            missing = [c for c in needed_cols if c not in df.columns]
            if missing:
                raise ValueError(f"Missing columns {missing} in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")][needed_cols].copy()
            if d2.empty:
                continue

            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            if d2.empty:
                continue

            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2["strike_num"] = pd.to_numeric(d2["strike"], errors="coerce")
            d2["strike_int"] = d2["strike_num"].round().astype("Int64")  # safer than truncation
            d2["option_type"] = d2["option_type"].astype(str).str.upper()

            d2 = d2.dropna(subset=["day", "underlying", "expiry_date", "strike_int", "close"])
            d2["strike_int"] = d2["strike_int"].astype(int)

            # SAFETY: ignore stale rows where expiry is already before the trading day
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            # window filter
            d2 = d2[(d2["day"] >= window_start) & (d2["day"] <= window_end)]
            if d2.empty:
                continue

            # group by (und, day, expiry)
            for (und, dy, ex), g in d2.groupby(["underlying", "day", "expiry_date"], sort=False):
                key_ud = (und, dy)
                if key_ud not in min_expiry_map:
                    continue
                if min_expiry_map[key_ud] != ex:
                    continue

                day_key = (und, dy, ex)
                if day_key in processed_day_keys:
                    skipped_rows.append({
                        "day": dy, "underlying": und, "expiry": ex,
                        "reason": "Duplicate (underlying,day,expiry) encountered in multiple pickles; skipped to avoid double-count"
                    })
                    continue
                processed_day_keys.add(day_key)

                uday = underlying_data.get(und)
                if uday is None:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "No underlying series downloaded"})
                    continue
                uday = uday[uday["day"] == dy]
                if uday.empty:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "Underlying missing for day"})
                    continue

                trades, skips = simulate_day_multi_trades(
                    und=und,
                    dy=dy,
                    expiry=ex,
                    day_opt=g,
                    underlying_day=uday,
                )
                all_trades.extend([t.__dict__ for t in trades])
                skipped_rows.extend(skips)

            print(f"[PASS2 OK] {os.path.basename(p)} processed")

        except Exception as e:
            msg = f"[PASS2 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    all_df = pd.DataFrame(all_trades)
    if not all_df.empty:
        all_df = all_df.sort_values(["day", "underlying", "trade_seq"]).reset_index(drop=True)

    skip_df = pd.DataFrame(skipped_rows)
    if not skip_df.empty:
        if "day" not in skip_df.columns:
            skip_df["day"] = pd.NaT
        if "underlying" not in skip_df.columns:
            skip_df["underlying"] = pd.NA
        skip_df = skip_df.sort_values(["day", "underlying"], na_position="last").reset_index(drop=True)

    return all_df, skip_df


# =============================================================================
# Actual trades: one underlying per day (nearest expiry), include all re-entries for that underlying/day
# =============================================================================
def pick_actual_underlying_by_day(min_expiry_map: Dict[Tuple[str, date], date]) -> Dict[date, str]:
    by_day: Dict[date, List[Tuple[date, str]]] = {}
    for (und, dy), ex in min_expiry_map.items():
        if und not in TRADEABLE:
            continue

        dte = int((ex - dy).days)
        if dte not in ALLOWED_DTE:
            continue

        by_day.setdefault(dy, []).append((ex, und))

    out: Dict[date, str] = {}
    for dy, lst in by_day.items():
        # nearest expiry first; if tied, prefer NIFTY
        lst_sorted = sorted(lst, key=lambda t: (t[0], 0 if t[1] == "NIFTY" else 1))
        out[dy] = lst_sorted[0][1]
    return out

def build_actual_trades_df(all_trades_df: pd.DataFrame, min_expiry_map: Dict[Tuple[str, date], date]) -> pd.DataFrame:
    if all_trades_df.empty:
        return pd.DataFrame()

    actual_underlying = pick_actual_underlying_by_day(min_expiry_map)

    m = all_trades_df.copy()
    m["actual_underlying_for_day"] = m["day"].map(actual_underlying)

    # keep only days for which a 0/1-DTE actual underlying exists
    m = m[m["actual_underlying_for_day"].notna()]

    # keep only the selected underlying for that day
    m = m[m["underlying"] == m["actual_underlying_for_day"]]

    # keep only 0- and 1-DTE rows
    # keep only 0- and 1-DTE rows
    m = m[m["days_to_expiry"].isin(ALLOWED_DTE)]

    # keep all reattempts for the one selected underlying on that day
    m = m.drop(columns=["actual_underlying_for_day"])
    m = m.sort_values(["day", "trade_seq"]).reset_index(drop=True)

    # 1 if net exit PnL is positive, else 0
    m["is_exit_pnl_positive"] = (m["exit_pnl"] > 0).astype(int)

    return m


# =============================================================================
# Excel output
# =============================================================================
def _autosize_columns_safe(ws) -> None:
    # Safe autosize even when the sheet is "empty-ish"
    try:
        max_col = ws.max_column or 0
        if max_col <= 0:
            return
        for col_idx in range(1, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = 0
            for row_idx in range(1, min(ws.max_row or 1, 2000) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
    except Exception:
        # Never fail the whole run just because autosize misbehaved
        return

def _color_actual_trades_by_date(wb, actual_trades_df) -> None:
    """Shade rows so all attempts on the same calendar date share one colour,
    alternating between two soft fills as the date changes (visual grouping)."""
    if actual_trades_df is None or actual_trades_df.empty:
        return
    if "actual_trades" not in wb.sheetnames:
        return
    cols = list(actual_trades_df.columns)
    if "day" not in cols:
        return
    from openpyxl.styles import PatternFill
    ws = wb["actual_trades"]
    ncols = len(cols)
    fills = [
        PatternFill(fill_type="solid", fgColor="E8F0FE"),  # light blue
        PatternFill(fill_type="solid", fgColor="FFF3E0"),  # light amber
    ]
    days = actual_trades_df["day"].tolist()
    color_idx = 0
    prev_day = None
    first = True
    for i, d in enumerate(days):
        if first:
            first = False
        elif d != prev_day:
            color_idx ^= 1
        prev_day = d
        fill = fills[color_idx]
        excel_row = i + 2  # header occupies row 1
        for c in range(1, ncols + 1):
            ws.cell(row=excel_row, column=c).fill = fill


def write_excel(all_trades_df: pd.DataFrame, actual_trades_df: pd.DataFrame, skipped_df: pd.DataFrame) -> None:
    out_dir = os.path.dirname(os.path.abspath(OUTPUT_XLSX))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    piv_exit = pd.DataFrame()
    piv_eod_first = pd.DataFrame()
    monthwise_summary = pd.DataFrame()
    if not all_trades_df.empty:
        piv_exit = all_trades_df.pivot_table(index="day", columns="underlying", values="exit_pnl", aggfunc="sum").reset_index()

        first = all_trades_df[all_trades_df["trade_seq"] == 1]
        piv_eod_first = first.pivot_table(index="day", columns="underlying", values="eod_pnl", aggfunc="sum").reset_index()

        inst = all_trades_df.copy()
        inst["is_win_exit"] = inst["exit_pnl"] > 0
        inst["is_stoploss"] = inst["exit_reason"].astype(str).str.upper().eq("STOPLOSS")
        inst["is_profit_protect"] = inst["exit_reason"].astype(str).str.upper().eq("PROFIT_PROTECT")
        instrument_summary = (
            inst.groupby("underlying", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                win_rate_exit_pct=("is_win_exit", lambda s: 100.0 * s.mean()),
                stoploss_rate_pct=("is_stoploss", lambda s: 100.0 * s.mean()),
                profit_protect_rate_pct=("is_profit_protect", lambda s: 100.0 * s.mean()),
                avg_max_profit=("max_profit", "mean"),
                avg_max_loss=("max_loss", "mean"),
                worst_max_loss=("max_loss", "min"),
            )
            .sort_values("total_exit_pnl", ascending=False)
            .reset_index(drop=True)
        )
    else:
        instrument_summary = pd.DataFrame()

    if not actual_trades_df.empty:
        tmp = actual_trades_df.copy()
        tmp["month"] = pd.to_datetime(tmp["day"]).dt.to_period("M").astype(str)

        # Existing trade-level monthly summary
        monthwise_summary = (
            tmp.groupby("month", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                winning_trades=("is_exit_pnl_positive", "sum"),
            )
        )
        monthwise_summary["losing_trades"] = monthwise_summary["trades"] - monthwise_summary["winning_trades"]
        monthwise_summary["win_rate_pct"] = (
                100.0 * monthwise_summary["winning_trades"] / monthwise_summary["trades"]
        ).round(2)

        # New: daily PnL inside each month
        daily_tmp = (
            tmp.groupby(["month", "day"], as_index=False)
            .agg(daily_pnl=("exit_pnl", "sum"))
        )

        loss_day_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                avg_loss_on_loss_days=(
                    "daily_pnl",
                    lambda s: float(s[s < 0].mean()) if (s < 0).any() else 0.0
                ),
                max_loss_in_a_day=(
                    "daily_pnl",
                    lambda s: float(s.min()) if len(s) else 0.0
                ),
            )
        )

        # Count of profitable vs loss days within each month
        day_count_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                profitable_days=("daily_pnl", lambda s: int((s > 0).sum())),
                loss_days=("daily_pnl", lambda s: int((s < 0).sum())),
            )
        )

        # Date on which the worst (maximum-loss) day occurred, per month
        worst_rows = daily_tmp.loc[daily_tmp.groupby("month")["daily_pnl"].idxmin()]
        worst_day = worst_rows[["month", "day"]].rename(columns={"day": "max_loss_day_date"})

        monthwise_summary = monthwise_summary.merge(loss_day_stats, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(worst_day, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(day_count_stats, on="month", how="left")

        # Place the date column right after the max-loss value column
        _cols = list(monthwise_summary.columns)
        if "max_loss_day_date" in _cols and "max_loss_in_a_day" in _cols:
            _cols.remove("max_loss_day_date")
            _cols.insert(_cols.index("max_loss_in_a_day") + 1, "max_loss_day_date")
            monthwise_summary = monthwise_summary[_cols]

        # ---- Grand-total row across all months (totals for every column) ----
        if not monthwise_summary.empty:
            tot_trades = int(monthwise_summary["trades"].sum())
            tot_win = int(monthwise_summary["winning_trades"].sum())
            tot_lose = int(monthwise_summary["losing_trades"].sum())
            tot_pnl = float(monthwise_summary["total_exit_pnl"].sum())
            tot_prof_days = int(monthwise_summary["profitable_days"].sum())
            tot_loss_days = int(monthwise_summary["loss_days"].sum())
            neg_days = daily_tmp[daily_tmp["daily_pnl"] < 0]
            overall_worst_idx = daily_tmp["daily_pnl"].idxmin()
            total_row = {
                "month": "TOTAL",
                "trades": tot_trades,
                "total_exit_pnl": round(tot_pnl, 2),
                "avg_exit_pnl": round(tot_pnl / tot_trades, 2) if tot_trades else 0.0,
                "winning_trades": tot_win,
                "losing_trades": tot_lose,
                "win_rate_pct": round(100.0 * tot_win / tot_trades, 2) if tot_trades else 0.0,
                "avg_loss_on_loss_days": round(float(neg_days["daily_pnl"].mean()), 2) if len(neg_days) else 0.0,
                "max_loss_in_a_day": round(float(daily_tmp["daily_pnl"].min()), 2),
                "max_loss_day_date": daily_tmp.loc[overall_worst_idx, "day"],
                "profitable_days": tot_prof_days,
                "loss_days": tot_loss_days,
            }
            total_df = pd.DataFrame([total_row]).reindex(columns=monthwise_summary.columns)
            monthwise_summary = pd.concat([monthwise_summary, total_df], ignore_index=True)
    else:
        monthwise_summary = pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as xw:
        all_trades_df.to_excel(xw, sheet_name="all_trades_backtested", index=False)
        actual_trades_df.to_excel(xw, sheet_name="actual_trades", index=False)
        monthwise_summary.to_excel(xw, sheet_name="monthwise_summary", index=False)
        piv_exit.to_excel(xw, sheet_name="exit_pnl_pivot", index=False)
        piv_eod_first.to_excel(xw, sheet_name="eod_pnl_first_trade_pivot", index=False)
        instrument_summary.to_excel(xw, sheet_name="instrument_summary", index=False)
        skipped_df.to_excel(xw, sheet_name="skipped", index=False)

        wb = xw.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _autosize_columns_safe(ws)

        _color_actual_trades_by_date(wb, actual_trades_df)

    print(f"[DONE] Excel written: {OUTPUT_XLSX}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")

    print(f"[INFO] Pickles found: {len(paths)}")

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)

    lookback_label = "AUTO/full pickle range" if LOOKBACK_MONTHS is None else f"{LOOKBACK_MONTHS} months cap"

    print(f"[INFO] Data day-range seen: {min_day_seen} -> {end_day}")
    print(f"[INFO] Backtest window: {window_start} -> {end_day} ({lookback_label})")
    print(f"[INFO] Stoploss %/attempt: {_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Per-attempt stop cap: Rs {_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Daily max loss: Rs {_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES)} | "
          f"ProfitProtect trigger/giveback %: {_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES)} | "
          f"Re-entry delay min/attempt: {REENTRY_DELAY_BY_ATTEMPT} | Allowed DTE: {ALLOWED_DTE}")
    print(f"[INFO] Entry time: {ENTRY_TIME_IST} | Strategy exit/filter cutoff: {EXIT_TIME_IST}")
    print(f"[INFO] Day profit target: {PROFIT_TARGET_PCT:.0%} of premium (0 = disabled)")
    print(f"[INFO] Tradeables: {sorted(TRADEABLE)}")
    print(f"[INFO] Output: {OUTPUT_XLSX}")

    print("[STEP] Initializing Kite ...")
    kite = oUtils.intialize_kite_api()
    print("[OK] Kite ready.")

    underlying_data = download_underlyings(kite, window_start, end_day)

    all_trades_df, skipped_df = process_pickles_generate_trades(
        paths, min_expiry_map, underlying_data, window_start, end_day
    )

    actual_trades_df = build_actual_trades_df(all_trades_df, min_expiry_map)

    write_excel(all_trades_df, actual_trades_df, skipped_df)

    if not all_trades_df.empty:
        print(all_trades_df.groupby("underlying")[["exit_pnl"]].describe())
    else:
        print("[WARN] No completed trades. Check 'skipped' sheet for reasons.")


if __name__ == "__main__":
    main()