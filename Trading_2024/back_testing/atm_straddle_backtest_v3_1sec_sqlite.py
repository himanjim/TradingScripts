#!/usr/bin/env python3
"""
ATM short-straddle backtest v3 on one-second SQLite databases.

Purpose
-------
This script ports the strategy rules from ``atm_straddle_backtest_v3.py`` to
SQLite databases produced by ``KiteOptions1SecSpikeCollector_v3_1_oUtils_target.py``.
It does not download option or underlying data and does not require a Kite API
session. Every trading day is read from the collector's compact SQLite store.

Strategy parity with the one-minute engine
------------------------------------------
The following rules are retained:

* ATM selection from the underlying LTP at each entry/re-entry;
* percentage stop-loss with an absolute rupee cap;
* stop-loss detection from adverse CE/PE high-low combinations;
* profit target detection from the favourable CE-low + PE-low combination;
* close-based profit protection and breakeven ratchet;
* re-entry after stop-loss, profit-protect and optionally profit-target;
* re-entry premium-ratio gate;
* daily maximum loss and day-level realised-profit trail;
* transaction-cost model and one-index-per-day actual-trade selection.

Resolution changes
------------------
The simulation clock is one second rather than one minute:

* ``ENTRY_TIME_IST=09:16`` means 09:16:00 IST;
* monitoring starts one second after entry;
* exits can occur at any recorded/reconstructed second;
* re-entry delays remain expressed in minutes and are added to the exact
  second of the preceding exit.

Sparse database reconstruction
------------------------------
The collector normally stores only seconds in which LTP changed. This script
reconstructs a continuous one-second series as follows:

* close/LTP is forward-filled from the most recent stored observation;
* for an omitted second, open=high=low=close=the carried LTP;
* for a stored second, the collector's recorded one-second OHLC is retained;
* values before the first observation of an instrument remain unavailable.

Thus, a leg that does not change during a second contributes its latest known
LTP while the other leg can still record a spike. This is the correct
reconstruction for a change-only LTP database, subject to the feed limitations
recorded by the collector.

Accuracy boundary
-----------------
The collector timestamps bars by local receipt time because Kite QUOTE packets
do not contain an exchange timestamp. CE and PE lows inside the same second can
still occur at different milliseconds. The one-second low-low model is much
more tightly synchronized than one-minute low-low, but it is not an order-book
or guaranteed-fill simulation.

Run
---
    python atm_straddle_backtest_v3_1sec_sqlite.py

Optional explicit config:
    python atm_straddle_backtest_v3_1sec_sqlite.py --config path/to/file.properties

The ``STRADDLE_CONFIG`` environment variable is also supported.
"""

from __future__ import annotations

import argparse
import glob
import math
import os
import sqlite3
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime, time as dtime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd

try:
    from zoneinfo import ZoneInfo
except ImportError as exc:  # pragma: no cover
    raise RuntimeError("Python 3.9+ is required") from exc


# =============================================================================
# CONFIGURATION LOADING
# =============================================================================


def _config_path_from_argv() -> Optional[str]:
    """Read --config without consuming the main argument parser."""
    argv = sys.argv[1:]
    for index, value in enumerate(argv):
        if value == "--config" and index + 1 < len(argv):
            return argv[index + 1]
        if value.startswith("--config="):
            return value.split("=", 1)[1]
    return None


def _load_property_file() -> str:
    """Load KEY=VALUE settings into os.environ.

    Existing environment variables take precedence, matching the behaviour of
    the original one-minute strategy.
    """

    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_path = os.path.join(
        script_dir,
        "configs/straddle_config_DTE_1_v3_1sec_sqlite.properties",
    )
    cfg_path = (
        _config_path_from_argv()
        or os.getenv("STRADDLE_CONFIG")
        or default_path
    )
    cfg_path = os.path.abspath(os.path.expanduser(cfg_path))

    if not os.path.exists(cfg_path):
        print(f"[CONFIG] Property file not found at {cfg_path}; using defaults/environment.")
        return cfg_path

    loaded = 0
    with open(cfg_path, "r", encoding="utf-8") as handle:
        for raw in handle:
            line = raw.strip()
            if not line or line.startswith(("#", ";")) or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
                value = value[1:-1]
            if key and key not in os.environ:
                os.environ[key] = value
                loaded += 1

    print(f"[CONFIG] Loaded {loaded} setting(s) from {cfg_path}")
    return cfg_path


PROPERTY_FILE_PATH = _load_property_file()
IST = ZoneInfo("Asia/Kolkata")
PRICE_SCALE = 100.0
TOKEN_MASK = 0xFFFFFFFF


# =============================================================================
# SMALL CONFIG HELPERS
# =============================================================================


def _parse_bool(raw: Optional[str], default: bool = False) -> bool:
    if raw is None or str(raw).strip() == "":
        return default
    value = str(raw).strip().lower()
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    raise ValueError(f"Invalid boolean value: {raw!r}")


def _parse_int_list(raw: Optional[str], default: Sequence[int]) -> List[int]:
    if raw:
        try:
            values = [
                int(round(float(item)))
                for item in raw.replace(" ", "").split(",")
                if item != ""
            ]
            if values:
                return values
        except Exception:
            pass
    return list(default)


def _parse_pct_value(raw: Any) -> float:
    text = str(raw).strip().replace("%", "")
    if not text:
        raise ValueError("blank percentage")
    value = float(text)
    if abs(value) > 1.0:
        value /= 100.0
    if value < 0:
        raise ValueError("percentage cannot be negative")
    return float(value)


def _parse_pct_list(raw: Optional[str], default: Sequence[float]) -> List[float]:
    if raw:
        try:
            values = [
                _parse_pct_value(item)
                for item in raw.replace(" ", "").split(",")
                if item != ""
            ]
            if values:
                return values
        except Exception:
            pass
    return [_parse_pct_value(item) for item in default]


def _env_float(name: str, default: float) -> float:
    raw = os.getenv(name)
    if raw is None or str(raw).strip() == "":
        return float(default)
    return float(str(raw).replace(",", "").strip())


def _env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None or str(raw).strip() == "":
        return int(default)
    return int(round(float(raw)))


def parse_clock(value: str) -> dtime:
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(value.strip(), fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Invalid time {value!r}; expected HH:MM[:SS]")


def _safe_filename_part(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value)


def _downloads_folder() -> str:
    path = Path.home() / "Downloads"
    return str(path if path.exists() else Path.home())


# =============================================================================
# USER SETTINGS
# =============================================================================

# SQLite discovery. One collector database normally represents one target index
# and one trading day.
SQLITE_DIR = os.path.abspath(
    os.path.expanduser(os.getenv("SQLITE_DIR", "./kite_spike_data"))
)
SQLITE_GLOB = os.getenv("SQLITE_GLOB", "kite_option_spikes_v3_*.sqlite3")
SQLITE_RECURSIVE = _parse_bool(os.getenv("SQLITE_RECURSIVE"), True)
FAIL_ON_DB_ERROR = _parse_bool(
    os.getenv("FAIL_ON_DB_ERROR", os.getenv("FAIL_ON_PICKLE_ERROR", "0")),
    False,
)

# Data-quality controls. Defaults do not add a new trading rule. They reject
# only days that cannot support an entry or are structurally incompatible.
REQUIRE_COLLECTOR_SCHEMA_VERSION = _env_int("REQUIRE_COLLECTOR_SCHEMA_VERSION", 3)
SKIP_DAY_ON_RECONNECT = _parse_bool(os.getenv("SKIP_DAY_ON_RECONNECT"), False)
SKIP_DAY_ON_FEED_ERROR = _parse_bool(os.getenv("SKIP_DAY_ON_FEED_ERROR"), False)
MAX_ENTRY_PRICE_AGE_SECONDS = _env_int("MAX_ENTRY_PRICE_AGE_SECONDS", 0)
# 0 means no age limit. This is intentionally disabled by default because the
# change-only collector may omit many unchanged seconds.

ENTRY_TIME_IST = os.getenv("ENTRY_TIME_IST", "09:16")
EXIT_TIME_IST = os.getenv("EXIT_TIME_IST", "15:29")
ENTRY_TIME = parse_clock(ENTRY_TIME_IST)
EXIT_TIME = parse_clock(EXIT_TIME_IST)

SESSION_START = parse_clock(os.getenv("SESSION_START", "09:15:00"))
SESSION_END = parse_clock(os.getenv("SESSION_END", "15:30:00"))
if SESSION_START >= SESSION_END:
    raise ValueError("SESSION_START must be earlier than SESSION_END")

ALLOWED_DTE = _parse_int_list(os.getenv("ALLOWED_DTE"), [1])

LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_pct_list(
    os.getenv("LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
    [0.2444, 0.2555, 0.2666, 0.2776, 0.2887, 0.2998, 0.3108],
)
MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT = _env_float(
    "MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT", 4500.0
)
MAX_DAILY_LOSS_RUPEES = _env_float("MAX_DAILY_LOSS_RUPEES", 20000.0)
MAX_REATTEMPTS = _env_int("MAX_REATTEMPTS", 15)

PROFIT_PROTECT_TRIGGER_RUPEES = _parse_pct_value(
    os.getenv("PROFIT_PROTECT_TRIGGER_RUPEES", "0.0816")
)
_pp_arm_raw = os.getenv("PROFIT_PROTECT_ARM_PCT")
_pp_give_raw = os.getenv("PROFIT_PROTECT_GIVEBACK_PCT")
PROFIT_PROTECT_ARM_PCT = (
    _parse_pct_value(_pp_arm_raw)
    if _pp_arm_raw
    else PROFIT_PROTECT_TRIGGER_RUPEES
)
PROFIT_PROTECT_GIVEBACK_PCT = (
    _parse_pct_value(_pp_give_raw)
    if _pp_give_raw
    else PROFIT_PROTECT_TRIGGER_RUPEES
)
BREAKEVEN_ARM_PCT = _parse_pct_value(os.getenv("BREAKEVEN_ARM_PCT", "0"))
BREAKEVEN_LOCK_PCT = _parse_pct_value(os.getenv("BREAKEVEN_LOCK_PCT", "0"))
REENTRY_MAX_PREMIUM_RATIO = _env_float("REENTRY_MAX_PREMIUM_RATIO", 0.0)

PROFIT_TARGET_PCT = _parse_pct_value(os.getenv("PROFIT_TARGET_PCT", "0.05"))
REENTRY_ON_PROFIT_TARGET = _parse_bool(
    os.getenv("REENTRY_ON_PROFIT_TARGET"), False
)
REENTRY_DELAY_AFTER_TARGET = _parse_int_list(
    os.getenv("REENTRY_DELAY_AFTER_TARGET"), []
)
REENTRY_DELAY_BY_ATTEMPT = _parse_int_list(
    os.getenv("REENTRY_DELAY_BY_ATTEMPT"),
    [6, 6, 6, 6, 6, 6],
)
DAILY_PROFIT_TRAIL_ARM_RUPEES = _env_float(
    "DAILY_PROFIT_TRAIL_ARM_RUPEES", 0.0
)
DAILY_PROFIT_TRAIL_GIVEBACK_RUPEES = _env_float(
    "DAILY_PROFIT_TRAIL_GIVEBACK_RUPEES", 0.0
)
MIN_MINUTES_LEFT_FOR_REENTRY = _env_int("MIN_MINUTES_LEFT_FOR_REENTRY", 0)

QTY_UNITS = {
    "NIFTY": _env_int("NIFTY_QTY_UNITS", 325),
    "SENSEX": _env_int("SENSEX_QTY_UNITS", 100),
}
STRIKE_STEP = {
    "NIFTY": _env_int("NIFTY_STRIKE_STEP", 50),
    "SENSEX": _env_int("SENSEX_STRIKE_STEP", 100),
}
TRADEABLE = set(QTY_UNITS)

# Same transaction-cost model as the uploaded one-minute engine.
BROKERAGE_PER_ORDER = _env_float("BROKERAGE_PER_ORDER", 20.0)
ORDERS_PER_TRADE = _env_int("ORDERS_PER_TRADE", 4)
STT_SELL_PCT = _env_float("STT_SELL_PCT", 0.001)
EXCHANGE_TXN_PCT = _env_float("EXCHANGE_TXN_PCT", 0.0003553)
SEBI_PER_CRORE = _env_float("SEBI_PER_CRORE", 10.0)
STAMP_BUY_PCT = _env_float("STAMP_BUY_PCT", 0.00003)
IPFT_PER_CRORE = _env_float("IPFT_PER_CRORE", 0.010)
GST_PCT = _env_float("GST_PCT", 0.18)
INCLUDE_TRANSACTION_COSTS = _parse_bool(
    os.getenv("INCLUDE_TRANSACTION_COSTS"), True
)

LOOKBACK_MONTHS_RAW = os.getenv("LOOKBACK_MONTHS", "AUTO").strip()
if LOOKBACK_MONTHS_RAW.upper() in {"", "AUTO", "ALL", "MAX", "FULL"}:
    LOOKBACK_MONTHS: Optional[int] = None
else:
    LOOKBACK_MONTHS = int(round(float(LOOKBACK_MONTHS_RAW)))

_default_output = os.path.join(
    _downloads_folder(),
    f"short_straddle_v3_1sec_sqlite_entry{_safe_filename_part(ENTRY_TIME_IST)}"
    f"_exit{_safe_filename_part(EXIT_TIME_IST)}.xlsx",
)
OUTPUT_XLSX = os.path.abspath(
    os.path.expanduser(os.getenv("OUTPUT_XLSX", _default_output))
)


# =============================================================================
# STRATEGY HELPERS
# =============================================================================


def loss_limit_pct_for_attempt(attempt_index: int) -> float:
    values = LOSS_LIMIT_RUPEES_BY_ATTEMPT
    if not values:
        return 0.0
    return float(values[attempt_index]) if attempt_index < len(values) else float(values[-1])


def reentry_delay_for_attempt(attempt_index: int) -> int:
    values = REENTRY_DELAY_BY_ATTEMPT
    if not values:
        return 0
    return int(values[attempt_index]) if attempt_index < len(values) else int(values[-1])


def reentry_delay_after_target_for_attempt(attempt_index: int) -> int:
    values = REENTRY_DELAY_AFTER_TARGET
    if not values:
        return reentry_delay_for_attempt(attempt_index)
    return int(values[attempt_index]) if attempt_index < len(values) else int(values[-1])


def round_to_step(value: float, step: int) -> int:
    # Same rounding form as the one-minute engine.
    return int(round(value / step) * step)


def build_second_index(day_value: date) -> pd.DatetimeIndex:
    start = pd.Timestamp(datetime.combine(day_value, SESSION_START), tz=IST)
    end = pd.Timestamp(datetime.combine(day_value, SESSION_END), tz=IST)
    return pd.date_range(start=start, end=end, freq="1s")


def timestamp_for_day(day_value: date, clock: dtime) -> pd.Timestamp:
    return pd.Timestamp(datetime.combine(day_value, clock), tz=IST)


def compute_window_start(end_day: date, months: int) -> date:
    # Month-end exactness is not important for a user-supplied optional cap.
    return (pd.Timestamp(end_day) - pd.DateOffset(months=months)).date()


def determine_window_start(min_day: date, max_day: date) -> date:
    if LOOKBACK_MONTHS is None:
        return min_day
    return max(min_day, compute_window_start(max_day, LOOKBACK_MONTHS))


# =============================================================================
# TRANSACTION COSTS
# =============================================================================


def compute_trade_charges(
    entry_ce: float,
    entry_pe: float,
    exit_ce: float,
    exit_pe: float,
    qty: int,
) -> float:
    if not INCLUDE_TRANSACTION_COSTS:
        return 0.0

    entry_turnover = (entry_ce + entry_pe) * qty
    exit_turnover = (exit_ce + exit_pe) * qty
    total_turnover = entry_turnover + exit_turnover

    brokerage = BROKERAGE_PER_ORDER * ORDERS_PER_TRADE
    stt = entry_turnover * STT_SELL_PCT
    txn = total_turnover * EXCHANGE_TXN_PCT
    sebi = total_turnover * SEBI_PER_CRORE / 1_00_00_000
    stamp = exit_turnover * STAMP_BUY_PCT
    ipft = total_turnover * IPFT_PER_CRORE / 1_00_00_000
    gst = (brokerage + txn + sebi) * GST_PCT
    return round(brokerage + stt + txn + sebi + stamp + ipft + gst, 2)


# =============================================================================
# SQLITE CATALOGUE AND DATA STRUCTURES
# =============================================================================


@dataclass(frozen=True)
class DatabaseRecord:
    path: str
    filename: str
    trading_day: date
    underlying: str
    expiry: date
    days_to_expiry: int
    schema_version: int
    script_version: str
    time_basis: str
    save_only_price_changes: bool
    bar_count: int
    option_count: int
    first_second: Optional[int]
    last_second: Optional[int]
    reconnect_events: int
    error_events: int


@dataclass
class DataQualityRow:
    source_db: str
    day: date
    underlying: str
    expiry: date
    bars_loaded: int
    option_bars_loaded: int
    underlying_bars_loaded: int
    first_underlying_time: Optional[str]
    last_underlying_time: Optional[str]
    reconnect_events: int
    error_events: int
    anchors: int
    reconnect_anchors: int
    schema_version: int
    time_basis: str
    save_only_price_changes: bool
    usable: bool
    note: str


@dataclass
class TradeRow:
    source_db: str
    data_resolution: str
    day: date
    underlying: str
    trade_seq: int
    expiry: date
    days_to_expiry: int
    atm_strike: int
    qty_units: int
    entry_time: str
    exit_time: str
    exit_reason: str
    entry_underlying: float
    ce_symbol: str
    pe_symbol: str
    entry_ce: float
    entry_pe: float
    exit_ce: float
    exit_pe: float
    close_pnl_at_exit: float
    trigger_pnl_at_exit: float
    exit_pnl_gross: float
    txn_charges: float
    exit_pnl: float
    eod_pnl: float
    max_profit: float
    max_loss: float
    max_profit_before_exit: float
    entry_premium_sum: float
    stop_pct: float
    uncapped_stop_rupees: float
    stop_cap_rupees: float
    stop_rupees: float
    profit_protect_trigger_pct: float
    profit_protect_trigger_rupees: float
    daily_realized_pnl_after_trade: float
    daily_loss_limit_rupees: float
    daily_loss_limit_hit: bool


@dataclass
class DayData:
    record: DatabaseRecord
    underlying_rows: pd.DataFrame
    option_rows: pd.DataFrame
    feed_events: pd.DataFrame
    second_index: pd.DatetimeIndex

    def __post_init__(self) -> None:
        self._underlying_close_cache: Optional[pd.Series] = None
        self._leg_cache: Dict[str, pd.DataFrame] = {}
        self._symbol_map: Dict[Tuple[int, str], str] = {}
        if not self.option_rows.empty:
            grouped = self.option_rows.groupby(["strike_int", "option_type"], sort=False)
            for (strike, option_type), group in grouped:
                symbols = sorted(group["instrument"].dropna().astype(str).unique().tolist())
                if symbols:
                    self._symbol_map[(int(strike), str(option_type).upper())] = symbols[0]

    def pick_symbol(self, strike: int, option_type: str) -> Optional[str]:
        return self._symbol_map.get((int(strike), option_type.upper()))

    def underlying_close(self) -> pd.Series:
        if self._underlying_close_cache is None:
            sparse = _deduplicated_sparse_series(self.underlying_rows, "close")
            self._underlying_close_cache = sparse.reindex(self.second_index).ffill()
        return self._underlying_close_cache

    def leg_ohlc(self, symbol: str) -> pd.DataFrame:
        cached = self._leg_cache.get(symbol)
        if cached is not None:
            return cached

        rows = self.option_rows[
            self.option_rows["instrument"].astype(str).eq(str(symbol))
        ].copy()
        if rows.empty:
            empty = pd.DataFrame(
                index=self.second_index,
                columns=["open", "high", "low", "close", "source_age_seconds"],
                dtype="float64",
            )
            self._leg_cache[symbol] = empty
            return empty

        rows = rows.sort_values("date").drop_duplicates("date", keep="last").set_index("date")
        close_sparse = pd.to_numeric(rows["close"], errors="coerce")
        close = close_sparse.reindex(self.second_index).ffill()

        # Missing seconds represent no stored LTP transition. Their one-second
        # OHLC is therefore the carried LTP. Stored seconds retain the actual
        # one-second high and low captured by the collector.
        result = pd.DataFrame(index=self.second_index)
        for column in ("open", "high", "low"):
            sparse = pd.to_numeric(rows[column], errors="coerce").reindex(self.second_index)
            result[column] = sparse.where(sparse.notna(), close)
        result["close"] = close

        source_seconds = pd.Series(
            self.second_index.view("int64") // 1_000_000_000,
            index=self.second_index,
            dtype="float64",
        )
        event_marker = pd.Series(float("nan"), index=self.second_index)
        event_times = rows.index.intersection(self.second_index)
        if len(event_times):
            event_marker.loc[event_times] = (
                event_times.view("int64") // 1_000_000_000
            ).astype(float)
        last_event = event_marker.ffill()
        result["source_age_seconds"] = source_seconds - last_event

        self._leg_cache[symbol] = result
        return result



def _connect_readonly(path: str) -> sqlite3.Connection:
    uri = Path(path).resolve().as_uri() + "?mode=ro"
    connection = sqlite3.connect(uri, uri=True, timeout=30.0)
    connection.row_factory = sqlite3.Row
    return connection


def _table_exists(connection: sqlite3.Connection, name: str) -> bool:
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type IN ('table','view') AND name=?",
        (name,),
    ).fetchone()
    return row is not None


def _read_metadata(connection: sqlite3.Connection) -> Dict[str, str]:
    if not _table_exists(connection, "run_metadata"):
        return {}
    rows = connection.execute("SELECT key,value FROM run_metadata").fetchall()
    return {str(row[0]): str(row[1]) for row in rows}


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    return datetime.fromisoformat(text).date()


def _normalise_underlying(value: Any) -> Optional[str]:
    text = str(value or "").upper().strip()
    if "SENSEX" in text:
        return "SENSEX"
    if "BANKNIFTY" in text or "NIFTY BANK" in text:
        return "BANKNIFTY"
    if "NIFTY" in text:
        return "NIFTY"
    return None


def _parse_bool_metadata(value: Optional[str], default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def catalog_database(path: str) -> DatabaseRecord:
    connection = _connect_readonly(path)
    try:
        required = {"run_metadata", "instruments", "bars"}
        missing = [name for name in required if not _table_exists(connection, name)]
        if missing:
            raise RuntimeError(f"Missing collector tables: {missing}")

        metadata = _read_metadata(connection)
        trading_day_raw = metadata.get("trading_day")
        if not trading_day_raw:
            raise RuntimeError("run_metadata.trading_day is missing")
        trading_day = _parse_date(trading_day_raw)

        schema_version = int(metadata.get("schema_version", "0") or 0)
        if REQUIRE_COLLECTOR_SCHEMA_VERSION > 0 and schema_version != REQUIRE_COLLECTOR_SCHEMA_VERSION:
            raise RuntimeError(
                f"Unsupported collector schema_version={schema_version}; "
                f"required={REQUIRE_COLLECTOR_SCHEMA_VERSION}"
            )

        instrument_rows = connection.execute(
            """
            SELECT token,index_name,symbol,kind,option_type,strike,expiry,strike_step
            FROM instruments
            ORDER BY token
            """
        ).fetchall()
        if not instrument_rows:
            raise RuntimeError("instruments table is empty")

        underlying = _normalise_underlying(metadata.get("target_index"))
        if underlying is None:
            for row in instrument_rows:
                underlying = _normalise_underlying(row[1])
                if underlying:
                    break
        if underlying not in TRADEABLE:
            raise RuntimeError(f"Unsupported target index: {underlying!r}")

        expiries = sorted(
            {
                _parse_date(row[6])
                for row in instrument_rows
                if int(row[3]) == 2 and row[6]
            }
        )
        expiries = [expiry for expiry in expiries if expiry >= trading_day]
        if not expiries:
            raise RuntimeError("No non-expired option expiry in instruments table")
        expiry = expiries[0]

        bar_stats = connection.execute(
            """
            SELECT COUNT(*) AS n,
                   MIN(k >> 32) AS first_second,
                   MAX(k >> 32) AS last_second
            FROM bars
            """
        ).fetchone()
        bar_count = int(bar_stats[0] or 0)
        first_second = int(bar_stats[1]) if bar_stats[1] is not None else None
        last_second = int(bar_stats[2]) if bar_stats[2] is not None else None

        option_count = int(
            connection.execute("SELECT COUNT(*) FROM instruments WHERE kind=2").fetchone()[0]
        )

        reconnect_events = 0
        error_events = 0
        if _table_exists(connection, "feed_events"):
            reconnect_events = int(
                connection.execute(
                    "SELECT COUNT(*) FROM feed_events WHERE event_type IN ('RECONNECT','CLOSE')"
                ).fetchone()[0]
            )
            error_events = int(
                connection.execute(
                    "SELECT COUNT(*) FROM feed_events WHERE event_type IN ('ERROR','NORECONNECT')"
                ).fetchone()[0]
            )

        return DatabaseRecord(
            path=os.path.abspath(path),
            filename=os.path.basename(path),
            trading_day=trading_day,
            underlying=underlying,
            expiry=expiry,
            days_to_expiry=(expiry - trading_day).days,
            schema_version=schema_version,
            script_version=metadata.get("script_version", ""),
            time_basis=metadata.get("time_basis", "unknown"),
            save_only_price_changes=_parse_bool_metadata(
                metadata.get("save_only_price_changes"), True
            ),
            bar_count=bar_count,
            option_count=option_count,
            first_second=first_second,
            last_second=last_second,
            reconnect_events=reconnect_events,
            error_events=error_events,
        )
    finally:
        connection.close()


def discover_sqlite_paths() -> List[str]:
    base = Path(SQLITE_DIR)
    if not base.exists():
        raise FileNotFoundError(f"SQLITE_DIR does not exist: {base}")

    if base.is_file():
        candidates = [base]
    elif SQLITE_RECURSIVE:
        candidates = list(base.rglob(SQLITE_GLOB))
    else:
        candidates = list(base.glob(SQLITE_GLOB))

    paths = sorted(
        {
            str(path.resolve())
            for path in candidates
            if path.is_file()
            and not str(path).endswith(("-wal", "-shm", ".lock"))
        }
    )
    if not paths:
        raise FileNotFoundError(
            f"No SQLite databases matched {SQLITE_GLOB!r} under {SQLITE_DIR}"
        )
    return paths


def scan_database_catalog(paths: Sequence[str]) -> Tuple[List[DatabaseRecord], List[Dict[str, Any]]]:
    records: List[DatabaseRecord] = []
    skipped: List[Dict[str, Any]] = []
    for path in paths:
        try:
            record = catalog_database(path)
            records.append(record)
            print(
                f"[CATALOG OK] {record.filename}: {record.trading_day} "
                f"{record.underlying} expiry={record.expiry} DTE={record.days_to_expiry} "
                f"bars={record.bar_count}"
            )
        except Exception as exc:
            row = {"source_db": path, "reason": f"Catalogue failure: {exc}"}
            skipped.append(row)
            if FAIL_ON_DB_ERROR:
                raise RuntimeError(row["reason"]) from exc
            print(f"[CATALOG WARN] {os.path.basename(path)}: {exc}")

    if not records:
        raise RuntimeError("No usable collector databases were found")
    return records, skipped


def choose_unique_records(records: Sequence[DatabaseRecord]) -> Tuple[List[DatabaseRecord], List[Dict[str, Any]]]:
    """Avoid double-counting duplicate databases for the same day/index/expiry.

    The collector normally reopens the same file on restart. If duplicate files
    exist, retain the one with the largest bar count, then newest modification
    time as a deterministic tie-breaker.
    """

    groups: Dict[Tuple[date, str, date], List[DatabaseRecord]] = {}
    for record in records:
        groups.setdefault(
            (record.trading_day, record.underlying, record.expiry), []
        ).append(record)

    selected: List[DatabaseRecord] = []
    skipped: List[Dict[str, Any]] = []
    for key, group in groups.items():
        ranked = sorted(
            group,
            key=lambda item: (
                item.bar_count,
                os.path.getmtime(item.path),
                item.path,
            ),
            reverse=True,
        )
        selected.append(ranked[0])
        for duplicate in ranked[1:]:
            skipped.append(
                {
                    "source_db": duplicate.path,
                    "day": duplicate.trading_day,
                    "underlying": duplicate.underlying,
                    "expiry": duplicate.expiry,
                    "reason": (
                        "Duplicate day/index/expiry database; a larger or newer "
                        f"database was selected: {ranked[0].filename}"
                    ),
                }
            )
    return sorted(selected, key=lambda row: (row.trading_day, row.underlying)), skipped


def actual_underlying_by_day(records: Sequence[DatabaseRecord]) -> Dict[date, str]:
    by_day: Dict[date, List[Tuple[date, str]]] = {}
    for record in records:
        if record.underlying not in TRADEABLE:
            continue
        if record.days_to_expiry not in ALLOWED_DTE:
            continue
        by_day.setdefault(record.trading_day, []).append(
            (record.expiry, record.underlying)
        )

    result: Dict[date, str] = {}
    for day_value, candidates in by_day.items():
        # Preserve the original selector: nearest expiry; NIFTY on an expiry tie.
        selected = sorted(
            candidates,
            key=lambda item: (item[0], 0 if item[1] == "NIFTY" else 1),
        )[0]
        result[day_value] = selected[1]
    return result


# =============================================================================
# SQLITE DAY LOADING AND SPARSE-TO-SECOND RECONSTRUCTION
# =============================================================================


def _decode_rows_query() -> str:
    return """
        SELECT
            (b.k >> 32) AS second_of_day,
            (b.k & 4294967295) AS instrument_token,
            i.index_name,
            i.exchange,
            i.symbol,
            i.kind,
            i.option_type,
            i.strike,
            i.expiry,
            i.strike_step,
            i.lot_size,
            b.p AS previous_paise,
            b.o AS open_paise,
            b.h AS high_paise,
            b.l AS low_paise,
            b.c AS close_paise,
            b.off AS strike_offset,
            b.f AS flags
        FROM bars AS b
        JOIN instruments AS i
          ON i.token = (b.k & 4294967295)
        ORDER BY second_of_day, instrument_token
    """


def _deduplicated_sparse_series(rows: pd.DataFrame, column: str) -> pd.Series:
    if rows.empty:
        return pd.Series(dtype="float64")
    data = rows[["date", column]].dropna().copy()
    if data.empty:
        return pd.Series(dtype="float64")
    data = data.sort_values("date").drop_duplicates("date", keep="last").set_index("date")
    return pd.to_numeric(data[column], errors="coerce")


def load_day_data(record: DatabaseRecord) -> Tuple[DayData, DataQualityRow]:
    connection = _connect_readonly(record.path)
    try:
        frame = pd.read_sql_query(_decode_rows_query(), connection)
        if _table_exists(connection, "feed_events"):
            feed_events = pd.read_sql_query(
                "SELECT id,event_time,event_type,details FROM feed_events ORDER BY id",
                connection,
            )
        else:
            feed_events = pd.DataFrame(
                columns=["id", "event_time", "event_type", "details"]
            )
    finally:
        connection.close()

    if frame.empty:
        raise RuntimeError("bars table is empty")

    base = pd.Timestamp(record.trading_day, tz=IST)
    frame["date"] = base + pd.to_timedelta(frame["second_of_day"], unit="s")
    frame["name"] = frame["index_name"].astype(str).map(_normalise_underlying)
    frame["type"] = frame["kind"].map({1: "UNDERLYING", 2: "OPTION"})
    frame["option_type"] = frame["option_type"].fillna("").astype(str).str.upper()
    frame["strike_int"] = pd.to_numeric(frame["strike"], errors="coerce").round().astype("Int64")
    frame["instrument"] = frame["symbol"].astype(str)
    frame["expiry_date"] = pd.to_datetime(frame["expiry"], errors="coerce").dt.date

    for source, target in (
        ("previous_paise", "previous_price"),
        ("open_paise", "open"),
        ("high_paise", "high"),
        ("low_paise", "low"),
        ("close_paise", "close"),
    ):
        frame[target] = pd.to_numeric(frame[source], errors="coerce") / PRICE_SCALE

    frame["is_anchor"] = (pd.to_numeric(frame["flags"], errors="coerce").fillna(0).astype(int) & 1) != 0
    frame["is_reconnect_anchor"] = (
        pd.to_numeric(frame["flags"], errors="coerce").fillna(0).astype(int) & 4
    ) != 0

    underlying_rows = frame[
        (frame["type"] == "UNDERLYING") & (frame["name"] == record.underlying)
    ].copy()
    option_rows = frame[
        (frame["type"] == "OPTION")
        & (frame["name"] == record.underlying)
        & (frame["expiry_date"] == record.expiry)
    ].copy()
    option_rows = option_rows.dropna(subset=["strike_int", "option_type", "instrument", "close"])
    option_rows["strike_int"] = option_rows["strike_int"].astype(int)

    first_underlying = (
        underlying_rows["date"].min() if not underlying_rows.empty else pd.NaT
    )
    last_underlying = (
        underlying_rows["date"].max() if not underlying_rows.empty else pd.NaT
    )

    note_parts: List[str] = []
    usable = True
    entry_ts = timestamp_for_day(record.trading_day, ENTRY_TIME)

    if underlying_rows.empty:
        usable = False
        note_parts.append("No underlying rows")
    elif pd.isna(first_underlying) or first_underlying > entry_ts:
        usable = False
        note_parts.append(
            f"First underlying observation {first_underlying} is after entry {entry_ts}"
        )

    if option_rows.empty:
        usable = False
        note_parts.append("No option rows for selected expiry")

    if SKIP_DAY_ON_RECONNECT and record.reconnect_events > 0:
        usable = False
        note_parts.append(f"Reconnect/close events={record.reconnect_events}")
    if SKIP_DAY_ON_FEED_ERROR and record.error_events > 0:
        usable = False
        note_parts.append(f"Feed error events={record.error_events}")

    quality = DataQualityRow(
        source_db=record.path,
        day=record.trading_day,
        underlying=record.underlying,
        expiry=record.expiry,
        bars_loaded=int(len(frame)),
        option_bars_loaded=int(len(option_rows)),
        underlying_bars_loaded=int(len(underlying_rows)),
        first_underlying_time=(
            pd.Timestamp(first_underlying).strftime("%H:%M:%S")
            if pd.notna(first_underlying)
            else None
        ),
        last_underlying_time=(
            pd.Timestamp(last_underlying).strftime("%H:%M:%S")
            if pd.notna(last_underlying)
            else None
        ),
        reconnect_events=record.reconnect_events,
        error_events=record.error_events,
        anchors=int(frame["is_anchor"].sum()),
        reconnect_anchors=int(frame["is_reconnect_anchor"].sum()),
        schema_version=record.schema_version,
        time_basis=record.time_basis,
        save_only_price_changes=record.save_only_price_changes,
        usable=usable,
        note="; ".join(note_parts) if note_parts else "OK",
    )

    day_data = DayData(
        record=record,
        underlying_rows=underlying_rows,
        option_rows=option_rows,
        feed_events=feed_events,
        second_index=build_second_index(record.trading_day),
    )
    return day_data, quality


def _value_age_ok(age_seconds: float) -> bool:
    if pd.isna(age_seconds):
        return False
    return MAX_ENTRY_PRICE_AGE_SECONDS <= 0 or age_seconds <= MAX_ENTRY_PRICE_AGE_SECONDS


def _underlying_asof(day_data: DayData, timestamp: pd.Timestamp) -> float:
    series = day_data.underlying_close()
    if timestamp not in series.index:
        return float("nan")
    value = series.loc[timestamp]
    return float(value) if pd.notna(value) else float("nan")


# =============================================================================
# ONE-SECOND STRATEGY SIMULATION
# =============================================================================


def simulate_day_multi_trades(day_data: DayData) -> Tuple[List[TradeRow], List[Dict[str, Any]]]:
    record = day_data.record
    und = record.underlying
    dy = record.trading_day
    expiry = record.expiry

    results: List[TradeRow] = []
    skipped: List[Dict[str, Any]] = []

    if und not in TRADEABLE:
        return results, [
            {
                "source_db": record.path,
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "reason": f"Unsupported underlying {und}",
            }
        ]

    idx_all = day_data.second_index
    session_end_ts = idx_all[-1]
    configured_exit_ts = timestamp_for_day(dy, EXIT_TIME)
    trade_end_ts = min(session_end_ts, configured_exit_ts)

    qty = int(QTY_UNITS[und])
    expected_step = int(STRIKE_STEP[und])

    instrument_steps = pd.to_numeric(
        day_data.option_rows.get("strike_step"), errors="coerce"
    ).dropna()
    if not instrument_steps.empty:
        actual_step = int(round(float(instrument_steps.mode().iloc[0])))
        if actual_step != expected_step:
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "reason": (
                        f"Strike-step mismatch: DB={actual_step}, "
                        f"config={expected_step}"
                    ),
                }
            )
            return results, skipped

    cur_entry_ts = timestamp_for_day(dy, ENTRY_TIME)
    trade_seq = 1
    if cur_entry_ts >= trade_end_ts:
        skipped.append(
            {
                "source_db": record.path,
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq,
                "reason": (
                    f"No entry: ENTRY_TIME_IST {ENTRY_TIME_IST} is at/after "
                    f"EXIT_TIME_IST {EXIT_TIME_IST}"
                ),
            }
        )
        return results, skipped

    daily_realized_pnl = 0.0
    daily_realized_peak = 0.0
    daily_loss_enabled = MAX_DAILY_LOSS_RUPEES > 0
    previous_entry_premium_per_unit: Optional[float] = None
    profit_protect_enabled = PROFIT_PROTECT_ARM_PCT > 0.0

    while cur_entry_ts < trade_end_ts:
        if daily_loss_enabled and daily_realized_pnl <= -MAX_DAILY_LOSS_RUPEES:
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "reason": (
                        "Daily loss limit hit before entry: "
                        f"realized={daily_realized_pnl:.2f}, "
                        f"limit={MAX_DAILY_LOSS_RUPEES:.2f}"
                    ),
                }
            )
            break

        underlying_price = _underlying_asof(day_data, cur_entry_ts)
        if pd.isna(underlying_price):
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "reason": f"No underlying LTP at {cur_entry_ts.strftime('%H:%M:%S')}",
                }
            )
            break

        atm = round_to_step(float(underlying_price), expected_step)
        ce_symbol = day_data.pick_symbol(atm, "CE")
        pe_symbol = day_data.pick_symbol(atm, "PE")
        if not ce_symbol or not pe_symbol:
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "atm_strike": atm,
                    "reason": "ATM CE/PE not available in SQLite collector band",
                }
            )
            break

        ce = day_data.leg_ohlc(ce_symbol)
        pe = day_data.leg_ohlc(pe_symbol)
        if cur_entry_ts not in ce.index or cur_entry_ts not in pe.index:
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "reason": "Entry timestamp outside reconstructed second index",
                }
            )
            break

        ce_entry = ce.at[cur_entry_ts, "close"]
        pe_entry = pe.at[cur_entry_ts, "close"]
        ce_age = ce.at[cur_entry_ts, "source_age_seconds"]
        pe_age = pe.at[cur_entry_ts, "source_age_seconds"]
        if pd.isna(ce_entry) or pd.isna(pe_entry):
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "atm_strike": atm,
                    "reason": "No CE/PE LTP at entry after sparse-series reconstruction",
                }
            )
            break
        if not _value_age_ok(float(ce_age)) or not _value_age_ok(float(pe_age)):
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "atm_strike": atm,
                    "reason": (
                        "Entry price too stale: "
                        f"CE age={ce_age:.0f}s, PE age={pe_age:.0f}s, "
                        f"limit={MAX_ENTRY_PRICE_AGE_SECONDS}s"
                    ),
                }
            )
            break

        entry_premium_per_unit = float(ce_entry) + float(pe_entry)
        if (
            trade_seq > 1
            and REENTRY_MAX_PREMIUM_RATIO > 0
            and previous_entry_premium_per_unit is not None
            and entry_premium_per_unit
            > previous_entry_premium_per_unit * REENTRY_MAX_PREMIUM_RATIO
        ):
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "atm_strike": atm,
                    "reason": (
                        f"No re-entry: ATM premium {entry_premium_per_unit:.2f} > "
                        f"{REENTRY_MAX_PREMIUM_RATIO:.2f}x previous "
                        f"{previous_entry_premium_per_unit:.2f}"
                    ),
                }
            )
            break
        previous_entry_premium_per_unit = entry_premium_per_unit

        monitor_start_ts = cur_entry_ts + pd.Timedelta(seconds=1)
        if monitor_start_ts > trade_end_ts:
            break

        entry_premium_sum = entry_premium_per_unit * qty
        loss_pct = loss_limit_pct_for_attempt(trade_seq - 1)
        uncapped_stop = loss_pct * entry_premium_sum
        stop_cap = float(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
        stop_rupees = min(uncapped_stop, stop_cap) if stop_cap > 0 else uncapped_stop

        ce_close = ce["close"]
        pe_close = pe["close"]
        ce_high = ce["high"]
        ce_low = ce["low"]
        pe_high = pe["high"]
        pe_low = pe["low"]

        pnl_close_all = (
            (float(ce_entry) - ce_close) * qty
            + (float(pe_entry) - pe_close) * qty
        )
        pnl = pnl_close_all.loc[monitor_start_ts:trade_end_ts].dropna()

        # Same adverse intrabar stop model as A, now within each second.
        pnl_ce_high_pe_low_all = (
            (float(ce_entry) - ce_high) * qty
            + (float(pe_entry) - pe_low) * qty
        )
        pnl_ce_low_pe_high_all = (
            (float(ce_entry) - ce_low) * qty
            + (float(pe_entry) - pe_high) * qty
        )
        pnl_sl_all = pd.concat(
            [pnl_close_all, pnl_ce_high_pe_low_all, pnl_ce_low_pe_high_all],
            axis=1,
        ).min(axis=1)
        pnl_sl = pnl_sl_all.loc[monitor_start_ts:trade_end_ts].dropna()

        if pnl.empty or pnl_sl.empty:
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "atm_strike": atm,
                    "reason": "PnL series empty after entry",
                }
            )
            break

        eod_ts = pnl.index[-1]
        eod_pnl = float(pnl.iloc[-1])
        default_exit_reason = "TIME_EXIT" if trade_end_ts < session_end_ts else "EOD"
        max_profit = float(max(0.0, pnl.max()))
        max_loss = float(min(0.0, pnl.min()))

        stop_floor = pd.Series(-float(stop_rupees), index=pnl_sl.index)
        if BREAKEVEN_ARM_PCT > 0:
            running_peak = pnl.cummax().reindex(pnl_sl.index).ffill()
            breakeven_armed = running_peak >= BREAKEVEN_ARM_PCT * entry_premium_sum
            breakeven_floor = BREAKEVEN_LOCK_PCT * entry_premium_sum
            stop_floor = stop_floor.mask(
                breakeven_armed.fillna(False),
                max(-float(stop_rupees), float(breakeven_floor)),
            )

        stop_hit = pnl_sl <= stop_floor
        stop_ts = (
            pnl_sl.index[stop_hit.to_numpy().argmax()] if stop_hit.any() else None
        )

        protect_ts = None
        if profit_protect_enabled:
            peak = pnl.cummax()
            arm_rupees = PROFIT_PROTECT_ARM_PCT * entry_premium_sum
            giveback_rupees = PROFIT_PROTECT_GIVEBACK_PCT * entry_premium_sum
            armed = peak >= arm_rupees
            trail = peak - giveback_rupees
            protect_hit = armed & (pnl <= trail)
            protect_ts = (
                pnl.index[protect_hit.to_numpy().argmax()]
                if protect_hit.any()
                else None
            )

        target_ts = None
        target_rupees: Optional[float] = None
        pnl_tp = pd.Series(dtype="float64")
        if PROFIT_TARGET_PCT > 0:
            target_rupees = PROFIT_TARGET_PCT * entry_premium_sum
            # Exact parity with uploaded A: favourable CE-low + PE-low, but now
            # the two lows must belong to the same one-second bucket.
            pnl_best_all = (
                (float(ce_entry) - ce_low) * qty
                + (float(pe_entry) - pe_low) * qty
            )
            pnl_tp = pd.concat([pnl_close_all, pnl_best_all], axis=1).max(axis=1)
            pnl_tp = pnl_tp.loc[monitor_start_ts:trade_end_ts].dropna()
            target_hit = pnl_tp >= target_rupees
            target_ts = (
                pnl_tp.index[target_hit.to_numpy().argmax()]
                if target_hit.any()
                else None
            )

        exit_ts = eod_ts
        exit_reason = default_exit_reason
        candidates: List[Tuple[pd.Timestamp, int, str]] = []
        if stop_ts is not None:
            candidates.append((stop_ts, 0, "STOPLOSS"))
        if target_ts is not None:
            candidates.append((target_ts, 1, "PROFIT_TARGET"))
        if protect_ts is not None:
            candidates.append((protect_ts, 2, "PROFIT_PROTECT"))
        if candidates:
            candidates.sort(key=lambda item: (item[0], item[1]))
            exit_ts, _, exit_reason = candidates[0]

        close_pnl_at_exit = float(pnl_close_all.loc[exit_ts])
        if exit_reason == "STOPLOSS":
            trigger_pnl_at_exit = float(pnl_sl_all.loc[exit_ts])
            exit_pnl_gross = float(stop_floor.loc[exit_ts])
        elif exit_reason == "PROFIT_TARGET":
            trigger_pnl_at_exit = float(pnl_tp.loc[exit_ts])
            # Preserve A's exact-target booking convention.
            exit_pnl_gross = float(target_rupees)
        else:
            trigger_pnl_at_exit = close_pnl_at_exit
            exit_pnl_gross = close_pnl_at_exit

        pnl_pre_exit = pnl.loc[:exit_ts]
        max_profit_before_exit = (
            float(max(0.0, pnl_pre_exit.max())) if not pnl_pre_exit.empty else 0.0
        )

        exit_ce = float(ce_close.loc[exit_ts])
        exit_pe = float(pe_close.loc[exit_ts])
        charges = compute_trade_charges(
            float(ce_entry),
            float(pe_entry),
            exit_ce,
            exit_pe,
            qty,
        )
        exit_pnl = exit_pnl_gross - charges

        daily_realized_pnl += exit_pnl
        daily_realized_peak = max(daily_realized_peak, daily_realized_pnl)
        daily_loss_hit = bool(
            daily_loss_enabled
            and daily_realized_pnl <= -MAX_DAILY_LOSS_RUPEES
        )

        results.append(
            TradeRow(
                source_db=record.path,
                data_resolution="1sec_ltp_receipt_time",
                day=dy,
                underlying=und,
                trade_seq=trade_seq,
                expiry=expiry,
                days_to_expiry=record.days_to_expiry,
                atm_strike=int(atm),
                qty_units=qty,
                entry_time=cur_entry_ts.strftime("%H:%M:%S"),
                exit_time=exit_ts.strftime("%H:%M:%S"),
                exit_reason=exit_reason,
                entry_underlying=float(underlying_price),
                ce_symbol=ce_symbol,
                pe_symbol=pe_symbol,
                entry_ce=float(ce_entry),
                entry_pe=float(pe_entry),
                exit_ce=exit_ce,
                exit_pe=exit_pe,
                close_pnl_at_exit=close_pnl_at_exit,
                trigger_pnl_at_exit=trigger_pnl_at_exit,
                exit_pnl_gross=exit_pnl_gross,
                txn_charges=charges,
                exit_pnl=exit_pnl,
                eod_pnl=eod_pnl,
                max_profit=max_profit,
                max_loss=max_loss,
                max_profit_before_exit=max_profit_before_exit,
                entry_premium_sum=entry_premium_sum,
                stop_pct=loss_pct,
                uncapped_stop_rupees=uncapped_stop,
                stop_cap_rupees=stop_cap,
                stop_rupees=stop_rupees,
                profit_protect_trigger_pct=PROFIT_PROTECT_TRIGGER_RUPEES,
                profit_protect_trigger_rupees=(
                    PROFIT_PROTECT_TRIGGER_RUPEES * entry_premium_sum
                ),
                daily_realized_pnl_after_trade=daily_realized_pnl,
                daily_loss_limit_rupees=MAX_DAILY_LOSS_RUPEES,
                daily_loss_limit_hit=daily_loss_hit,
            )
        )

        if daily_loss_hit:
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq + 1,
                    "reason": (
                        f"No re-entry: daily loss limit hit; "
                        f"realized={daily_realized_pnl:.2f}, "
                        f"limit={MAX_DAILY_LOSS_RUPEES:.2f}"
                    ),
                }
            )
            break

        if (
            DAILY_PROFIT_TRAIL_GIVEBACK_RUPEES > 0
            and daily_realized_peak >= DAILY_PROFIT_TRAIL_ARM_RUPEES
            and daily_realized_pnl
            <= daily_realized_peak - DAILY_PROFIT_TRAIL_GIVEBACK_RUPEES
        ):
            skipped.append(
                {
                    "source_db": record.path,
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq + 1,
                    "reason": (
                        f"No re-entry: day profit trail hit; "
                        f"realized={daily_realized_pnl:.2f}, "
                        f"peak={daily_realized_peak:.2f}, "
                        f"giveback={DAILY_PROFIT_TRAIL_GIVEBACK_RUPEES:.2f}"
                    ),
                }
            )
            break

        reenter_reasons = (
            ("STOPLOSS", "PROFIT_PROTECT", "PROFIT_TARGET")
            if REENTRY_ON_PROFIT_TARGET
            else ("STOPLOSS", "PROFIT_PROTECT")
        )
        if exit_reason in reenter_reasons and (trade_seq - 1) < MAX_REATTEMPTS:
            if exit_reason == "PROFIT_TARGET":
                delay_minutes = reentry_delay_after_target_for_attempt(trade_seq - 1)
            else:
                delay_minutes = reentry_delay_for_attempt(trade_seq - 1)

            trade_seq += 1
            cur_entry_ts = exit_ts + pd.Timedelta(minutes=delay_minutes)

            if (
                MIN_MINUTES_LEFT_FOR_REENTRY > 0
                and trade_end_ts - cur_entry_ts
                < pd.Timedelta(minutes=MIN_MINUTES_LEFT_FOR_REENTRY)
            ):
                skipped.append(
                    {
                        "source_db": record.path,
                        "day": dy,
                        "underlying": und,
                        "expiry": expiry,
                        "trade_seq": trade_seq,
                        "reason": (
                            f"No re-entry: fewer than "
                            f"{MIN_MINUTES_LEFT_FOR_REENTRY} minutes remain"
                        ),
                    }
                )
                break
            if cur_entry_ts >= trade_end_ts:
                skipped.append(
                    {
                        "source_db": record.path,
                        "day": dy,
                        "underlying": und,
                        "expiry": expiry,
                        "trade_seq": trade_seq,
                        "reason": (
                            f"No re-entry: next entry "
                            f"{cur_entry_ts.strftime('%H:%M:%S')} is at/after "
                            f"EXIT_TIME_IST {EXIT_TIME_IST}"
                        ),
                    }
                )
                break
            continue

        break

    return results, skipped


# =============================================================================
# MULTI-DATABASE PROCESSING
# =============================================================================


def process_databases(
    records: Sequence[DatabaseRecord],
    window_start: date,
    window_end: date,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    trade_rows: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []
    quality_rows: List[Dict[str, Any]] = []

    for record in records:
        if not (window_start <= record.trading_day <= window_end):
            continue
        if record.days_to_expiry not in ALLOWED_DTE:
            skipped_rows.append(
                {
                    "source_db": record.path,
                    "day": record.trading_day,
                    "underlying": record.underlying,
                    "expiry": record.expiry,
                    "reason": (
                        f"DTE {record.days_to_expiry} is not in ALLOWED_DTE={ALLOWED_DTE}"
                    ),
                }
            )
            continue

        try:
            day_data, quality = load_day_data(record)
            quality_rows.append(asdict(quality))
            if not quality.usable:
                skipped_rows.append(
                    {
                        "source_db": record.path,
                        "day": record.trading_day,
                        "underlying": record.underlying,
                        "expiry": record.expiry,
                        "reason": f"Data-quality rejection: {quality.note}",
                    }
                )
                continue

            trades, skips = simulate_day_multi_trades(day_data)
            trade_rows.extend(asdict(trade) for trade in trades)
            skipped_rows.extend(skips)
            print(
                f"[SIM OK] {record.filename}: trades={len(trades)} "
                f"skips={len(skips)}"
            )
        except Exception as exc:
            row = {
                "source_db": record.path,
                "day": record.trading_day,
                "underlying": record.underlying,
                "expiry": record.expiry,
                "reason": f"Simulation failure: {exc}",
            }
            skipped_rows.append(row)
            if FAIL_ON_DB_ERROR:
                raise RuntimeError(row["reason"]) from exc
            print(f"[SIM WARN] {record.filename}: {exc}")

    all_trades = pd.DataFrame(trade_rows)
    if not all_trades.empty:
        all_trades = all_trades.sort_values(
            ["day", "underlying", "trade_seq"]
        ).reset_index(drop=True)

    skipped = pd.DataFrame(skipped_rows)
    if not skipped.empty:
        sort_columns = [column for column in ("day", "underlying", "trade_seq") if column in skipped]
        if sort_columns:
            skipped = skipped.sort_values(sort_columns, na_position="last").reset_index(drop=True)

    quality_df = pd.DataFrame(quality_rows)
    if not quality_df.empty:
        quality_df = quality_df.sort_values(["day", "underlying"]).reset_index(drop=True)

    return all_trades, skipped, quality_df


def build_actual_trades_df(
    all_trades: pd.DataFrame,
    actual_selector: Mapping[date, str],
) -> pd.DataFrame:
    if all_trades.empty:
        return pd.DataFrame()

    result = all_trades.copy()
    result["actual_underlying_for_day"] = result["day"].map(actual_selector)
    result = result[result["actual_underlying_for_day"].notna()]
    result = result[result["underlying"] == result["actual_underlying_for_day"]]
    result = result[result["days_to_expiry"].isin(ALLOWED_DTE)]
    result = result.drop(columns=["actual_underlying_for_day"])
    result = result.sort_values(["day", "trade_seq"]).reset_index(drop=True)
    result["is_exit_pnl_positive"] = (result["exit_pnl"] > 0).astype(int)
    return result


# =============================================================================
# EXCEL REPORT
# =============================================================================


def _autosize_columns_safe(worksheet: Any) -> None:
    try:
        max_column = worksheet.max_column or 0
        max_row = worksheet.max_row or 0
        for column_index in range(1, max_column + 1):
            letter = worksheet.cell(row=1, column=column_index).column_letter
            width = 0
            for row_index in range(1, min(max_row, 2000) + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is not None:
                    width = max(width, len(str(value)))
            worksheet.column_dimensions[letter].width = min(60, max(10, width + 2))
    except Exception:
        return


def _colour_actual_trades_by_date(workbook: Any, actual_trades: pd.DataFrame) -> None:
    if actual_trades.empty or "actual_trades" not in workbook.sheetnames:
        return
    from openpyxl.styles import PatternFill

    worksheet = workbook["actual_trades"]
    fills = [
        PatternFill(fill_type="solid", fgColor="E8F0FE"),
        PatternFill(fill_type="solid", fgColor="FFF3E0"),
    ]
    colour_index = 0
    previous_day: Any = None
    for row_offset, day_value in enumerate(actual_trades["day"].tolist(), start=2):
        if previous_day is not None and day_value != previous_day:
            colour_index ^= 1
        previous_day = day_value
        for column_index in range(1, len(actual_trades.columns) + 1):
            worksheet.cell(row=row_offset, column=column_index).fill = fills[colour_index]


def _monthwise_summary(actual_trades: pd.DataFrame) -> pd.DataFrame:
    if actual_trades.empty:
        return pd.DataFrame()

    data = actual_trades.copy()
    data["month"] = pd.to_datetime(data["day"]).dt.to_period("M").astype(str)
    summary = (
        data.groupby("month", as_index=False)
        .agg(
            trades=("exit_pnl", "count"),
            total_exit_pnl=("exit_pnl", "sum"),
            avg_exit_pnl=("exit_pnl", "mean"),
            winning_trades=("is_exit_pnl_positive", "sum"),
        )
    )
    summary["losing_trades"] = summary["trades"] - summary["winning_trades"]
    summary["win_rate_pct"] = (
        100.0 * summary["winning_trades"] / summary["trades"]
    ).round(2)

    daily = data.groupby(["month", "day"], as_index=False).agg(
        daily_pnl=("exit_pnl", "sum")
    )
    loss_stats = daily.groupby("month", as_index=False).agg(
        avg_loss_on_loss_days=(
            "daily_pnl",
            lambda series: float(series[series < 0].mean())
            if (series < 0).any()
            else 0.0,
        ),
        max_loss_in_a_day=("daily_pnl", "min"),
        profitable_days=("daily_pnl", lambda series: int((series > 0).sum())),
        loss_days=("daily_pnl", lambda series: int((series < 0).sum())),
    )
    worst_rows = daily.loc[daily.groupby("month")["daily_pnl"].idxmin()]
    worst_dates = worst_rows[["month", "day"]].rename(
        columns={"day": "max_loss_day_date"}
    )
    summary = summary.merge(loss_stats, on="month", how="left")
    summary = summary.merge(worst_dates, on="month", how="left")

    worst_index = daily["daily_pnl"].idxmin()
    total_trades = int(summary["trades"].sum())
    total_wins = int(summary["winning_trades"].sum())
    total_pnl = float(summary["total_exit_pnl"].sum())
    negative_days = daily[daily["daily_pnl"] < 0]
    total_row = {
        "month": "TOTAL",
        "trades": total_trades,
        "total_exit_pnl": round(total_pnl, 2),
        "avg_exit_pnl": round(total_pnl / total_trades, 2) if total_trades else 0.0,
        "winning_trades": total_wins,
        "losing_trades": total_trades - total_wins,
        "win_rate_pct": round(100.0 * total_wins / total_trades, 2) if total_trades else 0.0,
        "avg_loss_on_loss_days": (
            round(float(negative_days["daily_pnl"].mean()), 2)
            if not negative_days.empty
            else 0.0
        ),
        "max_loss_in_a_day": round(float(daily["daily_pnl"].min()), 2),
        "profitable_days": int((daily["daily_pnl"] > 0).sum()),
        "loss_days": int((daily["daily_pnl"] < 0).sum()),
        "max_loss_day_date": daily.loc[worst_index, "day"],
    }
    total_df = pd.DataFrame([total_row]).reindex(columns=summary.columns)
    return pd.concat([summary, total_df], ignore_index=True)


def write_excel(
    all_trades: pd.DataFrame,
    actual_trades: pd.DataFrame,
    skipped: pd.DataFrame,
    quality: pd.DataFrame,
    catalog: pd.DataFrame,
) -> None:
    output = Path(OUTPUT_XLSX)
    output.parent.mkdir(parents=True, exist_ok=True)

    if not all_trades.empty:
        exit_pivot = all_trades.pivot_table(
            index="day", columns="underlying", values="exit_pnl", aggfunc="sum"
        ).reset_index()
        first = all_trades[all_trades["trade_seq"] == 1]
        eod_pivot = first.pivot_table(
            index="day", columns="underlying", values="eod_pnl", aggfunc="sum"
        ).reset_index()

        instrument_data = all_trades.copy()
        instrument_data["is_win"] = instrument_data["exit_pnl"] > 0
        instrument_data["is_stoploss"] = instrument_data["exit_reason"].eq("STOPLOSS")
        instrument_data["is_target"] = instrument_data["exit_reason"].eq("PROFIT_TARGET")
        instrument_data["is_protect"] = instrument_data["exit_reason"].eq("PROFIT_PROTECT")
        instrument_summary = (
            instrument_data.groupby("underlying", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                win_rate_pct=("is_win", lambda series: 100.0 * series.mean()),
                stoploss_rate_pct=("is_stoploss", lambda series: 100.0 * series.mean()),
                target_rate_pct=("is_target", lambda series: 100.0 * series.mean()),
                profit_protect_rate_pct=("is_protect", lambda series: 100.0 * series.mean()),
                avg_max_profit=("max_profit", "mean"),
                avg_max_loss=("max_loss", "mean"),
                worst_max_loss=("max_loss", "min"),
            )
            .sort_values("total_exit_pnl", ascending=False)
            .reset_index(drop=True)
        )
    else:
        exit_pivot = pd.DataFrame()
        eod_pivot = pd.DataFrame()
        instrument_summary = pd.DataFrame()

    monthwise = _monthwise_summary(actual_trades)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        all_trades.to_excel(writer, sheet_name="all_trades_backtested", index=False)
        actual_trades.to_excel(writer, sheet_name="actual_trades", index=False)
        monthwise.to_excel(writer, sheet_name="monthwise_summary", index=False)
        exit_pivot.to_excel(writer, sheet_name="exit_pnl_pivot", index=False)
        eod_pivot.to_excel(writer, sheet_name="eod_pnl_first_trade_pivot", index=False)
        instrument_summary.to_excel(writer, sheet_name="instrument_summary", index=False)
        catalog.to_excel(writer, sheet_name="db_catalog", index=False)
        quality.to_excel(writer, sheet_name="data_quality", index=False)
        skipped.to_excel(writer, sheet_name="skipped", index=False)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            _autosize_columns_safe(worksheet)
        _colour_actual_trades_by_date(workbook, actual_trades)

    print(f"[DONE] Excel written: {output}")


# =============================================================================
# VALIDATION AND MAIN
# =============================================================================


def validate_configuration() -> None:
    if ENTRY_TIME >= EXIT_TIME:
        raise RuntimeError("ENTRY_TIME_IST must be earlier than EXIT_TIME_IST")
    if not ALLOWED_DTE:
        raise RuntimeError("ALLOWED_DTE cannot be empty")
    if MAX_REATTEMPTS < 0:
        raise RuntimeError("MAX_REATTEMPTS cannot be negative")
    if PROFIT_TARGET_PCT < 0:
        raise RuntimeError("PROFIT_TARGET_PCT cannot be negative")
    if MAX_ENTRY_PRICE_AGE_SECONDS < 0:
        raise RuntimeError("MAX_ENTRY_PRICE_AGE_SECONDS cannot be negative")


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backtest the v3 ATM short straddle on one-second collector SQLite databases."
    )
    parser.add_argument(
        "--config",
        help="Properties file. STRADDLE_CONFIG is used when this is omitted.",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Catalogue and validate databases without running the strategy.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    validate_configuration()

    print("=" * 92)
    print("ATM short-straddle v3 — one-second SQLite backtest")
    print(f"Config: {PROPERTY_FILE_PATH}")
    print(f"SQLite directory: {SQLITE_DIR}")
    print(f"SQLite glob: {SQLITE_GLOB} | recursive={SQLITE_RECURSIVE}")
    print(f"Entry/exit: {ENTRY_TIME_IST} -> {EXIT_TIME_IST}")
    print(f"Allowed DTE: {ALLOWED_DTE}")
    print(f"Profit target: {PROFIT_TARGET_PCT:.2%} | re-enter target={REENTRY_ON_PROFIT_TARGET}")
    print(f"Max reattempts: {MAX_REATTEMPTS}")
    print(f"Output: {OUTPUT_XLSX}")
    print("=" * 92)

    paths = discover_sqlite_paths()
    print(f"[INFO] SQLite databases found: {len(paths)}")

    records, catalog_skips = scan_database_catalog(paths)
    records, duplicate_skips = choose_unique_records(records)

    min_day = min(record.trading_day for record in records)
    max_day = max(record.trading_day for record in records)
    window_start = determine_window_start(min_day, max_day)
    print(f"[INFO] Data day range: {min_day} -> {max_day}")
    print(f"[INFO] Backtest window: {window_start} -> {max_day}")

    selector = actual_underlying_by_day(records)
    catalog_df = pd.DataFrame(asdict(record) for record in records)
    if not catalog_df.empty:
        catalog_df["selected_as_actual"] = catalog_df.apply(
            lambda row: selector.get(row["trading_day"]) == row["underlying"],
            axis=1,
        )

    if args.validate_only:
        quality_rows: List[Dict[str, Any]] = []
        skipped_rows = catalog_skips + duplicate_skips
        for record in records:
            try:
                _, quality = load_day_data(record)
                quality_rows.append(asdict(quality))
            except Exception as exc:
                skipped_rows.append(
                    {
                        "source_db": record.path,
                        "day": record.trading_day,
                        "underlying": record.underlying,
                        "reason": f"Validation failure: {exc}",
                    }
                )
        quality_df = pd.DataFrame(quality_rows)
        skipped_df = pd.DataFrame(skipped_rows)
        print(quality_df.to_string(index=False) if not quality_df.empty else "No quality rows")
        if not skipped_df.empty:
            print("\nSkipped/errors:\n" + skipped_df.to_string(index=False))
        return 0

    all_trades, skipped_df, quality_df = process_databases(
        records,
        window_start,
        max_day,
    )
    extra_skips = pd.DataFrame(catalog_skips + duplicate_skips)
    if not extra_skips.empty:
        skipped_df = pd.concat([extra_skips, skipped_df], ignore_index=True, sort=False)

    actual_trades = build_actual_trades_df(all_trades, selector)
    write_excel(all_trades, actual_trades, skipped_df, quality_df, catalog_df)

    if actual_trades.empty:
        print("[WARN] No actual trades. Inspect db_catalog, data_quality and skipped sheets.")
    else:
        total = float(actual_trades["exit_pnl"].sum())
        days = int(actual_trades["day"].nunique())
        trades = int(len(actual_trades))
        print(f"[SUMMARY] actual trades={trades}, days={days}, net P&L={total:.2f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
