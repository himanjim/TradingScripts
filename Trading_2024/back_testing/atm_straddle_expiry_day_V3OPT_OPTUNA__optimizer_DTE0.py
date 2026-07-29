"""
V3-OPT ATM short-straddle backtester + Optuna profit optimizer.

This file is the optimizer-enabled version of:
    A. atm_straddle_expiry_day_V3OPT.py
    B. straddle_config_v3opt_DTE0.properties

It preserves the V3 strategy rules:
    * hard entry/exit cut-offs;
    * momentum-gated first entries and re-entries;
    * percentage stop-loss with an absolute rupee cap;
    * per-underlying profit targets, late targets, profit-protect settings,
      late-session tighter giveback, and daily-loss circuit breakers;
    * target-based re-entry and stop/protect re-entry;
    * transaction charges and actual-trade selection.

RUN_MODE is read from the same property file:
    RUN_MODE=backtest  -> one normal backtest and Excel report.
    RUN_MODE=optimize  -> load data once, cache each trading day, then run Optuna.

The optimizer maximizes total net profit by default. Set OPT_CV_FOLDS above 1
only when you deliberately want a walk-forward consistency penalty.

Dependencies:
    pandas, openpyxl, optuna
"""

# =============================================================================
# V3-OPT = V2-OPT + per-underlying profiles:
#   Any of these keys may be prefixed NIFTY_ / SENSEX_ to override per index:
#     PROFIT_TARGET_PCT, PROFIT_TARGET_PCT_LATE, PROFIT_TARGET_LATE_FROM_IST,
#     PROFIT_PROTECT_TRIGGER_RUPEES, MAX_DAILY_LOSS_RUPEES,
#     PROFIT_PROTECT_LATE_GIVEBACK, PROFIT_PROTECT_LATE_FROM_IST
#   New: PROFIT_PROTECT_LATE_GIVEBACK - once armed, from
#   PROFIT_PROTECT_LATE_FROM_IST onward the trailing giveback tightens to this
#   fraction of premium (captures late-day peaks the wide trail returns).
# -----------------------------------------------------------------------------
# V2-OPT (expiry-day profit optimization) - changes vs original:
#   1. ENTRY_MOMENTUM_GATE: first entry, stop/protect re-entries and target
#      re-entries wait (minute by minute) until the current-ATM straddle
#      premium is LOWER than it was MOMENTUM_LOOKBACK_MIN minutes earlier,
#      i.e. entries only happen while premium is already decaying.
#   2. PROFIT_TARGET_REENTRY_ENABLED: hitting the profit target no longer ends
#      the day; a fresh ATM straddle is opened after
#      PROFIT_TARGET_REENTRY_DELAY minutes (plus the momentum gate).
#   3. PROFIT_TARGET_PCT_LATE: attempts ENTERED at/after
#      PROFIT_TARGET_LATE_FROM_IST use a smaller profit target (the last-hour
#      premium crush cannot deliver the full-day percentage).
#   Everything else (stop-loss model, profit-protect, charges, daily circuit
#   breaker, quantities) is IDENTICAL to the original engine.
# =============================================================================
import os
from pathlib import Path
import glob
import time
import json
from dataclasses import dataclass
from datetime import datetime, date, time as dtime, timedelta
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd

import Trading_2024.OptionTradeUtils as oUtils

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore

try:
    import pytz  # type: ignore
except Exception:
    pytz = None  # type: ignore

try:
    from dateutil.relativedelta import relativedelta  # type: ignore
except Exception:
    relativedelta = None  # type: ignore


# =============================================================================
# USER CONFIG
# =============================================================================
# ---------------------------------------------------------------------------
# CONFIGURATION SOURCE: external property file
# ---------------------------------------------------------------------------
# Every tunable setting lives in a simple KEY=VALUE property file so it can be
# changed WITHOUT editing this script. Path defaults to
# "straddle_config.properties" next to this file; override with the
# STRADDLE_CONFIG environment variable. Values are pushed into the process
# environment so all the os.getenv(...) reads below pick them up. A real
# environment variable that is already set takes precedence over the file.
def _load_property_file() -> str:
    cfg_path = os.getenv(
        "STRADDLE_CONFIG",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "configs/straddle_config_v3opt_optuna.properties"),
    )
    if not os.path.exists(cfg_path):
        print(f"[CONFIG] Property file not found at {cfg_path}; using built-in defaults.")
        return cfg_path
    loaded = 0
    with open(cfg_path, "r", encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith(";"):
                continue
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            key, val = key.strip(), val.strip()
            # strip optional surrounding quotes
            if len(val) >= 2 and val[0] == val[-1] and val[0] in ("'", '"'):
                val = val[1:-1]
            if key and key not in os.environ:   # real env vars win over the file
                os.environ[key] = val
                loaded += 1
    print(f"[CONFIG] Loaded {loaded} setting(s) from {cfg_path}")
    return cfg_path

PROPERTY_FILE_PATH = _load_property_file()

# =============================================================================
# RUN CONTROL AND OPTIMIZER SETTINGS
# =============================================================================
# All settings can be placed in straddle_config_v3opt_DTE0.properties. A real process
# environment variable still takes precedence because the property loader above
# does not overwrite existing environment variables.

RUN_MODE = os.getenv("RUN_MODE", "optimize").strip().lower()

# Number of Optuna trials. V3 has a wider search space than the earlier six-
# parameter optimizer; 800-1500 trials is a more realistic production search.
OPT_TRIALS = int(float(os.getenv("OPT_TRIALS", "1000")))

# 1 = maximize total net profit over the complete sample.
# >1 = contiguous monthly folds; objective becomes mean fold profit minus a
#      dispersion penalty. This is optional and is not pure total-profit scoring.
OPT_CV_FOLDS = int(float(os.getenv("OPT_CV_FOLDS", "1")))
OPT_CV_PENALTY = float(os.getenv("OPT_CV_PENALTY", "0.15"))

OPT_SEED = int(float(os.getenv("OPT_SEED", "42")))
OPT_STARTUP_TRIALS = int(float(os.getenv("OPT_STARTUP_TRIALS", "50")))
OPT_PROGRESS_EVERY = max(1, int(float(os.getenv("OPT_PROGRESS_EVERY", "1"))))

# Trial results are flushed after every completed trial.
_OPT_DEFAULT_PARENT = Path.home() / "Downloads"
if not _OPT_DEFAULT_PARENT.exists():
    _OPT_DEFAULT_PARENT = Path.home()
OPT_OUTPUT_DIR = os.getenv(
    "OPT_OUTPUT_DIR",
    str(_OPT_DEFAULT_PARENT / "v3opt_optimizer_runs"),
)
OPT_STUDY_NAME = os.getenv("OPT_STUDY_NAME", "short_straddle_v3opt_profit")
OPT_SAVE_DB = os.getenv("OPT_SAVE_DB", "1").strip() == "1"

# Smoke-test controls. Leave blank/0 for the full data set.
_SAMPLE_PICKLES_RAW = os.getenv("SAMPLE_MAX_PICKLES", "").strip()
_SAMPLE_DAYS_RAW = os.getenv("SAMPLE_MAX_DAYS", "").strip()
SAMPLE_MAX_PICKLES = int(float(_SAMPLE_PICKLES_RAW)) if _SAMPLE_PICKLES_RAW else None
SAMPLE_MAX_DAYS = int(float(_SAMPLE_DAYS_RAW)) if _SAMPLE_DAYS_RAW else None

# Minimum activity guards. These stop a broken/thin configuration from winning
# merely because it produced only a few lucky trades.
OPT_MIN_DAYS = int(float(os.getenv("OPT_MIN_DAYS", "100")))
OPT_MIN_MONTHS = int(float(os.getenv("OPT_MIN_MONTHS", "6")))
_OPT_DISQUALIFY = -1.0e18

# Search-group switches. Setting a switch to 0 freezes that group at the values
# already present in straddle_config_v3opt_DTE0.properties.
OPT_TUNE_TIMES = os.getenv("OPT_TUNE_TIMES", "1").strip() == "1"
OPT_TUNE_STOPLOSS = os.getenv("OPT_TUNE_STOPLOSS", "1").strip() == "1"
OPT_TUNE_REENTRY = os.getenv("OPT_TUNE_REENTRY", "1").strip() == "1"
OPT_TUNE_PROFILES = os.getenv("OPT_TUNE_PROFILES", "1").strip() == "1"
OPT_TUNE_DAILY_LOSS = os.getenv("OPT_TUNE_DAILY_LOSS", "1").strip() == "1"

# Search bounds. These defaults are intentionally centred around the supplied
# V3 configuration rather than using extremely broad, wasteful ranges.
OPT_ENTRY_FROM_IST = os.getenv("OPT_ENTRY_FROM_IST", "09:15")
OPT_ENTRY_TO_IST = os.getenv("OPT_ENTRY_TO_IST", "10:15")
OPT_ENTRY_STEP_MIN = max(1, int(float(os.getenv("OPT_ENTRY_STEP_MIN", "1"))))

OPT_EXIT_FROM_IST = os.getenv("OPT_EXIT_FROM_IST", "15:00")
OPT_EXIT_TO_IST = os.getenv("OPT_EXIT_TO_IST", "15:30")
OPT_EXIT_STEP_MIN = max(1, int(float(os.getenv("OPT_EXIT_STEP_MIN", "1"))))

OPT_SL_BASE_MIN = float(os.getenv("OPT_SL_BASE_MIN", "0.20"))
OPT_SL_BASE_MAX = float(os.getenv("OPT_SL_BASE_MAX", "0.60"))
OPT_SL_STEP_MIN = float(os.getenv("OPT_SL_STEP_MIN", "0.00"))
OPT_SL_STEP_MAX = float(os.getenv("OPT_SL_STEP_MAX", "0.08"))
OPT_STOP_CAP_MIN = int(float(os.getenv("OPT_STOP_CAP_MIN", "2000")))
OPT_STOP_CAP_MAX = int(float(os.getenv("OPT_STOP_CAP_MAX", "5000")))
OPT_STOP_CAP_STEP = max(1, int(float(os.getenv("OPT_STOP_CAP_STEP", "250"))))

OPT_REATTEMPTS_MIN = int(float(os.getenv("OPT_REATTEMPTS_MIN", "3")))
OPT_REATTEMPTS_MAX = int(float(os.getenv("OPT_REATTEMPTS_MAX", "12")))
OPT_REENTRY_DELAY_MIN = int(float(os.getenv("OPT_REENTRY_DELAY_MIN", "1")))
OPT_REENTRY_DELAY_MAX = int(float(os.getenv("OPT_REENTRY_DELAY_MAX", "15")))
OPT_REENTRY_STEP_MIN = int(float(os.getenv("OPT_REENTRY_STEP_MIN", "0")))
OPT_REENTRY_STEP_MAX = int(float(os.getenv("OPT_REENTRY_STEP_MAX", "3")))
OPT_TARGET_REENTRY_DELAY_MIN = int(float(os.getenv("OPT_TARGET_REENTRY_DELAY_MIN", "1")))
OPT_TARGET_REENTRY_DELAY_MAX = int(float(os.getenv("OPT_TARGET_REENTRY_DELAY_MAX", "10")))
OPT_MOMENTUM_LOOKBACK_MIN = int(float(os.getenv("OPT_MOMENTUM_LOOKBACK_MIN", "1")))
OPT_MOMENTUM_LOOKBACK_MAX = int(float(os.getenv("OPT_MOMENTUM_LOOKBACK_MAX", "8")))

OPT_PROTECT_MIN = float(os.getenv("OPT_PROTECT_MIN", "0.20"))
OPT_PROTECT_MAX = float(os.getenv("OPT_PROTECT_MAX", "0.50"))
OPT_LATE_GIVEBACK_MIN = float(os.getenv("OPT_LATE_GIVEBACK_MIN", "0.00"))
OPT_LATE_GIVEBACK_MAX = float(os.getenv("OPT_LATE_GIVEBACK_MAX", "0.45"))
OPT_TARGET_MIN = float(os.getenv("OPT_TARGET_MIN", "0.55"))
OPT_TARGET_MAX = float(os.getenv("OPT_TARGET_MAX", "0.90"))
OPT_LATE_TARGET_MIN = float(os.getenv("OPT_LATE_TARGET_MIN", "0.30"))
OPT_LATE_TARGET_MAX = float(os.getenv("OPT_LATE_TARGET_MAX", "0.65"))
OPT_LATE_FROM_IST = os.getenv("OPT_LATE_FROM_IST", "13:30")
OPT_LATE_TO_IST = os.getenv("OPT_LATE_TO_IST", "15:00")
OPT_LATE_TIME_STEP_MIN = max(1, int(float(os.getenv("OPT_LATE_TIME_STEP_MIN", "5"))))
OPT_DAILY_LOSS_MIN = int(float(os.getenv("OPT_DAILY_LOSS_MIN", "15000")))
OPT_DAILY_LOSS_MAX = int(float(os.getenv("OPT_DAILY_LOSS_MAX", "40000")))
OPT_DAILY_LOSS_STEP = max(1, int(float(os.getenv("OPT_DAILY_LOSS_STEP", "1000"))))


# PICKLES_DIR = r"G:\My Drive\Trading\Historical_Options_Data"
PICKLES_DIR = os.getenv("PICKLES_DIR", r"G:\My Drive\Trading\Dhan_Historical_Options_Data_New")
ENTRY_TIME_IST = os.getenv("ENTRY_TIME_IST", "11:55")  # "HH:MM"
# EXIT_TIME_IST is the strategy time filter / square-off cutoff.
# No fresh entry or re-entry is initiated at or after this time. If an attempt is
# still open at this time, it is closed at the cutoff with exit_reason="TIME_EXIT".
EXIT_TIME_IST = os.getenv("EXIT_TIME_IST", "15:30")  # "HH:MM"

def _safe_fname_part(s: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in s)

def _get_downloads_folder() -> str:
    """
    Returns the current user's default Downloads folder.
    Falls back to home directory if Downloads is not found.
    """
    downloads = Path.home() / "Downloads"
    return str(downloads if downloads.exists() else Path.home())

# --- Generic integer-list parser used for DTE and re-entry delay settings ---
def _parse_int_list(env_val, default_list):
    """
    Parse a comma-separated integer list from an environment variable.

    Examples:
        ALLOWED_DTE="0,1"              -> [0, 1]
        REENTRY_DELAY_BY_ATTEMPT="1,5" -> [1, 5]

    If parsing fails or the env var is blank, the supplied default is used.
    """
    if env_val:
        try:
            vals = [int(round(float(x))) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass
    return list(default_list)


def _parse_pct_value(x) -> float:
    """
    Parse a percentage value into decimal form.

    Accepted user formats:
        10      -> 0.10  (10%)
        "10%"   -> 0.10  (10%)
        0.10    -> 0.10  (10%)
        "0.10"  -> 0.10  (10%)

    This lets you configure percentages naturally from environment variables
    while keeping calculations internally consistent.
    """
    s = str(x).strip().replace("%", "")
    if s == "":
        raise ValueError("blank percentage value")

    v = float(s)

    # If user entered 10, treat it as 10%; if user entered 0.10, keep it as 10%.
    if abs(v) > 1.0:
        v = v / 100.0

    if v < 0:
        raise ValueError("percentage cannot be negative")

    return float(v)


def _parse_pct_list(env_val, default_list):
    """
    Parse comma-separated percentages into decimal form.

    Examples:
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15" -> [0.10, 0.12, 0.15]
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10"     -> [0.10]
    """
    if env_val:
        try:
            vals = [_parse_pct_value(x) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass

    return [_parse_pct_value(x) for x in default_list]


def _fmt_int_list(lst) -> str:
    """Format integer-list settings for the output filename."""
    return "-".join(str(int(v)) for v in lst) if lst else "off"


def _fmt_pct_value(v: float) -> str:
    """Format a decimal percentage such as 0.10 as '10pct' for filenames/logs."""
    return f"{v * 100:.2f}".rstrip("0").rstrip(".") + "pct"


def _fmt_pct_list(lst) -> str:
    """Format a list of decimal percentages for the output filename."""
    return "-".join(_fmt_pct_value(float(v)) for v in lst) if lst else "off"


def _parse_float_env(env_name: str, default_value: float) -> float:
    """
    Parse a positive/zero floating-point rupee setting from an environment variable.

    If parsing fails, the supplied default is used. A value <= 0 is treated by
    the relevant logic as disabled where applicable.
    """
    raw = os.getenv(env_name)
    if raw is None or str(raw).strip() == "":
        return float(default_value)
    try:
        return float(str(raw).replace(",", "").strip())
    except Exception:
        return float(default_value)


def _fmt_rupee_value(v: float) -> str:
    """Format rupee values compactly for output filenames/logs."""
    return str(int(round(float(v))))


LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_pct_list(
    os.getenv("LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
    [0.2487, 0.2824, 0.3162, 0.3499, 0.3837, 0.4174, 0.4512],
)


def loss_limit_pct_for_attempt(attempt_idx: int) -> float:
    """Return the stop-loss percentage, in decimal form, for the given attempt."""
    s = LOSS_LIMIT_RUPEES_BY_ATTEMPT
    if not s:
        return 0.0
    return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])


# --- Allowed days-to-expiry to trade: [0,1]=expiry day + day before; [0]=expiry only ---
ALLOWED_DTE = _parse_int_list(os.getenv("ALLOWED_DTE"), [0])

# --- Profit-protect threshold/giveback as % of premium collected on that attempt ---
# Default: 30%.
#
# Env examples:
#     PROFIT_PROTECT_TRIGGER_RUPEES="30"
#     PROFIT_PROTECT_TRIGGER_RUPEES="0.30"
#
# Current logic uses the same rupee amount for:
#     1. arming profit-protect once peak P&L reaches G
#     2. exiting when current P&L falls to peak - G
PROFIT_PROTECT_TRIGGER_RUPEES = _parse_pct_value(os.getenv("PROFIT_PROTECT_TRIGGER_RUPEES", 0.254741))

# --- Absolute daily circuit breaker -------------------------------------------------
# Once cumulative realized NET P&L for the current underlying/day reaches this
# loss, no further re-entry is allowed for that day.
#
# Default: Rs. 30,000 loss. Set MAX_DAILY_LOSS_RUPEES=0 to disable.
MAX_DAILY_LOSS_RUPEES = _parse_float_env("MAX_DAILY_LOSS_RUPEES", 30000.0)

# --- Absolute cap on the percentage-based per-attempt stop-loss ----------------------
# The stop-loss is still calculated as:
#     LOSS_LIMIT_% * entry_premium_sum
# But it is capped at this absolute rupee value.
#
# Effective stop-loss per attempt:
#     min(LOSS_LIMIT_% * entry_premium_sum, MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
#
# Default: Rs. 3,000 loss per attempt. Set to 0 to disable the cap.
MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_float_env("MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT", 3000.0)

MAX_REATTEMPTS = int(os.getenv("MAX_REATTEMPTS", "10"))  # 1 = only one re-entry

# --- Per-DAY profit target as a fraction of premium collected on the CURRENT attempt ---
# When an attempt's profit reaches PROFIT_TARGET_PCT * (CE+PE)*qty, it exits at the
# target and NO further trades are taken that day. 0 disables. e.g. 0.70 = 70%.
PROFIT_TARGET_PCT = float(os.getenv("PROFIT_TARGET_PCT", 0.7))

# --- V2-OPT: late-attempt profit target -------------------------------------
# Attempts ENTERED at/after PROFIT_TARGET_LATE_FROM_IST use this target %.
# 0 disables (all attempts use PROFIT_TARGET_PCT).
PROFIT_TARGET_PCT_LATE = float(os.getenv("PROFIT_TARGET_PCT_LATE", "0") or 0)
PROFIT_TARGET_LATE_FROM_IST = os.getenv("PROFIT_TARGET_LATE_FROM_IST", "14:15")

# --- V2-OPT: re-entry after PROFIT_TARGET ------------------------------------
PROFIT_TARGET_REENTRY_ENABLED = os.getenv("PROFIT_TARGET_REENTRY_ENABLED", "0").strip() == "1"
PROFIT_TARGET_REENTRY_DELAY = int(float(os.getenv("PROFIT_TARGET_REENTRY_DELAY", "3")))

# --- V2-OPT: premium-momentum entry gate -------------------------------------
ENTRY_MOMENTUM_GATE = os.getenv("ENTRY_MOMENTUM_GATE", "0").strip() == "1"
MOMENTUM_LOOKBACK_MIN = int(float(os.getenv("MOMENTUM_LOOKBACK_MIN", "3")))

# --- V3-OPT: late-session tighter profit-protect giveback --------------------
# 0 disables. Example 0.30 = once armed, giveback shrinks to 30% of premium
# from PROFIT_PROTECT_LATE_FROM_IST onward.
PROFIT_PROTECT_LATE_GIVEBACK = float(os.getenv("PROFIT_PROTECT_LATE_GIVEBACK", "0") or 0)
PROFIT_PROTECT_LATE_FROM_IST = os.getenv("PROFIT_PROTECT_LATE_FROM_IST", "14:30")


def _und_env(und: str, key: str, fallback):
    """Per-underlying override: NIFTY_<key> / SENSEX_<key> wins over <key>."""
    raw = os.getenv(f"{und}_{key}")
    if raw is None or str(raw).strip() == "":
        return fallback
    return str(raw).strip()


def resolve_underlying_profile(und: str) -> Dict[str, Any]:
    """Resolve the per-underlying V3 settings dictionary."""
    prof: Dict[str, Any] = {}
    prof["pt_pct"] = float(_und_env(und, "PROFIT_TARGET_PCT", PROFIT_TARGET_PCT))
    prof["pt_pct_late"] = float(_und_env(und, "PROFIT_TARGET_PCT_LATE", PROFIT_TARGET_PCT_LATE))
    prof["pt_late_from"] = parse_hhmm(str(_und_env(und, "PROFIT_TARGET_LATE_FROM_IST", PROFIT_TARGET_LATE_FROM_IST)))
    prof["g_pct"] = _parse_pct_value(_und_env(und, "PROFIT_PROTECT_TRIGGER_RUPEES", PROFIT_PROTECT_TRIGGER_RUPEES))
    prof["dll"] = float(_und_env(und, "MAX_DAILY_LOSS_RUPEES", MAX_DAILY_LOSS_RUPEES))
    prof["g2_pct"] = float(_und_env(und, "PROFIT_PROTECT_LATE_GIVEBACK", PROFIT_PROTECT_LATE_GIVEBACK))
    prof["g2_from"] = parse_hhmm(str(_und_env(und, "PROFIT_PROTECT_LATE_FROM_IST", PROFIT_PROTECT_LATE_FROM_IST)))
    return prof
# --- Per-attempt RE-ENTRY GAP in minutes (index 0 = gap before 1st re-entry, 1 = before 2nd, ...) ---
# Attempts beyond the list reuse the LAST value. Override via env comma list, e.g.
# REENTRY_DELAY_BY_ATTEMPT="10,15,20".
REENTRY_DELAY_BY_ATTEMPT = _parse_int_list(
    os.getenv("REENTRY_DELAY_BY_ATTEMPT"),
    [6, 8, 10, 12, 14, 16, 18, 15, 15, 15, 15],
)

def reentry_delay_for_attempt(attempt_idx: int) -> int:
    s = REENTRY_DELAY_BY_ATTEMPT
    if not s:
        return 0
    return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])

_DEFAULT_OUT = os.path.join(
    _get_downloads_folder(),
    f"short_straddle_V3OPT_{_safe_fname_part(ENTRY_TIME_IST)}"
    f"_exit{_safe_fname_part(EXIT_TIME_IST)}"
    # f"_SLpct_{_safe_fname_part(_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    # f"_DTE_{_safe_fname_part('-'.join(str(d) for d in ALLOWED_DTE))}"
    f"_PPTpct_{_safe_fname_part(_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES))}"
    f"_DailyMaxLoss_{_safe_fname_part(_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES))}"
    f"_StopCap_{_safe_fname_part(_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    f"_MR_{_safe_fname_part(str(MAX_REATTEMPTS))}"
    # f"_PT_{_safe_fname_part(str(int(round(PROFIT_TARGET_PCT * 100))))}pct"
    f"_RDM_{_safe_fname_part(_fmt_int_list(REENTRY_DELAY_BY_ATTEMPT))}.xlsx"
)

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", _DEFAULT_OUT)

FAIL_ON_PICKLE_ERROR = os.getenv("FAIL_ON_PICKLE_ERROR", "0").strip() == "1"

SESSION_START_IST = dtime(9, 15)
SESSION_END_IST = dtime(15, 30)

# LOOKBACK_MONTHS is AUTO by default: the script backtests the full date range
# present in the option pickles. If you explicitly set LOOKBACK_MONTHS to a
# number, that number acts as an optional manual cap.
LOOKBACK_MONTHS_RAW = os.getenv("LOOKBACK_MONTHS", "AUTO").strip()
LOOKBACK_MONTHS: Optional[int]
if LOOKBACK_MONTHS_RAW.upper() in ("", "AUTO", "ALL", "MAX", "FULL"):
    LOOKBACK_MONTHS = None
else:
    LOOKBACK_MONTHS = int(float(LOOKBACK_MONTHS_RAW))

QTY_UNITS = {"NIFTY": 325, "SENSEX": 100}
TRADEABLE = set(QTY_UNITS.keys())

STRIKE_STEP = {"NIFTY": 50, "SENSEX": 100}

# =============================================================================
# TRANSACTION CHARGES (Zerodha F&O Options — NSE)
# =============================================================================
# Each short-straddle attempt = 4 executed orders (sell CE, sell PE, buy CE, buy PE)
BROKERAGE_PER_ORDER       = 20.0       # ₹20 flat per executed order
ORDERS_PER_TRADE          = 4          # sell CE + sell PE + buy CE + buy PE
STT_SELL_PCT              = 0.001      # 0.1% on sell-side premium
EXCHANGE_TXN_PCT          = 0.0003553  # 0.03553% on premium (NSE options)
SEBI_PER_CRORE            = 10.0       # ₹10 per crore of turnover
STAMP_BUY_PCT             = 0.00003    # 0.003% on buy-side premium
IPFT_PER_CRORE            = 0.010      # ₹0.01 per crore (on premium)
GST_PCT                   = 0.18       # 18% on (brokerage + txn charges + SEBI)
INCLUDE_TRANSACTION_COSTS = True       # set False to disable

UNDERLYING_KITE = {
    "NIFTY": {"exchange": "NSE", "tradingsymbol": "NIFTY 50"},
    "SENSEX": {"exchange": "BSE", "tradingsymbol": "SENSEX"},
}

MAX_DAYS_PER_CHUNK = 25
MAX_ATTEMPTS = 5
SLEEP_BETWEEN_CALLS_SEC = 0.20


# =============================================================================
# HELPERS
# =============================================================================
def parse_hhmm(s: str) -> dtime:
    hh, mm = s.strip().split(":")
    return dtime(int(hh), int(mm))

ENTRY_TIME = parse_hhmm(ENTRY_TIME_IST)
PT_LATE_FROM_TIME = parse_hhmm(PROFIT_TARGET_LATE_FROM_IST)
# Strategy cutoff time. The simulator will not initiate a new attempt at or
# after this time, and the current attempt will not be monitored beyond it.
EXIT_TIME = parse_hhmm(EXIT_TIME_IST)

def ist_tz():
    if ZoneInfo is not None:
        return ZoneInfo("Asia/Kolkata")
    if pytz is not None:
        return pytz.timezone("Asia/Kolkata")
    return "Asia/Kolkata"

def ensure_ist(series_or_scalar) -> Any:
    tz = ist_tz()
    dt = pd.to_datetime(series_or_scalar, errors="coerce")
    if isinstance(dt, pd.Series):
        if dt.dt.tz is None:
            return dt.dt.tz_localize(tz)
        return dt.dt.tz_convert(tz)
    if getattr(dt, "tzinfo", None) is None:
        return dt.tz_localize(tz)
    return dt.tz_convert(tz)

def normalize_underlying(name: str) -> Optional[str]:
    if not isinstance(name, str):
        return None
    u = name.upper().strip()
    if "SENSEX" in u:
        return "SENSEX"
    if "BANKNIFTY" in u or "NIFTY BANK" in u:
        return "BANKNIFTY"
    if "NIFTY" in u:
        return "NIFTY"
    return None

def round_to_step(x: float, step: int) -> int:
    return int(round(x / step) * step)

def build_minute_index(day_d: date, start_t: dtime, end_t: dtime) -> pd.DatetimeIndex:
    tz = ist_tz()
    start = pd.Timestamp(datetime.combine(day_d, start_t), tz=tz)
    end = pd.Timestamp(datetime.combine(day_d, end_t), tz=tz)
    return pd.date_range(start=start, end=end, freq="1min")

def asof_close(df: pd.DataFrame, ts: pd.Timestamp) -> float:
    if df.empty:
        return float("nan")
    d = df[["date", "close"]].dropna().copy()
    d["date"] = ensure_ist(d["date"])
    d = d.sort_values("date").set_index("date")
    loc = d.index.get_indexer([ts], method="pad")
    if loc[0] == -1:
        return float("nan")
    return float(d.iloc[loc[0]]["close"])

def compute_window_start(end_day: date, months: int) -> date:
    if relativedelta is not None:
        return (pd.Timestamp(end_day) - relativedelta(months=months)).date()
    return (pd.Timestamp(end_day) - pd.Timedelta(days=30 * months)).date()


def determine_backtest_window_start(min_day_seen: date, end_day: date) -> date:
    """
    Determine the backtest start date.

    Default behaviour: use the earliest usable option-data date found in the
    pickles, i.e. the maximum available backtest period.

    Optional override: if LOOKBACK_MONTHS is set to a numeric value through the
    environment, use the later of:
        1. earliest option-data date; and
        2. end_day - LOOKBACK_MONTHS
    so the script never requests data before the pickles actually start.
    """
    if LOOKBACK_MONTHS is None:
        return min_day_seen

    capped_start = compute_window_start(end_day, LOOKBACK_MONTHS)
    return max(min_day_seen, capped_start)

# =============================================================================
# TRANSACTION COST CALCULATOR
# =============================================================================
def compute_trade_charges(
    entry_ce: float, entry_pe: float,
    exit_ce: float, exit_pe: float,
    qty: int,
) -> float:
    """
    Compute total Zerodha transaction charges for one short-straddle attempt.

    Entry = SELL CE + SELL PE  (2 orders, sell side)
    Exit  = BUY  CE + BUY  PE (2 orders, buy side)

    Returns total charges in rupees (always positive).
    """
    if not INCLUDE_TRANSACTION_COSTS:
        return 0.0

    # Turnover values (in rupees)
    entry_turnover = (entry_ce + entry_pe) * qty   # sell side
    exit_turnover  = (exit_ce + exit_pe) * qty     # buy side
    total_turnover = entry_turnover + exit_turnover

    # 1. Brokerage: ₹20 × 4 orders
    brokerage = BROKERAGE_PER_ORDER * ORDERS_PER_TRADE

    # 2. STT: 0.1% on sell-side premium only (entry for short straddle)
    stt = entry_turnover * STT_SELL_PCT

    # 3. Exchange transaction charges: 0.03553% on both sides
    txn_charges = total_turnover * EXCHANGE_TXN_PCT

    # 4. SEBI charges: ₹10 per crore on total turnover
    sebi = total_turnover * SEBI_PER_CRORE / 1_00_00_000

    # 5. Stamp duty: 0.003% on buy side only (exit for short straddle)
    stamp = exit_turnover * STAMP_BUY_PCT

    # 6. IPFT: ₹0.01 per crore on premium (both sides)
    ipft = total_turnover * IPFT_PER_CRORE / 1_00_00_000

    # 7. GST: 18% on (brokerage + transaction charges + SEBI charges)
    gst = (brokerage + txn_charges + sebi) * GST_PCT

    total_charges = brokerage + stt + txn_charges + sebi + stamp + ipft + gst
    return round(total_charges, 2)

# =============================================================================
# Kite historical helpers
# =============================================================================
def _iter_chunks_by_date(from_dt: datetime, to_dt: datetime, days_per_chunk: int) -> List[Tuple[datetime, datetime]]:
    if from_dt > to_dt:
        raise ValueError("from_dt must be <= to_dt")
    chunks: List[Tuple[datetime, datetime]] = []
    cur = from_dt.date()
    end_d = to_dt.date()
    while cur <= end_d:
        chunk_end = min(cur + timedelta(days=days_per_chunk - 1), end_d)
        c_from = from_dt if cur == from_dt.date() else datetime.combine(cur, SESSION_START_IST)
        c_to = to_dt if chunk_end == end_d else datetime.combine(chunk_end, SESSION_END_IST)
        chunks.append((c_from, c_to))
        cur = chunk_end + timedelta(days=1)
    return chunks

def _kite_instruments_cached(kite, exchange: str, cache: Dict[str, List[Dict]]) -> List[Dict]:
    ex = exchange.upper().strip()
    if ex not in cache:
        print(f"[STEP] Loading instruments dump for {ex} ...")
        cache[ex] = kite.instruments(ex)
        print(f"[INFO] {ex} instruments: {len(cache[ex])}")
    return cache[ex]

def get_instrument_token(kite, exchange: str, tradingsymbol: str, cache: Dict[str, List[Dict]]) -> int:
    ex = exchange.upper().strip()
    wanted = tradingsymbol.strip().upper()
    for r in _kite_instruments_cached(kite, ex, cache):
        if str(r.get("tradingsymbol", "")).upper() == wanted:
            return int(r["instrument_token"])
    raise ValueError(f"Instrument not found on {ex}: '{tradingsymbol}'")

def fetch_history_minute(kite, instrument_token: int, from_dt: datetime, to_dt: datetime, label: str) -> List[Dict]:
    interval = "minute"
    chunks = _iter_chunks_by_date(from_dt, to_dt, MAX_DAYS_PER_CHUNK)
    rows_all: List[Dict] = []
    print(f"[INFO] Fetch {label} token={instrument_token} chunks={len(chunks)} {from_dt} -> {to_dt}")
    for i, (c_from, c_to) in enumerate(chunks, start=1):
        last_err = None
        for attempt in range(1, MAX_ATTEMPTS + 1):
            try:
                rows = kite.historical_data(
                    instrument_token=instrument_token,
                    from_date=c_from,
                    to_date=c_to,
                    interval=interval,
                    continuous=False,
                    oi=False,
                )
                rows_all.extend(rows)
                last_err = None
                break
            except Exception as e:
                last_err = e
                time.sleep(min(8.0, 1.5 * attempt))
        if last_err is not None:
            print(f"[ERROR] {label} chunk {i}/{len(chunks)} failed: {c_from}->{c_to}: {last_err}")
        time.sleep(SLEEP_BETWEEN_CALLS_SEC)
    return rows_all

def rows_to_df(rows: List[Dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["date", "open", "high", "low", "close", "volume"])
    df = pd.DataFrame(rows)
    df["date"] = ensure_ist(df["date"])
    df = df.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    return df


# =============================================================================
# DATA STRUCTURES
# =============================================================================
@dataclass
class TradeRow:
    day: date
    underlying: str
    trade_seq: int
    expiry: date
    days_to_expiry: int
    atm_strike: int
    qty_units: int
    entry_time: str
    exit_time: str
    exit_reason: str
    entry_underlying: float
    ce_symbol: str
    pe_symbol: str
    entry_ce: float
    entry_pe: float
    exit_ce: float
    exit_pe: float
    exit_pnl_gross: float   # P&L before charges
    txn_charges: float      # total transaction charges for this attempt
    exit_pnl: float         # net P&L after deducting charges
    eod_pnl: float
    max_profit: float
    max_loss: float
    max_profit_before_exit: float   # peak profit reached before this trade exited

    # Premium/risk diagnostics for percentage-based risk rules
    entry_premium_sum: float                 # (entry_ce + entry_pe) * qty
    stop_pct: float                          # stop-loss % of entry premium, decimal form; 0.10 = 10%
    uncapped_stop_rupees: float              # percentage-based stop before absolute cap
    stop_cap_rupees: float                   # configured absolute cap; <=0 means cap disabled
    stop_rupees: float                       # effective rupee stop after cap
    profit_protect_trigger_pct: float         # profit-protect % of entry premium, decimal form; 0.30 = 30%
    profit_protect_trigger_rupees: float      # computed rupee profit-protect trigger/giveback
    daily_realized_pnl_after_trade: float     # cumulative net P&L after this attempt
    daily_loss_limit_rupees: float            # configured daily loss circuit breaker
    daily_loss_limit_hit: bool                # True means no further trades for that day


# =============================================================================
# PASS-1: nearest expiry per (underlying, day)
# =============================================================================
def scan_pickles_pass1(pickle_paths: List[str]) -> Tuple[date, Dict[Tuple[str, date], date], date]:
    max_day_seen: Optional[date] = None
    min_day_seen: Optional[date] = None
    min_expiry_map: Dict[Tuple[str, date], date] = {}

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            for c in ("date", "name", "expiry", "type"):
                if c not in df.columns:
                    raise ValueError(f"Missing column '{c}' in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")]
            if d2.empty:
                continue

            d2 = d2[["date", "name", "expiry"]].copy()
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2 = d2.dropna(subset=["underlying", "day", "expiry_date"])

            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            file_min_day = d2["day"].min()
            file_max_day = d2["day"].max()
            max_day_seen = file_max_day if (max_day_seen is None or file_max_day > max_day_seen) else max_day_seen
            min_day_seen = file_min_day if (min_day_seen is None or file_min_day < min_day_seen) else min_day_seen

            grp = d2.groupby(["underlying", "day"], sort=False)["expiry_date"].min()
            for (und, dy), ex in grp.items():
                key = (und, dy)
                if key not in min_expiry_map or ex < min_expiry_map[key]:
                    min_expiry_map[key] = ex

            print(f"[PASS1 OK] {os.path.basename(p)} option_days={d2['day'].nunique()}")

        except Exception as e:
            msg = f"[PASS1 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    if max_day_seen is None or min_day_seen is None:
        raise RuntimeError("No usable option data found in pickles (PASS1) for tradeable underlyings.")

    return max_day_seen, min_expiry_map, min_day_seen


# =============================================================================
# Underlying download
# =============================================================================
def download_underlyings(kite, day_start: date, day_end: date) -> Dict[str, pd.DataFrame]:
    cache: Dict[str, List[Dict]] = {}
    from_dt = datetime.combine(day_start, SESSION_START_IST)
    to_dt = datetime.combine(day_end, SESSION_END_IST)

    out: Dict[str, pd.DataFrame] = {}
    for und, meta in UNDERLYING_KITE.items():
        token = get_instrument_token(kite, meta["exchange"], meta["tradingsymbol"], cache)
        rows = fetch_history_minute(kite, token, from_dt, to_dt, label=f"{meta['exchange']}:{meta['tradingsymbol']}")
        df = rows_to_df(rows)
        df["day"] = df["date"].dt.tz_convert(ist_tz()).dt.date
        out[und] = df
        print(f"[UNDERLYING OK] {und}: candles={len(df)} days={df['day'].nunique()}")
    return out


# =============================================================================
# Simulation helpers
# =============================================================================
def _pick_symbol(day_opt: pd.DataFrame, strike: int, opt_type: str) -> Optional[str]:
    sub = day_opt[(day_opt["strike_int"] == strike) & (day_opt["option_type"] == opt_type)]
    if sub.empty:
        return None
    syms = sorted(sub["instrument"].astype(str).unique().tolist())
    return syms[0] if syms else None

def _build_leg_series(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex,
                      strike: int, opt_type: str, symbol: str,
                      price_col: str = "close", do_ffill: bool = True) -> pd.Series:
    sub = day_opt[
        (day_opt["strike_int"] == strike) &
        (day_opt["option_type"] == opt_type) &
        (day_opt["instrument"].astype(str) == symbol)
    ][["date", price_col]].dropna()

    if sub.empty:
        return pd.Series(index=idx_all, dtype="float64")

    sub = sub.copy()
    sub["date"] = ensure_ist(sub["date"])
    sub = sub.sort_values("date").drop_duplicates(subset=["date"], keep="last").set_index("date")
    s = sub[price_col].astype(float).reindex(idx_all)
    return s.ffill() if do_ffill else s


def build_price_book(
    day_opt: pd.DataFrame,
    idx_all: pd.DatetimeIndex,
) -> Tuple[Dict[Tuple[int, str, str], pd.Series], Dict[Tuple[int, str], str]]:
    """
    Build all parameter-independent option-price series once for one day.

    The old single-backtest path rebuilt and reindexed CE/PE close/high/low
    series every time a strategy parameter changed. That is unacceptable inside
    an optimizer. This cache preserves the exact symbol-selection and reindexing
    rules while making later trials dictionary lookups.
    """
    book: Dict[Tuple[int, str, str], pd.Series] = {}
    symbols: Dict[Tuple[int, str], str] = {}
    if day_opt is None or day_opt.empty:
        return book, symbols

    for (strike, opt_type), sub in day_opt.groupby(["strike_int", "option_type"], sort=False):
        available = sorted(sub["instrument"].astype(str).unique().tolist())
        if not available:
            continue

        symbol = available[0]  # identical rule to _pick_symbol()
        strike_i = int(strike)
        opt_s = str(opt_type)
        symbols[(strike_i, opt_s)] = symbol

        selected = sub[sub["instrument"].astype(str) == symbol][
            ["date", "close", "high", "low"]
        ].copy()
        selected["date"] = ensure_ist(selected["date"])
        selected = (
            selected.sort_values("date")
            .drop_duplicates(subset=["date"], keep="last")
            .set_index("date")
        )
        for col in ("close", "high", "low"):
            book[(strike_i, opt_s, col)] = selected[col].astype(float).reindex(idx_all)

    return book, symbols


def _leg_from_book(
    price_book: Dict[Tuple[int, str, str], pd.Series],
    idx_all: pd.DatetimeIndex,
    strike: int,
    opt_type: str,
    price_col: str,
) -> pd.Series:
    """Return a cached raw minute series; missing legs become an empty series."""
    series = price_book.get((int(strike), str(opt_type), str(price_col)))
    if series is None:
        return pd.Series(index=idx_all, dtype="float64")
    return series


@dataclass(frozen=True)
class UnderlyingProfile:
    """All V3 exit/risk settings that may differ between NIFTY and SENSEX."""

    profit_target_pct: float
    profit_target_pct_late: float
    profit_target_late_from: dtime
    profit_protect_pct: float
    max_daily_loss_rupees: float
    profit_protect_late_giveback_pct: float
    profit_protect_late_from: dtime


@dataclass
class Params:
    """
    Complete parameter set consumed by the simulator.

    Passing this object explicitly is critical: an Optuna trial can run a fresh
    configuration in the same Python process without mutating module globals or
    re-importing the strategy.
    """

    entry_time: dtime
    exit_time: dtime
    loss_limit_pct_by_attempt: List[float]
    max_loss_limit_cap_rupees: float
    max_reattempts: int
    reentry_delay_by_attempt: List[int]

    entry_momentum_gate: bool
    momentum_lookback_min: int
    profit_target_reentry_enabled: bool
    profit_target_reentry_delay_min: int

    profiles: Dict[str, UnderlyingProfile]

    def loss_limit_pct_for_attempt(self, attempt_idx: int) -> float:
        values = self.loss_limit_pct_by_attempt
        if not values:
            return 0.0
        return float(values[attempt_idx]) if attempt_idx < len(values) else float(values[-1])

    def reentry_delay_for_attempt(self, attempt_idx: int) -> int:
        values = self.reentry_delay_by_attempt
        if not values:
            return 0
        return int(values[attempt_idx]) if attempt_idx < len(values) else int(values[-1])

    def profile_for(self, underlying: str) -> UnderlyingProfile:
        try:
            return self.profiles[underlying]
        except KeyError as exc:
            raise KeyError(f"No strategy profile configured for {underlying}") from exc


def default_params() -> Params:
    """Build a parameter object from the loaded V3 property file."""
    profiles: Dict[str, UnderlyingProfile] = {}
    for underlying in sorted(TRADEABLE):
        raw = resolve_underlying_profile(underlying)
        profiles[underlying] = UnderlyingProfile(
            profit_target_pct=float(raw["pt_pct"]),
            profit_target_pct_late=float(raw["pt_pct_late"]),
            profit_target_late_from=raw["pt_late_from"],
            profit_protect_pct=float(raw["g_pct"]),
            max_daily_loss_rupees=float(raw["dll"]),
            profit_protect_late_giveback_pct=float(raw["g2_pct"]),
            profit_protect_late_from=raw["g2_from"],
        )

    return Params(
        entry_time=ENTRY_TIME,
        exit_time=EXIT_TIME,
        loss_limit_pct_by_attempt=list(LOSS_LIMIT_RUPEES_BY_ATTEMPT),
        max_loss_limit_cap_rupees=float(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT),
        max_reattempts=int(MAX_REATTEMPTS),
        reentry_delay_by_attempt=list(REENTRY_DELAY_BY_ATTEMPT),
        entry_momentum_gate=bool(ENTRY_MOMENTUM_GATE),
        momentum_lookback_min=int(MOMENTUM_LOOKBACK_MIN),
        profit_target_reentry_enabled=bool(PROFIT_TARGET_REENTRY_ENABLED),
        profit_target_reentry_delay_min=int(PROFIT_TARGET_REENTRY_DELAY),
        profiles=profiles,
    )


@dataclass
class DayGroup:
    """
    One parameter-independent (underlying, date, nearest-expiry) simulation unit.

    Pickle parsing, date normalization, day slicing and option-series reindexing
    are done once. Every Optuna trial reuses these cached objects.
    """

    und: str
    dy: date
    expiry: date
    day_opt: pd.DataFrame
    underlying_day: pd.DataFrame
    price_book: Optional[Dict[Tuple[int, str, str], pd.Series]] = None
    symbols: Optional[Dict[Tuple[int, str], str]] = None
    idx_all: Optional[pd.DatetimeIndex] = None


def simulate_day_multi_trades(
    *,
    und: str,
    dy: date,
    expiry: date,
    day_opt: pd.DataFrame,
    underlying_day: pd.DataFrame,
    params: Params,
    price_book: Optional[Dict[Tuple[int, str, str], pd.Series]] = None,
    symbols: Optional[Dict[Tuple[int, str], str]] = None,
    idx_all: Optional[pd.DatetimeIndex] = None,
) -> Tuple[List[TradeRow], List[Dict[str, Any]]]:

    results: List[TradeRow] = []
    skipped: List[Dict[str, Any]] = []

    if idx_all is None:
        idx_all = build_minute_index(dy, SESSION_START_IST, SESSION_END_IST)
    if price_book is None or symbols is None:
        price_book, symbols = build_price_book(day_opt, idx_all)
    session_end_ts = idx_all[-1]

    # EXIT_TIME_IST is the strategy time filter.
    #
    # Earlier versions used EXIT_TIME_IST only as a "last new entry" gate while
    # still allowing the currently-open trade to run till SESSION_END_IST. That
    # made the Excel output look as if the filter was not working because trade
    # rows could still show exit_time after EXIT_TIME_IST.
    #
    # In this corrected version:
    #   1. no fresh entry/re-entry is initiated at or after EXIT_TIME_IST; and
    #   2. the active attempt is monitored only up to EXIT_TIME_IST.
    #
    # Therefore, when EXIT_TIME_IST is earlier than 15:30, the trade exits at the
    # cutoff if STOPLOSS / PROFIT_TARGET / PROFIT_PROTECT has not already fired.
    configured_exit_cutoff_ts = pd.Timestamp(datetime.combine(dy, params.exit_time), tz=ist_tz())
    trade_end_ts = min(session_end_ts, configured_exit_cutoff_ts)

    qty = int(QTY_UNITS[und])
    step = int(STRIKE_STEP[und])

    # --- V3-OPT: per-underlying profile --------------------------------------
    _prof = params.profile_for(und)

    # --- V2-OPT: ATM-straddle premium series cache + momentum gate ----------
    _prem_cache: Dict[int, Optional[pd.Series]] = {}

    def _atm_prem_series(strike: int) -> Optional[pd.Series]:
        if strike not in _prem_cache:
            ce_symbol = symbols.get((int(strike), "CE"))
            pe_symbol = symbols.get((int(strike), "PE"))
            if not ce_symbol or not pe_symbol:
                _prem_cache[strike] = None
            else:
                ce = _leg_from_book(price_book, idx_all, strike, "CE", "close").ffill()
                pe = _leg_from_book(price_book, idx_all, strike, "PE", "close").ffill()
                _prem_cache[strike] = ce + pe
        return _prem_cache[strike]

    def _momentum_wait(ts: pd.Timestamp) -> pd.Timestamp:
        """Advance ts minute-by-minute until the current-ATM straddle premium
        is lower than MOMENTUM_LOOKBACK_MIN minutes earlier (i.e. premium is
        decaying), or until the strategy cutoff. Uses only information
        available at ts (close of already-completed minutes)."""
        if not params.entry_momentum_gate:
            return ts
        session_start_ts = idx_all[0]
        while ts < trade_end_ts:
            u = asof_close(underlying_day, ts)
            if pd.isna(u):
                return ts
            atm_g = round_to_step(float(u), step)
            s = _atm_prem_series(atm_g)
            if s is None:
                return ts
            now = s.loc[ts] if ts in s.index else float("nan")
            ts_back = ts - pd.Timedelta(minutes=params.momentum_lookback_min)
            if ts_back < session_start_ts:
                ts_back = session_start_ts
            before = s.loc[ts_back] if ts_back in s.index else float("nan")
            if pd.notna(now) and pd.notna(before) and float(now) < float(before):
                return ts
            ts = ts + pd.Timedelta(minutes=1)
        return ts

    # Profit-protect is now percentage-based, so the actual rupee value is not
    # known until CE/PE entry prices are available for the current attempt.
    profit_protect_pct = float(_prof.profit_protect_pct)
    profit_protect_enabled = profit_protect_pct > 0.0

    cur_entry_ts = pd.Timestamp(datetime.combine(dy, params.entry_time), tz=ist_tz())
    trade_seq = 1

    # V2-OPT: momentum-gate the FIRST entry of the day.
    cur_entry_ts = _momentum_wait(cur_entry_ts)

    # If the configured first entry itself is at/after EXIT_TIME_IST, the day
    # is skipped cleanly. This is intentional: EXIT_TIME_IST is the hard strategy
    # time filter, so a trade needs at least one minute of monitoring before it.
    if cur_entry_ts >= trade_end_ts:
        skipped.append({
            "day": dy,
            "underlying": und,
            "expiry": expiry,
            "trade_seq": trade_seq,
            "reason": (
                f"No entry: entry {params.entry_time.strftime('%H:%M')} is at/after "
                f"exit cutoff {params.exit_time.strftime('%H:%M')}"
            ),
        })
        return results, skipped

    # Cumulative realized NET P&L for this underlying/day. Used for the daily
    # loss circuit breaker. Charges are included through exit_pnl.
    daily_realized_pnl = 0.0
    max_daily_loss_rupees = float(_prof.max_daily_loss_rupees)
    daily_loss_limit_enabled = max_daily_loss_rupees > 0

    while cur_entry_ts < trade_end_ts:
        if daily_loss_limit_enabled and daily_realized_pnl <= -float(max_daily_loss_rupees):
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq,
                "reason": (
                    f"Daily loss limit hit before next entry: "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={max_daily_loss_rupees:.2f}"
                ),
            })
            break

        u_px = asof_close(underlying_day, cur_entry_ts)
        if pd.isna(u_px):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": f"No underlying price at entry {cur_entry_ts.strftime('%H:%M')}"})
            break

        atm = round_to_step(float(u_px), step)

        ce_sym = symbols.get((int(atm), "CE"))
        pe_sym = symbols.get((int(atm), "PE"))
        if not ce_sym or not pe_sym:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "ATM CE/PE not available in pickle band"})
            break

        # Raw entry series and forward-filled close series come from the
        # parameter-independent day cache.
        ce_close_raw = _leg_from_book(price_book, idx_all, atm, "CE", "close")
        pe_close_raw = _leg_from_book(price_book, idx_all, atm, "PE", "close")
        ce_close = ce_close_raw.ffill()
        pe_close = pe_close_raw.ffill()

        # High/low are intentionally not forward-filled: they are used only for
        # intraminute stop-loss and profit-target detection.
        ce_high = _leg_from_book(price_book, idx_all, atm, "CE", "high")
        ce_low = _leg_from_book(price_book, idx_all, atm, "CE", "low")
        pe_high = _leg_from_book(price_book, idx_all, atm, "PE", "high")
        pe_low = _leg_from_book(price_book, idx_all, atm, "PE", "low")

        if cur_entry_ts not in idx_all:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": "Entry timestamp not in session index"})
            break

        ce_entry = ce_close_raw.loc[cur_entry_ts]
        pe_entry = pe_close_raw.loc[cur_entry_ts]
        monitor_start_ts = pd.Timestamp(cur_entry_ts) + pd.Timedelta(minutes=1)
        if monitor_start_ts > trade_end_ts:
            break

        if pd.isna(ce_entry) or pd.isna(pe_entry):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "No CE/PE price at entry (after ffill)"})
            break

        # ---------------------------------------------------------------------
        # Percentage-based risk basis for THIS attempt
        # ---------------------------------------------------------------------
        # For every entry/re-entry, compute the premium collected in rupees.
        # Stop-loss and profit-protect thresholds are derived from this value.
        #
        # Example:
        #   entry_ce=70, entry_pe=50, qty=325
        #   entry_premium_sum = (70 + 50) * 325 = 39,000
        #   10% stop-loss = 3,900
        #   30% profit-protect threshold/giveback = 11,700
        # ---------------------------------------------------------------------
        entry_premium_sum = (float(ce_entry) + float(pe_entry)) * qty

        loss_limit_pct = params.loss_limit_pct_for_attempt(trade_seq - 1)
        uncapped_loss_limit_rupees = float(loss_limit_pct * entry_premium_sum)

        # Absolute cap on the percentage-based stop-loss.
        # Example: 10% of premium may be Rs. 4,500, but with a Rs. 3,000 cap
        # the effective stop used by the simulator is Rs. 3,000.
        stop_cap_rupees = float(params.max_loss_limit_cap_rupees)
        if stop_cap_rupees > 0:
            loss_limit_rupees = float(min(uncapped_loss_limit_rupees, stop_cap_rupees))
        else:
            loss_limit_rupees = float(uncapped_loss_limit_rupees)

        # G is the same variable used by the existing profit-protect logic:
        #   - profit-protect arms when peak P&L >= G
        #   - profit-protect exits when current P&L <= peak - G
        G = float(profit_protect_pct * entry_premium_sum)

        # Close-based PnL (same as before)
        pnl_close_all = (float(ce_entry) - ce_close) * qty + (float(pe_entry) - pe_close) * qty
        pnl = pnl_close_all.loc[monitor_start_ts:trade_end_ts].dropna()  # keep 'pnl' as close-based for profit-protect

        # STOPLOSS worst-case PnL candidates within each minute:
        #  A) CE high, PE low
        pnl_ceHigh_peLow_all = (float(ce_entry) - ce_high) * qty + (float(pe_entry) - pe_low) * qty
        #  B) CE low, PE high
        pnl_ceLow_peHigh_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_high) * qty

        # Worst-case PnL per minute among (close, A, B)
        pnl_sl_all = pd.concat([pnl_close_all, pnl_ceHigh_peLow_all, pnl_ceLow_peHigh_all], axis=1).min(axis=1)
        pnl_sl = pnl_sl_all.loc[monitor_start_ts:trade_end_ts].dropna()

        if pnl.empty:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "PnL series empty after entry"})
            break

        eod_ts = pnl.index[-1]
        eod_pnl = float(pnl.iloc[-1])

        # If EXIT_TIME_IST is earlier than market close and no risk/profit event
        # triggers before that, the attempt is closed at the configured cutoff.
        # The old "EOD" label is retained only when the monitoring horizon is
        # the real session end.
        default_exit_reason = "TIME_EXIT" if trade_end_ts < session_end_ts else "EOD"

        max_profit = float(max(0.0, pnl.max()))
        max_loss = float(min(0.0, pnl.min()))

        # STOPLOSS uses the attempt-specific rupee value after applying the
        # absolute per-attempt cap.
        stop_hit = pnl_sl <= -loss_limit_rupees
        stop_ts = pnl_sl.index[stop_hit.to_numpy().argmax()] if stop_hit.any() else None

        protect_ts = None
        if profit_protect_enabled:
            peak = pnl.cummax()
            armed = peak >= G
            trail = peak - G
            # V3-OPT: tighter giveback from PROFIT_PROTECT_LATE_FROM_IST onward
            if _prof.profit_protect_late_giveback_pct > 0.0:
                _g2_ts = pd.Timestamp(datetime.combine(dy, _prof.profit_protect_late_from), tz=ist_tz())
                _late_mask = pnl.index >= _g2_ts
                if _late_mask.any():
                    _trail_late = peak - float(_prof.profit_protect_late_giveback_pct) * entry_premium_sum
                    trail = trail.where(~_late_mask, trail.combine(_trail_late, max))
            protect_hit = armed & (pnl <= trail)
            protect_ts = pnl.index[protect_hit.to_numpy().argmax()] if protect_hit.any() else None

        # --- Per-day PROFIT TARGET: % of premium collected on this attempt ---
        # When reached, this trade exits at the target AND no further trades are
        # taken for the day (PROFIT_TARGET is excluded from the re-entry rule below).
        target_ts = None
        target_rupees = None
        # V2-OPT: attempts entered at/after PROFIT_TARGET_LATE_FROM_IST use
        # the (smaller) late-session profit target.
        _pt_pct_eff = _prof.profit_target_pct
        if _prof.profit_target_pct_late > 0.0 and cur_entry_ts.timetz().replace(tzinfo=None) >= _prof.profit_target_late_from:
            _pt_pct_eff = _prof.profit_target_pct_late
        if _pt_pct_eff > 0.0:
            target_rupees = _pt_pct_eff * entry_premium_sum
            # best-case (favourable) intrabar profit: both legs bought back at their lows
            pnl_best_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_low) * qty
            pnl_tp = pd.concat([pnl_close_all, pnl_best_all], axis=1).max(axis=1)
            pnl_tp = pnl_tp.loc[monitor_start_ts:trade_end_ts].dropna()
            tp_hit = pnl_tp >= float(target_rupees)
            target_ts = pnl_tp.index[tp_hit.to_numpy().argmax()] if tp_hit.any() else None

        # Earliest triggered exit wins; on identical timestamps prefer the more
        # conservative outcome: STOPLOSS, then PROFIT_TARGET, then PROFIT_PROTECT.
        exit_ts = eod_ts
        exit_reason = default_exit_reason
        _candidates = []
        if stop_ts is not None:
            _candidates.append((stop_ts, 0, "STOPLOSS"))
        if target_ts is not None:
            _candidates.append((target_ts, 1, "PROFIT_TARGET"))
        if protect_ts is not None:
            _candidates.append((protect_ts, 2, "PROFIT_PROTECT"))
        if _candidates:
            _candidates.sort(key=lambda c: (c[0], c[1]))
            exit_ts, _, exit_reason = _candidates[0]

        if exit_reason == "STOPLOSS":
            exit_pnl_gross = -float(loss_limit_rupees)
        elif exit_reason == "PROFIT_TARGET":
            exit_pnl_gross = float(target_rupees)
        else:
            exit_pnl_gross = float(pnl.loc[exit_ts])

        # Peak (close-based) profit reached during this trade's life, up to its exit
        pnl_pre_exit = pnl.loc[:exit_ts]
        max_profit_before_exit = float(max(0.0, pnl_pre_exit.max())) if len(pnl_pre_exit) else 0.0

        exit_ce = float(ce_close.loc[exit_ts]) if pd.notna(ce_close.loc[exit_ts]) else float("nan")
        exit_pe = float(pe_close.loc[exit_ts]) if pd.notna(pe_close.loc[exit_ts]) else float("nan")

        txn_charges = compute_trade_charges(
            entry_ce=float(ce_entry), entry_pe=float(pe_entry),
            exit_ce=exit_ce if not pd.isna(exit_ce) else 0.0,
            exit_pe=exit_pe if not pd.isna(exit_pe) else 0.0,
            qty=qty,
        )
        exit_pnl = exit_pnl_gross - txn_charges

        # Update cumulative realized NET P&L for the day. This is checked
        # before allowing any further re-entry.
        daily_realized_pnl += float(exit_pnl)
        daily_loss_limit_hit = bool(
            daily_loss_limit_enabled and daily_realized_pnl <= -float(max_daily_loss_rupees)
        )

        dte = int((expiry - dy).days)

        results.append(
            TradeRow(
                day=dy,
                underlying=und,
                trade_seq=trade_seq,
                expiry=expiry,
                days_to_expiry=dte,
                atm_strike=int(atm),
                qty_units=qty,
                entry_time=pd.Timestamp(cur_entry_ts).strftime("%H:%M"),
                exit_time=pd.Timestamp(exit_ts).strftime("%H:%M"),
                exit_reason=exit_reason,
                entry_underlying=float(u_px),
                ce_symbol=ce_sym,
                pe_symbol=pe_sym,
                entry_ce=float(ce_entry),
                entry_pe=float(pe_entry),
                exit_ce=exit_ce,
                exit_pe=exit_pe,
                exit_pnl_gross=exit_pnl_gross,
                txn_charges=txn_charges,
                exit_pnl=exit_pnl,
                eod_pnl=eod_pnl,
                max_profit=max_profit,
                max_loss=max_loss,
                max_profit_before_exit=max_profit_before_exit,
                entry_premium_sum=float(entry_premium_sum),
                stop_pct=float(loss_limit_pct),
                uncapped_stop_rupees=float(uncapped_loss_limit_rupees),
                stop_cap_rupees=float(stop_cap_rupees),
                stop_rupees=float(loss_limit_rupees),
                profit_protect_trigger_pct=float(profit_protect_pct),
                profit_protect_trigger_rupees=float(G),
                daily_realized_pnl_after_trade=float(daily_realized_pnl),
                daily_loss_limit_rupees=float(max_daily_loss_rupees),
                daily_loss_limit_hit=bool(daily_loss_limit_hit),
            )
        )

        if daily_loss_limit_hit:
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq + 1,
                "reason": (
                    f"No re-entry: daily loss limit hit after trade_seq={trade_seq}; "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={max_daily_loss_rupees:.2f}"
                ),
            })
            break

        if exit_reason in ("STOPLOSS", "PROFIT_PROTECT") and (trade_seq - 1) < params.max_reattempts:
            delay_min = params.reentry_delay_for_attempt(trade_seq - 1)  # gap before this re-entry
            trade_seq += 1
            cur_entry_ts = pd.Timestamp(exit_ts) + pd.Timedelta(minutes=delay_min)
            # V2-OPT: wait for premium decay confirmation before re-entering.
            cur_entry_ts = _momentum_wait(cur_entry_ts)

            # Do not initiate any further trade at or after EXIT_TIME_IST.
            # Because trade_end_ts is also the monitoring end, this keeps the
            # output strictly filtered by EXIT_TIME_IST.
            if cur_entry_ts >= trade_end_ts:
                skipped.append({
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "reason": (
                        f"No re-entry: next entry time {pd.Timestamp(cur_entry_ts).strftime('%H:%M')} "
                        f"is at/after exit cutoff {params.exit_time.strftime('%H:%M')}"
                    ),
                })
                break
            continue

        # V2-OPT: PROFIT_TARGET no longer necessarily ends the day. Recycle
        # into a fresh ATM straddle after PROFIT_TARGET_REENTRY_DELAY minutes
        # (plus the momentum gate). Position stays sequential: one straddle
        # at a time, same quantities.
        if (
            exit_reason == "PROFIT_TARGET"
            and params.profit_target_reentry_enabled
            and (trade_seq - 1) < params.max_reattempts
        ):
            trade_seq += 1
            cur_entry_ts = pd.Timestamp(exit_ts) + pd.Timedelta(minutes=params.profit_target_reentry_delay_min)
            cur_entry_ts = _momentum_wait(cur_entry_ts)
            if cur_entry_ts >= trade_end_ts:
                skipped.append({
                    "day": dy,
                    "underlying": und,
                    "expiry": expiry,
                    "trade_seq": trade_seq,
                    "reason": (
                        f"No target re-entry: next entry time "
                        f"{pd.Timestamp(cur_entry_ts).strftime('%H:%M')} is at/after "
                        f"exit cutoff {params.exit_time.strftime('%H:%M')}"
                    ),
                })
                break
            continue

        break

    return results, skipped


# =============================================================================
# PASS-2: process each pickle and simulate trades for days where this expiry is nearest
# =============================================================================

def build_day_groups(
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
    max_pickles: Optional[int] = None,
    max_days: Optional[int] = None,
) -> Tuple[List[DayGroup], List[Dict[str, Any]]]:
    """
    Parse all option pickles once and cache each nearest-expiry trading day.

    `max_pickles` and `max_days` are smoke-test controls only. A production
    optimization should leave both as None.
    """
    groups: List[DayGroup] = []
    skipped_rows: List[Dict[str, Any]] = []
    processed_day_keys: set[Tuple[str, date, date]] = set()

    paths = list(pickle_paths)
    if max_pickles is not None and max_pickles > 0:
        paths = paths[:max_pickles]

    total_files = len(paths)
    for file_index, path in enumerate(paths, start=1):
        try:
            df = pd.read_pickle(path)
            if not isinstance(df, pd.DataFrame) or df.empty:
                print(f"[LOAD {file_index}/{total_files}] {os.path.basename(path)}: empty")
                continue

            needed_cols = [
                "date", "name", "type", "option_type", "strike", "expiry",
                "instrument", "high", "low", "close",
            ]
            missing = [column for column in needed_cols if column not in df.columns]
            if missing:
                raise ValueError(f"Missing columns {missing} in {path}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")][needed_cols].copy()
            if d2.empty:
                continue

            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            if d2.empty:
                continue

            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2["strike_num"] = pd.to_numeric(d2["strike"], errors="coerce")
            d2["strike_int"] = d2["strike_num"].round().astype("Int64")
            d2["option_type"] = d2["option_type"].astype(str).str.upper()
            d2 = d2.dropna(
                subset=["day", "underlying", "expiry_date", "strike_int", "close"]
            )
            d2["strike_int"] = d2["strike_int"].astype(int)

            d2 = d2[d2["expiry_date"] >= d2["day"]]
            d2 = d2[(d2["day"] >= window_start) & (d2["day"] <= window_end)]
            if d2.empty:
                continue

            for (und, dy, expiry), group in d2.groupby(
                ["underlying", "day", "expiry_date"], sort=False
            ):
                if min_expiry_map.get((und, dy)) != expiry:
                    continue

                day_key = (und, dy, expiry)
                if day_key in processed_day_keys:
                    skipped_rows.append({
                        "day": dy,
                        "underlying": und,
                        "expiry": expiry,
                        "reason": (
                            "Duplicate (underlying, day, expiry) across pickles; "
                            "skipped to prevent double-counting"
                        ),
                    })
                    continue
                processed_day_keys.add(day_key)

                underlying_all = underlying_data.get(und)
                if underlying_all is None:
                    skipped_rows.append({
                        "day": dy, "underlying": und, "expiry": expiry,
                        "reason": "No underlying series downloaded",
                    })
                    continue

                underlying_day = underlying_all[underlying_all["day"] == dy]
                if underlying_day.empty:
                    skipped_rows.append({
                        "day": dy, "underlying": und, "expiry": expiry,
                        "reason": "Underlying minute data missing for day",
                    })
                    continue

                groups.append(
                    DayGroup(
                        und=und,
                        dy=dy,
                        expiry=expiry,
                        day_opt=group.copy(),
                        underlying_day=underlying_day.copy(),
                    )
                )

            print(
                f"[LOAD {file_index}/{total_files}] {os.path.basename(path)} "
                f"(day-groups={len(groups)})",
                flush=True,
            )

        except Exception as exc:
            message = (
                f"[LOAD {file_index}/{total_files} WARN] "
                f"{os.path.basename(path)} failed: {exc}"
            )
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(message) from exc
            print(message, flush=True)

    groups.sort(key=lambda item: (item.dy, item.und))

    if max_days is not None and max_days > 0:
        unique_days = sorted({item.dy for item in groups})
        keep_days = set(unique_days[-max_days:])
        groups = [item for item in groups if item.dy in keep_days]
        print(
            f"[LOAD] smoke sample: retained {len(keep_days)} dates / "
            f"{len(groups)} day-groups",
            flush=True,
        )

    if groups:
        print(
            f"[LOAD] precomputing option price books for {len(groups)} day-groups ...",
            flush=True,
        )
        for group_index, item in enumerate(groups, start=1):
            item.idx_all = build_minute_index(item.dy, SESSION_START_IST, SESSION_END_IST)
            item.price_book, item.symbols = build_price_book(item.day_opt, item.idx_all)
            if group_index % 50 == 0 or group_index == len(groups):
                print(
                    f"[LOAD] price books {group_index}/{len(groups)}",
                    flush=True,
                )

    print(f"[LOAD] ready: {len(groups)} day-groups", flush=True)
    return groups, skipped_rows


def simulate_groups(
    params: Params,
    groups: List[DayGroup],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Simulate one complete parameter set over the already-cached day-groups."""
    all_trades: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []

    for item in groups:
        if item.idx_all is None or item.price_book is None or item.symbols is None:
            item.idx_all = build_minute_index(item.dy, SESSION_START_IST, SESSION_END_IST)
            item.price_book, item.symbols = build_price_book(item.day_opt, item.idx_all)

        trades, skips = simulate_day_multi_trades(
            und=item.und,
            dy=item.dy,
            expiry=item.expiry,
            day_opt=item.day_opt,
            underlying_day=item.underlying_day,
            params=params,
            price_book=item.price_book,
            symbols=item.symbols,
            idx_all=item.idx_all,
        )
        all_trades.extend(trade.__dict__ for trade in trades)
        skipped_rows.extend(skips)

    all_df = pd.DataFrame(all_trades)
    if not all_df.empty:
        all_df = all_df.sort_values(
            ["day", "underlying", "trade_seq"]
        ).reset_index(drop=True)

    return all_df, pd.DataFrame(skipped_rows)


def process_pickles_generate_trades(
    params: Params,
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Normal single-backtest path, using the same cached engine as Optuna."""
    groups, parse_skips = build_day_groups(
        pickle_paths,
        min_expiry_map,
        underlying_data,
        window_start,
        window_end,
    )
    all_df, sim_skips = simulate_groups(params, groups)

    frames: List[pd.DataFrame] = []
    if parse_skips:
        frames.append(pd.DataFrame(parse_skips))
    if not sim_skips.empty:
        frames.append(sim_skips)
    skip_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    if not skip_df.empty:
        if "day" not in skip_df.columns:
            skip_df["day"] = pd.NaT
        if "underlying" not in skip_df.columns:
            skip_df["underlying"] = pd.NA
        skip_df = skip_df.sort_values(
            ["day", "underlying"], na_position="last"
        ).reset_index(drop=True)

    return all_df, skip_df


# =============================================================================
# Actual trades: one underlying per day (nearest expiry), include all re-entries for that underlying/day
# =============================================================================
def pick_actual_underlying_by_day(min_expiry_map: Dict[Tuple[str, date], date]) -> Dict[date, str]:
    by_day: Dict[date, List[Tuple[date, str]]] = {}
    for (und, dy), ex in min_expiry_map.items():
        if und not in TRADEABLE:
            continue

        dte = int((ex - dy).days)
        if dte not in ALLOWED_DTE:
            continue

        by_day.setdefault(dy, []).append((ex, und))

    out: Dict[date, str] = {}
    for dy, lst in by_day.items():
        # nearest expiry first; if tied, prefer NIFTY
        lst_sorted = sorted(lst, key=lambda t: (t[0], 0 if t[1] == "NIFTY" else 1))
        out[dy] = lst_sorted[0][1]
    return out

def build_actual_trades_df(all_trades_df: pd.DataFrame, min_expiry_map: Dict[Tuple[str, date], date]) -> pd.DataFrame:
    if all_trades_df.empty:
        return pd.DataFrame()

    actual_underlying = pick_actual_underlying_by_day(min_expiry_map)

    m = all_trades_df.copy()
    m["actual_underlying_for_day"] = m["day"].map(actual_underlying)

    # keep only days for which a 0/1-DTE actual underlying exists
    m = m[m["actual_underlying_for_day"].notna()]

    # keep only the selected underlying for that day
    m = m[m["underlying"] == m["actual_underlying_for_day"]]

    # keep only 0- and 1-DTE rows
    # keep only 0- and 1-DTE rows
    m = m[m["days_to_expiry"].isin(ALLOWED_DTE)]

    # keep all reattempts for the one selected underlying on that day
    m = m.drop(columns=["actual_underlying_for_day"])
    m = m.sort_values(["day", "trade_seq"]).reset_index(drop=True)

    # 1 if net exit PnL is positive, else 0
    m["is_exit_pnl_positive"] = (m["exit_pnl"] > 0).astype(int)

    return m


# =============================================================================
# Excel output
# =============================================================================
def _autosize_columns_safe(ws) -> None:
    # Safe autosize even when the sheet is "empty-ish"
    try:
        max_col = ws.max_column or 0
        if max_col <= 0:
            return
        for col_idx in range(1, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = 0
            for row_idx in range(1, min(ws.max_row or 1, 2000) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
    except Exception:
        # Never fail the whole run just because autosize misbehaved
        return

def _color_actual_trades_by_date(wb, actual_trades_df) -> None:
    """Shade rows so all attempts on the same calendar date share one colour,
    alternating between two soft fills as the date changes (visual grouping)."""
    if actual_trades_df is None or actual_trades_df.empty:
        return
    if "actual_trades" not in wb.sheetnames:
        return
    cols = list(actual_trades_df.columns)
    if "day" not in cols:
        return
    from openpyxl.styles import PatternFill
    ws = wb["actual_trades"]
    ncols = len(cols)
    fills = [
        PatternFill(fill_type="solid", fgColor="E8F0FE"),  # light blue
        PatternFill(fill_type="solid", fgColor="FFF3E0"),  # light amber
    ]
    days = actual_trades_df["day"].tolist()
    color_idx = 0
    prev_day = None
    first = True
    for i, d in enumerate(days):
        if first:
            first = False
        elif d != prev_day:
            color_idx ^= 1
        prev_day = d
        fill = fills[color_idx]
        excel_row = i + 2  # header occupies row 1
        for c in range(1, ncols + 1):
            ws.cell(row=excel_row, column=c).fill = fill


def write_excel(all_trades_df: pd.DataFrame, actual_trades_df: pd.DataFrame, skipped_df: pd.DataFrame) -> None:
    out_dir = os.path.dirname(os.path.abspath(OUTPUT_XLSX))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    piv_exit = pd.DataFrame()
    piv_eod_first = pd.DataFrame()
    monthwise_summary = pd.DataFrame()
    if not all_trades_df.empty:
        piv_exit = all_trades_df.pivot_table(index="day", columns="underlying", values="exit_pnl", aggfunc="sum").reset_index()

        first = all_trades_df[all_trades_df["trade_seq"] == 1]
        piv_eod_first = first.pivot_table(index="day", columns="underlying", values="eod_pnl", aggfunc="sum").reset_index()

        inst = all_trades_df.copy()
        inst["is_win_exit"] = inst["exit_pnl"] > 0
        inst["is_stoploss"] = inst["exit_reason"].astype(str).str.upper().eq("STOPLOSS")
        inst["is_profit_protect"] = inst["exit_reason"].astype(str).str.upper().eq("PROFIT_PROTECT")
        instrument_summary = (
            inst.groupby("underlying", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                win_rate_exit_pct=("is_win_exit", lambda s: 100.0 * s.mean()),
                stoploss_rate_pct=("is_stoploss", lambda s: 100.0 * s.mean()),
                profit_protect_rate_pct=("is_profit_protect", lambda s: 100.0 * s.mean()),
                avg_max_profit=("max_profit", "mean"),
                avg_max_loss=("max_loss", "mean"),
                worst_max_loss=("max_loss", "min"),
            )
            .sort_values("total_exit_pnl", ascending=False)
            .reset_index(drop=True)
        )
    else:
        instrument_summary = pd.DataFrame()

    if not actual_trades_df.empty:
        tmp = actual_trades_df.copy()
        tmp["month"] = pd.to_datetime(tmp["day"]).dt.to_period("M").astype(str)

        # Existing trade-level monthly summary
        monthwise_summary = (
            tmp.groupby("month", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                winning_trades=("is_exit_pnl_positive", "sum"),
            )
        )
        monthwise_summary["losing_trades"] = monthwise_summary["trades"] - monthwise_summary["winning_trades"]
        monthwise_summary["win_rate_pct"] = (
                100.0 * monthwise_summary["winning_trades"] / monthwise_summary["trades"]
        ).round(2)

        # New: daily PnL inside each month
        daily_tmp = (
            tmp.groupby(["month", "day"], as_index=False)
            .agg(daily_pnl=("exit_pnl", "sum"))
        )

        loss_day_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                avg_loss_on_loss_days=(
                    "daily_pnl",
                    lambda s: float(s[s < 0].mean()) if (s < 0).any() else 0.0
                ),
                max_loss_in_a_day=(
                    "daily_pnl",
                    lambda s: float(s.min()) if len(s) else 0.0
                ),
            )
        )

        # Count of profitable vs loss days within each month
        day_count_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                profitable_days=("daily_pnl", lambda s: int((s > 0).sum())),
                loss_days=("daily_pnl", lambda s: int((s < 0).sum())),
            )
        )

        # Date on which the worst (maximum-loss) day occurred, per month
        worst_rows = daily_tmp.loc[daily_tmp.groupby("month")["daily_pnl"].idxmin()]
        worst_day = worst_rows[["month", "day"]].rename(columns={"day": "max_loss_day_date"})

        monthwise_summary = monthwise_summary.merge(loss_day_stats, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(worst_day, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(day_count_stats, on="month", how="left")

        # Place the date column right after the max-loss value column
        _cols = list(monthwise_summary.columns)
        if "max_loss_day_date" in _cols and "max_loss_in_a_day" in _cols:
            _cols.remove("max_loss_day_date")
            _cols.insert(_cols.index("max_loss_in_a_day") + 1, "max_loss_day_date")
            monthwise_summary = monthwise_summary[_cols]

        # ---- Grand-total row across all months (totals for every column) ----
        if not monthwise_summary.empty:
            tot_trades = int(monthwise_summary["trades"].sum())
            tot_win = int(monthwise_summary["winning_trades"].sum())
            tot_lose = int(monthwise_summary["losing_trades"].sum())
            tot_pnl = float(monthwise_summary["total_exit_pnl"].sum())
            tot_prof_days = int(monthwise_summary["profitable_days"].sum())
            tot_loss_days = int(monthwise_summary["loss_days"].sum())
            neg_days = daily_tmp[daily_tmp["daily_pnl"] < 0]
            overall_worst_idx = daily_tmp["daily_pnl"].idxmin()
            total_row = {
                "month": "TOTAL",
                "trades": tot_trades,
                "total_exit_pnl": round(tot_pnl, 2),
                "avg_exit_pnl": round(tot_pnl / tot_trades, 2) if tot_trades else 0.0,
                "winning_trades": tot_win,
                "losing_trades": tot_lose,
                "win_rate_pct": round(100.0 * tot_win / tot_trades, 2) if tot_trades else 0.0,
                "avg_loss_on_loss_days": round(float(neg_days["daily_pnl"].mean()), 2) if len(neg_days) else 0.0,
                "max_loss_in_a_day": round(float(daily_tmp["daily_pnl"].min()), 2),
                "max_loss_day_date": daily_tmp.loc[overall_worst_idx, "day"],
                "profitable_days": tot_prof_days,
                "loss_days": tot_loss_days,
            }
            total_df = pd.DataFrame([total_row]).reindex(columns=monthwise_summary.columns)
            monthwise_summary = pd.concat([monthwise_summary, total_df], ignore_index=True)
    else:
        monthwise_summary = pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as xw:
        all_trades_df.to_excel(xw, sheet_name="all_trades_backtested", index=False)
        actual_trades_df.to_excel(xw, sheet_name="actual_trades", index=False)
        monthwise_summary.to_excel(xw, sheet_name="monthwise_summary", index=False)
        piv_exit.to_excel(xw, sheet_name="exit_pnl_pivot", index=False)
        piv_eod_first.to_excel(xw, sheet_name="eod_pnl_first_trade_pivot", index=False)
        instrument_summary.to_excel(xw, sheet_name="instrument_summary", index=False)
        skipped_df.to_excel(xw, sheet_name="skipped", index=False)

        wb = xw.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _autosize_columns_safe(ws)

        _color_actual_trades_by_date(wb, actual_trades_df)

    print(f"[DONE] Excel written: {OUTPUT_XLSX}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")

    print(f"[INFO] Pickles found: {len(paths)}")

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)

    lookback_label = "AUTO/full pickle range" if LOOKBACK_MONTHS is None else f"{LOOKBACK_MONTHS} months cap"

    print(f"[INFO] Data day-range seen: {min_day_seen} -> {end_day}")
    print(f"[INFO] Backtest window: {window_start} -> {end_day} ({lookback_label})")
    print(f"[INFO] Stoploss %/attempt: {_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Per-attempt stop cap: Rs {_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Daily max loss: Rs {_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES)} | "
          f"ProfitProtect trigger/giveback %: {_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES)} | "
          f"Re-entry delay min/attempt: {REENTRY_DELAY_BY_ATTEMPT} | Allowed DTE: {ALLOWED_DTE}")
    print(f"[INFO] Entry time: {ENTRY_TIME_IST} | Strategy exit/filter cutoff: {EXIT_TIME_IST}")
    print(f"[INFO] Day profit target: {PROFIT_TARGET_PCT:.0%} of premium (0 = disabled)")
    print(f"[INFO] Tradeables: {sorted(TRADEABLE)}")
    print(f"[INFO] Output: {OUTPUT_XLSX}")

    print("[STEP] Initializing Kite ...")
    kite = oUtils.intialize_kite_api()
    print("[OK] Kite ready.")

    underlying_data = download_underlyings(kite, window_start, end_day)

    all_trades_df, skipped_df = process_pickles_generate_trades(
        default_params(), paths, min_expiry_map, underlying_data, window_start, end_day
    )

    actual_trades_df = build_actual_trades_df(all_trades_df, min_expiry_map)

    write_excel(all_trades_df, actual_trades_df, skipped_df)

    if not all_trades_df.empty:
        print(all_trades_df.groupby("underlying")[["exit_pnl"]].describe())
    else:
        print("[WARN] No completed trades. Check 'skipped' sheet for reasons.")



# =============================================================================
# OPTUNA PROFIT OPTIMIZER
# =============================================================================
def _time_to_minutes(value: dtime) -> int:
    return int(value.hour * 60 + value.minute)


def _minutes_to_time(total_minutes: int) -> dtime:
    total_minutes = int(total_minutes)
    hour, minute = divmod(total_minutes, 60)
    if not (0 <= hour <= 23):
        raise ValueError(f"Invalid minute-of-day value: {total_minutes}")
    return dtime(hour, minute)


def _bounded_time_suggestion(
    trial,
    *,
    name: str,
    from_ist: str,
    to_ist: str,
    step_minutes: int,
    base_value: dtime,
    enabled: bool,
) -> dtime:
    """Suggest a time on a fixed minute grid, or return the property-file value."""
    if not enabled:
        return base_value

    start_minute = _time_to_minutes(parse_hhmm(from_ist))
    end_minute = _time_to_minutes(parse_hhmm(to_ist))
    if end_minute < start_minute:
        raise ValueError(f"{name}: end time must not precede start time")

    offset_max = end_minute - start_minute
    offset = trial.suggest_int(name, 0, offset_max, step=max(1, step_minutes))
    return _minutes_to_time(start_minute + offset)


def _linear_schedule(
    base_value: float,
    step_value: float,
    length: int,
    *,
    growth_slots: Optional[int],
    upper: float,
    decimals: int = 4,
) -> List[float]:
    """
    Create a non-decreasing schedule that can plateau after `growth_slots`.

    This preserves the original list semantics: attempts beyond a configured
    list reuse its final value instead of continuing to increase indefinitely.
    """
    length_i = max(1, int(length))
    plateau_slots = length_i if growth_slots is None else max(1, int(growth_slots))
    return [
        round(
            min(
                float(upper),
                float(base_value)
                + min(index, plateau_slots - 1) * float(step_value),
            ),
            decimals,
        )
        for index in range(length_i)
    ]


def _integer_schedule(
    base_value: int,
    step_value: int,
    length: int,
    *,
    growth_slots: Optional[int],
) -> List[int]:
    """Integer version of `_linear_schedule`, with the same plateau semantics."""
    length_i = max(1, int(length))
    plateau_slots = length_i if growth_slots is None else max(1, int(growth_slots))
    return [
        int(base_value + min(index, plateau_slots - 1) * step_value)
        for index in range(length_i)
    ]


def _profile_from_trial(
    trial,
    *,
    underlying: str,
    base: UnderlyingProfile,
) -> UnderlyingProfile:
    """Create one NIFTY/SENSEX profile from independent Optuna suggestions."""
    if not OPT_TUNE_PROFILES:
        if OPT_TUNE_DAILY_LOSS:
            daily_loss = trial.suggest_int(
                f"{underlying}_max_daily_loss_rupees",
                OPT_DAILY_LOSS_MIN,
                OPT_DAILY_LOSS_MAX,
                step=OPT_DAILY_LOSS_STEP,
            )
            return UnderlyingProfile(
                profit_target_pct=base.profit_target_pct,
                profit_target_pct_late=base.profit_target_pct_late,
                profit_target_late_from=base.profit_target_late_from,
                profit_protect_pct=base.profit_protect_pct,
                max_daily_loss_rupees=float(daily_loss),
                profit_protect_late_giveback_pct=base.profit_protect_late_giveback_pct,
                profit_protect_late_from=base.profit_protect_late_from,
            )
        return base

    protect_pct = trial.suggest_float(
        f"{underlying}_profit_protect_pct",
        OPT_PROTECT_MIN,
        OPT_PROTECT_MAX,
    )
    late_giveback_raw = trial.suggest_float(
        f"{underlying}_late_giveback_pct",
        OPT_LATE_GIVEBACK_MIN,
        OPT_LATE_GIVEBACK_MAX,
    )
    # A late giveback wider than the normal giveback cannot tighten the trail.
    late_giveback_pct = min(float(late_giveback_raw), float(protect_pct))

    target_pct = trial.suggest_float(
        f"{underlying}_profit_target_pct",
        OPT_TARGET_MIN,
        OPT_TARGET_MAX,
    )
    late_target_raw = trial.suggest_float(
        f"{underlying}_late_profit_target_pct",
        OPT_LATE_TARGET_MIN,
        OPT_LATE_TARGET_MAX,
    )
    # Late-session target is intended to be no harder than the normal target.
    late_target_pct = min(float(late_target_raw), float(target_pct))

    target_late_from = _bounded_time_suggestion(
        trial,
        name=f"{underlying}_target_late_from_offset",
        from_ist=OPT_LATE_FROM_IST,
        to_ist=OPT_LATE_TO_IST,
        step_minutes=OPT_LATE_TIME_STEP_MIN,
        base_value=base.profit_target_late_from,
        enabled=True,
    )
    protect_late_from = _bounded_time_suggestion(
        trial,
        name=f"{underlying}_protect_late_from_offset",
        from_ist=OPT_LATE_FROM_IST,
        to_ist=OPT_LATE_TO_IST,
        step_minutes=OPT_LATE_TIME_STEP_MIN,
        base_value=base.profit_protect_late_from,
        enabled=True,
    )

    if OPT_TUNE_DAILY_LOSS:
        daily_loss = trial.suggest_int(
            f"{underlying}_max_daily_loss_rupees",
            OPT_DAILY_LOSS_MIN,
            OPT_DAILY_LOSS_MAX,
            step=OPT_DAILY_LOSS_STEP,
        )
    else:
        daily_loss = int(round(base.max_daily_loss_rupees))

    return UnderlyingProfile(
        profit_target_pct=float(target_pct),
        profit_target_pct_late=float(late_target_pct),
        profit_target_late_from=target_late_from,
        profit_protect_pct=float(protect_pct),
        max_daily_loss_rupees=float(daily_loss),
        profit_protect_late_giveback_pct=float(late_giveback_pct),
        profit_protect_late_from=protect_late_from,
    )


def _params_from_trial(trial, base: Params) -> Params:
    """Map one Optuna trial to a complete simulator parameter object."""
    entry_time = _bounded_time_suggestion(
        trial,
        name="entry_time_offset",
        from_ist=OPT_ENTRY_FROM_IST,
        to_ist=OPT_ENTRY_TO_IST,
        step_minutes=OPT_ENTRY_STEP_MIN,
        base_value=base.entry_time,
        enabled=OPT_TUNE_TIMES,
    )
    exit_time = _bounded_time_suggestion(
        trial,
        name="exit_time_offset",
        from_ist=OPT_EXIT_FROM_IST,
        to_ist=OPT_EXIT_TO_IST,
        step_minutes=OPT_EXIT_STEP_MIN,
        base_value=base.exit_time,
        enabled=OPT_TUNE_TIMES,
    )

    if entry_time >= exit_time:
        # This should not occur with the default search windows, but retaining a
        # guard makes custom user ranges safe.
        raise ValueError(
            f"Invalid trial: entry {entry_time.strftime('%H:%M')} must precede "
            f"exit {exit_time.strftime('%H:%M')}"
        )

    if OPT_TUNE_REENTRY:
        max_reattempts = trial.suggest_int(
            "max_reattempts",
            OPT_REATTEMPTS_MIN,
            OPT_REATTEMPTS_MAX,
        )
        reentry_base = trial.suggest_int(
            "reentry_delay_base_min",
            OPT_REENTRY_DELAY_MIN,
            OPT_REENTRY_DELAY_MAX,
        )
        reentry_step = trial.suggest_int(
            "reentry_delay_step_min",
            OPT_REENTRY_STEP_MIN,
            OPT_REENTRY_STEP_MAX,
        )
        reentry_growth_slots = trial.suggest_int(
            "reentry_delay_growth_slots",
            1,
            OPT_REATTEMPTS_MAX + 1,
        )
        target_reentry_delay = trial.suggest_int(
            "profit_target_reentry_delay_min",
            OPT_TARGET_REENTRY_DELAY_MIN,
            OPT_TARGET_REENTRY_DELAY_MAX,
        )
        momentum_lookback = trial.suggest_int(
            "momentum_lookback_min",
            OPT_MOMENTUM_LOOKBACK_MIN,
            OPT_MOMENTUM_LOOKBACK_MAX,
        )
        momentum_enabled = bool(trial.suggest_int("entry_momentum_gate", 0, 1))
        target_reentry_enabled = bool(
            trial.suggest_int("profit_target_reentry_enabled", 0, 1)
        )
        delay_schedule = _integer_schedule(
            reentry_base,
            reentry_step,
            max_reattempts + 1,
            growth_slots=reentry_growth_slots,
        )
    else:
        max_reattempts = base.max_reattempts
        target_reentry_delay = base.profit_target_reentry_delay_min
        momentum_lookback = base.momentum_lookback_min
        momentum_enabled = base.entry_momentum_gate
        target_reentry_enabled = base.profit_target_reentry_enabled
        delay_schedule = list(base.reentry_delay_by_attempt)

    if OPT_TUNE_STOPLOSS:
        sl_base = trial.suggest_float(
            "stop_loss_base_pct",
            OPT_SL_BASE_MIN,
            OPT_SL_BASE_MAX,
        )
        sl_step = trial.suggest_float(
            "stop_loss_step_pct",
            OPT_SL_STEP_MIN,
            OPT_SL_STEP_MAX,
        )
        stop_growth_slots = trial.suggest_int(
            "stop_loss_growth_slots",
            1,
            OPT_REATTEMPTS_MAX + 1,
        )
        stop_cap = trial.suggest_int(
            "max_stop_loss_rupees",
            OPT_STOP_CAP_MIN,
            OPT_STOP_CAP_MAX,
            step=OPT_STOP_CAP_STEP,
        )
        stop_schedule = _linear_schedule(
            sl_base,
            sl_step,
            max_reattempts + 1,
            growth_slots=stop_growth_slots,
            upper=0.95,
        )
    else:
        stop_cap = int(round(base.max_loss_limit_cap_rupees))
        stop_schedule = list(base.loss_limit_pct_by_attempt)

    profiles = {
        underlying: _profile_from_trial(
            trial,
            underlying=underlying,
            base=base.profile_for(underlying),
        )
        for underlying in sorted(base.profiles)
    }

    return Params(
        entry_time=entry_time,
        exit_time=exit_time,
        loss_limit_pct_by_attempt=stop_schedule,
        max_loss_limit_cap_rupees=float(stop_cap),
        max_reattempts=int(max_reattempts),
        reentry_delay_by_attempt=delay_schedule,
        entry_momentum_gate=bool(momentum_enabled),
        momentum_lookback_min=int(momentum_lookback),
        profit_target_reentry_enabled=bool(target_reentry_enabled),
        profit_target_reentry_delay_min=int(target_reentry_delay),
        profiles=profiles,
    )


def _inr(value: Any) -> str:
    """Indian-grouped rupee formatting used in console progress."""
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = 0.0

    sign = "-" if number < 0 else ""
    integer = str(int(round(abs(number))))
    if len(integer) <= 3:
        grouped = integer
    else:
        tail = integer[-3:]
        head = integer[:-3]
        pairs: List[str] = []
        while len(head) > 2:
            pairs.append(head[-2:])
            head = head[:-2]
        if head:
            pairs.append(head)
        grouped = ",".join(reversed(pairs)) + "," + tail
    return f"{sign}Rs {grouped}"


def robustness_metrics(actual_df: pd.DataFrame) -> Dict[str, Any]:
    """Compute daily/monthly diagnostics for one trial's actually-traded book."""
    if actual_df is None or actual_df.empty:
        return {
            "n_days": 0,
            "n_months": 0,
            "profitable_day_ratio": 0.0,
            "profitable_month_ratio": 0.0,
            "total_pnl": 0.0,
            "mean_month": 0.0,
            "median_month": 0.0,
            "worst_day": 0.0,
            "worst_month": 0.0,
            "monthly": pd.Series(dtype="float64"),
        }

    trades = actual_df.copy()
    day_values = pd.to_datetime(trades["day"])
    daily = trades.groupby(day_values.dt.date)["exit_pnl"].sum().sort_index()
    monthly = (
        trades.groupby(day_values.dt.to_period("M"))["exit_pnl"]
        .sum()
        .sort_index()
    )

    return {
        "n_days": int(len(daily)),
        "n_months": int(len(monthly)),
        "profitable_day_ratio": float((daily > 0).mean()),
        "profitable_month_ratio": float((monthly > 0).mean()),
        "total_pnl": float(daily.sum()),
        "mean_month": float(monthly.mean()),
        "median_month": float(monthly.median()),
        "worst_day": float(daily.min()),
        "worst_month": float(monthly.min()),
        "monthly": monthly,
    }


def _activity_minima() -> Tuple[int, int]:
    """Relax guards for an explicitly requested smoke sample."""
    if SAMPLE_MAX_DAYS is not None or SAMPLE_MAX_PICKLES is not None:
        return 1, 1
    return max(1, OPT_MIN_DAYS), max(1, OPT_MIN_MONTHS)


def _score_from_metrics(metrics: Dict[str, Any]) -> float:
    """Default objective: total net profit in rupees."""
    min_days, min_months = _activity_minima()
    if metrics["n_days"] < min_days or metrics["n_months"] < min_months:
        return _OPT_DISQUALIFY
    return float(metrics["total_pnl"])


def _cv_score(actual_df: pd.DataFrame, folds: int) -> float:
    """
    Optional contiguous-month walk-forward objective.

    This is not pure total profit: it rewards profit across multiple regimes by
    subtracting OPT_CV_PENALTY times the fold-profit standard deviation.
    """
    if actual_df is None or actual_df.empty:
        return _OPT_DISQUALIFY

    trades = actual_df.copy()
    day_values = pd.to_datetime(trades["day"])
    daily = trades.groupby(day_values.dt.date)["exit_pnl"].sum()
    monthly = (
        trades.groupby(day_values.dt.to_period("M"))["exit_pnl"]
        .sum()
        .sort_index()
    )

    min_days, min_months = _activity_minima()
    if len(daily) < min_days or len(monthly) < max(min_months, folds):
        return _OPT_DISQUALIFY

    monthly_items = list(monthly.items())
    block_profits: List[float] = []
    count = len(monthly_items)
    for fold_index in range(folds):
        lower = (fold_index * count) // folds
        upper = ((fold_index + 1) * count) // folds
        block = monthly_items[lower:upper]
        if block:
            block_profits.append(float(sum(value for _, value in block)))

    if not block_profits:
        return _OPT_DISQUALIFY

    series = pd.Series(block_profits, dtype="float64")
    return float(series.mean() - OPT_CV_PENALTY * series.std(ddof=0))


def params_to_properties(params: Params, *, include_run_control: bool = True) -> str:
    """Render a complete, unambiguous property block for a winning trial."""
    rows: List[str] = [
        "# Generated by atm_straddle_expiry_day_V3OPT_OPTUNA__optimizer_DTE0.py",
        "# Verify with RUN_MODE=backtest before any live use.",
    ]
    if include_run_control:
        rows.append("RUN_MODE=backtest")

    rows.extend([
        f"PICKLES_DIR={PICKLES_DIR}",
        f"ENTRY_TIME_IST={params.entry_time.strftime('%H:%M')}",
        f"EXIT_TIME_IST={params.exit_time.strftime('%H:%M')}",
        f"ENTRY_MOMENTUM_GATE={1 if params.entry_momentum_gate else 0}",
        f"MOMENTUM_LOOKBACK_MIN={params.momentum_lookback_min}",
        "ALLOWED_DTE=" + ", ".join(str(value) for value in ALLOWED_DTE),
        "LOSS_LIMIT_RUPEES_BY_ATTEMPT="
        + ", ".join(f"{value:.4f}" for value in params.loss_limit_pct_by_attempt),
        f"MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT={int(round(params.max_loss_limit_cap_rupees))}",
        f"MAX_REATTEMPTS={params.max_reattempts}",
        "REENTRY_DELAY_BY_ATTEMPT="
        + ", ".join(str(value) for value in params.reentry_delay_by_attempt),
        f"PROFIT_TARGET_REENTRY_ENABLED={1 if params.profit_target_reentry_enabled else 0}",
        f"PROFIT_TARGET_REENTRY_DELAY={params.profit_target_reentry_delay_min}",
    ])

    for underlying in sorted(params.profiles):
        profile = params.profile_for(underlying)
        rows.extend([
            "",
            f"# ---- {underlying} optimized profile ----",
            f"{underlying}_PROFIT_PROTECT_TRIGGER_RUPEES={profile.profit_protect_pct:.6f}",
            f"{underlying}_PROFIT_PROTECT_LATE_GIVEBACK="
            f"{profile.profit_protect_late_giveback_pct:.6f}",
            f"{underlying}_PROFIT_PROTECT_LATE_FROM_IST="
            f"{profile.profit_protect_late_from.strftime('%H:%M')}",
            f"{underlying}_PROFIT_TARGET_PCT={profile.profit_target_pct:.6f}",
            f"{underlying}_PROFIT_TARGET_PCT_LATE={profile.profit_target_pct_late:.6f}",
            f"{underlying}_PROFIT_TARGET_LATE_FROM_IST="
            f"{profile.profit_target_late_from.strftime('%H:%M')}",
            f"{underlying}_MAX_DAILY_LOSS_RUPEES="
            f"{int(round(profile.max_daily_loss_rupees))}",
        ])

    rows.extend([
        "",
        f"LOOKBACK_MONTHS={LOOKBACK_MONTHS_RAW}",
        f"FAIL_ON_PICKLE_ERROR={1 if FAIL_ON_PICKLE_ERROR else 0}",
    ])
    return "\n".join(rows) + "\n"


class _FrozenTrialView:
    """Re-run the same suggestion code against a completed Optuna trial."""

    def __init__(self, trial) -> None:
        self._params = dict(trial.params)

    def suggest_int(self, name: str, *args, **kwargs) -> int:
        return int(self._params[name])

    def suggest_float(self, name: str, *args, **kwargs) -> float:
        return float(self._params[name])


def _base_trial_values(base: Params) -> Dict[str, Any]:
    """
    Convert the supplied V3 configuration into the optimizer's compact search
    parameterization.

    The stop/re-entry schedules are represented by a base, slope and plateau, so
    this is a close search seed rather than a byte-for-byte copy of every list
    element. The exact property-file baseline is evaluated separately and is
    retained automatically when no Optuna trial beats it.
    """
    values: Dict[str, Any] = {}

    def time_offset(value: dtime, start_text: str) -> int:
        return _time_to_minutes(value) - _time_to_minutes(parse_hhmm(start_text))

    if OPT_TUNE_TIMES:
        values["entry_time_offset"] = time_offset(base.entry_time, OPT_ENTRY_FROM_IST)
        values["exit_time_offset"] = time_offset(base.exit_time, OPT_EXIT_FROM_IST)

    if OPT_TUNE_REENTRY:
        values["max_reattempts"] = int(base.max_reattempts)
        delays = list(base.reentry_delay_by_attempt) or [0]
        values["reentry_delay_base_min"] = int(delays[0])
        values["reentry_delay_step_min"] = (
            int(delays[1] - delays[0]) if len(delays) > 1 else 0
        )
        values["reentry_delay_growth_slots"] = min(
            len(delays), OPT_REATTEMPTS_MAX + 1
        )
        values["profit_target_reentry_delay_min"] = int(
            base.profit_target_reentry_delay_min
        )
        values["momentum_lookback_min"] = int(base.momentum_lookback_min)
        values["entry_momentum_gate"] = 1 if base.entry_momentum_gate else 0
        values["profit_target_reentry_enabled"] = (
            1 if base.profit_target_reentry_enabled else 0
        )

    if OPT_TUNE_STOPLOSS:
        stops = list(base.loss_limit_pct_by_attempt) or [0.0]
        values["stop_loss_base_pct"] = float(stops[0])
        values["stop_loss_step_pct"] = (
            float(stops[1] - stops[0]) if len(stops) > 1 else 0.0
        )
        values["stop_loss_growth_slots"] = min(
            len(stops), OPT_REATTEMPTS_MAX + 1
        )
        values["max_stop_loss_rupees"] = int(
            round(base.max_loss_limit_cap_rupees / OPT_STOP_CAP_STEP)
            * OPT_STOP_CAP_STEP
        )

    for underlying in sorted(base.profiles):
        profile = base.profile_for(underlying)
        if OPT_TUNE_PROFILES:
            values[f"{underlying}_profit_protect_pct"] = float(
                profile.profit_protect_pct
            )
            values[f"{underlying}_late_giveback_pct"] = float(
                profile.profit_protect_late_giveback_pct
            )
            values[f"{underlying}_profit_target_pct"] = float(
                profile.profit_target_pct
            )
            values[f"{underlying}_late_profit_target_pct"] = float(
                profile.profit_target_pct_late
            )
            values[f"{underlying}_target_late_from_offset"] = time_offset(
                profile.profit_target_late_from,
                OPT_LATE_FROM_IST,
            )
            values[f"{underlying}_protect_late_from_offset"] = time_offset(
                profile.profit_protect_late_from,
                OPT_LATE_FROM_IST,
            )
        if OPT_TUNE_DAILY_LOSS:
            values[f"{underlying}_max_daily_loss_rupees"] = int(
                round(profile.max_daily_loss_rupees / OPT_DAILY_LOSS_STEP)
                * OPT_DAILY_LOSS_STEP
            )

    return values


def _value_within_search_space(name: str, value: Any) -> bool:
    """Prevent an out-of-range property-file baseline from breaking enqueue."""
    numeric_ranges: Dict[str, Tuple[float, float]] = {
        "entry_time_offset": (
            0,
            _time_to_minutes(parse_hhmm(OPT_ENTRY_TO_IST))
            - _time_to_minutes(parse_hhmm(OPT_ENTRY_FROM_IST)),
        ),
        "exit_time_offset": (
            0,
            _time_to_minutes(parse_hhmm(OPT_EXIT_TO_IST))
            - _time_to_minutes(parse_hhmm(OPT_EXIT_FROM_IST)),
        ),
        "max_reattempts": (OPT_REATTEMPTS_MIN, OPT_REATTEMPTS_MAX),
        "reentry_delay_base_min": (OPT_REENTRY_DELAY_MIN, OPT_REENTRY_DELAY_MAX),
        "reentry_delay_step_min": (OPT_REENTRY_STEP_MIN, OPT_REENTRY_STEP_MAX),
        "reentry_delay_growth_slots": (1, OPT_REATTEMPTS_MAX + 1),
        "profit_target_reentry_delay_min": (
            OPT_TARGET_REENTRY_DELAY_MIN,
            OPT_TARGET_REENTRY_DELAY_MAX,
        ),
        "momentum_lookback_min": (
            OPT_MOMENTUM_LOOKBACK_MIN,
            OPT_MOMENTUM_LOOKBACK_MAX,
        ),
        "entry_momentum_gate": (0, 1),
        "profit_target_reentry_enabled": (0, 1),
        "stop_loss_base_pct": (OPT_SL_BASE_MIN, OPT_SL_BASE_MAX),
        "stop_loss_step_pct": (OPT_SL_STEP_MIN, OPT_SL_STEP_MAX),
        "stop_loss_growth_slots": (1, OPT_REATTEMPTS_MAX + 1),
        "max_stop_loss_rupees": (OPT_STOP_CAP_MIN, OPT_STOP_CAP_MAX),
    }

    if name.endswith("_profit_protect_pct"):
        bounds = (OPT_PROTECT_MIN, OPT_PROTECT_MAX)
    elif name.endswith("_late_giveback_pct"):
        bounds = (OPT_LATE_GIVEBACK_MIN, OPT_LATE_GIVEBACK_MAX)
    elif name.endswith("_profit_target_pct") and not name.endswith("_late_profit_target_pct"):
        bounds = (OPT_TARGET_MIN, OPT_TARGET_MAX)
    elif name.endswith("_late_profit_target_pct"):
        bounds = (OPT_LATE_TARGET_MIN, OPT_LATE_TARGET_MAX)
    elif name.endswith("_target_late_from_offset") or name.endswith(
        "_protect_late_from_offset"
    ):
        bounds = (
            0,
            _time_to_minutes(parse_hhmm(OPT_LATE_TO_IST))
            - _time_to_minutes(parse_hhmm(OPT_LATE_FROM_IST)),
        )
    elif name.endswith("_max_daily_loss_rupees"):
        bounds = (OPT_DAILY_LOSS_MIN, OPT_DAILY_LOSS_MAX)
    else:
        bounds = numeric_ranges.get(name)

    if bounds is None:
        return True
    return float(bounds[0]) <= float(value) <= float(bounds[1])


def _trial_record(
    trial,
    params: Params,
    metrics: Dict[str, Any],
    run_index: int,
    elapsed_seconds: float,
) -> Dict[str, Any]:
    """Compact durable record written after every completed trial."""
    return {
        "run_index": int(run_index),
        "trial_number": int(trial.number),
        "state": str(getattr(trial, "state", "")),
        "objective_score": trial.value,
        "net_pnl": round(float(metrics.get("total_pnl", 0.0)), 2),
        "mean_month": round(float(metrics.get("mean_month", 0.0)), 2),
        "median_month": round(float(metrics.get("median_month", 0.0)), 2),
        "worst_month": round(float(metrics.get("worst_month", 0.0)), 2),
        "worst_day": round(float(metrics.get("worst_day", 0.0)), 2),
        "profitable_month_ratio": round(
            float(metrics.get("profitable_month_ratio", 0.0)), 6
        ),
        "profitable_day_ratio": round(
            float(metrics.get("profitable_day_ratio", 0.0)), 6
        ),
        "n_months": int(metrics.get("n_months", 0)),
        "n_days": int(metrics.get("n_days", 0)),
        "entry_time": params.entry_time.strftime("%H:%M"),
        "exit_time": params.exit_time.strftime("%H:%M"),
        "max_reattempts": params.max_reattempts,
        "stop_schedule": ";".join(
            f"{value:.4f}" for value in params.loss_limit_pct_by_attempt
        ),
        "reentry_schedule": ";".join(
            str(value) for value in params.reentry_delay_by_attempt
        ),
        "params_json": json.dumps(trial.params, sort_keys=True),
        "properties_json": json.dumps(
            {
                line.split("=", 1)[0]: line.split("=", 1)[1]
                for line in params_to_properties(
                    params, include_run_control=False
                ).splitlines()
                if line and not line.startswith("#") and "=" in line
            },
            sort_keys=True,
        ),
        "elapsed_seconds": round(float(elapsed_seconds), 1),
    }


def optimize(
    groups: List[DayGroup],
    min_expiry_map: Dict[Tuple[str, date], date],
    *,
    n_trials: int,
    cv_folds: int,
    seed: int,
    progress_every: int,
):
    """Run Optuna over cached day-groups and persist every result."""
    import csv
    import datetime as dt
    import time as time_module

    try:
        import optuna
    except ImportError as exc:
        raise RuntimeError(
            "Optuna is required. Install it in the active interpreter with: "
            "pip install optuna"
        ) from exc

    optuna.logging.set_verbosity(optuna.logging.WARNING)
    os.makedirs(OPT_OUTPUT_DIR, exist_ok=True)

    run_stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    trial_csv_path = os.path.join(
        OPT_OUTPUT_DIR,
        f"{OPT_STUDY_NAME}_{run_stamp}_trials.csv",
    )
    csv_fields = [
        "run_index", "trial_number", "state", "objective_score", "net_pnl",
        "mean_month", "median_month", "worst_month", "worst_day",
        "profitable_month_ratio", "profitable_day_ratio", "n_months",
        "n_days", "entry_time", "exit_time", "max_reattempts",
        "stop_schedule", "reentry_schedule", "params_json",
        "properties_json", "elapsed_seconds",
    ]
    csv_file = open(trial_csv_path, "w", newline="", encoding="utf-8")
    csv_writer = csv.DictWriter(csv_file, fieldnames=csv_fields)
    csv_writer.writeheader()
    csv_file.flush()

    base = default_params()

    # Evaluate the exact property-file configuration before starting Optuna.
    # This is the hard benchmark and also the fallback winner if the search does
    # not improve the requested objective.
    baseline_all_df, _baseline_skips_df = simulate_groups(base, groups)
    baseline_actual_df = build_actual_trades_df(baseline_all_df, min_expiry_map)
    baseline_metrics = robustness_metrics(baseline_actual_df)
    baseline_objective = (
        _cv_score(baseline_actual_df, cv_folds)
        if cv_folds > 1
        else _score_from_metrics(baseline_metrics)
    )
    print(
        f"[OPT] exact property-file baseline: "
        f"net={_inr(baseline_metrics['total_pnl'])}, "
        f"objective={_inr(baseline_objective)}",
        flush=True,
    )

    storage = None
    if OPT_SAVE_DB:
        db_path = os.path.abspath(
            os.path.join(OPT_OUTPUT_DIR, f"{OPT_STUDY_NAME}.db")
        )
        storage = "sqlite:///" + db_path.replace("\\", "/")
        print(f"[OPT] resumable study: {db_path}", flush=True)

    sampler = optuna.samplers.TPESampler(
        seed=seed,
        n_startup_trials=max(1, OPT_STARTUP_TRIALS),
    )
    study = optuna.create_study(
        direction="maximize",
        sampler=sampler,
        study_name=OPT_STUDY_NAME,
        storage=storage,
        load_if_exists=bool(storage),
    )

    search_signature = json.dumps(
        {
            "version": "V3OPT_OPTUNA_2",
            "tune_times": OPT_TUNE_TIMES,
            "tune_stoploss": OPT_TUNE_STOPLOSS,
            "tune_reentry": OPT_TUNE_REENTRY,
            "tune_profiles": OPT_TUNE_PROFILES,
            "tune_daily_loss": OPT_TUNE_DAILY_LOSS,
            "entry_range": [OPT_ENTRY_FROM_IST, OPT_ENTRY_TO_IST, OPT_ENTRY_STEP_MIN],
            "exit_range": [OPT_EXIT_FROM_IST, OPT_EXIT_TO_IST, OPT_EXIT_STEP_MIN],
            "stop_range": [
                OPT_SL_BASE_MIN, OPT_SL_BASE_MAX, OPT_SL_STEP_MIN,
                OPT_SL_STEP_MAX, OPT_STOP_CAP_MIN, OPT_STOP_CAP_MAX,
                OPT_STOP_CAP_STEP,
            ],
            "reentry_range": [
                OPT_REATTEMPTS_MIN, OPT_REATTEMPTS_MAX,
                OPT_REENTRY_DELAY_MIN, OPT_REENTRY_DELAY_MAX,
                OPT_REENTRY_STEP_MIN, OPT_REENTRY_STEP_MAX,
                1, OPT_REATTEMPTS_MAX + 1,
                OPT_TARGET_REENTRY_DELAY_MIN, OPT_TARGET_REENTRY_DELAY_MAX,
                OPT_MOMENTUM_LOOKBACK_MIN, OPT_MOMENTUM_LOOKBACK_MAX,
            ],
            "profile_range": [
                OPT_PROTECT_MIN, OPT_PROTECT_MAX,
                OPT_LATE_GIVEBACK_MIN, OPT_LATE_GIVEBACK_MAX,
                OPT_TARGET_MIN, OPT_TARGET_MAX,
                OPT_LATE_TARGET_MIN, OPT_LATE_TARGET_MAX,
                OPT_LATE_FROM_IST, OPT_LATE_TO_IST, OPT_LATE_TIME_STEP_MIN,
                OPT_DAILY_LOSS_MIN, OPT_DAILY_LOSS_MAX, OPT_DAILY_LOSS_STEP,
            ],
        },
        sort_keys=True,
    )
    existing_signature = study.user_attrs.get("search_signature")
    if existing_signature is not None and existing_signature != search_signature:
        raise RuntimeError(
            "The saved Optuna study was created with a different search space. "
            "Change OPT_STUDY_NAME (recommended) or restore the original ranges "
            "before resuming it."
        )
    if existing_signature is None:
        study.set_user_attr("search_signature", search_signature)

    if len(study.trials) == 0:
        baseline_values = {
            name: value
            for name, value in _base_trial_values(base).items()
            if _value_within_search_space(name, value)
        }
        if baseline_values:
            study.enqueue_trial(baseline_values)
            print("[OPT] V3-shaped seed configuration queued as the first trial.", flush=True)

    print(f"[OPT] trial log: {trial_csv_path}", flush=True)
    print(
        f"[OPT] starting {n_trials} trial(s), day-groups={len(groups)}, "
        f"cv_folds={cv_folds}",
        flush=True,
    )

    start_time = time_module.time()
    completed_this_run = {"count": 0}

    def objective(trial):
        params = _params_from_trial(trial, base)
        all_df, _ = simulate_groups(params, groups)
        actual_df = build_actual_trades_df(all_df, min_expiry_map)
        metrics = robustness_metrics(actual_df)

        for key in (
            "n_days", "n_months", "profitable_day_ratio",
            "profitable_month_ratio", "total_pnl", "mean_month",
            "median_month", "worst_day", "worst_month",
        ):
            trial.set_user_attr(key, metrics[key])
        trial.set_user_attr(
            "monthly_pnl",
            {str(period): float(value) for period, value in metrics["monthly"].items()},
        )

        if cv_folds > 1:
            return _cv_score(actual_df, cv_folds)
        return _score_from_metrics(metrics)

    def progress_callback(study_obj, trial):
        completed_this_run["count"] += 1
        run_index = completed_this_run["count"]
        elapsed = time_module.time() - start_time

        try:
            params = _params_from_trial(_FrozenTrialView(trial), base)
            metrics = {
                key: trial.user_attrs.get(key, 0.0)
                for key in (
                    "n_days", "n_months", "profitable_day_ratio",
                    "profitable_month_ratio", "total_pnl", "mean_month",
                    "median_month", "worst_day", "worst_month",
                )
            }
            csv_writer.writerow(
                _trial_record(trial, params, metrics, run_index, elapsed)
            )
            csv_file.flush()
        except Exception as exc:
            print(
                f"[OPT WARN] failed to persist trial {trial.number}: {exc}",
                flush=True,
            )

        if run_index % progress_every != 0 and run_index != n_trials:
            return

        attrs = trial.user_attrs
        try:
            best_value = study_obj.best_value
            best_number = study_obj.best_trial.number
        except Exception:
            best_value = float("nan")
            best_number = -1

        print(
            f"[TRIAL {run_index:>4}/{n_trials}] "
            f"net={_inr(attrs.get('total_pnl', 0.0))} "
            f"mean_month={_inr(attrs.get('mean_month', 0.0))} "
            f"prof_month={float(attrs.get('profitable_month_ratio', 0.0))*100:5.1f}% "
            f"prof_day={float(attrs.get('profitable_day_ratio', 0.0))*100:5.1f}% "
            f"worst_month={_inr(attrs.get('worst_month', 0.0))} "
            f"| best_objective={_inr(best_value)}",
            flush=True,
        )

        if trial.number == best_number:
            monthly = attrs.get("monthly_pnl", {})
            if monthly:
                print("   >>> NEW BEST: month-wise net P/L", flush=True)
                cells = [
                    f"{month}:{_inr(value)}"
                    for month, value in sorted(monthly.items())
                ]
                for start in range(0, len(cells), 4):
                    print("       " + "   ".join(cells[start:start + 4]), flush=True)

    try:
        study.optimize(
            objective,
            n_trials=n_trials,
            callbacks=[progress_callback],
            show_progress_bar=False,
        )
    finally:
        csv_file.close()
        try:
            full_csv_path = os.path.join(
                OPT_OUTPUT_DIR,
                f"{OPT_STUDY_NAME}_{run_stamp}_optuna_full.csv",
            )
            study.trials_dataframe().to_csv(full_csv_path, index=False)
            print(f"[OPT] full Optuna table: {full_csv_path}", flush=True)
        except Exception as exc:
            print(f"[OPT WARN] full Optuna table not written: {exc}", flush=True)

    best_trial = study.best_trial
    optuna_best_params = _params_from_trial(_FrozenTrialView(best_trial), base)

    # Re-simulate the Optuna winner once, then compare it with the exact baseline.
    optuna_all_df, optuna_skips_df = simulate_groups(optuna_best_params, groups)
    optuna_actual_df = build_actual_trades_df(optuna_all_df, min_expiry_map)
    optuna_metrics = robustness_metrics(optuna_actual_df)
    optuna_objective = (
        _cv_score(optuna_actual_df, cv_folds)
        if cv_folds > 1
        else _score_from_metrics(optuna_metrics)
    )

    if baseline_objective >= optuna_objective:
        best_params = base
        best_all_df = baseline_all_df
        best_actual_df = baseline_actual_df
        best_skips_df = _baseline_skips_df
        best_metrics = baseline_metrics
        selected_objective = baseline_objective
        selected_source = "exact property-file baseline"
    else:
        best_params = optuna_best_params
        best_all_df = optuna_all_df
        best_actual_df = optuna_actual_df
        best_skips_df = optuna_skips_df
        best_metrics = optuna_metrics
        selected_objective = optuna_objective
        selected_source = f"Optuna trial {best_trial.number}"

    properties_path = os.path.join(
        OPT_OUTPUT_DIR,
        f"{OPT_STUDY_NAME}_{run_stamp}_BEST.properties",
    )
    with open(properties_path, "w", encoding="utf-8") as handle:
        handle.write(params_to_properties(best_params))

    best_trades_path = os.path.join(
        OPT_OUTPUT_DIR,
        f"{OPT_STUDY_NAME}_{run_stamp}_BEST_actual_trades.csv",
    )
    best_actual_df.to_csv(best_trades_path, index=False)

    best_monthly_path = os.path.join(
        OPT_OUTPUT_DIR,
        f"{OPT_STUDY_NAME}_{run_stamp}_BEST_monthly_pnl.csv",
    )
    best_metrics["monthly"].rename("net_pnl").to_csv(best_monthly_path)

    if not best_skips_df.empty:
        best_skips_path = os.path.join(
            OPT_OUTPUT_DIR,
            f"{OPT_STUDY_NAME}_{run_stamp}_BEST_skipped.csv",
        )
        best_skips_df.to_csv(best_skips_path, index=False)

    print("\n================ BEST V3-OPT CONFIG ================", flush=True)
    print(f"Selected source    : {selected_source}", flush=True)
    print(f"Optuna best trial  : {best_trial.number}", flush=True)
    print(f"Selected objective : {_inr(selected_objective)}", flush=True)
    print(f"Baseline objective : {_inr(baseline_objective)}", flush=True)
    print(f"Total net profit   : {_inr(best_metrics['total_pnl'])}", flush=True)
    print(f"Mean monthly P/L   : {_inr(best_metrics['mean_month'])}", flush=True)
    print(f"Worst month        : {_inr(best_metrics['worst_month'])}", flush=True)
    print(
        f"Profitable months  : "
        f"{best_metrics['profitable_month_ratio']*100:.1f}%",
        flush=True,
    )
    print(
        f"Profitable days    : "
        f"{best_metrics['profitable_day_ratio']*100:.1f}%",
        flush=True,
    )
    print(f"Best properties    : {properties_path}", flush=True)
    print(f"Best actual trades : {best_trades_path}", flush=True)
    print(f"Best monthly P/L   : {best_monthly_path}", flush=True)
    print("\n---- READY-TO-PASTE PROPERTY BLOCK ----", flush=True)
    print(params_to_properties(best_params), flush=True)

    return study, best_params


def run_optimizer(
    *,
    n_trials: int,
    cv_folds: int,
    max_pickles: Optional[int],
    max_days: Optional[int],
    progress_every: int,
    seed: int,
):
    """End-to-end optimizer entry point: load once, cache once, simulate many."""
    print("[PHASE 1] Scanning option pickles ...", flush=True)
    paths = sorted(
        glob.glob(os.path.join(PICKLES_DIR, "*.pkl"))
        + glob.glob(os.path.join(PICKLES_DIR, "*.pickle"))
    )
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")
    if max_pickles is not None and max_pickles > 0:
        paths = paths[:max_pickles]
    print(f"[PHASE 1] pickle files in scope: {len(paths)}", flush=True)

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)
    print(
        f"[PHASE 1] backtest window: {window_start} -> {end_day}",
        flush=True,
    )

    print("[PHASE 2] Downloading NIFTY/SENSEX underlying minute data ...", flush=True)
    kite = oUtils.intialize_kite_api()
    underlying_data = download_underlyings(kite, window_start, end_day)

    print("[PHASE 3] Building cached day-groups ...", flush=True)
    groups, parse_skips = build_day_groups(
        paths,
        min_expiry_map,
        underlying_data,
        window_start,
        end_day,
        max_pickles=max_pickles,
        max_days=max_days,
    )
    if not groups:
        raise RuntimeError(
            "No day-groups were built. Check the pickle path, date window, "
            "allowed underlyings and underlying data."
        )
    if parse_skips:
        print(
            f"[PHASE 3] non-fatal parse/data skips: {len(parse_skips)}",
            flush=True,
        )

    print("[PHASE 4] Starting Optuna search ...", flush=True)
    return optimize(
        groups,
        min_expiry_map,
        n_trials=n_trials,
        cv_folds=cv_folds,
        seed=seed,
        progress_every=progress_every,
    )


# =============================================================================
# ENTRYPOINT
# =============================================================================
if __name__ == "__main__":
    if RUN_MODE == "optimize":
        run_optimizer(
            n_trials=OPT_TRIALS,
            cv_folds=OPT_CV_FOLDS,
            max_pickles=SAMPLE_MAX_PICKLES,
            max_days=SAMPLE_MAX_DAYS,
            progress_every=OPT_PROGRESS_EVERY,
            seed=OPT_SEED,
        )
    elif RUN_MODE == "backtest":
        main()
    else:
        raise ValueError(
            f"Unknown RUN_MODE={RUN_MODE!r}; use 'optimize' or 'backtest'."
        )
