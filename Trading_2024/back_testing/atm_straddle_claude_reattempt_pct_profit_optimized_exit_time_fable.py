"""
ATM short-straddle backtester + robustness optimizer (NIFTY / SENSEX, DTE=0/1).
================================================================================

This file does two things, selected by the RUN_MODE setting in the RUN CONTROL
block at the top (no command line -- just edit and press Run in PyCharm):

  1. SINGLE BACKTEST   -> RUN_MODE = "backtest"
     Runs one backtest using the parameters/env-vars below and writes a detailed
     multi-sheet Excel report. Identical to the original workflow.

  2. ROBUSTNESS OPTIMIZER -> RUN_MODE = "optimize"
     Searches the seven tunables with Optuna (TPE), maximizing ROBUSTNESS = the
     total net profit (profit only). Loads the option
     data ONCE and re-simulates per trial, printing progress every trial. EVERY
     tested config + its results is saved to a CSV in OPT_OUTPUT_DIR (flushed per
     trial, so an interrupted run keeps everything). Set OPT_SAVE_DB = True to
     also persist a resumable Optuna SQLite study.

Optimized parameters:
     ENTRY_TIME_IST, EXIT_TIME_IST, LOSS_LIMIT_RUPEES_BY_ATTEMPT,
     PROFIT_PROTECT_TRIGGER_RUPEES, MAX_REATTEMPTS, PROFIT_TARGET_PCT,
     REENTRY_DELAY_BY_ATTEMPT

EXIT_TIME_IST means "last fresh entry/re-entry time", not forced square-off.
Trades already open after EXIT_TIME_IST continue to exit using the existing
STOPLOSS / PROFIT_TARGET / PROFIT_PROTECT / EOD logic.

--------------------------------------------------------------------------------
QUICK SMALL-SAMPLE RUN (do this first to confirm everything works end-to-end):
--------------------------------------------------------------------------------
In the RUN CONTROL block at the top, set:
     RUN_MODE          = "optimize"
     OPT_TRIALS        = 5
     SAMPLE_MAX_PICKLES = 3      # read only 3 pickle files (also shrinks the
                                 # Kite underlying download)
     SAMPLE_MAX_DAYS    = 20     # simulate only the 20 most recent day-groups
     OPT_PROGRESS_EVERY = 1      # print after every trial
...then press Run. It should finish in a minute or two and print load progress,
all five trial lines, and a BEST CONFIG block.

For the real search, set:
     OPT_TRIALS = 300
     OPT_CV_FOLDS = 5
     SAMPLE_MAX_PICKLES = None
     SAMPLE_MAX_DAYS = None

Dependencies: pandas, openpyxl, optuna  (pip install optuna).
"""

import os
from pathlib import Path
import glob
import time
from dataclasses import dataclass
from datetime import datetime, date, time as dtime, timedelta
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd

import Trading_2024.OptionTradeUtils as oUtils

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore

try:
    import pytz  # type: ignore
except Exception:
    pytz = None  # type: ignore

try:
    from dateutil.relativedelta import relativedelta  # type: ignore
except Exception:
    relativedelta = None  # type: ignore


# =============================================================================
# RUN CONTROL  -- EDIT THESE, THEN JUST RUN THE FILE (no command line needed)
# =============================================================================
# What to do when you run this file:
#   "backtest" -> one backtest with the params below + Excel report (original flow)
#   "optimize" -> Optuna robustness search over the seven tunables
RUN_MODE = "optimize"

# --- Optimizer settings (used only when RUN_MODE == "optimize") ---
OPT_TRIALS = 400            # number of optimization trials
OPT_CV_FOLDS = 5            # 1 = score on full sample; >1 = walk-forward block robustness
OPT_PROGRESS_EVERY = 5      # (retained for compatibility) per-trial stats now print EVERY trial
OPT_SEED = 42              # RNG seed for reproducible searches

# --- DIG-DEEPER controls (all new; see the optimizer section for details) ---
# Stage-2 'zoom' refinement: after the broad search, a second study re-searches a
# NARROWED box around the stage-1 winner on a FINER grid (1-minute entry/exit).
# 0 disables. This is what finds the extra profit sitting between the 5-min grid
# points and at the edges of the broad optimum.
OPT_STAGE2_TRIALS = 150
# Fraction of each parameter's ORIGINAL range kept around the winner in stage 2
# (0.15 = +-7.5% of the original range on each side, clipped to the original box).
OPT_STAGE2_SHRINK = 0.15

# Early pruning: evaluate each trial on chronological chunks of days and let
# Optuna KILL clearly-losing configs early (PercentilePruner). This typically
# saves 30-50% of the wall time, which buys MORE TRIALS in the same time -- i.e.
# a deeper search. Caveat: pruning judges on partial-sample profit, so it can
# very occasionally kill a config that only wins late; the 25th-percentile rule
# and warmup make that unlikely. Set False for exhaustive (slower) evaluation.
OPT_PRUNE = True
OPT_PRUNE_CHUNKS = 4        # chronological checkpoints per trial (>=2 to prune)

# Warm-start: enqueue the CURRENT config (module defaults / property values) and
# two curated variants as the first trials, so TPE starts from the best-known
# region instead of from random points.
OPT_ENQUEUE_SEEDS = True

# Also optimize the two risk gates that were previously FIXED. Both directly cap
# profit, so exposing them widens the reachable optimum. Set False to pin either
# to its module default.
OPT_TUNE_DAILY_LOSS = True   # tunes MAX_DAILY_LOSS_RUPEES (daily circuit breaker)
OPT_TUNE_STOP_CAP = True     # tunes MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT (per-attempt rupee cap)

# --- Where to SAVE every tested config + its results ---
# A CSV row is written (and flushed) after EVERY trial, so an interrupted run
# still keeps everything tested so far. A timestamped file is created per run.
OPT_OUTPUT_DIR = r"G:\My Drive\Trading\optimizer_runs"
OPT_STUDY_NAME = "atm_straddle_robust"
# Set True to ALSO persist the Optuna study to a SQLite DB in OPT_OUTPUT_DIR.
# That makes the study RESUMABLE: re-running appends more trials to the same
# study (TPE keeps learning) instead of starting over. False = in-memory only.
OPT_SAVE_DB = False

# --- Small-sample smoke test (set BOTH to None for a real run) ---
# For a quick end-to-end check, e.g. SAMPLE_MAX_PICKLES = 3 and SAMPLE_MAX_DAYS = 20.
# SAMPLE_MAX_PICKLES also shrinks the Kite underlying download, so the smoke test
# stays fast without touching anything else.
SAMPLE_MAX_PICKLES = None   # e.g. 3  -> read only the first N pickle files
SAMPLE_MAX_DAYS = None      # e.g. 20 -> simulate only the most recent N day-groups

# (Objective config lives with the optimizer further down: the score is TOTAL NET
#  PROFIT; see OPT_CV_PENALTY, OPT_MIN_DAYS, OPT_MIN_MONTHS.)

# =============================================================================
# USER CONFIG
# =============================================================================
# PICKLES_DIR = r"G:\My Drive\Trading\Dhan_Historical_Options_Data_New_0_1"
PICKLES_DIR = r"G:\My Drive\Trading\Historical_Options_Data"
ENTRY_TIME_IST = os.getenv("ENTRY_TIME_IST", "10:00")  # "HH:MM"

# EXIT_TIME_IST is a fresh-entry / re-entry cutoff, not a forced exit time.
# A trade opened before or exactly at this cutoff continues to be monitored and
# exits under the original logic: STOPLOSS, PROFIT_TARGET, PROFIT_PROTECT, or EOD.
# Optimizer mode also searches this value for maximum profit.
EXIT_TIME_IST = os.getenv("EXIT_TIME_IST", "15:30")  # "HH:MM"

def _safe_fname_part(s: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in s)

def _get_downloads_folder() -> str:
    """
    Returns the current user's default Downloads folder.
    Falls back to home directory if Downloads is not found.
    """
    downloads = Path.home() / "Downloads"
    return str(downloads if downloads.exists() else Path.home())

# --- Generic integer-list parser used for DTE and re-entry delay settings ---
def _parse_int_list(env_val, default_list):
    """
    Parse a comma-separated integer list from an environment variable.

    Examples:
        ALLOWED_DTE="0,1"              -> [0, 1]
        REENTRY_DELAY_BY_ATTEMPT="1,5" -> [1, 5]

    If parsing fails or the env var is blank, the supplied default is used.
    """
    if env_val:
        try:
            vals = [int(round(float(x))) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass
    return list(default_list)


def _parse_pct_value(x) -> float:
    """
    Parse a percentage value into decimal form.

    Accepted user formats:
        10      -> 0.10  (10%)
        "10%"   -> 0.10  (10%)
        0.10    -> 0.10  (10%)
        "0.10"  -> 0.10  (10%)

    This lets you configure percentages naturally from environment variables
    while keeping calculations internally consistent.
    """
    s = str(x).strip().replace("%", "")
    if s == "":
        raise ValueError("blank percentage value")

    v = float(s)

    # If user entered 10, treat it as 10%; if user entered 0.10, keep it as 10%.
    if abs(v) > 1.0:
        v = v / 100.0

    if v < 0:
        raise ValueError("percentage cannot be negative")

    return float(v)


def _parse_pct_list(env_val, default_list):
    """
    Parse comma-separated percentages into decimal form.

    Examples:
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15" -> [0.10, 0.12, 0.15]
        LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10"     -> [0.10]
    """
    if env_val:
        try:
            vals = [_parse_pct_value(x) for x in env_val.replace(" ", "").split(",") if x != ""]
            if vals:
                return vals
        except Exception:
            pass

    return [_parse_pct_value(x) for x in default_list]


def _fmt_int_list(lst) -> str:
    """Format integer-list settings for the output filename."""
    return "-".join(str(int(v)) for v in lst) if lst else "off"


def _fmt_pct_value(v: float) -> str:
    """Format a decimal percentage such as 0.10 as '10pct' for filenames/logs."""
    return f"{v * 100:.2f}".rstrip("0").rstrip(".") + "pct"


def _fmt_pct_list(lst) -> str:
    """Format a list of decimal percentages for the output filename."""
    return "-".join(_fmt_pct_value(float(v)) for v in lst) if lst else "off"


def _parse_float_env(env_name: str, default_value: float) -> float:
    """
    Parse a positive/zero floating-point rupee setting from an environment variable.

    If parsing fails, the supplied default is used. A value <= 0 is treated by
    the relevant logic as disabled where applicable.
    """
    raw = os.getenv(env_name)
    if raw is None or str(raw).strip() == "":
        return float(default_value)
    try:
        return float(str(raw).replace(",", "").strip())
    except Exception:
        return float(default_value)


def _fmt_rupee_value(v: float) -> str:
    """Format rupee values compactly for output filenames/logs."""
    return str(int(round(float(v))))


# =============================================================================
# RISK CONFIGURATION AS % OF ENTRY PREMIUM
# =============================================================================
# IMPORTANT:
# The original version used absolute rupee values in these two variables.
# This version intentionally keeps the same variable names for compatibility
# with your existing env-var workflow, but the meaning is now PERCENTAGE.
#
# Base for percentage calculation:
#     entry_premium_sum_rupees = (entry_ce + entry_pe) * qty
#
# Stop-loss threshold:
#     loss_limit_rupees = LOSS_LIMIT_% * entry_premium_sum_rupees
#
# Profit-protect threshold/giveback:
#     G = PROFIT_PROTECT_% * entry_premium_sum_rupees
#
# Example:
#     If CE+PE premium collected = 120 and qty = 325,
#     entry_premium_sum_rupees = 39,000.
#     10% stop-loss = 3,900.
#     30% profit-protect trigger/giveback = 11,700.
# =============================================================================

# --- Per-attempt STOP-LOSS as % of premium collected on that attempt ---
# Index 0 = first entry, 1 = first re-entry, etc.
# Attempts beyond the list reuse the LAST value.
#
# Default: 10% for every attempt.
#
# Env examples:
#     LOSS_LIMIT_RUPEES_BY_ATTEMPT="10"
#     LOSS_LIMIT_RUPEES_BY_ATTEMPT="10,12,15"
#     LOSS_LIMIT_RUPEES_BY_ATTEMPT="0.10,0.12,0.15"
LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_pct_list(
    os.getenv("LOSS_LIMIT_RUPEES_BY_ATTEMPT"),
    [0.3716, 0.3969, 0.4222, 0.4475, 0.4728, 0.4982, 0.5235, 0.5488],
)


def loss_limit_pct_for_attempt(attempt_idx: int) -> float:
    """Return the stop-loss percentage, in decimal form, for the given attempt."""
    s = LOSS_LIMIT_RUPEES_BY_ATTEMPT
    if not s:
        return 0.0
    return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])


# --- Allowed days-to-expiry to trade: [0,1]=expiry day + day before; [0]=expiry only ---
ALLOWED_DTE = _parse_int_list(os.getenv("ALLOWED_DTE"), [0])

# --- Profit-protect threshold/giveback as % of premium collected on that attempt ---
# Default: 30%.
#
# Env examples:
#     PROFIT_PROTECT_TRIGGER_RUPEES="30"
#     PROFIT_PROTECT_TRIGGER_RUPEES="0.30"
#
# Current logic uses the same rupee amount for:
#     1. arming profit-protect once peak P&L reaches G
#     2. exiting when current P&L falls to peak - G
PROFIT_PROTECT_TRIGGER_RUPEES = _parse_pct_value(os.getenv("PROFIT_PROTECT_TRIGGER_RUPEES", "0.3755"))

# --- Absolute daily circuit breaker -------------------------------------------------
# Once cumulative realized NET P&L for the current underlying/day reaches this
# loss, no further re-entry is allowed for that day.
#
# Default: Rs. 30,000 loss. Set MAX_DAILY_LOSS_RUPEES=0 to disable.
MAX_DAILY_LOSS_RUPEES = _parse_float_env("MAX_DAILY_LOSS_RUPEES", 30000.0)

# --- Absolute cap on the percentage-based per-attempt stop-loss ----------------------
# The stop-loss is still calculated as:
#     LOSS_LIMIT_% * entry_premium_sum
# But it is capped at this absolute rupee value.
#
# Effective stop-loss per attempt:
#     min(LOSS_LIMIT_% * entry_premium_sum, MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)
#
# Default: Rs. 3,000 loss per attempt. Set to 0 to disable the cap.
MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT = _parse_float_env("MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT", 3000.0)

MAX_REATTEMPTS = int(os.getenv("MAX_REATTEMPTS", "7"))  # 1 = only one re-entry

# --- Per-DAY profit target as a fraction of premium collected on the CURRENT attempt ---
# When an attempt's profit reaches PROFIT_TARGET_PCT * (CE+PE)*qty, it exits at the
# target and NO further trades are taken that day. 0 disables. e.g. 0.70 = 70%.
PROFIT_TARGET_PCT = float(os.getenv("PROFIT_TARGET_PCT", "0.8090"))
# --- Per-attempt RE-ENTRY GAP in minutes (index 0 = gap before 1st re-entry, 1 = before 2nd, ...) ---
# Attempts beyond the list reuse the LAST value. Override via env comma list, e.g.
# REENTRY_DELAY_BY_ATTEMPT="10,15,20".
REENTRY_DELAY_BY_ATTEMPT = _parse_int_list(
    os.getenv("REENTRY_DELAY_BY_ATTEMPT"),
    [8, 8, 8, 8, 8, 8, 8, 8],
)

def reentry_delay_for_attempt(attempt_idx: int) -> int:
    s = REENTRY_DELAY_BY_ATTEMPT
    if not s:
        return 0
    return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])

_DEFAULT_OUT = os.path.join(
    _get_downloads_folder(),
    f"short_straddle_backtest_reattempt{_safe_fname_part(ENTRY_TIME_IST)}"
    f"_EXIT{_safe_fname_part(EXIT_TIME_IST)}"
    f"_SLpct_{_safe_fname_part(_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    f"_DTE_{_safe_fname_part('-'.join(str(d) for d in ALLOWED_DTE))}"
    f"_PPTpct_{_safe_fname_part(_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES))}"
    f"_DailyMaxLoss_{_safe_fname_part(_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES))}"
    f"_StopCap_{_safe_fname_part(_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT))}"
    f"_MR_{_safe_fname_part(str(MAX_REATTEMPTS))}"
    f"_PT_{_safe_fname_part(str(int(round(PROFIT_TARGET_PCT * 100))))}pct"
    f"_RDM_{_safe_fname_part(_fmt_int_list(REENTRY_DELAY_BY_ATTEMPT))}.xlsx"
)

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", _DEFAULT_OUT)

FAIL_ON_PICKLE_ERROR = os.getenv("FAIL_ON_PICKLE_ERROR", "0").strip() == "1"

SESSION_START_IST = dtime(9, 15)
SESSION_END_IST = dtime(15, 30)

# LOOKBACK_MONTHS is AUTO by default: the script backtests the full date range
# present in the option pickles. If you explicitly set LOOKBACK_MONTHS to a
# number, that number acts as an optional manual cap.
LOOKBACK_MONTHS_RAW = os.getenv("LOOKBACK_MONTHS", "AUTO").strip()
LOOKBACK_MONTHS: Optional[int]
if LOOKBACK_MONTHS_RAW.upper() in ("", "AUTO", "ALL", "MAX", "FULL"):
    LOOKBACK_MONTHS = None
else:
    LOOKBACK_MONTHS = int(float(LOOKBACK_MONTHS_RAW))

QTY_UNITS = {"NIFTY": 325, "SENSEX": 100}
TRADEABLE = set(QTY_UNITS.keys())

STRIKE_STEP = {"NIFTY": 50, "SENSEX": 100}

# =============================================================================
# TRANSACTION CHARGES (Zerodha F&O Options — NSE)
# =============================================================================
# Each short-straddle attempt = 4 executed orders (sell CE, sell PE, buy CE, buy PE)
BROKERAGE_PER_ORDER       = 20.0       # ₹20 flat per executed order
ORDERS_PER_TRADE          = 4          # sell CE + sell PE + buy CE + buy PE
STT_SELL_PCT              = 0.001      # 0.1% on sell-side premium
EXCHANGE_TXN_PCT          = 0.0003553  # 0.03553% on premium (NSE options)
SEBI_PER_CRORE            = 10.0       # ₹10 per crore of turnover
STAMP_BUY_PCT             = 0.00003    # 0.003% on buy-side premium
IPFT_PER_CRORE            = 0.010      # ₹0.01 per crore (on premium)
GST_PCT                   = 0.18       # 18% on (brokerage + txn charges + SEBI)
INCLUDE_TRANSACTION_COSTS = True       # set False to disable

UNDERLYING_KITE = {
    "NIFTY": {"exchange": "NSE", "tradingsymbol": "NIFTY 50"},
    "SENSEX": {"exchange": "BSE", "tradingsymbol": "SENSEX"},
}

MAX_DAYS_PER_CHUNK = 25
MAX_ATTEMPTS = 5
SLEEP_BETWEEN_CALLS_SEC = 0.20


# =============================================================================
# HELPERS
# =============================================================================
def parse_hhmm(s: str) -> dtime:
    hh, mm = s.strip().split(":")
    return dtime(int(hh), int(mm))

ENTRY_TIME = parse_hhmm(ENTRY_TIME_IST)
EXIT_TIME = parse_hhmm(EXIT_TIME_IST)

def ist_tz():
    if ZoneInfo is not None:
        return ZoneInfo("Asia/Kolkata")
    if pytz is not None:
        return pytz.timezone("Asia/Kolkata")
    return "Asia/Kolkata"

def ensure_ist(series_or_scalar) -> Any:
    tz = ist_tz()
    dt = pd.to_datetime(series_or_scalar, errors="coerce")
    if isinstance(dt, pd.Series):
        if dt.dt.tz is None:
            return dt.dt.tz_localize(tz)
        return dt.dt.tz_convert(tz)
    if getattr(dt, "tzinfo", None) is None:
        return dt.tz_localize(tz)
    return dt.tz_convert(tz)

def normalize_underlying(name: str) -> Optional[str]:
    if not isinstance(name, str):
        return None
    u = name.upper().strip()
    if "SENSEX" in u:
        return "SENSEX"
    if "BANKNIFTY" in u or "NIFTY BANK" in u:
        return "BANKNIFTY"
    if "NIFTY" in u:
        return "NIFTY"
    return None

def round_to_step(x: float, step: int) -> int:
    return int(round(x / step) * step)

def build_minute_index(day_d: date, start_t: dtime, end_t: dtime) -> pd.DatetimeIndex:
    tz = ist_tz()
    start = pd.Timestamp(datetime.combine(day_d, start_t), tz=tz)
    end = pd.Timestamp(datetime.combine(day_d, end_t), tz=tz)
    return pd.date_range(start=start, end=end, freq="1min")

def asof_close(df: pd.DataFrame, ts: pd.Timestamp) -> float:
    if df.empty:
        return float("nan")
    d = df[["date", "close"]].dropna().copy()
    d["date"] = ensure_ist(d["date"])
    d = d.sort_values("date").set_index("date")
    loc = d.index.get_indexer([ts], method="pad")
    if loc[0] == -1:
        return float("nan")
    return float(d.iloc[loc[0]]["close"])

def compute_window_start(end_day: date, months: int) -> date:
    if relativedelta is not None:
        return (pd.Timestamp(end_day) - relativedelta(months=months)).date()
    return (pd.Timestamp(end_day) - pd.Timedelta(days=30 * months)).date()


def determine_backtest_window_start(min_day_seen: date, end_day: date) -> date:
    """
    Determine the backtest start date.

    Default behaviour: use the earliest usable option-data date found in the
    pickles, i.e. the maximum available backtest period.

    Optional override: if LOOKBACK_MONTHS is set to a numeric value through the
    environment, use the later of:
        1. earliest option-data date; and
        2. end_day - LOOKBACK_MONTHS
    so the script never requests data before the pickles actually start.
    """
    if LOOKBACK_MONTHS is None:
        return min_day_seen

    capped_start = compute_window_start(end_day, LOOKBACK_MONTHS)
    return max(min_day_seen, capped_start)

# =============================================================================
# TRANSACTION COST CALCULATOR
# =============================================================================
def compute_trade_charges(
    entry_ce: float, entry_pe: float,
    exit_ce: float, exit_pe: float,
    qty: int,
) -> float:
    """
    Compute total Zerodha transaction charges for one short-straddle attempt.

    Entry = SELL CE + SELL PE  (2 orders, sell side)
    Exit  = BUY  CE + BUY  PE (2 orders, buy side)

    Returns total charges in rupees (always positive).
    """
    if not INCLUDE_TRANSACTION_COSTS:
        return 0.0

    # Turnover values (in rupees)
    entry_turnover = (entry_ce + entry_pe) * qty   # sell side
    exit_turnover  = (exit_ce + exit_pe) * qty     # buy side
    total_turnover = entry_turnover + exit_turnover

    # 1. Brokerage: ₹20 × 4 orders
    brokerage = BROKERAGE_PER_ORDER * ORDERS_PER_TRADE

    # 2. STT: 0.1% on sell-side premium only (entry for short straddle)
    stt = entry_turnover * STT_SELL_PCT

    # 3. Exchange transaction charges: 0.03553% on both sides
    txn_charges = total_turnover * EXCHANGE_TXN_PCT

    # 4. SEBI charges: ₹10 per crore on total turnover
    sebi = total_turnover * SEBI_PER_CRORE / 1_00_00_000

    # 5. Stamp duty: 0.003% on buy side only (exit for short straddle)
    stamp = exit_turnover * STAMP_BUY_PCT

    # 6. IPFT: ₹0.01 per crore on premium (both sides)
    ipft = total_turnover * IPFT_PER_CRORE / 1_00_00_000

    # 7. GST: 18% on (brokerage + transaction charges + SEBI charges)
    gst = (brokerage + txn_charges + sebi) * GST_PCT

    total_charges = brokerage + stt + txn_charges + sebi + stamp + ipft + gst
    return round(total_charges, 2)

# =============================================================================
# Kite historical helpers
# =============================================================================
def _iter_chunks_by_date(from_dt: datetime, to_dt: datetime, days_per_chunk: int) -> List[Tuple[datetime, datetime]]:
    if from_dt > to_dt:
        raise ValueError("from_dt must be <= to_dt")
    chunks: List[Tuple[datetime, datetime]] = []
    cur = from_dt.date()
    end_d = to_dt.date()
    while cur <= end_d:
        chunk_end = min(cur + timedelta(days=days_per_chunk - 1), end_d)
        c_from = from_dt if cur == from_dt.date() else datetime.combine(cur, SESSION_START_IST)
        c_to = to_dt if chunk_end == end_d else datetime.combine(chunk_end, SESSION_END_IST)
        chunks.append((c_from, c_to))
        cur = chunk_end + timedelta(days=1)
    return chunks

def _kite_instruments_cached(kite, exchange: str, cache: Dict[str, List[Dict]]) -> List[Dict]:
    ex = exchange.upper().strip()
    if ex not in cache:
        print(f"[STEP] Loading instruments dump for {ex} ...")
        cache[ex] = kite.instruments(ex)
        print(f"[INFO] {ex} instruments: {len(cache[ex])}")
    return cache[ex]

def get_instrument_token(kite, exchange: str, tradingsymbol: str, cache: Dict[str, List[Dict]]) -> int:
    ex = exchange.upper().strip()
    wanted = tradingsymbol.strip().upper()
    for r in _kite_instruments_cached(kite, ex, cache):
        if str(r.get("tradingsymbol", "")).upper() == wanted:
            return int(r["instrument_token"])
    raise ValueError(f"Instrument not found on {ex}: '{tradingsymbol}'")

def fetch_history_minute(kite, instrument_token: int, from_dt: datetime, to_dt: datetime, label: str) -> List[Dict]:
    interval = "minute"
    chunks = _iter_chunks_by_date(from_dt, to_dt, MAX_DAYS_PER_CHUNK)
    rows_all: List[Dict] = []
    print(f"[INFO] Fetch {label} token={instrument_token} chunks={len(chunks)} {from_dt} -> {to_dt}")
    for i, (c_from, c_to) in enumerate(chunks, start=1):
        last_err = None
        for attempt in range(1, MAX_ATTEMPTS + 1):
            try:
                rows = kite.historical_data(
                    instrument_token=instrument_token,
                    from_date=c_from,
                    to_date=c_to,
                    interval=interval,
                    continuous=False,
                    oi=False,
                )
                rows_all.extend(rows)
                last_err = None
                break
            except Exception as e:
                last_err = e
                time.sleep(min(8.0, 1.5 * attempt))
        if last_err is not None:
            print(f"[ERROR] {label} chunk {i}/{len(chunks)} failed: {c_from}->{c_to}: {last_err}")
        time.sleep(SLEEP_BETWEEN_CALLS_SEC)
    return rows_all

def rows_to_df(rows: List[Dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["date", "open", "high", "low", "close", "volume"])
    df = pd.DataFrame(rows)
    df["date"] = ensure_ist(df["date"])
    df = df.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    return df


# =============================================================================
# DATA STRUCTURES
# =============================================================================
@dataclass
class TradeRow:
    day: date
    underlying: str
    trade_seq: int
    expiry: date
    days_to_expiry: int
    atm_strike: int
    qty_units: int
    entry_time: str
    exit_time: str
    exit_reason: str
    entry_underlying: float
    ce_symbol: str
    pe_symbol: str
    entry_ce: float
    entry_pe: float
    exit_ce: float
    exit_pe: float
    exit_pnl_gross: float   # P&L before charges
    txn_charges: float      # total transaction charges for this attempt
    exit_pnl: float         # net P&L after deducting charges
    eod_pnl: float
    max_profit: float
    max_loss: float
    max_profit_before_exit: float   # peak profit reached before this trade exited

    # Premium/risk diagnostics for percentage-based risk rules
    entry_premium_sum: float                 # (entry_ce + entry_pe) * qty
    stop_pct: float                          # stop-loss % of entry premium, decimal form; 0.10 = 10%
    uncapped_stop_rupees: float              # percentage-based stop before absolute cap
    stop_cap_rupees: float                   # configured absolute cap; <=0 means cap disabled
    stop_rupees: float                       # effective rupee stop after cap
    profit_protect_trigger_pct: float         # profit-protect % of entry premium, decimal form; 0.30 = 30%
    profit_protect_trigger_rupees: float      # computed rupee profit-protect trigger/giveback
    daily_realized_pnl_after_trade: float     # cumulative net P&L after this attempt
    daily_loss_limit_rupees: float            # configured daily loss circuit breaker
    daily_loss_limit_hit: bool                # True means no further trades for that day


# =============================================================================
# PASS-1: nearest expiry per (underlying, day)
# =============================================================================
def scan_pickles_pass1(pickle_paths: List[str]) -> Tuple[date, Dict[Tuple[str, date], date], date]:
    max_day_seen: Optional[date] = None
    min_day_seen: Optional[date] = None
    min_expiry_map: Dict[Tuple[str, date], date] = {}

    for p in pickle_paths:
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            for c in ("date", "name", "expiry", "type"):
                if c not in df.columns:
                    raise ValueError(f"Missing column '{c}' in {p}")

            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")]
            if d2.empty:
                continue

            d2 = d2[["date", "name", "expiry"]].copy()
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2 = d2.dropna(subset=["underlying", "day", "expiry_date"])

            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            d2 = d2[d2["expiry_date"] >= d2["day"]]
            if d2.empty:
                continue

            file_min_day = d2["day"].min()
            file_max_day = d2["day"].max()
            max_day_seen = file_max_day if (max_day_seen is None or file_max_day > max_day_seen) else max_day_seen
            min_day_seen = file_min_day if (min_day_seen is None or file_min_day < min_day_seen) else min_day_seen

            grp = d2.groupby(["underlying", "day"], sort=False)["expiry_date"].min()
            for (und, dy), ex in grp.items():
                key = (und, dy)
                if key not in min_expiry_map or ex < min_expiry_map[key]:
                    min_expiry_map[key] = ex

            print(f"[PASS1 OK] {os.path.basename(p)} option_days={d2['day'].nunique()}")

        except Exception as e:
            msg = f"[PASS1 WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    if max_day_seen is None or min_day_seen is None:
        raise RuntimeError("No usable option data found in pickles (PASS1) for tradeable underlyings.")

    return max_day_seen, min_expiry_map, min_day_seen


# =============================================================================
# Underlying download
# =============================================================================
def download_underlyings(kite, day_start: date, day_end: date) -> Dict[str, pd.DataFrame]:
    cache: Dict[str, List[Dict]] = {}
    from_dt = datetime.combine(day_start, SESSION_START_IST)
    to_dt = datetime.combine(day_end, SESSION_END_IST)

    out: Dict[str, pd.DataFrame] = {}
    for und, meta in UNDERLYING_KITE.items():
        token = get_instrument_token(kite, meta["exchange"], meta["tradingsymbol"], cache)
        rows = fetch_history_minute(kite, token, from_dt, to_dt, label=f"{meta['exchange']}:{meta['tradingsymbol']}")
        df = rows_to_df(rows)
        df["day"] = df["date"].dt.tz_convert(ist_tz()).dt.date
        out[und] = df
        print(f"[UNDERLYING OK] {und}: candles={len(df)} days={df['day'].nunique()}")
    return out


# =============================================================================
# Simulation helpers
# =============================================================================
def _pick_symbol(day_opt: pd.DataFrame, strike: int, opt_type: str) -> Optional[str]:
    sub = day_opt[(day_opt["strike_int"] == strike) & (day_opt["option_type"] == opt_type)]
    if sub.empty:
        return None
    syms = sorted(sub["instrument"].astype(str).unique().tolist())
    return syms[0] if syms else None

def _build_leg_series(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex,
                      strike: int, opt_type: str, symbol: str,
                      price_col: str = "close", do_ffill: bool = True) -> pd.Series:
    sub = day_opt[
        (day_opt["strike_int"] == strike) &
        (day_opt["option_type"] == opt_type) &
        (day_opt["instrument"].astype(str) == symbol)
    ][["date", price_col]].dropna()

    if sub.empty:
        return pd.Series(index=idx_all, dtype="float64")

    sub = sub.copy()
    sub["date"] = ensure_ist(sub["date"])
    sub = sub.sort_values("date").drop_duplicates(subset=["date"], keep="last").set_index("date")
    s = sub[price_col].astype(float).reindex(idx_all)
    return s.ffill() if do_ffill else s


def build_price_book(day_opt: pd.DataFrame, idx_all: pd.DatetimeIndex):
    """
    Precompute, ONCE per trading day, the minute-indexed (close/high/low) series
    for every (strike, option_type) present -- exactly as _build_leg_series would
    produce them with do_ffill=False. The optimizer rebuilds the same parameter-
    independent series on every trial; caching them here turns ~300x of pandas
    reindex+tz work into a single dict lookup, which is the whole speed-up.

    Returns:
        book    : {(strike_int, opt_type, price_col): pd.Series on idx_all (raw)}
        symbols : {(strike_int, opt_type): chosen instrument symbol}
    Numerically identical to _build_leg_series (verified by regression test).
    """
    book: Dict[Tuple[int, str, str], pd.Series] = {}
    symbols: Dict[Tuple[int, str], str] = {}
    if day_opt is None or day_opt.empty:
        return book, symbols

    for (strike, opt), sub in day_opt.groupby(["strike_int", "option_type"], sort=False):
        # Same instrument-selection rule as _pick_symbol (first sorted symbol).
        syms = sorted(sub["instrument"].astype(str).unique().tolist())
        if not syms:
            continue
        sym = syms[0]
        symbols[(int(strike), str(opt))] = sym

        sub_sym = sub[sub["instrument"].astype(str) == sym][["date", "close", "high", "low"]].copy()
        sub_sym["date"] = ensure_ist(sub_sym["date"])
        sub_sym = (sub_sym.sort_values("date")
                          .drop_duplicates(subset=["date"], keep="last")
                          .set_index("date"))
        for col in ("close", "high", "low"):
            book[(int(strike), str(opt), col)] = sub_sym[col].astype(float).reindex(idx_all)
    return book, symbols


def _leg_from_book(book, idx_all, strike: int, opt_type: str, price_col: str) -> pd.Series:
    """Cached-series accessor mirroring _build_leg_series(..., do_ffill=False)."""
    s = book.get((int(strike), str(opt_type), price_col))
    if s is None:
        return pd.Series(index=idx_all, dtype="float64")
    return s

# =============================================================================
# PARAMS (everything the optimizer can vary, threaded explicitly into the sim
# instead of being read from module globals -- so we can run many backtests in
# one process without re-importing). The module globals above are still used to
# build the DEFAULT params for the normal single-run main() workflow.
# =============================================================================
@dataclass
class Params:
    entry_time: dtime                       # ENTRY_TIME_IST, as a datetime.time
    exit_time: dtime                        # EXIT_TIME_IST, last fresh entry/re-entry cutoff
    loss_limit_pct_by_attempt: List[float]  # LOSS_LIMIT_RUPEES_BY_ATTEMPT (decimals)
    profit_protect_pct: float               # PROFIT_PROTECT_TRIGGER_RUPEES (decimal)
    max_reattempts: int                     # MAX_REATTEMPTS
    profit_target_pct: float                # PROFIT_TARGET_PCT (decimal)
    reentry_delay_by_attempt: List[int]     # REENTRY_DELAY_BY_ATTEMPT (minutes)
    # not optimized here, but consumed by the sim -- kept on Params so the sim
    # never reaches back into globals:
    max_daily_loss_rupees: float
    max_loss_limit_cap_rupees: float

    def loss_limit_pct_for_attempt(self, attempt_idx: int) -> float:
        s = self.loss_limit_pct_by_attempt
        if not s:
            return 0.0
        return float(s[attempt_idx]) if attempt_idx < len(s) else float(s[-1])

    def reentry_delay_for_attempt(self, attempt_idx: int) -> int:
        s = self.reentry_delay_by_attempt
        if not s:
            return 0
        return int(s[attempt_idx]) if attempt_idx < len(s) else int(s[-1])


def default_params() -> "Params":
    """Build Params from the module-level globals (preserves env-var behaviour)."""
    return Params(
        entry_time=ENTRY_TIME,
        exit_time=EXIT_TIME,
        loss_limit_pct_by_attempt=list(LOSS_LIMIT_RUPEES_BY_ATTEMPT),
        profit_protect_pct=float(PROFIT_PROTECT_TRIGGER_RUPEES),
        max_reattempts=int(MAX_REATTEMPTS),
        profit_target_pct=float(PROFIT_TARGET_PCT),
        reentry_delay_by_attempt=list(REENTRY_DELAY_BY_ATTEMPT),
        max_daily_loss_rupees=float(MAX_DAILY_LOSS_RUPEES),
        max_loss_limit_cap_rupees=float(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT),
    )


def simulate_day_multi_trades(
    *,
    und: str,
    dy: date,
    expiry: date,
    day_opt: pd.DataFrame,
    underlying_day: pd.DataFrame,
    params: "Params",
    price_book: Optional[dict] = None,   # precomputed per-day series (perf cache)
    symbols: Optional[dict] = None,      # precomputed (strike,opt)->symbol
    idx_all: Optional[pd.DatetimeIndex] = None,
) -> Tuple[List[TradeRow], List[Dict[str, Any]]]:

    results: List[TradeRow] = []
    skipped: List[Dict[str, Any]] = []

    # Use the precomputed per-day price book if the caller supplied one (the
    # optimizer does, so the heavy series construction happens once). If not
    # (e.g. main()'s single run, or a direct call), build it on the fly here so
    # behaviour is identical either way.
    if idx_all is None:
        idx_all = build_minute_index(dy, SESSION_START_IST, SESSION_END_IST)
    if price_book is None or symbols is None:
        price_book, symbols = build_price_book(day_opt, idx_all)
    session_end_ts = idx_all[-1]

    # EXIT_TIME_IST is only a cutoff for NEW entries/re-entries.
    # It does NOT truncate the monitoring series and does NOT force an exit.
    # Existing open trades can still exit later by the original EOD/SL/target logic.
    entry_cutoff_ts = pd.Timestamp(datetime.combine(dy, params.exit_time), tz=ist_tz())
    if entry_cutoff_ts > session_end_ts:
        entry_cutoff_ts = session_end_ts

    qty = int(QTY_UNITS[und])
    step = int(STRIKE_STEP[und])

    # Profit-protect is now percentage-based, so the actual rupee value is not
    # known until CE/PE entry prices are available for the current attempt.
    profit_protect_pct = float(params.profit_protect_pct)
    profit_protect_enabled = profit_protect_pct > 0.0

    cur_entry_ts = pd.Timestamp(datetime.combine(dy, params.entry_time), tz=ist_tz())
    trade_seq = 1

    # Cumulative realized NET P&L for this underlying/day. Used for the daily
    # loss circuit breaker. Charges are included through exit_pnl.
    daily_realized_pnl = 0.0
    daily_loss_limit_enabled = params.max_daily_loss_rupees > 0

    while cur_entry_ts <= entry_cutoff_ts:
        if daily_loss_limit_enabled and daily_realized_pnl <= -float(params.max_daily_loss_rupees):
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq,
                "reason": (
                    f"Daily loss limit hit before next entry: "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={params.max_daily_loss_rupees:.2f}"
                ),
            })
            break

        u_px = asof_close(underlying_day, cur_entry_ts)
        if pd.isna(u_px):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": f"No underlying price at entry {cur_entry_ts.strftime('%H:%M')}"})
            break

        atm = round_to_step(float(u_px), step)

        ce_sym = symbols.get((atm, "CE"))
        pe_sym = symbols.get((atm, "PE"))
        if not ce_sym or not pe_sym:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "ATM CE/PE not available in pickle band"})
            break

        # Close series (used for entry pricing, profit-protect tracking, and reporting)
        # Raw close series for exact entry validation -- from the precomputed book.
        ce_close_raw = _leg_from_book(price_book, idx_all, atm, "CE", "close")
        pe_close_raw = _leg_from_book(price_book, idx_all, atm, "PE", "close")

        # Forward-filled close series for post-entry tracking/reporting
        ce_close = ce_close_raw.ffill()
        pe_close = pe_close_raw.ffill()

        # High/Low series (used only to detect STOPLOSS intraminute extremes)
        ce_high = _leg_from_book(price_book, idx_all, atm, "CE", "high")
        ce_low = _leg_from_book(price_book, idx_all, atm, "CE", "low")
        pe_high = _leg_from_book(price_book, idx_all, atm, "PE", "high")
        pe_low = _leg_from_book(price_book, idx_all, atm, "PE", "low")

        if cur_entry_ts not in idx_all:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "reason": "Entry timestamp not in session index"})
            break

        ce_entry = ce_close_raw.loc[cur_entry_ts]
        pe_entry = pe_close_raw.loc[cur_entry_ts]
        monitor_start_ts = pd.Timestamp(cur_entry_ts) + pd.Timedelta(minutes=1)
        if monitor_start_ts > session_end_ts:
            break

        if pd.isna(ce_entry) or pd.isna(pe_entry):
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "No CE/PE price at entry (after ffill)"})
            break

        # ---------------------------------------------------------------------
        # Percentage-based risk basis for THIS attempt
        # ---------------------------------------------------------------------
        # For every entry/re-entry, compute the premium collected in rupees.
        # Stop-loss and profit-protect thresholds are derived from this value.
        #
        # Example:
        #   entry_ce=70, entry_pe=50, qty=325
        #   entry_premium_sum = (70 + 50) * 325 = 39,000
        #   10% stop-loss = 3,900
        #   30% profit-protect threshold/giveback = 11,700
        # ---------------------------------------------------------------------
        entry_premium_sum = (float(ce_entry) + float(pe_entry)) * qty

        loss_limit_pct = params.loss_limit_pct_for_attempt(trade_seq - 1)
        uncapped_loss_limit_rupees = float(loss_limit_pct * entry_premium_sum)

        # Absolute cap on the percentage-based stop-loss.
        # Example: 10% of premium may be Rs. 4,500, but with a Rs. 3,000 cap
        # the effective stop used by the simulator is Rs. 3,000.
        stop_cap_rupees = float(params.max_loss_limit_cap_rupees)
        if stop_cap_rupees > 0:
            loss_limit_rupees = float(min(uncapped_loss_limit_rupees, stop_cap_rupees))
        else:
            loss_limit_rupees = float(uncapped_loss_limit_rupees)

        # G is the same variable used by the existing profit-protect logic:
        #   - profit-protect arms when peak P&L >= G
        #   - profit-protect exits when current P&L <= peak - G
        G = float(profit_protect_pct * entry_premium_sum)

        # Close-based PnL (same as before)
        pnl_close_all = (float(ce_entry) - ce_close) * qty + (float(pe_entry) - pe_close) * qty
        pnl = pnl_close_all.loc[monitor_start_ts:].dropna()  # keep 'pnl' as close-based for profit-protect

        # STOPLOSS worst-case PnL candidates within each minute:
        #  A) CE high, PE low
        pnl_ceHigh_peLow_all = (float(ce_entry) - ce_high) * qty + (float(pe_entry) - pe_low) * qty
        #  B) CE low, PE high
        pnl_ceLow_peHigh_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_high) * qty

        # Worst-case PnL per minute among (close, A, B)
        pnl_sl_all = pd.concat([pnl_close_all, pnl_ceHigh_peLow_all, pnl_ceLow_peHigh_all], axis=1).min(axis=1)
        pnl_sl = pnl_sl_all.loc[monitor_start_ts:].dropna()

        if pnl.empty:
            skipped.append({"day": dy, "underlying": und, "expiry": expiry, "trade_seq": trade_seq,
                            "atm_strike": atm, "reason": "PnL series empty after entry"})
            break

        eod_ts = pnl.index[-1]
        eod_pnl = float(pnl.iloc[-1])

        max_profit = float(max(0.0, pnl.max()))
        max_loss = float(min(0.0, pnl.min()))

        # STOPLOSS uses the attempt-specific rupee value after applying the
        # absolute per-attempt cap.
        stop_hit = pnl_sl <= -loss_limit_rupees
        stop_ts = pnl_sl.index[stop_hit.to_numpy().argmax()] if stop_hit.any() else None

        protect_ts = None
        if profit_protect_enabled:
            peak = pnl.cummax()
            armed = peak >= G
            trail = peak - G
            protect_hit = armed & (pnl <= trail)
            protect_ts = pnl.index[protect_hit.to_numpy().argmax()] if protect_hit.any() else None

        # --- Per-day PROFIT TARGET: % of premium collected on this attempt ---
        # When reached, this trade exits at the target AND no further trades are
        # taken for the day (PROFIT_TARGET is excluded from the re-entry rule below).
        target_ts = None
        target_rupees = None
        if params.profit_target_pct > 0.0:
            target_rupees = params.profit_target_pct * entry_premium_sum
            # best-case (favourable) intrabar profit: both legs bought back at their lows
            pnl_best_all = (float(ce_entry) - ce_low) * qty + (float(pe_entry) - pe_low) * qty
            pnl_tp = pd.concat([pnl_close_all, pnl_best_all], axis=1).max(axis=1)
            pnl_tp = pnl_tp.loc[monitor_start_ts:].dropna()
            tp_hit = pnl_tp >= float(target_rupees)
            target_ts = pnl_tp.index[tp_hit.to_numpy().argmax()] if tp_hit.any() else None

        # Earliest triggered exit wins; on identical timestamps prefer the more
        # conservative outcome: STOPLOSS, then PROFIT_TARGET, then PROFIT_PROTECT.
        exit_ts = eod_ts
        exit_reason = "EOD"
        _candidates = []
        if stop_ts is not None:
            _candidates.append((stop_ts, 0, "STOPLOSS"))
        if target_ts is not None:
            _candidates.append((target_ts, 1, "PROFIT_TARGET"))
        if protect_ts is not None:
            _candidates.append((protect_ts, 2, "PROFIT_PROTECT"))
        if _candidates:
            _candidates.sort(key=lambda c: (c[0], c[1]))
            exit_ts, _, exit_reason = _candidates[0]

        if exit_reason == "STOPLOSS":
            exit_pnl_gross = -float(loss_limit_rupees)
        elif exit_reason == "PROFIT_TARGET":
            exit_pnl_gross = float(target_rupees)
        else:
            exit_pnl_gross = float(pnl.loc[exit_ts])

        # Peak (close-based) profit reached during this trade's life, up to its exit
        pnl_pre_exit = pnl.loc[:exit_ts]
        max_profit_before_exit = float(max(0.0, pnl_pre_exit.max())) if len(pnl_pre_exit) else 0.0

        exit_ce = float(ce_close.loc[exit_ts]) if pd.notna(ce_close.loc[exit_ts]) else float("nan")
        exit_pe = float(pe_close.loc[exit_ts]) if pd.notna(pe_close.loc[exit_ts]) else float("nan")

        txn_charges = compute_trade_charges(
            entry_ce=float(ce_entry), entry_pe=float(pe_entry),
            exit_ce=exit_ce if not pd.isna(exit_ce) else 0.0,
            exit_pe=exit_pe if not pd.isna(exit_pe) else 0.0,
            qty=qty,
        )
        exit_pnl = exit_pnl_gross - txn_charges

        # Update cumulative realized NET P&L for the day. This is checked
        # before allowing any further re-entry.
        daily_realized_pnl += float(exit_pnl)
        daily_loss_limit_hit = bool(
            daily_loss_limit_enabled and daily_realized_pnl <= -float(params.max_daily_loss_rupees)
        )

        dte = int((expiry - dy).days)

        results.append(
            TradeRow(
                day=dy,
                underlying=und,
                trade_seq=trade_seq,
                expiry=expiry,
                days_to_expiry=dte,
                atm_strike=int(atm),
                qty_units=qty,
                entry_time=pd.Timestamp(cur_entry_ts).strftime("%H:%M"),
                exit_time=pd.Timestamp(exit_ts).strftime("%H:%M"),
                exit_reason=exit_reason,
                entry_underlying=float(u_px),
                ce_symbol=ce_sym,
                pe_symbol=pe_sym,
                entry_ce=float(ce_entry),
                entry_pe=float(pe_entry),
                exit_ce=exit_ce,
                exit_pe=exit_pe,
                exit_pnl_gross=exit_pnl_gross,
                txn_charges=txn_charges,
                exit_pnl=exit_pnl,
                eod_pnl=eod_pnl,
                max_profit=max_profit,
                max_loss=max_loss,
                max_profit_before_exit=max_profit_before_exit,
                entry_premium_sum=float(entry_premium_sum),
                stop_pct=float(loss_limit_pct),
                uncapped_stop_rupees=float(uncapped_loss_limit_rupees),
                stop_cap_rupees=float(stop_cap_rupees),
                stop_rupees=float(loss_limit_rupees),
                profit_protect_trigger_pct=float(profit_protect_pct),
                profit_protect_trigger_rupees=float(G),
                daily_realized_pnl_after_trade=float(daily_realized_pnl),
                daily_loss_limit_rupees=float(params.max_daily_loss_rupees),
                daily_loss_limit_hit=bool(daily_loss_limit_hit),
            )
        )

        if daily_loss_limit_hit:
            skipped.append({
                "day": dy,
                "underlying": und,
                "expiry": expiry,
                "trade_seq": trade_seq + 1,
                "reason": (
                    f"No re-entry: daily loss limit hit after trade_seq={trade_seq}; "
                    f"realized_pnl={daily_realized_pnl:.2f}, "
                    f"limit={params.max_daily_loss_rupees:.2f}"
                ),
            })
            break

        if exit_reason in ("STOPLOSS", "PROFIT_PROTECT") and (trade_seq - 1) < params.max_reattempts:
            delay_min = params.reentry_delay_for_attempt(trade_seq - 1)  # gap before this re-entry
            trade_seq += 1
            cur_entry_ts = pd.Timestamp(exit_ts) + pd.Timedelta(minutes=delay_min)

            # No fresh re-entry is allowed after EXIT_TIME_IST.
            # The just-closed trade is unaffected; only the next trade is blocked.
            if cur_entry_ts > entry_cutoff_ts:
                break
            continue

        break

    return results, skipped


# =============================================================================
# PASS-2: process each pickle and simulate trades for days where this expiry is nearest
# =============================================================================
@dataclass
class DayGroup:
    """One simulatable (underlying, day, expiry) unit with its option + underlying
    slices already extracted. Built ONCE; reused across every optimizer trial.
    The price_book/symbols/idx_all are filled lazily on first simulation and then
    reused, so the expensive per-day series construction happens only once."""
    und: str
    dy: date
    expiry: date
    day_opt: pd.DataFrame
    underlying_day: pd.DataFrame
    price_book: Optional[dict] = None
    symbols: Optional[dict] = None
    idx_all: Optional[pd.DatetimeIndex] = None


def build_day_groups(
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
    max_pickles: Optional[int] = None,   # small-sample: only read this many files
    max_days: Optional[int] = None,      # small-sample: keep only the most recent N day-groups
) -> Tuple[List[DayGroup], List[Dict[str, Any]]]:
    """
    Parse pickles and slice them into per-day groups. This is the EXPENSIVE part
    (disk I/O + parsing) and is parameter-independent, so the optimizer runs it
    ONCE and re-simulates the returned groups for every trial.

    Progress: prints a counter after each pickle so a long load is visible.
    Small-sample: `max_pickles` limits how many files are read; `max_days` keeps
    only the most recent N day-groups after loading. Both are for quick smoke
    tests and should be left as None for a real run.
    """
    groups: List[DayGroup] = []
    skipped_rows: List[Dict[str, Any]] = []
    processed_day_keys: set = set()  # prevent double-count of the same (und,day,expiry) across files

    # Small-sample shortcut: only look at the first `max_pickles` files.
    if max_pickles is not None and max_pickles > 0:
        pickle_paths = pickle_paths[:max_pickles]

    total_files = len(pickle_paths)
    for fi, p in enumerate(pickle_paths, start=1):
        try:
            df = pd.read_pickle(p)
            if not isinstance(df, pd.DataFrame) or df.empty:
                print(f"[LOAD {fi}/{total_files}] {os.path.basename(p)}: empty (skipped)")
                continue

            needed_cols = ["date", "name", "type", "option_type", "strike", "expiry", "instrument", "high", "low", "close"]
            missing = [c for c in needed_cols if c not in df.columns]
            if missing:
                raise ValueError(f"Missing columns {missing} in {p}")

            # Keep only OPTION rows and the columns we actually use.
            d2 = df[df["type"].astype(str).str.upper().eq("OPTION")][needed_cols].copy()
            if d2.empty:
                continue

            # Normalize types/keys used downstream.
            d2["date"] = ensure_ist(d2["date"])
            d2["day"] = d2["date"].dt.date
            d2["underlying"] = d2["name"].astype(str).map(normalize_underlying)
            d2 = d2[d2["underlying"].isin(TRADEABLE)]
            if d2.empty:
                continue

            d2["expiry_date"] = pd.to_datetime(d2["expiry"], errors="coerce").dt.date
            d2["strike_num"] = pd.to_numeric(d2["strike"], errors="coerce")
            d2["strike_int"] = d2["strike_num"].round().astype("Int64")
            d2["option_type"] = d2["option_type"].astype(str).str.upper()

            d2 = d2.dropna(subset=["day", "underlying", "expiry_date", "strike_int", "close"])
            d2["strike_int"] = d2["strike_int"].astype(int)
            d2 = d2[d2["expiry_date"] >= d2["day"]]          # drop stale (already-expired) rows
            if d2.empty:
                continue
            d2 = d2[(d2["day"] >= window_start) & (d2["day"] <= window_end)]  # date window
            if d2.empty:
                continue

            # One simulatable unit per (underlying, day, expiry), but only when this
            # expiry is the NEAREST expiry for that (underlying, day) -- i.e. the
            # contract we'd actually trade that day.
            for (und, dy, ex), g in d2.groupby(["underlying", "day", "expiry_date"], sort=False):
                if min_expiry_map.get((und, dy)) != ex:
                    continue
                day_key = (und, dy, ex)
                if day_key in processed_day_keys:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex,
                                         "reason": "Duplicate (underlying,day,expiry) across pickles; skipped to avoid double-count"})
                    continue
                processed_day_keys.add(day_key)

                # Attach the underlying minute-bars for this day so the simulator
                # doesn't have to re-filter the full series on every trial.
                uday = underlying_data.get(und)
                if uday is None:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "No underlying series downloaded"})
                    continue
                uday = uday[uday["day"] == dy]
                if uday.empty:
                    skipped_rows.append({"day": dy, "underlying": und, "expiry": ex, "reason": "Underlying missing for day"})
                    continue

                groups.append(DayGroup(und=und, dy=dy, expiry=ex,
                                       day_opt=g.copy(), underlying_day=uday.copy()))

            print(f"[LOAD {fi}/{total_files}] {os.path.basename(p)} grouped "
                  f"(day-groups so far: {len(groups)})", flush=True)

        except Exception as e:
            msg = f"[LOAD {fi}/{total_files} WARN] {os.path.basename(p)} failed: {e}"
            if FAIL_ON_PICKLE_ERROR:
                raise RuntimeError(msg) from e
            print(msg)

    # Deterministic order so re-simulation is reproducible across trials.
    groups.sort(key=lambda gr: (gr.dy, gr.und))

    # Small-sample shortcut: keep only the most recent N day-groups.
    if max_days is not None and max_days > 0 and len(groups) > max_days:
        groups = groups[-max_days:]
        print(f"[LOAD] small-sample: keeping most recent {len(groups)} day-groups")

    # Pre-build each day's price book NOW (one-time, parameter-independent), so
    # every optimizer trial is uniformly fast and the progress ETA is honest from
    # the very first trial instead of trial 1 being anomalously slow.
    if groups:
        print(f"[LOAD] precomputing per-day price books for {len(groups)} day-groups ...", flush=True)
        for gi, gr in enumerate(groups, start=1):
            gr.idx_all = build_minute_index(gr.dy, SESSION_START_IST, SESSION_END_IST)
            gr.price_book, gr.symbols = build_price_book(gr.day_opt, gr.idx_all)
            if gi % 50 == 0 or gi == len(groups):
                print(f"[LOAD] price book {gi}/{len(groups)}", flush=True)

    print(f"[LOAD] done: {len(groups)} day-groups ready", flush=True)
    return groups, skipped_rows


def simulate_groups(params: "Params", groups: List[DayGroup]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Run the per-day simulation over pre-built groups for ONE parameter set.
    The per-day price book is built on the first trial and cached on the group,
    so every subsequent trial only does the (cheap) parameter-dependent work."""
    all_trades: List[Dict[str, Any]] = []
    sim_skips: List[Dict[str, Any]] = []
    for gr in groups:
        # Build the parameter-INDEPENDENT price book once, then reuse it forever.
        if gr.price_book is None:
            gr.idx_all = build_minute_index(gr.dy, SESSION_START_IST, SESSION_END_IST)
            gr.price_book, gr.symbols = build_price_book(gr.day_opt, gr.idx_all)
        trades, skips = simulate_day_multi_trades(
            und=gr.und, dy=gr.dy, expiry=gr.expiry,
            day_opt=gr.day_opt, underlying_day=gr.underlying_day, params=params,
            price_book=gr.price_book, symbols=gr.symbols, idx_all=gr.idx_all,
        )
        all_trades.extend([t.__dict__ for t in trades])
        sim_skips.extend(skips)

    all_df = pd.DataFrame(all_trades)
    if not all_df.empty:
        all_df = all_df.sort_values(["day", "underlying", "trade_seq"]).reset_index(drop=True)
    return all_df, pd.DataFrame(sim_skips)


def process_pickles_generate_trades(
    params: "Params",
    pickle_paths: List[str],
    min_expiry_map: Dict[Tuple[str, date], date],
    underlying_data: Dict[str, pd.DataFrame],
    window_start: date,
    window_end: date,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Single-run path used by main(): build groups, then simulate once."""
    groups, parse_skips = build_day_groups(pickle_paths, min_expiry_map, underlying_data, window_start, window_end)
    all_df, sim_skips = simulate_groups(params, groups)

    skip_df = pd.concat([pd.DataFrame(parse_skips), sim_skips], ignore_index=True) \
        if (parse_skips or not sim_skips.empty) else pd.DataFrame()
    if not skip_df.empty:
        if "day" not in skip_df.columns:
            skip_df["day"] = pd.NaT
        if "underlying" not in skip_df.columns:
            skip_df["underlying"] = pd.NA
        skip_df = skip_df.sort_values(["day", "underlying"], na_position="last").reset_index(drop=True)

    return all_df, skip_df


# =============================================================================
# Actual trades: one underlying per day (nearest expiry), include all re-entries for that underlying/day
# =============================================================================
def pick_actual_underlying_by_day(min_expiry_map: Dict[Tuple[str, date], date]) -> Dict[date, str]:
    by_day: Dict[date, List[Tuple[date, str]]] = {}
    for (und, dy), ex in min_expiry_map.items():
        if und not in TRADEABLE:
            continue

        dte = int((ex - dy).days)
        if dte not in ALLOWED_DTE:
            continue

        by_day.setdefault(dy, []).append((ex, und))

    out: Dict[date, str] = {}
    for dy, lst in by_day.items():
        # nearest expiry first; if tied, prefer NIFTY
        lst_sorted = sorted(lst, key=lambda t: (t[0], 0 if t[1] == "NIFTY" else 1))
        out[dy] = lst_sorted[0][1]
    return out

def build_actual_trades_df(all_trades_df: pd.DataFrame, min_expiry_map: Dict[Tuple[str, date], date]) -> pd.DataFrame:
    if all_trades_df.empty:
        return pd.DataFrame()

    actual_underlying = pick_actual_underlying_by_day(min_expiry_map)

    m = all_trades_df.copy()
    m["actual_underlying_for_day"] = m["day"].map(actual_underlying)

    # keep only days for which a 0/1-DTE actual underlying exists
    m = m[m["actual_underlying_for_day"].notna()]

    # keep only the selected underlying for that day
    m = m[m["underlying"] == m["actual_underlying_for_day"]]

    # keep only 0- and 1-DTE rows
    # keep only 0- and 1-DTE rows
    m = m[m["days_to_expiry"].isin(ALLOWED_DTE)]

    # keep all reattempts for the one selected underlying on that day
    m = m.drop(columns=["actual_underlying_for_day"])
    m = m.sort_values(["day", "trade_seq"]).reset_index(drop=True)

    # 1 if net exit PnL is positive, else 0
    m["is_exit_pnl_positive"] = (m["exit_pnl"] > 0).astype(int)

    return m


# =============================================================================
# Excel output
# =============================================================================
def _autosize_columns_safe(ws) -> None:
    # Safe autosize even when the sheet is "empty-ish"
    try:
        max_col = ws.max_column or 0
        if max_col <= 0:
            return
        for col_idx in range(1, max_col + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = 0
            for row_idx in range(1, min(ws.max_row or 1, 2000) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
    except Exception:
        # Never fail the whole run just because autosize misbehaved
        return

def _color_actual_trades_by_date(wb, actual_trades_df) -> None:
    """Shade rows so all attempts on the same calendar date share one colour,
    alternating between two soft fills as the date changes (visual grouping)."""
    if actual_trades_df is None or actual_trades_df.empty:
        return
    if "actual_trades" not in wb.sheetnames:
        return
    cols = list(actual_trades_df.columns)
    if "day" not in cols:
        return
    from openpyxl.styles import PatternFill
    ws = wb["actual_trades"]
    ncols = len(cols)
    fills = [
        PatternFill(fill_type="solid", fgColor="E8F0FE"),  # light blue
        PatternFill(fill_type="solid", fgColor="FFF3E0"),  # light amber
    ]
    days = actual_trades_df["day"].tolist()
    color_idx = 0
    prev_day = None
    first = True
    for i, d in enumerate(days):
        if first:
            first = False
        elif d != prev_day:
            color_idx ^= 1
        prev_day = d
        fill = fills[color_idx]
        excel_row = i + 2  # header occupies row 1
        for c in range(1, ncols + 1):
            ws.cell(row=excel_row, column=c).fill = fill


def write_excel(all_trades_df: pd.DataFrame, actual_trades_df: pd.DataFrame, skipped_df: pd.DataFrame) -> None:
    out_dir = os.path.dirname(os.path.abspath(OUTPUT_XLSX))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    piv_exit = pd.DataFrame()
    piv_eod_first = pd.DataFrame()
    monthwise_summary = pd.DataFrame()
    if not all_trades_df.empty:
        piv_exit = all_trades_df.pivot_table(index="day", columns="underlying", values="exit_pnl", aggfunc="sum").reset_index()

        first = all_trades_df[all_trades_df["trade_seq"] == 1]
        piv_eod_first = first.pivot_table(index="day", columns="underlying", values="eod_pnl", aggfunc="sum").reset_index()

        inst = all_trades_df.copy()
        inst["is_win_exit"] = inst["exit_pnl"] > 0
        inst["is_stoploss"] = inst["exit_reason"].astype(str).str.upper().eq("STOPLOSS")
        inst["is_profit_protect"] = inst["exit_reason"].astype(str).str.upper().eq("PROFIT_PROTECT")
        instrument_summary = (
            inst.groupby("underlying", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                win_rate_exit_pct=("is_win_exit", lambda s: 100.0 * s.mean()),
                stoploss_rate_pct=("is_stoploss", lambda s: 100.0 * s.mean()),
                profit_protect_rate_pct=("is_profit_protect", lambda s: 100.0 * s.mean()),
                avg_max_profit=("max_profit", "mean"),
                avg_max_loss=("max_loss", "mean"),
                worst_max_loss=("max_loss", "min"),
            )
            .sort_values("total_exit_pnl", ascending=False)
            .reset_index(drop=True)
        )
    else:
        instrument_summary = pd.DataFrame()

    if not actual_trades_df.empty:
        tmp = actual_trades_df.copy()
        tmp["month"] = pd.to_datetime(tmp["day"]).dt.to_period("M").astype(str)

        # Existing trade-level monthly summary
        monthwise_summary = (
            tmp.groupby("month", as_index=False)
            .agg(
                trades=("exit_pnl", "count"),
                total_exit_pnl=("exit_pnl", "sum"),
                avg_exit_pnl=("exit_pnl", "mean"),
                winning_trades=("is_exit_pnl_positive", "sum"),
            )
        )
        monthwise_summary["losing_trades"] = monthwise_summary["trades"] - monthwise_summary["winning_trades"]
        monthwise_summary["win_rate_pct"] = (
                100.0 * monthwise_summary["winning_trades"] / monthwise_summary["trades"]
        ).round(2)

        # New: daily PnL inside each month
        daily_tmp = (
            tmp.groupby(["month", "day"], as_index=False)
            .agg(daily_pnl=("exit_pnl", "sum"))
        )

        loss_day_stats = (
            daily_tmp.groupby("month", as_index=False)
            .agg(
                avg_loss_on_loss_days=(
                    "daily_pnl",
                    lambda s: float(s[s < 0].mean()) if (s < 0).any() else 0.0
                ),
                max_loss_in_a_day=(
                    "daily_pnl",
                    lambda s: float(s.min()) if len(s) else 0.0
                ),
            )
        )

        # Date on which the worst (maximum-loss) day occurred, per month
        worst_rows = daily_tmp.loc[daily_tmp.groupby("month")["daily_pnl"].idxmin()]
        worst_day = worst_rows[["month", "day"]].rename(columns={"day": "max_loss_day_date"})

        monthwise_summary = monthwise_summary.merge(loss_day_stats, on="month", how="left")
        monthwise_summary = monthwise_summary.merge(worst_day, on="month", how="left")

        # Place the date column right after the max-loss value column
        _cols = list(monthwise_summary.columns)
        if "max_loss_day_date" in _cols and "max_loss_in_a_day" in _cols:
            _cols.remove("max_loss_day_date")
            _cols.insert(_cols.index("max_loss_in_a_day") + 1, "max_loss_day_date")
            monthwise_summary = monthwise_summary[_cols]
    else:
        monthwise_summary = pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as xw:
        all_trades_df.to_excel(xw, sheet_name="all_trades_backtested", index=False)
        actual_trades_df.to_excel(xw, sheet_name="actual_trades", index=False)
        monthwise_summary.to_excel(xw, sheet_name="monthwise_summary", index=False)
        piv_exit.to_excel(xw, sheet_name="exit_pnl_pivot", index=False)
        piv_eod_first.to_excel(xw, sheet_name="eod_pnl_first_trade_pivot", index=False)
        instrument_summary.to_excel(xw, sheet_name="instrument_summary", index=False)
        skipped_df.to_excel(xw, sheet_name="skipped", index=False)

        wb = xw.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _autosize_columns_safe(ws)

        _color_actual_trades_by_date(wb, actual_trades_df)

    print(f"[DONE] Excel written: {OUTPUT_XLSX}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")

    print(f"[INFO] Pickles found: {len(paths)}")

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)

    lookback_label = "AUTO/full pickle range" if LOOKBACK_MONTHS is None else f"{LOOKBACK_MONTHS} months cap"

    print(f"[INFO] Data day-range seen: {min_day_seen} -> {end_day}")
    print(f"[INFO] Backtest window: {window_start} -> {end_day} ({lookback_label})")
    print(f"[INFO] Stoploss %/attempt: {_fmt_pct_list(LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Per-attempt stop cap: Rs {_fmt_rupee_value(MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT)} | "
          f"Daily max loss: Rs {_fmt_rupee_value(MAX_DAILY_LOSS_RUPEES)} | "
          f"ProfitProtect trigger/giveback %: {_fmt_pct_value(PROFIT_PROTECT_TRIGGER_RUPEES)} | "
          f"Re-entry delay min/attempt: {REENTRY_DELAY_BY_ATTEMPT} | Allowed DTE: {ALLOWED_DTE}")
    print(f"[INFO] Day profit target: {PROFIT_TARGET_PCT:.0%} of premium (0 = disabled)")
    print(f"[INFO] Entry time: {ENTRY_TIME_IST} | Fresh-entry/re-entry cutoff: {EXIT_TIME_IST}")
    print(f"[INFO] Tradeables: {sorted(TRADEABLE)}")
    print(f"[INFO] Output: {OUTPUT_XLSX}")

    print("[STEP] Initializing Kite ...")
    kite = oUtils.intialize_kite_api()
    print("[OK] Kite ready.")

    underlying_data = download_underlyings(kite, window_start, end_day)

    all_trades_df, skipped_df = process_pickles_generate_trades(
        default_params(), paths, min_expiry_map, underlying_data, window_start, end_day
    )

    actual_trades_df = build_actual_trades_df(all_trades_df, min_expiry_map)

    write_excel(all_trades_df, actual_trades_df, skipped_df)

    if not all_trades_df.empty:
        print(all_trades_df.groupby("underlying")[["exit_pnl"]].describe())
    else:
        print("[WARN] No completed trades. Check 'skipped' sheet for reasons.")


# =============================================================================
# ROBUSTNESS OPTIMIZER  (Optuna TPE + optional walk-forward CV)
# =============================================================================
# Goal: maximize TOTAL NET PROFIT, and profit only. The score is simply the total
# net P/L of the traded book over the whole sample. Profitable-month/day ratios,
# worst month and worst day are computed and printed as diagnostics but do NOT
# affect the score. (See the honest caveat next to OPT_* config below.)
#
# Tunables exposed to the optimizer (the 7 requested), with the variable-length
# per-attempt lists parameterized compactly as (base, step) so the search space
# stays low-dimensional and the per-attempt schedule stays monotone/sensible:
#   ENTRY_TIME_IST                -> minutes-from-open (discrete grid)
#   EXIT_TIME_IST                 -> last fresh entry/re-entry cutoff (discrete grid)
#   LOSS_LIMIT_RUPEES_BY_ATTEMPT  -> sl_base_pct + n*sl_step_pct  (rising stops)
#   PROFIT_PROTECT_TRIGGER_RUPEES -> single pct (0 disables)
#   MAX_REATTEMPTS                -> int
#   PROFIT_TARGET_PCT             -> single pct (0 disables)
#   REENTRY_DELAY_BY_ATTEMPT      -> delay_base + n*delay_step (minutes)
# =============================================================================

# ---- OBJECTIVE: maximize TOTAL NET PROFIT, and profit only ----
# The optimizer maximizes the total net P/L of the actually-traded book over the
# whole sample -- nothing else. Profitable-month / profitable-day ratios and the
# worst month / worst day are still COMPUTED and PRINTED as diagnostics so you can
# see the risk a config takes, but they do NOT influence the score.
#
# Honest caveat: pure-profit maximization tends to pick the most aggressive config
# the data allows and will accept a brutal worst day/month if that path made the
# most money in THIS sample. Keep an eye on the printed worst_day / worst_month of
# the winner. (If you later want profit balanced against downside, that is what the
# Sortino + worst-day-guardrail variant does; this build is profit-only by request.)
OPT_CV_PENALTY = _parse_float_env("OPT_CV_PENALTY", 0.50)  # CV mode only: penalty * std of per-fold net profit

# ---- SEARCH SPACE (stage 1). Tuples are (low, high) or (low, high, step). ----
# Stage 2 shrinks this box around the stage-1 winner and refines entry/exit to a
# 1-minute grid. Bounds were WIDENED vs the previous build: sl_base to 0.60 and
# profit_protect to 0.95, so profitable aggressive configs are reachable.
OPT_SPACE = {
    "entry":          (0, 250, 5),        # minutes from 09:20 -> 09:20..13:30
    "exit":           (0, 370, 5),        # minutes from 09:20 -> 09:20..15:30 (last-entry cutoff)
    "sl_base":        (0.05, 0.60),       # first-attempt stop, fraction of premium
    "sl_step":        (0.00, 0.10),       # per-attempt stop increment
    "max_reattempts": (0, 10, 1),
    "profit_protect": (0.00, 0.95),       # 0 disables profit-protect
    "profit_target":  (0.20, 0.95),       # of premium; > ~1.0 is unreachable for a short straddle
    "delay_base":     (1, 20, 1),         # re-entry gap base (minutes)
    "delay_step":     (0, 10, 1),         # re-entry gap per-attempt increment
    "daily_loss":     (10000, 60000, 5000),  # MAX_DAILY_LOSS_RUPEES (if OPT_TUNE_DAILY_LOSS)
    "stop_cap":       (1500, 8000, 250),     # per-attempt rupee stop cap (if OPT_TUNE_STOP_CAP)
}
OPT_MIN_DAYS = int(_parse_float_env("OPT_MIN_DAYS", 30))   # data guard: need enough days
OPT_MIN_MONTHS = int(_parse_float_env("OPT_MIN_MONTHS", 3))
_OPT_DISQUALIFY = -1.0e12   # score for data-guard failures (far below any real rupee P/L)


# Stable column order for the per-trial results CSV.
_TRIAL_COLUMNS = [
    "stage", "run_index", "trial_number", "state", "score",
    "entry_time", "exit_time_ist", "max_reattempts", "profit_protect_pct", "profit_target_pct",
    "sl_base_pct", "sl_step_pct", "loss_limit_schedule",
    "reentry_delay_base_min", "reentry_delay_step_min", "reentry_delay_schedule",
    "max_daily_loss_rupees", "stop_cap_rupees",
    "net_pnl", "mean_month", "median_month", "worst_month",
    "prof_month_ratio", "prof_day_ratio", "n_months", "n_days",
    "elapsed_s",
]


def _trial_record(trial_, base, run_index: int, elapsed: float,
                  stage: str = "stage1", space: Optional[dict] = None) -> Dict[str, Any]:
    """Flatten one finished trial (suggested params + derived schedules + result
    metrics) into a single CSV row. `base` supplies the non-optimized fields so we
    reconstruct the full Params via the same mapping the objective used."""
    p = dict(trial_.params)
    bp = _params_from_trial(_FrozenTrialView(trial_), base, space=space)
    ua = trial_.user_attrs
    return {
        "stage": stage,                          # stage1 (broad) / stage2 (zoom)
        "run_index": run_index,                  # 1..n_trials within THIS run
        "trial_number": trial_.number,           # global index within the study
        "state": str(getattr(trial_, "state", "")),
        "score": trial_.value,
        "entry_time": bp.entry_time.strftime("%H:%M"),
        "exit_time_ist": bp.exit_time.strftime("%H:%M"),
        "max_reattempts": bp.max_reattempts,
        "profit_protect_pct": round(bp.profit_protect_pct, 6),
        "profit_target_pct": round(bp.profit_target_pct, 6),
        "sl_base_pct": p.get("sl_base_pct"),
        "sl_step_pct": p.get("sl_step_pct"),
        "loss_limit_schedule": ";".join(str(round(x, 4)) for x in bp.loss_limit_pct_by_attempt),
        "reentry_delay_base_min": p.get("reentry_delay_base_min"),
        "reentry_delay_step_min": p.get("reentry_delay_step_min"),
        "reentry_delay_schedule": ";".join(str(x) for x in bp.reentry_delay_by_attempt),
        "max_daily_loss_rupees": int(bp.max_daily_loss_rupees),
        "stop_cap_rupees": int(bp.max_loss_limit_cap_rupees),
        "net_pnl": round(float(ua.get("total_pnl", 0.0)), 2),
        "mean_month": round(float(ua.get("mean_month", 0.0)), 2),
        "median_month": round(float(ua.get("median_month", 0.0)), 2),
        "worst_month": round(float(ua.get("worst_month", 0.0)), 2),
        "prof_month_ratio": round(float(ua.get("prof_month_ratio", 0.0)), 4),
        "prof_day_ratio": round(float(ua.get("prof_day_ratio", 0.0)), 4),
        "n_months": int(ua.get("n_months", 0)),
        "n_days": int(ua.get("n_days", 0)),
        "elapsed_s": round(elapsed, 1),
    }


def _inr(x: float) -> str:
    """Format a rupee amount with Indian digit grouping, ASCII-safe for Windows
    consoles (e.g. 1234567 -> 'Rs.12,34,567'). Avoids the unicode rupee sign so
    it never throws UnicodeEncodeError in a cp1252 terminal."""
    try:
        n = int(round(float(x)))
    except (ValueError, TypeError):
        return "Rs.0"
    sign = "-" if n < 0 else ""
    s = str(abs(n))
    if len(s) <= 3:
        body = s
    else:
        last3, rest = s[-3:], s[:-3]
        parts = []
        while len(rest) > 2:               # group remaining digits in pairs (Indian style)
            parts.insert(0, rest[-2:]); rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        body = ",".join(parts) + "," + last3
    return f"Rs.{sign}{body}"


def robustness_metrics(actual_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Per-day and per-month profitability stats from an actual_trades frame.

    Returns the headline robustness ratios PLUS rupee P/L summaries used in the
    per-trial print: net P/L, mean monthly P/L, median monthly P/L, worst month,
    and the full month->P/L series (`monthly`) for the best-trial breakdown.
    """
    if actual_df is None or actual_df.empty:
        return {"n_days": 0, "n_months": 0, "prof_day_ratio": 0.0,
                "prof_month_ratio": 0.0, "total_pnl": 0.0, "mean_month": 0.0,
                "median_month": 0.0, "worst_day": 0.0, "worst_month": 0.0,
                "monthly": pd.Series(dtype="float64")}
    t = actual_df.copy()
    d = pd.to_datetime(t["day"])
    daily = t.groupby(d.dt.date)["exit_pnl"].sum()
    monthly = t.groupby(d.dt.to_period("M"))["exit_pnl"].sum().sort_index()
    return {
        "n_days": int(len(daily)),
        "n_months": int(len(monthly)),
        "prof_day_ratio": float((daily > 0).mean()),
        "prof_month_ratio": float((monthly > 0).mean()),
        "total_pnl": float(daily.sum()),          # NET P/L across the whole sample
        "mean_month": float(monthly.mean()),      # MEAN monthly P/L
        "median_month": float(monthly.median()),  # MEDIAN monthly P/L
        "worst_day": float(daily.min()),
        "worst_month": float(monthly.min()),
        "monthly": monthly,                       # full month -> P/L (for breakdown)
    }


def _score_from_metrics(m: Dict[str, Any]) -> float:
    """Objective = TOTAL NET PROFIT (rupees) of the traded book. Profit only.
    Thin-data configs are disqualified so they cannot win by accident."""
    if m["n_days"] < OPT_MIN_DAYS or m["n_months"] < OPT_MIN_MONTHS:
        return _OPT_DISQUALIFY
    return float(m["total_pnl"])


def _cv_score(actual_df: pd.DataFrame, folds: int) -> float:
    """
    Walk-forward variant of the profit objective: split the months into `folds`
    CONTIGUOUS blocks and score mean(per-fold NET PROFIT) - penalty*std, so a config
    must make money across regimes, not just one lucky stretch. With cv_folds=1 the
    plain total-profit objective is used instead. Still profit-based.
    """
    if actual_df is None or actual_df.empty:
        return _OPT_DISQUALIFY
    t = actual_df.copy()
    d = pd.to_datetime(t["day"])
    monthly = t.groupby(d.dt.to_period("M"))["exit_pnl"].sum().sort_index()
    if len(monthly) < max(OPT_MIN_MONTHS, folds):
        return _OPT_DISQUALIFY
    # global activity guard
    daily = t.groupby(d.dt.date)["exit_pnl"].sum()
    if len(daily) < OPT_MIN_DAYS:
        return _OPT_DISQUALIFY

    months = list(monthly.items())
    block_profits: List[float] = []
    n = len(months)
    for k in range(folds):
        lo = (k * n) // folds
        hi = ((k + 1) * n) // folds
        block = months[lo:hi]
        if not block:
            continue
        block_profits.append(float(sum(v for _, v in block)))   # NET PROFIT in this block
    if not block_profits:
        return _OPT_DISQUALIFY
    s = pd.Series(block_profits)
    # mean per-fold profit, lightly penalized for uneven profit across regimes
    return float(s.mean() - OPT_CV_PENALTY * s.std(ddof=0))


def _params_from_trial(trial, base: "Params", space: Optional[dict] = None) -> "Params":
    """Map an Optuna trial to a Params object. Lists are built from (base, step).

    `space` selects the search box: OPT_SPACE for the broad stage-1 search, or a
    narrowed/fine-grid box for the stage-2 zoom. Parameter NAMES are identical in
    both stages so trials from either study can be rebuilt with the same code.
    """
    sp = space or OPT_SPACE

    # ENTRY_TIME: minutes-from-09:20 on the grid given by the space.
    e_lo, e_hi, e_st = sp["entry"]
    entry_min = trial.suggest_int("entry_minute_from_0920", e_lo, e_hi, step=e_st)
    entry_total_min = 9 * 60 + 20 + entry_min
    eh, em = divmod(entry_total_min, 60)
    entry_time = dtime(eh, em)

    # EXIT_TIME_IST: last fresh-entry cutoff. If suggested before the entry time,
    # clamp up so the config stays valid (entry <= cutoff); clamp to 15:30 max.
    x_lo, x_hi, x_st = sp["exit"]
    exit_min = trial.suggest_int("exit_minute_from_0920", x_lo, x_hi, step=x_st)
    exit_total_min = max(entry_total_min, 9 * 60 + 20 + exit_min)
    exit_total_min = min(exit_total_min, 15 * 60 + 30)
    xh, xm = divmod(exit_total_min, 60)
    exit_time = dtime(xh, xm)

    # Rising per-attempt stop-loss schedule (base + i*step).
    sl_base = trial.suggest_float("sl_base_pct", *sp["sl_base"])
    sl_step = trial.suggest_float("sl_step_pct", *sp["sl_step"])

    mr_lo, mr_hi, _ = sp["max_reattempts"]
    max_reattempts = trial.suggest_int("max_reattempts", mr_lo, mr_hi)
    profit_protect = trial.suggest_float("profit_protect_pct", *sp["profit_protect"])
    profit_target = trial.suggest_float("profit_target_pct", *sp["profit_target"])

    # Rising per-attempt re-entry delay schedule (minutes).
    d_lo, d_hi, _ = sp["delay_base"]
    delay_base = trial.suggest_int("reentry_delay_base_min", d_lo, d_hi)
    s_lo, s_hi, _ = sp["delay_step"]
    delay_step = trial.suggest_int("reentry_delay_step_min", s_lo, s_hi)

    # NEW: the two risk gates, previously fixed, now searchable. Both directly cap
    # profit (a tight stop cap truncates winners' risk budget; a tight daily loss
    # ends good days early after a bad start), so exposing them widens the
    # reachable optimum. When the flags are off they stay at the module defaults.
    if OPT_TUNE_DAILY_LOSS:
        dl_lo, dl_hi, dl_st = sp["daily_loss"]
        max_daily_loss = float(trial.suggest_int("max_daily_loss_rupees", dl_lo, dl_hi, step=dl_st))
    else:
        max_daily_loss = base.max_daily_loss_rupees
    if OPT_TUNE_STOP_CAP:
        c_lo, c_hi, c_st = sp["stop_cap"]
        stop_cap = float(trial.suggest_int("stop_cap_rupees", c_lo, c_hi, step=c_st))
    else:
        stop_cap = base.max_loss_limit_cap_rupees

    n_slots = max_reattempts + 1
    sl_list = [round(min(0.95, sl_base + i * sl_step), 4) for i in range(n_slots)]
    delay_list = [int(delay_base + i * delay_step) for i in range(n_slots)]

    return Params(
        entry_time=entry_time,
        exit_time=exit_time,
        loss_limit_pct_by_attempt=sl_list,
        profit_protect_pct=float(profit_protect),
        max_reattempts=int(max_reattempts),
        profit_target_pct=float(profit_target),
        reentry_delay_by_attempt=delay_list,
        max_daily_loss_rupees=max_daily_loss,
        max_loss_limit_cap_rupees=stop_cap,
    )


def _snap(v, lo, hi, step=None):
    """Clip v into [lo, hi] and snap onto the step grid anchored at lo."""
    v = max(lo, min(hi, v))
    if step:
        v = lo + int(round((v - lo) / step)) * step
        v = max(lo, min(hi, v))
    return v


def _seed_trials_from_defaults(base: "Params", space: dict) -> List[Dict[str, Any]]:
    """Build enqueue-able warm-start trials from the CURRENT config (module
    defaults / property values) plus two curated variants. Starting TPE from the
    best-known region typically saves dozens of random trials."""
    e_lo, e_hi, e_st = space["entry"]
    x_lo, x_hi, x_st = space["exit"]

    entry_min = _snap((base.entry_time.hour * 60 + base.entry_time.minute) - (9 * 60 + 20),
                      e_lo, e_hi, e_st)
    exit_min = _snap((base.exit_time.hour * 60 + base.exit_time.minute) - (9 * 60 + 20),
                     x_lo, x_hi, x_st)

    sl = base.loss_limit_pct_by_attempt or [0.30]
    sl_base = _snap(float(sl[0]), *space["sl_base"])
    diffs = [sl[i + 1] - sl[i] for i in range(len(sl) - 1)]
    sl_step = _snap(float(sum(diffs) / len(diffs)) if diffs else 0.0, *space["sl_step"])

    dl = base.reentry_delay_by_attempt or [10]
    d_base = int(_snap(int(dl[0]), *space["delay_base"]))
    ddiffs = [dl[i + 1] - dl[i] for i in range(len(dl) - 1)]
    d_step = int(_snap(int(round(sum(ddiffs) / len(ddiffs))) if ddiffs else 0, *space["delay_step"]))

    seed = {
        "entry_minute_from_0920": int(entry_min),
        "exit_minute_from_0920": int(exit_min),
        "sl_base_pct": float(sl_base),
        "sl_step_pct": float(sl_step),
        "max_reattempts": int(_snap(base.max_reattempts, space["max_reattempts"][0], space["max_reattempts"][1], 1)),
        "profit_protect_pct": float(_snap(base.profit_protect_pct, *space["profit_protect"])),
        "profit_target_pct": float(_snap(base.profit_target_pct, *space["profit_target"])),
        "reentry_delay_base_min": d_base,
        "reentry_delay_step_min": d_step,
    }
    if OPT_TUNE_DAILY_LOSS:
        seed["max_daily_loss_rupees"] = int(_snap(base.max_daily_loss_rupees, *space["daily_loss"]))
    if OPT_TUNE_STOP_CAP:
        seed["stop_cap_rupees"] = int(_snap(base.max_loss_limit_cap_rupees, *space["stop_cap"]))

    # Variant A: same config with profit-protect OFF (lets winners run to target).
    va = dict(seed); va["profit_protect_pct"] = 0.0
    # Variant B: same config with the profit target maxed (ride the full move).
    vb = dict(seed); vb["profit_target_pct"] = float(space["profit_target"][1])
    return [seed, va, vb]


def _shrink_space(best_params: Dict[str, Any], shrink: float) -> dict:
    """Build the stage-2 'zoom' box: each numeric bound narrowed to +-shrink/2 of
    the ORIGINAL range around the stage-1 winner (clipped to the original box),
    with entry/exit refined to a 1-MINUTE grid so the second pass can find optima
    between the coarse 5-minute grid points."""
    sp = {}
    def _num(name, key, step_override=None, as_int=False):
        lo, hi = OPT_SPACE[key][0], OPT_SPACE[key][1]
        st = step_override if step_override is not None else (OPT_SPACE[key][2] if len(OPT_SPACE[key]) > 2 else None)
        center = best_params.get(name, (lo + hi) / 2)
        half = (hi - lo) * shrink / 2.0
        nlo, nhi = max(lo, center - half), min(hi, center + half)
        if as_int:
            nlo, nhi = int(round(nlo)), int(round(nhi))
            if st: # keep the grid consistent: snap bounds onto the step grid
                nlo = int(_snap(nlo, lo, hi, st)); nhi = int(_snap(nhi, lo, hi, st))
                if nhi <= nlo:
                    nhi = min(hi, nlo + st)
            elif nhi <= nlo:
                nhi = min(hi, nlo + 1)
            return (nlo, nhi, st or 1)
        if nhi <= nlo:
            nhi = min(hi, nlo + 1e-6)
        return (nlo, nhi)
    sp["entry"] = _num("entry_minute_from_0920", "entry", step_override=1, as_int=True)   # 1-min grid
    sp["exit"] = _num("exit_minute_from_0920", "exit", step_override=1, as_int=True)      # 1-min grid
    sp["sl_base"] = _num("sl_base_pct", "sl_base")
    sp["sl_step"] = _num("sl_step_pct", "sl_step")
    sp["max_reattempts"] = _num("max_reattempts", "max_reattempts", as_int=True)
    sp["profit_protect"] = _num("profit_protect_pct", "profit_protect")
    sp["profit_target"] = _num("profit_target_pct", "profit_target")
    sp["delay_base"] = _num("reentry_delay_base_min", "delay_base", as_int=True)
    sp["delay_step"] = _num("reentry_delay_step_min", "delay_step", as_int=True)
    sp["daily_loss"] = _num("max_daily_loss_rupees", "daily_loss", as_int=True)
    sp["stop_cap"] = _num("stop_cap_rupees", "stop_cap", as_int=True)
    return sp


def optimize(
    groups: List[DayGroup],
    min_expiry_map: Dict[Tuple[str, date], date],
    n_trials: int = 100,
    cv_folds: int = 1,
    seed: int = 42,
    progress_every: int = 5,   # retained for signature compatibility
):
    """
    Two-stage Optuna search over pre-built groups.

    STAGE 1 (broad): multivariate TPE over the full OPT_SPACE box, warm-started
        from the current config (OPT_ENQUEUE_SEEDS), with optional early pruning
        of clearly-losing configs (OPT_PRUNE) so more trials fit in the same time.
    STAGE 2 (zoom): a second study over a box NARROWED around the stage-1 winner
        (OPT_STAGE2_SHRINK of each range) with entry/exit refined to a 1-MINUTE
        grid, warm-started from the stage-1 winner. This hunts the extra profit
        between the coarse grid points. Skipped when OPT_STAGE2_TRIALS = 0.

    The final winner (best of both stages) gets a NEIGHBOURHOOD STABILITY SCAN:
    each parameter is nudged one grid-step down/up and the full-sample profit of
    every neighbour is printed, so you can see whether the optimum is a robust
    plateau (neighbours earn similar money) or a knife-edge spike (they don't).

    cv_folds <= 1 : score = total net profit on the full sample.
    cv_folds  > 1 : contiguous-block walk-forward score (mean - penalty*std).
    Every tested config from BOTH stages is appended (flushed) to one CSV.
    """
    import optuna
    import time as _time
    import csv as _csv
    import types as _types
    import datetime as _dt
    optuna.logging.set_verbosity(optuna.logging.WARNING)  # we do our own printing

    base = default_params()  # supplies the non-optimized fields when flags are off

    # Chronologically ordered groups (and chunks for pruning checkpoints).
    groups_sorted = sorted(groups, key=lambda g: (g.dy, g.und))
    use_prune = bool(OPT_PRUNE) and OPT_PRUNE_CHUNKS >= 2 and len(groups_sorted) >= OPT_PRUNE_CHUNKS * 5
    if use_prune:
        n = len(groups_sorted)
        chunks = [groups_sorted[(k * n) // OPT_PRUNE_CHUNKS:((k + 1) * n) // OPT_PRUNE_CHUNKS]
                  for k in range(OPT_PRUNE_CHUNKS)]
        chunks = [c for c in chunks if c]
    else:
        chunks = [groups_sorted]

    # ---- results file: every tested config from both stages goes here ----
    os.makedirs(OPT_OUTPUT_DIR, exist_ok=True)
    run_ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(OPT_OUTPUT_DIR, f"{OPT_STUDY_NAME}_{run_ts}_trials.csv")
    csv_file = open(csv_path, "w", newline="")
    csv_writer = _csv.DictWriter(csv_file, fieldnames=_TRIAL_COLUMNS)
    csv_writer.writeheader()
    csv_file.flush()
    print(f"[OPT] saving every tested config to: {csv_path}", flush=True)
    if use_prune:
        print(f"[OPT] early pruning ON: {len(chunks)} chronological checkpoints, "
              f"25th-percentile rule (clearly-losing configs are cut short)", flush=True)

    # ---- one evaluation of a Params object (optionally prunable) ----
    def _evaluate(params: "Params", trial=None):
        """Simulate all day-groups for `params`. When `trial` is given and pruning
        is enabled, simulate chunk-by-chunk (chronological), report the running
        cumulative net profit at each checkpoint, and let Optuna prune."""
        if trial is None or not use_prune:
            all_df, _ = simulate_groups(params, groups_sorted)
            return build_actual_trades_df(all_df, min_expiry_map)
        parts: List[pd.DataFrame] = []
        for i, ch in enumerate(chunks):
            df_i, _ = simulate_groups(params, ch)
            if not df_i.empty:
                parts.append(df_i)
            if i < len(chunks) - 1:              # no report needed after the last chunk
                partial = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
                pa = build_actual_trades_df(partial, min_expiry_map)
                cum = float(pa["exit_pnl"].sum()) if (pa is not None and not pa.empty) else 0.0
                trial.report(cum, step=i)
                if trial.should_prune():
                    raise optuna.TrialPruned()
        full = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
        if not full.empty:
            full = full.sort_values(["day", "underlying", "trade_seq"]).reset_index(drop=True)
        return build_actual_trades_df(full, min_expiry_map)

    # ---- shared objective factory (stage-specific search box) ----
    def _make_objective(space: dict):
        def objective(trial):
            params = _params_from_trial(trial, base, space=space)
            actual_df = _evaluate(params, trial)
            m = robustness_metrics(actual_df)
            for k in ("n_days", "n_months", "prof_day_ratio", "prof_month_ratio",
                      "total_pnl", "mean_month", "median_month", "worst_day", "worst_month"):
                trial.set_user_attr(k, m[k])
            trial.set_user_attr("monthly_pnl", {str(p): float(v) for p, v in m["monthly"].items()})
            if cv_folds and cv_folds > 1:
                return _cv_score(actual_df, cv_folds)
            return _score_from_metrics(m)
        return objective

    # ---- shared per-trial progress + CSV callback ----
    start = _time.time()
    run_counter = {"n": 0}

    def _make_progress(stage: str, space: dict, stage_trials: int):
        def _progress(study_, trial_):
            run_counter["n"] += 1
            n_done = run_counter["n"]
            ua = trial_.user_attrs
            elapsed = _time.time() - start
            try:
                best_val = study_.best_value
                best_num = study_.best_trial.number
            except Exception:
                best_val, best_num = float("nan"), -1
            try:
                csv_writer.writerow(_trial_record(trial_, base, n_done, elapsed,
                                                  stage=stage, space=space))
                csv_file.flush()
            except Exception as e:
                print(f"[OPT WARN] could not write trial {trial_.number} to CSV: {e}", flush=True)

            state = str(getattr(trial_, "state", ""))
            if "PRUNED" in state:
                print(f"[{stage} {n_done:>4}] PRUNED early (unpromising) | "
                      f"MAX PROFIT={_inr(best_val)} | {elapsed:5.0f}s", flush=True)
                return
            n_mo = int(ua.get("n_months", 0))
            pmr = float(ua.get("prof_month_ratio", 0.0))
            prof_mo = int(round(pmr * n_mo))
            print(
                f"[{stage} {n_done:>4}] profit={_inr(ua.get('total_pnl', 0))} "
                f"mean/mo={_inr(ua.get('mean_month', 0))} "
                f"prof_mo={pmr*100:3.0f}%({prof_mo}/{n_mo}) "
                f"prof_day={float(ua.get('prof_day_ratio', 0.0))*100:3.0f}% "
                f"worst_mo={_inr(ua.get('worst_month', 0))} "
                f"worst_day={_inr(ua.get('worst_day', 0))} "
                f"| MAX PROFIT={_inr(best_val)} | {elapsed:5.0f}s",
                flush=True,
            )
            if trial_.number == best_num:
                mp = ua.get("monthly_pnl", {})
                if mp:
                    cells = [f"{k}:{_inr(v)}" for k, v in sorted(mp.items())]
                    print("   >>> NEW BEST -- month-wise net P/L:", flush=True)
                    for i in range(0, len(cells), 4):
                        print("       " + "   ".join(cells[i:i + 4]), flush=True)
        return _progress

    def _make_pruner():
        if not use_prune:
            return optuna.pruners.NopPruner()
        # 25th percentile: only the clearly-worst quarter at each checkpoint is
        # cut; warmup skips the first checkpoint so front-loaded noise can't kill
        # a config on chunk 1 alone; startup trials always run to completion.
        return optuna.pruners.PercentilePruner(25.0, n_startup_trials=12, n_warmup_steps=1)

    # ---- optional SQLite persistence (stage-1 study only; resumable) ----
    storage = None
    if OPT_SAVE_DB:
        db_path = os.path.join(OPT_OUTPUT_DIR, f"{OPT_STUDY_NAME}.db")
        storage = f"sqlite:///{db_path}"
        print(f"[OPT] stage-1 study persisted (resumable) at: {db_path}", flush=True)

    # ================= STAGE 1: broad multivariate search =================
    # multivariate+group TPE models parameter INTERACTIONS (entry x stop x target
    # move together in this strategy), which plain TPE treats independently.
    sampler1 = optuna.samplers.TPESampler(seed=seed, multivariate=True, group=True,
                                          n_startup_trials=max(15, n_trials // 10))
    study1 = optuna.create_study(direction="maximize", sampler=sampler1,
                                 pruner=_make_pruner(), study_name=OPT_STUDY_NAME,
                                 storage=storage, load_if_exists=bool(storage))
    if OPT_ENQUEUE_SEEDS:
        for s in _seed_trials_from_defaults(base, OPT_SPACE):
            study1.enqueue_trial(s)
        print(f"[OPT] warm-start: enqueued current config + 2 variants as the first trials", flush=True)

    print(f"[OPT] STAGE 1: {n_trials} trials over {len(groups_sorted)} day-groups "
          f"(cv_folds={cv_folds}) ...", flush=True)
    try:
        study1.optimize(_make_objective(OPT_SPACE), n_trials=n_trials,
                        callbacks=[_make_progress("stage1", OPT_SPACE, n_trials)],
                        show_progress_bar=False)
    except KeyboardInterrupt:
        print("[OPT] interrupted -- continuing with the best found so far.", flush=True)

    best_study, best_space = study1, OPT_SPACE

    # ================= STAGE 2: zoom refinement around the winner =================
    if OPT_STAGE2_TRIALS and OPT_STAGE2_TRIALS > 0 and study1.best_trial is not None:
        space2 = _shrink_space(dict(study1.best_trial.params), OPT_STAGE2_SHRINK)
        print(f"[OPT] STAGE 2: zooming {OPT_STAGE2_TRIALS} trials into a "
              f"{OPT_STAGE2_SHRINK:.0%}-of-range box around the stage-1 winner "
              f"(entry/exit now on a 1-minute grid) ...", flush=True)
        sampler2 = optuna.samplers.TPESampler(seed=seed + 1, multivariate=True, group=True,
                                              n_startup_trials=max(10, OPT_STAGE2_TRIALS // 10))
        study2 = optuna.create_study(direction="maximize", sampler=sampler2,
                                     pruner=_make_pruner())
        # Warm-start stage 2 from the stage-1 winner (snapped into the new box).
        s2_seed = {}
        for k, v in study1.best_trial.params.items():
            s2_seed[k] = v
        study2.enqueue_trial(s2_seed)
        try:
            study2.optimize(_make_objective(space2), n_trials=OPT_STAGE2_TRIALS,
                            callbacks=[_make_progress("stage2", space2, OPT_STAGE2_TRIALS)],
                            show_progress_bar=False)
        except KeyboardInterrupt:
            print("[OPT] stage 2 interrupted -- using the best found so far.", flush=True)
        try:
            if study2.best_value > study1.best_value:
                best_study, best_space = study2, space2
                print(f"[OPT] stage-2 zoom IMPROVED the winner: "
                      f"{_inr(study1.best_value)} -> {_inr(study2.best_value)}", flush=True)
            else:
                print(f"[OPT] stage-2 zoom did not beat stage 1 "
                      f"({_inr(study2.best_value)} vs {_inr(study1.best_value)}); "
                      f"keeping the stage-1 winner.", flush=True)
        except Exception:
            pass

    # ---- close CSV; dump the full Optuna tables for completeness ----
    try:
        csv_file.close()
    except Exception:
        pass
    try:
        full_path = os.path.join(OPT_OUTPUT_DIR, f"{OPT_STUDY_NAME}_{run_ts}_full.csv")
        best_study.trials_dataframe().to_csv(full_path, index=False)
        print(f"[OPT] full Optuna trials table (winning stage) saved to: {full_path}", flush=True)
    except Exception as e:
        print(f"[OPT WARN] could not write full trials table: {e}", flush=True)
    print(f"[OPT] per-trial results saved to: {csv_path}", flush=True)

    best = best_study.best_trial
    bp = _params_from_trial(_FrozenTrialView(best), base, space=best_space)

    # ================= TOP-10 table (both stages combined) =================
    all_trials = [(t, "s1") for t in study1.trials]
    if best_study is not study1:
        all_trials += [(t, "s2") for t in best_study.trials]
    ranked = [(t, tag) for (t, tag) in all_trials
              if t.value is not None and t.value > _OPT_DISQUALIFY / 2]
    ranked.sort(key=lambda x: x[0].value, reverse=True)
    print("\n================ TOP 10 CONFIGS ================")
    for t, tag in ranked[:10]:
        ua = t.user_attrs
        print(f"  [{tag}#{t.number:>3}] score={_inr(t.value)} pnl={_inr(ua.get('total_pnl', 0))} "
              f"worst_day={_inr(ua.get('worst_day', 0))} prof_day={float(ua.get('prof_day_ratio', 0))*100:.0f}% "
              f"| {t.params}")

    # ================= NEIGHBOURHOOD STABILITY SCAN =================
    # Nudge each parameter one grid-step down/up from the winner and re-simulate.
    # A trustworthy optimum is a PLATEAU: neighbours should earn broadly similar
    # profit. If they collapse, the winner is a knife-edge fit to this sample.
    print("\n================ NEIGHBOURHOOD STABILITY (1 step each way) ================")
    _grid_step = {
        "entry_minute_from_0920": best_space["entry"][2],
        "exit_minute_from_0920": best_space["exit"][2],
        "sl_base_pct": (best_space["sl_base"][1] - best_space["sl_base"][0]) * 0.05,
        "sl_step_pct": (best_space["sl_step"][1] - best_space["sl_step"][0]) * 0.05,
        "max_reattempts": 1,
        "profit_protect_pct": (best_space["profit_protect"][1] - best_space["profit_protect"][0]) * 0.05,
        "profit_target_pct": (best_space["profit_target"][1] - best_space["profit_target"][0]) * 0.05,
        "reentry_delay_base_min": 1,
        "reentry_delay_step_min": 1,
        "max_daily_loss_rupees": best_space["daily_loss"][2] if OPT_TUNE_DAILY_LOSS else None,
        "stop_cap_rupees": best_space["stop_cap"][2] if OPT_TUNE_STOP_CAP else None,
    }
    _bounds = {
        "entry_minute_from_0920": best_space["entry"][:2],
        "exit_minute_from_0920": best_space["exit"][:2],
        "sl_base_pct": best_space["sl_base"][:2],
        "sl_step_pct": best_space["sl_step"][:2],
        "max_reattempts": best_space["max_reattempts"][:2],
        "profit_protect_pct": best_space["profit_protect"][:2],
        "profit_target_pct": best_space["profit_target"][:2],
        "reentry_delay_base_min": best_space["delay_base"][:2],
        "reentry_delay_step_min": best_space["delay_step"][:2],
        "max_daily_loss_rupees": best_space["daily_loss"][:2] if OPT_TUNE_DAILY_LOSS else None,
        "stop_cap_rupees": best_space["stop_cap"][:2] if OPT_TUNE_STOP_CAP else None,
    }
    neighbour_profits: List[float] = []
    for pname, pval in best.params.items():
        st = _grid_step.get(pname)
        bnd = _bounds.get(pname)
        if st is None or bnd is None:
            continue
        row = [f"  {pname:<28} best={pval}"]
        for sign in (-1, +1):
            q = dict(best.params)
            nv = pval + sign * st
            nv = max(bnd[0], min(bnd[1], nv))
            q[pname] = int(round(nv)) if isinstance(pval, int) else float(nv)
            shim = _types.SimpleNamespace(params=q)
            try:
                np_params = _params_from_trial(_FrozenTrialView(shim), base, space=best_space)
                a = _evaluate(np_params, None)
                prof = robustness_metrics(a)["total_pnl"]
                neighbour_profits.append(prof)
                row.append(f"{'-' if sign < 0 else '+'}1step={_inr(prof)}")
            except Exception as e:
                row.append(f"{'-' if sign < 0 else '+'}1step=ERR({e})")
        print("  ".join(row), flush=True)
    if neighbour_profits:
        s = pd.Series(neighbour_profits)
        best_pnl = float(best.user_attrs.get("total_pnl", 0.0))
        print(f"  neighbours: median={_inr(s.median())} min={_inr(s.min())} "
              f"vs winner={_inr(best_pnl)}")
        if best_pnl > 0 and s.median() < 0.6 * best_pnl:
            print("  [WARN] neighbours earn <60% of the winner's profit on median: this "
                  "optimum looks like a KNIFE-EDGE fit. Prefer a flatter region from the "
                  "TOP-10 table above.")

    # ================= FINAL REPORT =================
    print("\n================ BEST CONFIG ================")
    print(f"MAX PROFIT (objective) = {_inr(best.value)}  (cv_folds={cv_folds})")
    print(f"ENTRY_TIME_IST              = {bp.entry_time.strftime('%H:%M')}")
    print(f"EXIT_TIME_IST               = {bp.exit_time.strftime('%H:%M')}")
    print(f"LOSS_LIMIT_RUPEES_BY_ATTEMPT= {[round(x,4) for x in bp.loss_limit_pct_by_attempt]}")
    print(f"PROFIT_PROTECT_TRIGGER      = {bp.profit_protect_pct:.4f}")
    print(f"MAX_REATTEMPTS              = {bp.max_reattempts}")
    print(f"PROFIT_TARGET_PCT           = {bp.profit_target_pct:.4f}")
    print(f"REENTRY_DELAY_BY_ATTEMPT    = {bp.reentry_delay_by_attempt}")
    print(f"MAX_DAILY_LOSS_RUPEES       = {int(bp.max_daily_loss_rupees)}"
          + ("" if OPT_TUNE_DAILY_LOSS else "   (not tuned)"))
    print(f"MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT (cap) = {int(bp.max_loss_limit_cap_rupees)}"
          + ("" if OPT_TUNE_STOP_CAP else "   (not tuned)"))
    print("---- robustness of best (full sample) ----")
    ba = best.user_attrs
    print(f"  net P/L            = {_inr(ba.get('total_pnl', 0))}")
    print(f"  mean monthly P/L   = {_inr(ba.get('mean_month', 0))}")
    print(f"  median monthly P/L = {_inr(ba.get('median_month', 0))}")
    print(f"  worst month        = {_inr(ba.get('worst_month', 0))}")
    n_mo = int(ba.get('n_months', 0))
    pmr = float(ba.get('prof_month_ratio', 0.0))
    print(f"  profitable months  = {int(round(pmr*n_mo))}/{n_mo} ({pmr*100:.1f}%)")
    print(f"  profitable days    = {float(ba.get('prof_day_ratio', 0.0))*100:.1f}% of {int(ba.get('n_days', 0))} days")
    mp = ba.get("monthly_pnl", {})
    if mp:
        print("  month-wise net P/L:")
        for k, v in sorted(mp.items()):
            flag = "" if v > 0 else "   <-- loss"
            print(f"     {k}: {_inr(v)}{flag}")

    # Ready-to-paste config block for a confirmation backtest.
    print("\n---- paste into your single-run params to verify ----")
    print(f"ENTRY_TIME_IST = \"{bp.entry_time.strftime('%H:%M')}\"")
    print(f"EXIT_TIME_IST = \"{bp.exit_time.strftime('%H:%M')}\"")
    print(f"# LOSS_LIMIT schedule (per attempt): {[round(x,4) for x in bp.loss_limit_pct_by_attempt]}")
    print(f"# PROFIT_PROTECT_TRIGGER_RUPEES (pct): {bp.profit_protect_pct:.4f}")
    print(f"# MAX_REATTEMPTS: {bp.max_reattempts}")
    print(f"# PROFIT_TARGET_PCT: {bp.profit_target_pct:.4f}")
    print(f"# REENTRY_DELAY_BY_ATTEMPT (min): {bp.reentry_delay_by_attempt}")
    print(f"# MAX_DAILY_LOSS_RUPEES: {int(bp.max_daily_loss_rupees)}")
    print(f"# MAX_LOSS_LIMIT_RUPEES_BY_ATTEMPT (cap): {int(bp.max_loss_limit_cap_rupees)}")
    return best_study, bp


class _FrozenTrialView:
    """Lets us rebuild Params from a finished trial's params via suggest_* calls."""
    def __init__(self, trial):
        self._p = dict(trial.params)

    def suggest_int(self, name, *a, **k):
        return int(self._p[name])

    def suggest_float(self, name, *a, **k):
        return float(self._p[name])


def run_optimizer(n_trials: int, cv_folds: int,
                  max_pickles: Optional[int] = None,
                  max_days: Optional[int] = None,
                  progress_every: int = 5,
                  seed: int = 42):
    """
    End-to-end optimizer entrypoint: load data ONCE (pickles + Kite underlyings),
    cache the day-groups, then run the Optuna search over them.

    The phases are printed as banners so you always know where it is:
        [PHASE 1] scan pickles      [PHASE 2] download underlyings
        [PHASE 3] build day-groups  [PHASE 4] optimize
    """
    print("[PHASE 1] Scanning pickles for date range and nearest expiries ...", flush=True)
    paths = sorted(glob.glob(os.path.join(PICKLES_DIR, "*.pkl")) + glob.glob(os.path.join(PICKLES_DIR, "*.pickle")))
    if not paths:
        raise FileNotFoundError(f"No .pkl/.pickle files found in: {PICKLES_DIR}")
    if max_pickles:
        # scan only the same subset we will load, so the window matches the sample
        paths = paths[:max_pickles]
    print(f"[PHASE 1] {len(paths)} pickle file(s) in scope", flush=True)

    end_day, min_expiry_map, min_day_seen = scan_pickles_pass1(paths)
    window_start = determine_backtest_window_start(min_day_seen, end_day)
    print(f"[PHASE 1] window: {window_start} -> {end_day}", flush=True)

    print("[PHASE 2] Initializing Kite and downloading underlyings ...", flush=True)
    kite = oUtils.intialize_kite_api()
    underlying_data = download_underlyings(kite, window_start, end_day)

    print("[PHASE 3] Building (and caching) day-groups ...", flush=True)
    groups, _ = build_day_groups(paths, min_expiry_map, underlying_data,
                                 window_start, end_day,
                                 max_pickles=max_pickles, max_days=max_days)
    if not groups:
        raise RuntimeError("No day-groups built; nothing to optimize. Check window / pickles.")

    print(f"[PHASE 4] Optimizing: {n_trials} trials, cv_folds={cv_folds} ...", flush=True)
    return optimize(groups, min_expiry_map, n_trials=n_trials, cv_folds=cv_folds,
                    progress_every=progress_every, seed=seed)


# =============================================================================
# ENTRYPOINT
# =============================================================================
# =============================================================================
# ENTRYPOINT  -- behaviour is driven entirely by the RUN CONTROL block at the
# top of this file. Just press Run in PyCharm; no command-line arguments.
# =============================================================================
if __name__ == "__main__":
    if RUN_MODE == "optimize":
        run_optimizer(
            n_trials=OPT_TRIALS,
            cv_folds=OPT_CV_FOLDS,
            max_pickles=SAMPLE_MAX_PICKLES,
            max_days=SAMPLE_MAX_DAYS,
            progress_every=OPT_PROGRESS_EVERY,
            seed=OPT_SEED,
        )
    elif RUN_MODE == "backtest":
        main()
    else:
        raise SystemExit(f"Unknown RUN_MODE={RUN_MODE!r}; set it to 'backtest' or 'optimize'.")