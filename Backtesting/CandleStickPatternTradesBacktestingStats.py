from openpyxl import Workbook


def print_statistics(pattern_recognition_results, market_previous_trend, file, correct_cs_pattern_excel_location):
    no_of_correct_patterns = 0
    no_of_tradable_patterns = 0
    no_of_strong_correct_trend = 0
    no_of_weak_correct_trend = 0
    no_of_high_volumes = 0
    no_of_correct_rsi = 0
    no_of_correct_rsi_14_9_period_sma = 0
    no_of_correct_resistance = 0
    no_of_correct_support = 0
    no_of_correct_risk_reward_ratio = 0
    no_of_trend_same_as_market = 0
    no_of_correct_candle_length = 0

    wb = Workbook ()

    # grab the active worksheet
    ws = wb.active

    start_char_ascii = 65
    ws[chr(start_char_ascii)+'2'] = 'STOCK'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'PATTERN'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'BGHT ON'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'SL/TG ON'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'TG(Y/N)'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'SL(Y/N)'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'BUY PRICE'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'EXIT PRICE'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'TG'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'SL'
    start_char_ascii += 1
    ws[chr(start_char_ascii) + '2'] = 'EXIT PROFIT'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'TRADABLE'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'PATTERN MATCH'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'CORRECT TREND'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'HIGH VOLUME'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'CORRECT RSI'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'CORRECT RSI SMA'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'CORRECT RES/SUP'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'SAME AS MARKET TREND'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'CORRECT RISK REWARD RATIO'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'CANDLE LENGTH OK'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'POINTS'
    start_char_ascii += 1
    ws[chr(start_char_ascii)+'2'] = 'RISK REWARD RATIO'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'SUPPORT'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'RESISTANCE'
    start_char_ascii += 1
    ws[chr (start_char_ascii) + '2'] = 'VOL'
    start_char_ascii += 1

    second_start_char_ascii = 65
    start_char_ascii = 65
    ws[chr (second_start_char_ascii) + chr (start_char_ascii) + '2'] = 'AV.VOL'
    start_char_ascii += 1
    ws[chr (second_start_char_ascii) + chr (start_char_ascii) + '2'] = 'RSI'
    start_char_ascii += 1
    ws[chr (second_start_char_ascii) + chr (start_char_ascii) + '2'] = 'RSI PAT OK'
    start_char_ascii += 1

    row_count = 3
    for pattern_recognition_result in pattern_recognition_results:

        start_char_ascii = 65
        stocks_pattern_recognition_response = pattern_recognition_result['res']

        if stocks_pattern_recognition_response.pattern_match:

            ws[chr(start_char_ascii)+str(row_count)] = stocks_pattern_recognition_response.stock_id
            start_char_ascii += 1
            ws[chr(start_char_ascii)+str(row_count)] = str(stocks_pattern_recognition_response.pattern_name.name)
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['buy_on']
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['trigged_on']
            start_char_ascii += 1
            ws[chr(start_char_ascii) + str (row_count)] = pattern_recognition_result['tg_trigged']
            start_char_ascii += 1
            ws[chr(start_char_ascii) + str (row_count)] = pattern_recognition_result['sl_trigged']
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['buy_price']
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['exit_price']
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['target']
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['stoploss']
            start_char_ascii += 1
            ws[chr (start_char_ascii) + str (row_count)] = pattern_recognition_result['earning']
            start_char_ascii += 1

            if stocks_pattern_recognition_response.is_pattern_tradable():
                no_of_tradable_patterns += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0

            start_char_ascii += 1

            no_of_correct_patterns += 1
            ws[chr(start_char_ascii) + str (row_count)] = 1

            start_char_ascii += 1

            if stocks_pattern_recognition_response.strong_correct_trend :
                no_of_strong_correct_trend += 1

            if stocks_pattern_recognition_response.weak_correct_trend :
                no_of_weak_correct_trend += 1

            if stocks_pattern_recognition_response.previous_trend is None or stocks_pattern_recognition_response.strong_correct_trend or stocks_pattern_recognition_response.weak_correct_trend:
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0

            start_char_ascii += 1

            if stocks_pattern_recognition_response.high_volumes :
                no_of_high_volumes += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0

            start_char_ascii += 1

            if stocks_pattern_recognition_response.correct_rsi :
                no_of_correct_rsi += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0

            start_char_ascii += 1

            if stocks_pattern_recognition_response.correct_rsi_14_9_period_SMA :
                no_of_correct_rsi_14_9_period_sma += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0

            start_char_ascii += 1

            if stocks_pattern_recognition_response.correct_resistance:
                no_of_correct_resistance += 1

            if stocks_pattern_recognition_response.correct_support:
                no_of_correct_support += 1

            if stocks_pattern_recognition_response.correct_resistance or stocks_pattern_recognition_response.correct_support:
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0

            start_char_ascii += 1

            if stocks_pattern_recognition_response.pattern_trend_same_as_market_trend:
                no_of_trend_same_as_market += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii) + str (row_count)] = 0
            start_char_ascii += 1

            if stocks_pattern_recognition_response.correct_risk_reward_ratio :
                no_of_correct_risk_reward_ratio += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii)+ str (row_count)] = 0
            start_char_ascii += 1

            if stocks_pattern_recognition_response.correct_candle_length :
                no_of_correct_candle_length += 1
                ws[chr(start_char_ascii) + str (row_count)] = 1
            else:
                ws[chr(start_char_ascii)+ str (row_count)] = 0
            start_char_ascii += 1

            ws[chr(start_char_ascii) + str (row_count)] = stocks_pattern_recognition_response.points
            start_char_ascii += 1

            ws[chr(start_char_ascii) + str (row_count)] = stocks_pattern_recognition_response.risk_reward_ratio
            start_char_ascii += 1

            ws[chr (start_char_ascii) + str (row_count)] = stocks_pattern_recognition_response.support
            start_char_ascii += 1

            ws[chr (start_char_ascii) + str (row_count)] = stocks_pattern_recognition_response.resistance
            start_char_ascii += 1

            ws[chr (start_char_ascii) + str (row_count)] = stocks_pattern_recognition_response.current_day_volumes
            start_char_ascii += 1

            second_start_char_ascii = 65
            start_char_ascii = 65

            ws[chr (second_start_char_ascii) + chr (start_char_ascii) + str (
                row_count)] = stocks_pattern_recognition_response.last_10_day_average_volumes
            start_char_ascii += 1

            ws[chr (second_start_char_ascii) + chr (start_char_ascii) + str (
                row_count)] = stocks_pattern_recognition_response.rsi
            start_char_ascii += 1

            ws[chr (second_start_char_ascii) + chr (start_char_ascii) + str (
                row_count)] = pattern_recognition_result['rsi_pat_ok']
            start_char_ascii += 1

            row_count += 1

    wb.save (correct_cs_pattern_excel_location)

    file.write('\nStats start @#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$')
    file.write('\nNo of patterns tested:'+str(len(pattern_recognition_results)))
    file.write('\nNo of tradable patterns:'+str(no_of_tradable_patterns))
    file.write('\nNo of correct patterns:'+str(no_of_correct_patterns))
    file.write('\nNo of strong correct trend:'+str(no_of_strong_correct_trend))
    file.write('\nNo of weak correct trend:'+str(no_of_weak_correct_trend))
    file.write('\nNo of high volumes:'+str(no_of_high_volumes))
    file.write('\nNo of correct rsi:'+str(no_of_correct_rsi))
    file.write('\nNo of correct rsi 14_9_period_sma:'+str(no_of_correct_rsi_14_9_period_sma))
    file.write('\nNo of correct resistance:'+str(no_of_correct_resistance))
    file.write('\nNo of correct support:'+str(no_of_correct_support))
    file.write('\nNo of trend same as market:'+str(no_of_trend_same_as_market))
    file.write('\nNo of correct risk reward ratio:'+str(no_of_correct_risk_reward_ratio))
    file.write ('\nNo of correct candle length:' + str (no_of_correct_candle_length))
    file.write('\nStats end @#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$@#$\n')


