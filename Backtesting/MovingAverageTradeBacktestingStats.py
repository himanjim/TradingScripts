from openpyxl import Workbook

def print_statistics(tradable_moving_average_strategy_responses, ma_strategy_excel_location):

    wb = Workbook ()

    # grab the active worksheet
    ws = wb.active

    ws['B2'] = 'STOCK'
    ws['C2'] = 'STRATEGY'
    ws['D2'] = 'BUY DATE'
    ws['E2'] = 'SELL DATE'
    ws['F2'] = 'BPRICE'
    ws['G2'] = 'SPRICE'
    ws['H2'] = 'PROFIT'
    ws['I2'] = 'PREV SMA LOW'
    ws['J2'] = 'WHEN SMA LOW'
    ws['K2'] = 'MACD SLOPE'
    ws['L2'] = 'SUCCESS'
    ws['M2'] = 'BOUGHT MAS'
    ws['N2'] = 'SOLD MAS'

    row_count = 3
    for tradable_moving_average_strategy_response in tradable_moving_average_strategy_responses:

        ws['B' + str(row_count)] = tradable_moving_average_strategy_response['stock_id']
        ws['C' + str(row_count)] = tradable_moving_average_strategy_response['strategy']
        ws['D' + str(row_count)] = tradable_moving_average_strategy_response['bought_when']
        ws['E' + str (row_count)] = tradable_moving_average_strategy_response['sold_when']
        ws['F' + str (row_count)] = tradable_moving_average_strategy_response['bought_at']
        ws['G' + str (row_count)] = tradable_moving_average_strategy_response['sold_at']
        ws['H' + str (row_count)] = tradable_moving_average_strategy_response['profit']
        ws['I' + str (row_count)] = tradable_moving_average_strategy_response['prev_sma_low']
        ws['J' + str (row_count)] = tradable_moving_average_strategy_response['when_sma_last_low']
        ws['K' + str (row_count)] = tradable_moving_average_strategy_response['macd_slope']
        ws['L' + str (row_count)] = tradable_moving_average_strategy_response['success']
        ws['M' + str (row_count)] = str (tradable_moving_average_strategy_response['bought_smas']) + str (
            tradable_moving_average_strategy_response['bought_lmas'])
        ws['N' + str (row_count)] = str (tradable_moving_average_strategy_response['sold_mas'])


        row_count += 1

    wb.save (ma_strategy_excel_location)



